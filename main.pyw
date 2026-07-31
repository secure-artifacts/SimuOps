import sys
import os
import json
import ctypes
from ctypes import wintypes
import math
import time
import csv
import re
import struct
import shutil
import threading
import uuid
import traceback
import pyautogui
import pyperclip
import pygetwindow as gw
import subprocess
import glob
from io import StringIO
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from pynput import mouse
    HAS_PYNPUT = True
except ImportError:
    mouse = None
    HAS_PYNPUT = False

try:
    import keyboard
    HAS_KEYBOARD = True
except ImportError:
    HAS_KEYBOARD = False
from datetime import datetime, timedelta
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QLineEdit, QTableWidget, QTableWidgetItem,
    QHeaderView, QMessageBox, QGroupBox, QSpinBox, QComboBox, QTextEdit,
    QProgressBar, QFileDialog, QTabWidget, QMenu, QInputDialog, QDialog, QListWidget,
    QSplitter, QTimeEdit, QCalendarWidget, QRadioButton, QButtonGroup, QDateTimeEdit,
    QAbstractItemView, QCheckBox, QListWidgetItem, QFormLayout, QTreeWidget, QTreeWidgetItem,
    QTreeWidgetItemIterator, QStackedWidget, QGridLayout, QStyle, QFrame, QSizePolicy,
    QStyledItemDelegate, QSlider, QScrollArea
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QTimer, QPoint, QSize, QDate, QTime, QDateTime, QRect, QItemSelectionModel
from PyQt5.QtGui import QColor, QPainter, QPen, QPixmap, QCursor, QIcon, QFont, QBrush, QDrag, QTextOption
from PyQt5.QtWidgets import QRubberBand

# --- High DPI Scaling Fix ---
# [修复] 顶部不再重复设置 DPI 属性，统一在 main() 入口处设置，避免前后矛盾导致字体模糊/重影

# --- PyAutoGUI Exception Compatibility ---
# Ensure PyAutoGUI returns None instead of raising ImageNotFoundException for backward compatibility
try:
    pyautogui.raisePyautoguiException = False
    pyautogui.FAILSAFE = False  # [新增] 禁用安全保护机制，防止鼠标移到角落导致程序异常退出
except Exception as e:
    print(f"[Startup] Failed to initialize pyautogui compatibility: {e}", file=sys.stderr, flush=True)

# --- Fix for Qt Platform Plugin ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
def fix_qt_plugin_path():
    venv_site_packages = os.path.join(BASE_DIR, "venv_simuops", "Lib", "site-packages")
    if not os.path.exists(venv_site_packages):
        for p in sys.path:
            if "site-packages" in p: venv_site_packages = p; break
    qt_plugin_path = os.path.join(venv_site_packages, "PyQt5", "Qt5", "plugins", "platforms")
    if os.path.exists(qt_plugin_path): os.environ["QT_QPA_PLATFORM_PLUGIN_PATH"] = os.path.join(venv_site_packages, "PyQt5", "Qt5", "plugins")

fix_qt_plugin_path()

DRAG_DEBUG_ENABLED = False
DRAG_DEBUG_SESSION_ID = datetime.now().strftime("%Y%m%d_%H%M%S")
DRAG_DEBUG_LOG_DIR = os.path.join(BASE_DIR, "logs")
DRAG_DEBUG_LOG_FILE = os.path.join(DRAG_DEBUG_LOG_DIR, f"drag_debug_{DRAG_DEBUG_SESSION_ID}.log")
DRAG_DEBUG_LATEST_FILE = os.path.join(DRAG_DEBUG_LOG_DIR, "drag_debug_latest.log")

def log_internal_issue(context, exc=None):
    """记录内部异常，避免关键路径静默失败。"""
    detail = context
    if exc is not None:
        detail = f"{context}: {type(exc).__name__}: {exc}"
    try:
        if DRAG_DEBUG_ENABLED:
            write_drag_debug(f"[内部诊断] {detail}")
        else:
            print(f"[内部诊断] {detail}", file=sys.stderr, flush=True)
    except Exception:
        try:
            print(f"[内部诊断] {detail}", file=sys.stderr, flush=True)
        except Exception:
            pass

def write_drag_debug(message):
    if not DRAG_DEBUG_ENABLED:
        return
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
    line = f"[{timestamp}] {message}"
    try:
        os.makedirs(DRAG_DEBUG_LOG_DIR, exist_ok=True)
        with open(DRAG_DEBUG_LOG_FILE, "a", encoding="utf-8") as f:
            f.write(line + "\n")
        with open(DRAG_DEBUG_LATEST_FILE, "a", encoding="utf-8") as f:
            f.write(line + "\n")
    except Exception:
        pass
    try:
        print(line, flush=True)
    except Exception:
        pass

def init_drag_debug_log():
    if not DRAG_DEBUG_ENABLED:
        return
    try:
        os.makedirs(DRAG_DEBUG_LOG_DIR, exist_ok=True)
        with open(DRAG_DEBUG_LATEST_FILE, "w", encoding="utf-8") as f:
            f.write("")
    except Exception as e:
        log_internal_issue("初始化拖拽调试日志失败", e)
    write_drag_debug("===== 拖拽调试会话启动 =====")
    write_drag_debug(f"BASE_DIR={BASE_DIR}")
    write_drag_debug(f"DEBUG_LOG_FILE={DRAG_DEBUG_LOG_FILE}")

def _load_pywin32_drag_modules():
    """按需加载 Windows 原生拖拽所需模块。"""
    if sys.platform != "win32":
        raise RuntimeError("真实拖拽仅支持 Windows")
    try:
        import pythoncom
        import win32con
        import winerror
        from win32com.server.util import NewEnum, wrap
        from win32com.server.exception import COMException
        from win32com.shell import shellcon
        return {
            "pythoncom": pythoncom,
            "win32con": win32con,
            "winerror": winerror,
            "NewEnum": NewEnum,
            "wrap": wrap,
            "COMException": COMException,
            "shellcon": shellcon,
        }
    except ImportError as e:
        raise ImportError("缺少 pywin32，真实拖拽不可用") from e

def _pack_dropfiles_structure(file_paths):
    """构造 CF_HDROP 所需的 DROPFILES 二进制结构。"""
    paths = [os.path.normpath(os.path.abspath(p)) for p in file_paths if p]
    if not paths:
        raise ValueError("没有可拖拽的文件")
    file_name_buffer = "\0".join(paths) + "\0\0"
    encoded = file_name_buffer.encode("utf-16le")
    header = struct.pack("IiiII", 20, 0, 0, 0, 1)  # pFiles, pt.x, pt.y, fNC, fWide
    return header + encoded

def perform_native_file_drag(file_paths, target_x, target_y, log_func=None):
    """执行更接近资源管理器行为的 OLE 文件拖拽。"""
    mods = _load_pywin32_drag_modules()
    pythoncom = mods["pythoncom"]
    win32con = mods["win32con"]
    winerror = mods["winerror"]
    NewEnum = mods["NewEnum"]
    wrap = mods["wrap"]
    COMException = mods["COMException"]
    shellcon = mods["shellcon"]

    file_paths = [os.path.normpath(os.path.abspath(p)) for p in file_paths if p]
    missing = [p for p in file_paths if not os.path.exists(p)]
    if missing:
        raise FileNotFoundError(f"文件不存在: {missing[0]}")
    if not file_paths:
        raise ValueError("没有可拖拽的文件")

    ole_inited = False
    try:
        pythoncom.OleInitialize()
        ole_inited = True
    except Exception:
        pythoncom.CoInitialize()

    def _drag_log(message):
        write_drag_debug(f"[NativeDrag] {message}")
        if log_func:
            try:
                log_func(message)
            except Exception:
                pass

    class IDropSource:
        _com_interfaces_ = [pythoncom.IID_IDropSource]
        _public_methods_ = ["QueryContinueDrag", "GiveFeedback"]

        def QueryContinueDrag(self, is_escape_pressed, modifier_key_state):
            if is_escape_pressed:
                return winerror.DRAGDROP_S_CANCEL
            if (modifier_key_state & win32con.MK_LBUTTON) == 0:
                return winerror.DRAGDROP_S_DROP
            return winerror.S_OK

        def GiveFeedback(self, effect):
            return winerror.DRAGDROP_S_USEDEFAULTCURSORS

    class IDataObject:
        _com_interfaces_ = [pythoncom.IID_IDataObject]
        _public_methods_ = [
            "GetData", "GetDataHere", "QueryGetData", "GetCanonicalFormatEtc",
            "SetData", "EnumFormatEtc", "DAdvise", "DUnadvise", "EnumDAdvise"
        ]

        def __init__(self, paths):
            self._paths = paths
            self._formatetc = [
                (win32con.CF_HDROP, None, pythoncom.DVASPECT_CONTENT, -1, pythoncom.TYMED_HGLOBAL)
            ]

        def _query_interface_(self, iid):
            if iid == pythoncom.IID_IEnumFORMATETC:
                return NewEnum(self._formatetc, iid=iid)

        def GetData(self, formatetc_in):
            cf_in, _target, aspect_in, _index, tymed_in = formatetc_in
            if (
                cf_in == win32con.CF_HDROP and
                (aspect_in & pythoncom.DVASPECT_CONTENT) and
                (tymed_in & pythoncom.TYMED_HGLOBAL)
            ):
                medium = pythoncom.STGMEDIUM()
                medium.set(pythoncom.TYMED_HGLOBAL, _pack_dropfiles_structure(self._paths))
                return medium
            raise COMException(hresult=winerror.DV_E_FORMATETC)

        def QueryGetData(self, formatetc_in):
            cf_in, _target, aspect_in, _index, tymed_in = formatetc_in
            if (
                cf_in == win32con.CF_HDROP and
                (aspect_in & pythoncom.DVASPECT_CONTENT) and
                (tymed_in & pythoncom.TYMED_HGLOBAL)
            ):
                return winerror.S_OK
            raise COMException(hresult=winerror.DV_E_FORMATETC)

        def EnumFormatEtc(self, direction):
            if direction != pythoncom.DATADIR_GET:
                raise COMException(hresult=winerror.E_NOTIMPL)
            return NewEnum(self._formatetc, iid=pythoncom.IID_IEnumFORMATETC)

        def GetDataHere(self, formatetc):
            raise COMException(hresult=winerror.E_NOTIMPL)

        def GetCanonicalFormatEtc(self, formatetc):
            raise COMException(hresult=winerror.DATA_S_SAMEFORMATETC)

        def SetData(self, formatetc, medium, fRelease):
            raise COMException(hresult=winerror.E_NOTIMPL)

        def DAdvise(self, formatetc, flags, sink):
            raise COMException(hresult=winerror.E_NOTIMPL)

        def DUnadvise(self, connection):
            raise COMException(hresult=winerror.E_NOTIMPL)

        def EnumDAdvise(self):
            raise COMException(hresult=winerror.E_NOTIMPL)

    screen_w, screen_h = pyautogui.size()
    start_x = max(5, min(screen_w - 5, target_x - max(120, min(220, screen_w // 8))))
    start_y = max(5, min(screen_h - 5, target_y))
    target_x = max(5, min(screen_w - 5, int(target_x)))
    target_y = max(5, min(screen_h - 5, int(target_y)))

    down_event = threading.Event()
    worker_error = []
    pyautogui.moveTo(start_x, start_y, duration=0.12)

    def _mouse_worker():
        try:
            time.sleep(0.08)
            pyautogui.mouseDown(button="left")
            down_event.set()
            time.sleep(0.06)
            pyautogui.moveRel(18, 0, duration=0.08)
            pyautogui.moveTo(target_x, target_y, duration=0.32)
            time.sleep(0.08)
            pyautogui.mouseUp(button="left")
        except Exception as e:
            worker_error.append(e)
            down_event.set()

    drag_thread = threading.Thread(target=_mouse_worker, daemon=True)
    drag_thread.start()

    if not down_event.wait(1.2):
        raise RuntimeError("鼠标按下超时，无法启动真实拖拽")

    try:
        source = wrap(IDropSource(), iid=pythoncom.IID_IDropSource, useDispatcher=0)
        data_object = wrap(IDataObject(file_paths), iid=pythoncom.IID_IDataObject, useDispatcher=0)
        _drag_log(f"start=({start_x},{start_y}) target=({target_x},{target_y}) files={file_paths}")
        effect = pythoncom.DoDragDrop(data_object, source, shellcon.DROPEFFECT_COPY)
    finally:
        if ole_inited:
            try:
                pythoncom.OleUninitialize()
            except Exception:
                pass
        else:
            pythoncom.CoUninitialize()
        drag_thread.join(timeout=2.0)

    if worker_error:
        raise worker_error[0]
    _drag_log(f"DoDragDrop effect={effect}")
    if effect != shellcon.DROPEFFECT_COPY:
        raise RuntimeError(f"目标未接受真实拖拽，effect={effect}")
    return effect

def _move_window_away_from_target(hwnd, target_x, target_y):
    """把窗口缩到对角小窗，尽量不遮挡目标区域。"""
    if sys.platform != "win32" or not hwnd:
        return
    screen_w, screen_h = pyautogui.size()
    win_w = max(220, min(300, screen_w // 6))
    win_h = max(160, min(220, screen_h // 4))
    margin = 10
    left = margin if target_x >= screen_w // 2 else max(margin, screen_w - win_w - margin)
    top = margin if target_y >= screen_h // 2 else max(margin, screen_h - win_h - margin)
    try:
        ctypes.windll.user32.MoveWindow(int(hwnd), int(left), int(top), int(win_w), int(win_h), True)
    except Exception as e:
        log_internal_issue(f"移动资源管理器窗口失败: hwnd={hwnd}", e)

def _close_window_silently(hwnd):
    """静默关闭窗口。"""
    if sys.platform != "win32" or not hwnd:
        return False
    try:
        hwnd = int(hwnd)
        user32 = ctypes.windll.user32
        WM_CLOSE = 0x0010
        WM_SYSCOMMAND = 0x0112
        SC_CLOSE = 0xF060

        def _window_alive():
            try:
                return bool(user32.IsWindow(hwnd))
            except Exception:
                return False

        if not _window_alive():
            return True

        user32.PostMessageW(hwnd, WM_CLOSE, 0, 0)
        for _ in range(6):
            time.sleep(0.12)
            if not _window_alive():
                return True

        user32.PostMessageW(hwnd, WM_SYSCOMMAND, SC_CLOSE, 0)
        for _ in range(6):
            time.sleep(0.12)
            if not _window_alive():
                return True

        try:
            user32.ShowWindow(hwnd, 9)
            user32.SetForegroundWindow(hwnd)
            time.sleep(0.1)
        except Exception:
            pass

        user32.PostMessageW(hwnd, WM_CLOSE, 0, 0)
        for _ in range(8):
            time.sleep(0.15)
            if not _window_alive():
                return True
        return False
    except Exception as e:
        log_internal_issue(f"关闭窗口失败: hwnd={hwnd}", e)
        return False

def _set_cursor_pos(x, y):
    if sys.platform == "win32":
        ctypes.windll.user32.SetCursorPos(int(x), int(y))
    else:
        pyautogui.moveTo(int(x), int(y))

def _mouse_left_down():
    if sys.platform == "win32":
        ctypes.windll.user32.mouse_event(0x0002, 0, 0, 0, 0)
    else:
        pyautogui.mouseDown(button="left")

def _mouse_left_up():
    if sys.platform == "win32":
        ctypes.windll.user32.mouse_event(0x0004, 0, 0, 0, 0)
    else:
        pyautogui.mouseUp(button="left")

def _wait_for_foreground_explorer_window(timeout=4.0):
    """等待资源管理器窗口成为前台窗口。"""
    end_time = time.time() + timeout
    while time.time() < end_time:
        try:
            hwnd = ctypes.windll.user32.GetForegroundWindow()
            if hwnd and get_window_class_name(hwnd) in ("CabinetWClass", "ExploreWClass"):
                return int(hwnd)
        except Exception as e:
            log_internal_issue("读取前台窗口失败", e)
        time.sleep(0.15)
    return None

def _find_explorer_item_center(hwnd, file_name, timeout=5.0):
    """从资源管理器窗口中找到指定文件项的中心点。"""
    from pywinauto import Desktop

    end_time = time.time() + timeout
    target_name = (file_name or "").strip().lower()
    last_err = None
    while time.time() < end_time:
        try:
            win = Desktop(backend="uia").window(handle=int(hwnd))
            candidates = []
            for item in win.descendants(control_type="ListItem"):
                try:
                    name = (item.window_text() or "").strip()
                    rect = item.rectangle()
                    if rect.width() <= 4 or rect.height() <= 4:
                        continue
                    lname = name.lower()
                    if lname == target_name:
                        return rect.mid_point().x, rect.mid_point().y
                    if target_name and target_name in lname:
                        candidates.append((rect.width() * rect.height(), rect))
                except Exception:
                    continue
            if candidates:
                candidates.sort(key=lambda x: x[0], reverse=True)
                rect = candidates[0][1]
                return rect.mid_point().x, rect.mid_point().y
        except Exception as e:
            last_err = e
        time.sleep(0.2)
    if last_err:
        raise RuntimeError(f"未能定位资源管理器中的文件项: {last_err}")
    raise RuntimeError("未能定位资源管理器中的文件项")

def perform_explorer_assisted_drag(file_path, target_x, target_y, log_func=None, target_hwnd=0):
    """通过资源管理器真实选中文件后，再把它拖到目标坐标。"""
    if sys.platform != "win32":
        raise RuntimeError("资源管理器拖拽仅支持 Windows")
    abs_path = os.path.normpath(os.path.abspath(file_path))
    if not os.path.isfile(abs_path):
        raise FileNotFoundError(f"文件不存在: {abs_path}")

    def _drag_log(message):
        write_drag_debug(f"[ExplorerDrag] {message}")
        if log_func:
            try:
                log_func(message)
            except Exception:
                pass

    hwnd = 0
    try:
        _drag_log(f"打开资源管理器并选中文件: {abs_path}")
        subprocess.Popen(f'explorer /select,"{abs_path}"', shell=True)
        time.sleep(0.9)
        hwnd = _wait_for_foreground_explorer_window(timeout=5.0)
        if not hwnd:
            raise RuntimeError("未等到资源管理器窗口")

        force_activate_window(hwnd)
        _move_window_away_from_target(hwnd, int(target_x), int(target_y))
        time.sleep(0.8)

        src_x, src_y = _find_explorer_item_center(hwnd, os.path.basename(abs_path), timeout=5.0)
        _drag_log(f"source=({src_x},{src_y}) target=({target_x},{target_y}) hwnd={hwnd}")

        pyautogui.moveTo(src_x, src_y, duration=0.22)
        time.sleep(0.2)
        _set_cursor_pos(src_x, src_y)
        _mouse_left_down()
        _drag_log("左键已按下，开始连续拖拽")
        time.sleep(0.35)
        pyautogui.moveRel(24, 0, duration=0.22)
        time.sleep(0.12)
        if target_hwnd:
            try:
                if ctypes.windll.user32.IsWindow(int(target_hwnd)):
                    force_activate_window(int(target_hwnd))
                    _drag_log(f"拖拽途中已切回目标窗口: hwnd={int(target_hwnd)}")
                    time.sleep(0.18)
            except Exception as e:
                log_internal_issue(f"拖拽途中切回目标窗口失败: hwnd={target_hwnd}", e)
        mid_x = int((src_x + int(target_x)) / 2)
        mid_y = int((src_y + int(target_y)) / 2)
        pyautogui.moveTo(mid_x, mid_y, duration=0.55)
        time.sleep(0.08)
        pyautogui.moveTo(int(target_x), int(target_y), duration=0.95)
        _drag_log("已到达目标坐标，保持按住等待网页响应")
        time.sleep(0.95)
        _mouse_left_up()
        _drag_log("左键已释放")
        _drag_log("资源管理器真实拖拽手势已完成")
        return True
    finally:
        if hwnd:
            time.sleep(0.2)
            if _close_window_silently(hwnd):
                _drag_log("资源管理器窗口已关闭")
            else:
                _drag_log("资源管理器窗口关闭失败")

SMART_FILL_IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".bmp", ".gif", ".webp", ".ico", ".tiff"}
SMART_FILL_TEXT_EXTS = {".txt", ".log", ".md", ".csv", ".ini", ".json", ".xml", ".py", ".bat"}
SMART_FILL_FILE_EXTS = SMART_FILL_IMAGE_EXTS | SMART_FILL_TEXT_EXTS | {
    ".pdf", ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx",
    ".zip", ".rar", ".7z", ".mp4", ".mov", ".avi", ".mkv",
    ".mp3", ".wav", ".flac", ".aac", ".ogg", ".m4a", ".wma"
}
SMART_FILL_IMAGE_EXTS_TEXT = ",".join(sorted(SMART_FILL_IMAGE_EXTS))
SMART_FILL_TEXT_EXTS_TEXT = ",".join(sorted(SMART_FILL_TEXT_EXTS))
SMART_FILL_FILE_EXTS_TEXT = ",".join(sorted(SMART_FILL_FILE_EXTS))
SMART_FILL_TXT_ONLY_EXTS_TEXT = ".txt"

def get_default_smart_fill_rules():
    return {
        "bundle_mode": "auto",               # auto / root_only / subdirs_only
        "scan_subdirs_recursive": True,
        "text_filename_mode": "pure_number", # pure_number / contains_number / non_chinese / non_chinese_no_number / all / regex
        "custom_text_regex": "",
        "upload_source": "images_only",      # images_only / images_then_files
        "file_source": "files_then_images",  # files_only / files_then_images / images_only
        "text_fill_mode": "content",         # content / path
        "text_shortage_action": "skip_row",  # skip_row / blank
        "image_exts_text": SMART_FILL_IMAGE_EXTS_TEXT,
        "text_exts_text": SMART_FILL_TEXT_EXTS_TEXT,
        "file_exts_text": SMART_FILL_FILE_EXTS_TEXT,
        "step_rules": {},
    }

def get_smart_fill_rules(config_obj=None):
    rules = get_default_smart_fill_rules()
    try:
        if config_obj and isinstance(config_obj.get("smart_fill_rules"), dict):
            for k, v in config_obj["smart_fill_rules"].items():
                if k == "step_rules" and isinstance(v, dict):
                    rules[k] = dict(v)
                elif k in rules:
                    rules[k] = v
    except Exception:
        pass
    return rules

def _parse_exts_text(exts_text, fallback_text):
    parts = []
    for raw in str(exts_text or fallback_text).replace("，", ",").replace("；", ",").split(","):
        p = raw.strip().lower()
        if not p:
            continue
        if not p.startswith("."):
            p = "." + p
        parts.append(p)
    if parts:
        return set(parts)
    if exts_text != fallback_text:
        return _parse_exts_text(fallback_text, fallback_text)
    return set()

def build_runtime_smart_fill_rules(raw_rules=None):
    rules = get_default_smart_fill_rules()
    if isinstance(raw_rules, dict):
        for k, v in raw_rules.items():
            if k in rules:
                rules[k] = v
    rules["image_exts"] = _parse_exts_text(rules.get("image_exts_text"), SMART_FILL_IMAGE_EXTS_TEXT)
    rules["text_exts"] = _parse_exts_text(rules.get("text_exts_text"), SMART_FILL_TEXT_EXTS_TEXT)
    rules["file_exts"] = _parse_exts_text(rules.get("file_exts_text"), SMART_FILL_FILE_EXTS_TEXT)
    return rules

def _get_smart_fill_ext_preset_texts(target_kind="file", act_type="", sub=None):
    presets = []
    if act_type in ["upload", "drag_file"]:
        presets.extend([SMART_FILL_IMAGE_EXTS_TEXT, SMART_FILL_FILE_EXTS_TEXT])
    elif act_type == "run_app":
        presets.extend([SMART_FILL_FILE_EXTS_TEXT, SMART_FILL_IMAGE_EXTS_TEXT])
    elif act_type == "clear_input_plus" and sub == "content":
        presets.extend([SMART_FILL_TXT_ONLY_EXTS_TEXT, SMART_FILL_TEXT_EXTS_TEXT])
    elif target_kind == "text":
        presets.extend([SMART_FILL_TEXT_EXTS_TEXT, SMART_FILL_TXT_ONLY_EXTS_TEXT])
    else:
        presets.extend([SMART_FILL_FILE_EXTS_TEXT, SMART_FILL_IMAGE_EXTS_TEXT])
    unique = []
    seen = set()
    for item in presets:
        text = str(item or "").strip()
        if text and text not in seen:
            seen.add(text)
            unique.append(text)
    return unique

def _create_exts_combo(current_text="", target_kind="file", act_type="", sub=None, placeholder=""):
    combo = QComboBox()
    combo.setEditable(True)
    combo.setInsertPolicy(QComboBox.NoInsert)
    for text in _get_smart_fill_ext_preset_texts(target_kind=target_kind, act_type=act_type, sub=sub):
        combo.addItem(text)
    if combo.lineEdit():
        combo.lineEdit().setPlaceholderText(placeholder)
    current_text = str(current_text or "").strip()
    if current_text:
        idx = combo.findText(current_text, Qt.MatchFixedString)
        if idx >= 0:
            combo.setCurrentIndex(idx)
        else:
            combo.setEditText(current_text)
    return combo

def _get_combo_text(combo):
    if not combo:
        return ""
    text = combo.currentText()
    if not text and combo.lineEdit():
        text = combo.lineEdit().text()
    return str(text or "").strip()

def natural_sort_key(text):
    text = "" if text is None else str(text)
    return [int(part) if part.isdigit() else part.lower() for part in re.split(r'([0-9]+)', text)]

def sorted_scandir_entries(path):
    entries = list(os.scandir(path))
    return sorted(entries, key=lambda e: (not e.is_dir(), natural_sort_key(e.name)))

def _read_text_for_smart_fill(path, max_bytes=1024 * 1024 * 2):
    try:
        if os.path.getsize(path) > max_bytes:
            return None
        with open(path, "rb") as f:
            raw = f.read()
        for enc in ("utf-8-sig", "utf-8", "gb18030", "gbk", "utf-16", "utf-16-le", "utf-16-be"):
            try:
                return raw.decode(enc).strip()
            except Exception:
                pass
        return raw.decode("utf-8", errors="ignore").strip()
    except Exception:
        return None

def _is_numbered_text_filename(path, rules=None):
    """智能填充时按规则匹配文本文件名。"""
    stem = os.path.splitext(os.path.basename(path))[0]
    rules = build_runtime_smart_fill_rules(rules)
    mode = str(rules.get("text_filename_mode", "pure_number"))
    has_chinese = bool(re.search(r'[\u4e00-\u9fff]', stem))
    has_number = bool(re.search(r'\d+', stem))
    has_letter = any(ch.isalpha() for ch in stem)
    if mode == "pure_number":
        return stem.strip().isdigit()
    if mode == "contains_number":
        return has_number
    if mode == "non_chinese":
        # 修复：纯数字文件名（如 1/2/3）不再算作“非中文”文本名，
        # 必须至少包含一个字母，避免英文文件被数字序号文件抢先命中。
        return (not has_chinese) and has_letter and bool(stem.strip())
    if mode == "non_chinese_no_number":
        return (not has_number) and (not has_chinese) and has_letter and bool(stem.strip())
    if mode == "all":
        return True
    if mode == "regex":
        pattern = str(rules.get("custom_text_regex", "")).strip()
        if not pattern:
            return False
        try:
            return bool(re.search(pattern, stem))
        except Exception:
            return False
    return stem.strip().isdigit()

def _smart_fill_item_identity(item):
    """为智能填充素材生成稳定标识，便于同一素材按匹配结果消耗。"""
    try:
        if isinstance(item, dict):
            path = item.get("path", "")
        else:
            path = str(item or "")
        return os.path.normcase(os.path.normpath(path)) if path else ""
    except Exception:
        return ""

def _smart_fill_stem_match_score(text_path, candidate_path):
    """返回文本文件与候选文件名的匹配分数，数值越小越优先。"""
    try:
        text_stem = os.path.splitext(os.path.basename(str(text_path or "")))[0].strip().lower()
        cand_stem = os.path.splitext(os.path.basename(str(candidate_path or "")))[0].strip().lower()
        if not text_stem or not cand_stem:
            return None
        if text_stem == cand_stem:
            return 0
        if cand_stem.startswith(text_stem) and len(cand_stem) > len(text_stem):
            next_char = cand_stem[len(text_stem)]
            if next_char in {".", "-", "_", " ", "(", "（", "[", "【"}:
                return 1
        text_num = re.fullmatch(r"\d+", text_stem)
        cand_num = re.match(r"(\d+)", cand_stem)
        if text_num and cand_num and text_num.group(0) == cand_num.group(1):
            return 2
    except Exception:
        return None
    return None

def _find_smart_fill_matched_item(step_items, cursor_holder, text_data):
    """根据当前文本文件，优先寻找同名/同前缀的文件素材。"""
    text_path = text_data.get("path", "") if isinstance(text_data, dict) else ""
    if not text_path or not step_items:
        return None
    start_idx = max(0, int((cursor_holder or [0])[0]))
    best_item = None
    best_score = None
    for idx in range(start_idx, len(step_items)):
        cand_path = step_items[idx].get("path", "") if isinstance(step_items[idx], dict) else str(step_items[idx] or "")
        score = _smart_fill_stem_match_score(text_path, cand_path)
        if score is None:
            continue
        if best_score is None or score < best_score:
            best_score = score
            best_item = step_items[idx]
            if score == 0:
                break
    return best_item

def _take_smart_value_with_preferred(items, cursor_holder, preferred_item=None, repeat_single=False):
    """取素材时优先命中指定素材，并同步推进游标。"""
    if not items:
        return None
    if repeat_single and len(items) == 1:
        return items[0]
    preferred_identity = _smart_fill_item_identity(preferred_item)
    if preferred_identity:
        for idx in range(max(0, cursor_holder[0]), len(items)):
            if _smart_fill_item_identity(items[idx]) == preferred_identity:
                cursor_holder[0] = idx + 1
                return items[idx]
    idx = cursor_holder[0]
    if idx >= len(items):
        return None
    cursor_holder[0] += 1
    return items[idx]

def _find_same_name_video_in_bundle(text_data, bundle):
    """查找当前文本素材在同目录下的同名视频。"""
    try:
        text_path = text_data.get("path", "") if isinstance(text_data, dict) else ""
        if not text_path:
            return None
        text_dir = os.path.normcase(os.path.normpath(os.path.dirname(text_path)))
        text_stem = os.path.splitext(os.path.basename(text_path))[0].strip().lower()
        if not text_stem:
            return None
        for path in bundle.get("files", []):
            ext = os.path.splitext(str(path).lower())[1]
            if ext not in {".mp4", ".mov", ".avi", ".mkv", ".flv", ".wmv", ".m4v", ".webm"}:
                continue
            cand_dir = os.path.normcase(os.path.normpath(os.path.dirname(str(path))))
            cand_stem = os.path.splitext(os.path.basename(str(path)))[0].strip().lower()
            if cand_dir == text_dir and cand_stem == text_stem:
                return path
    except Exception:
        return None
    return None

def _text_has_same_name_video_in_bundle(text_data, bundle):
    """判断当前文本素材是否已经存在同目录同名视频。"""
    return bool(_find_same_name_video_in_bundle(text_data, bundle))

def _get_pending_texts_from_bundle(bundle):
    """只保留还没有同名视频成品的文本素材。"""
    result = []
    for text_data in bundle.get("texts", []):
        if not _text_has_same_name_video_in_bundle(text_data, bundle):
            result.append(text_data)
    return result

def _scan_single_smart_fill_bundle(folder_path, recursive=True, rules=None):
    rules = build_runtime_smart_fill_rules(rules)
    bundle = {
        "folder": folder_path,
        "name": os.path.basename(folder_path) or folder_path,
        "images": [],
        "texts": [],
        "files": [],
    }

    try:
        if recursive:
            for root, dirs, files in os.walk(folder_path):
                dirs.sort(key=natural_sort_key)
                files.sort(key=natural_sort_key)
                for fname in files:
                    full_path = os.path.join(root, fname)
                    ext = os.path.splitext(fname.lower())[1]
                    if ext in rules["image_exts"]:
                        bundle["images"].append(full_path)
                    if ext in rules["text_exts"] and _is_numbered_text_filename(full_path, rules):
                        content = _read_text_for_smart_fill(full_path)
                        if content is not None:
                            bundle["texts"].append({"path": full_path, "content": content})
                    if ext in rules["file_exts"]:
                        bundle["files"].append(full_path)
        else:
            for entry in sorted_scandir_entries(folder_path):
                if not entry.is_file():
                    continue
                ext = os.path.splitext(entry.name.lower())[1]
                if ext in rules["image_exts"]:
                    bundle["images"].append(entry.path)
                if ext in rules["text_exts"] and _is_numbered_text_filename(entry.path, rules):
                    content = _read_text_for_smart_fill(entry.path)
                    if content is not None:
                        bundle["texts"].append({"path": entry.path, "content": content})
                if ext in rules["file_exts"]:
                    bundle["files"].append(entry.path)
    except Exception as e:
        log_internal_issue(f"扫描智能填充目录失败: {folder_path}", e)
        return None

    if not (bundle["images"] or bundle["texts"] or bundle["files"]):
        return None
    return bundle

def _collect_descendant_smart_fill_bundles(folder_path, recursive=True, rules=None):
    """收集子目录素材组。递归时每个实际子目录都单独成组，不再并入父目录。"""
    rules = build_runtime_smart_fill_rules(rules)
    bundles = []
    base_norm = os.path.normcase(os.path.normpath(folder_path))

    try:
        if recursive:
            for root, dirs, _files in os.walk(folder_path):
                dirs.sort(key=natural_sort_key)
                root_norm = os.path.normcase(os.path.normpath(root))
                if root_norm == base_norm:
                    continue
                child_bundle = _scan_single_smart_fill_bundle(root, recursive=False, rules=rules)
                if child_bundle:
                    bundles.append(child_bundle)
        else:
            for entry in sorted_scandir_entries(folder_path):
                if not entry.is_dir():
                    continue
                child_bundle = _scan_single_smart_fill_bundle(entry.path, recursive=False, rules=rules)
                if child_bundle:
                    bundles.append(child_bundle)
    except Exception as e:
        log_internal_issue(f"收集智能填充子目录素材组失败: {folder_path}", e)
        return bundles
    return bundles

def collect_smart_fill_bundles(folder_path, rules=None):
    rules = build_runtime_smart_fill_rules(rules)
    bundles = []
    try:
        entries = sorted_scandir_entries(folder_path)
    except Exception as e:
        log_internal_issue(f"枚举智能填充目录失败: {folder_path}", e)
        return bundles

    bundle_mode = str(rules.get("bundle_mode", "auto"))
    recursive = bool(rules.get("scan_subdirs_recursive", True))

    has_direct_files = any(
        entry.is_file() and os.path.splitext(entry.name.lower())[1] in rules["file_exts"]
        for entry in entries
    )
    has_subdirs = any(entry.is_dir() for entry in entries)

    if bundle_mode == "root_only":
        root_bundle = _scan_single_smart_fill_bundle(folder_path, recursive=recursive, rules=rules)
        if root_bundle:
            bundles.append(root_bundle)
        return bundles

    if bundle_mode == "subdirs_only":
        bundles.extend(_collect_descendant_smart_fill_bundles(folder_path, recursive=recursive, rules=rules))
        return bundles

    if has_direct_files:
        root_bundle = _scan_single_smart_fill_bundle(folder_path, recursive=False, rules=rules)
        if root_bundle:
            bundles.append(root_bundle)

    bundles.extend(_collect_descendant_smart_fill_bundles(folder_path, recursive=recursive, rules=rules))

    if not bundles and (has_direct_files or not has_subdirs):
        root_bundle = _scan_single_smart_fill_bundle(folder_path, recursive=recursive, rules=rules)
        if root_bundle:
            bundles.append(root_bundle)

    return bundles

def format_tree_snapshot(tree, is_folder_role_fn, limit=120):
    lines = []
    truncated = False

    def _walk(container, level=0):
        nonlocal truncated
        if hasattr(container, "topLevelItemCount"):
            count = container.topLevelItemCount()
            getter = container.topLevelItem
        else:
            count = container.childCount()
            getter = container.child

        for i in range(count):
            if len(lines) >= limit:
                truncated = True
                return
            item = getter(i)
            if not item:
                continue
            role = item.data(0, Qt.UserRole)
            kind = "FOLDER" if is_folder_role_fn(role) else "TASK"
            lines.append(f"{'  ' * level}{kind}:{item.text(0)}|role={role}")
            if item.childCount():
                _walk(item, level + 1)
                if truncated:
                    return

    try:
        _walk(tree)
    except Exception as e:
        return f"<snapshot_error:{e}>"

    if truncated:
        lines.append("...(truncated)")
    return " || ".join(lines) if lines else "<empty>"

init_drag_debug_log()

# --- Helper: Browser Profiles ---
# 全局缓存：允许外部将自定义 user-data-dir 注入扫描列表
_extra_scan_dirs = set()
CHROME_PROFILES_CACHE = {"data": [], "ts": 0.0, "scan_dirs": tuple()}
WINDOW_PROFILE_INFO_CACHE = {}

def clear_chrome_profile_cache():
    """清空 Chrome 账号扫描缓存。"""
    CHROME_PROFILES_CACHE["data"] = []
    CHROME_PROFILES_CACHE["ts"] = 0.0
    CHROME_PROFILES_CACHE["scan_dirs"] = tuple()

def clear_window_profile_caches():
    """清空窗口相关缓存，便于刷新后立即看到最新标签/备注。"""
    WINDOW_CLASS_NAME_CACHE.clear()
    WINDOW_ACCOUNT_MARKER_CACHE.clear()
    WINDOW_PROFILE_INFO_CACHE.clear()

def get_chrome_profiles(force_refresh=False):
    """终极唯一标识版：返回 (path, name, email, remark, raw_id) 元组。
    [修复] 深度扫描 Local State 和 Preferences，提取更多账户元数据。"""
    scan_dirs_snapshot = tuple(sorted(os.path.normpath(p) for p in _extra_scan_dirs if p))
    if (not force_refresh and CHROME_PROFILES_CACHE["data"] and
            time.time() - CHROME_PROFILES_CACHE["ts"] < 8 and
            CHROME_PROFILES_CACHE["scan_dirs"] == scan_dirs_snapshot):
        return list(CHROME_PROFILES_CACHE["data"])

    profiles = []
    if sys.platform != 'win32': return profiles
    
    found_profiles = {} # { canonical_path: { "name":..., "email":..., "remark":..., "id":..., "raw_path":... } }
    
    # 使用一个集合来存储所有已处理的规范化路径，确保唯一性
    processed_canonical_paths = set()
    
    # 1. 基础探测路径
    base_dirs = [
        os.path.expandvars(r'%LOCALAPPDATA%\Google\Chrome\User Data'),
        os.path.expandvars(r'%LOCALAPPDATA%\Google\Chrome Beta\User Data'),
        os.path.expandvars(r'%LOCALAPPDATA%\Google\Chrome SxS\User Data'),
    ]
    
    # [已移除] 不再硬编码 D:\Chrome_Data 等路径，改为由用户在 UI 中手动添加或自动识别运行中的进程
    pass
        
    for _extra in _extra_scan_dirs:
        if _extra: base_dirs.append(os.path.normpath(_extra))
    
    # 2. 增强进程扫描
    try:
        cmd = 'wmic process where "commandline like \'%%--user-data-dir=%%\'" get commandline'
        output = subprocess.check_output(cmd, shell=True).decode('gbk', errors='ignore')
        # [修复] 同样改为逐行解析，防止路径中的 .exe 干扰
        for line in output.splitlines():
            line = line.strip()
            if not line or '--user-data-dir=' not in line: continue
            p = extract_cmd_switch_value(line, "user-data-dir")
            if p and os.path.exists(p):
                base_dirs.append(os.path.normpath(p))
    except Exception as e:
        log_internal_issue("扫描 Chrome 进程命令行失败", e)

    # 3. 遍历 User Data 目录
    for u_dir in set(base_dirs):
        u_dir = os.path.normpath(u_dir)
        if not os.path.exists(u_dir): continue
        
        # [优化] 检查这是否是一个“独立账号根目录”（即目录下直接有 Default 文件夹）
        # 这种情况常见于 Google 助手的独立存储模式
        if os.path.exists(os.path.join(u_dir, "Default", "Preferences")):
            # 将该目录本身视为一个特殊的 User Data Dir 进行处理
            pass 

        state_info = {}
        state_path = os.path.join(u_dir, "Local State")
        if os.path.exists(state_path):
            try:
                with open(state_path, "r", encoding="utf-8", errors="ignore") as f:
                    state_data = json.load(f)
                info_cache = state_data.get("profile", {}).get("info_cache", {})
                for pid, info in info_cache.items():
                    state_info[pid] = {
                        "name": info.get("name", ""),
                        "email": info.get("user_name", "") or info.get("gaia_name", ""),
                        "remark": info.get("shortcut_name", "")
                    }
            except Exception as e:
                log_internal_issue(f"读取 Chrome Local State 失败: {state_path}", e)

        try:
            for entry in os.scandir(u_dir):
                if not entry.is_dir(): continue
                p_id = entry.name
                p_path = os.path.normpath(entry.path)
                
                pref_path = os.path.join(p_path, "Preferences")
                # [修复] 放宽过滤条件：只要有 Preferences、有锁文件，或者在 Local State 的缓存列表中，就认为是有效 Profile
                # 同时也检查子目录 Default（针对助手模式：root/Account/Default/Preferences）
                has_pref = os.path.exists(pref_path) or os.path.exists(os.path.join(p_path, "Default", "Preferences"))
                # [修复] 严格过滤：必须存在 Preferences 文件才认为是有效账户，不再依赖容易残留的 SingletonLock 或 Local State 缓存
                if not has_pref: continue
                
                # [新增] 助手模式适配：如果当前是 Account 目录且其下有 Default，则将 p_path 修正为 Default 路径
                if not os.path.exists(pref_path) and os.path.exists(os.path.join(p_path, "Default", "Preferences")):
                    p_path = os.path.join(p_path, "Default")
                    pref_path = os.path.join(p_path, "Preferences")
                
                # [核心修复] 统一规范化路径，并处理 Default 后缀，作为唯一标识进行去重
                canonical_profile_path = os.path.normpath(p_path).lower()
                # 如果路径是 Profile X\Default 形式，则将其规范化为 Profile X
                if canonical_profile_path.endswith("\\default"):
                    canonical_profile_path = os.path.dirname(canonical_profile_path)
                
                if canonical_profile_path in processed_canonical_paths: continue
                processed_canonical_paths.add(canonical_profile_path)

                # 提取详细信息
                info = state_info.get(p_id, {"name": "", "email": "", "remark": ""})
                is_logged_in = False
                if os.path.exists(pref_path):
                    try:
                        with open(pref_path, 'r', encoding='utf-8', errors='ignore') as f:
                            pref = json.load(f)
                        if not info["name"]: info["name"] = pref.get("profile", {}).get("name", "")
                        
                        # [修复] 提取登录状态：检查是否有有效的账号邮箱
                        acc_info = pref.get("account_info", [])
                        if isinstance(acc_info, list) and acc_info: 
                            info["email"] = acc_info[0].get("email", "")
                            if info["email"]: is_logged_in = True
                            
                        if not is_logged_in:
                            last_user = pref.get("google", {}).get("services", {}).get("last_username", "")
                            if last_user:
                                info["email"] = last_user
                                is_logged_in = True
                    except: pass
                
                # [核心修复] 极严格过滤：必须同时满足以下条件才认为是有效账户
                # 1. 必须有 Preferences 文件
                # 2. 必须有有效的邮箱信息（证明已登录）
                # 3. 排除那些虽然有文件但邮箱为空的干扰目录
                if not is_logged_in or not info["email"] or len(info["email"]) < 5: 
                    continue
                
                # [优化] 智能精简显示名，避免 L-01 (L-01) 这种重复堆叠
                base_name = info["name"] or info["email"] or p_id
                
                # 如果是 Default 或 Profile X，且有更好的 info["name"]，则不显示 p_id
                id_is_generic = p_id == "Default" or p_id.startswith("Profile ")
                
                display_name = base_name
                # [精简逻辑] 如果 base_name 已经包含了 p_id 或 email，就不再加括号
                if not id_is_generic and p_id.lower() not in base_name.lower():
                    display_name = f"{base_name} ({p_id})"
                
                if "User Data" not in u_dir:
                    # 对于助手目录，用父目录名作为前缀区分
                    parent_name = os.path.basename(u_dir)
                    if parent_name.lower() not in display_name.lower():
                        display_name = f"[{parent_name}] {display_name}"

                found_profiles[canonical_profile_path] = {
                    "name": display_name,
                    "email": info["email"],
                    "remark": info["remark"],
                    "id": p_id
                }
        except: pass

    # 4. 运行中进程补全逻辑 [核心优化]
    # 优先匹配已有账号，匹配不到再新增，确保不重复
    # [修复] 不再盲目将所有运行中的进程加入列表，仅当能确认其也是有效账户时才加入，防止出现大量冗余的临时/空白浏览器实例
    active_paths = get_active_chrome_profiles()
    for ap in active_paths:
        norm_ap = os.path.normpath(ap)
        
        # 使用深度归一化比对，检查是否已在扫描结果中
        exists = False
        for p_path in found_profiles:
            if is_same_path(p_path, norm_ap):
                exists = True; break
        
        if not exists:
            # [修复] 对于未扫描到的运行中实例，再次检查其 Preferences 文件，确认是否是已登录账户
            pref_path = os.path.join(norm_ap, "Preferences")
            if not os.path.exists(pref_path):
                alt_pref = os.path.join(norm_ap, "Default", "Preferences")
                if os.path.exists(alt_pref):
                    pref_path = alt_pref
                    norm_ap = os.path.join(norm_ap, "Default")
            
            is_valid_account = False
            p_email = ""
            p_name_from_pref = ""
            if os.path.exists(pref_path):
                try:
                    with open(pref_path, 'r', encoding='utf-8', errors='ignore') as f:
                        pref = json.load(f)
                    p_name_from_pref = pref.get("profile", {}).get("name", "")
                    acc_info = pref.get("account_info", [])
                    if isinstance(acc_info, list) and acc_info:
                        p_email = acc_info[0].get("email", "")
                    if not p_email:
                        p_email = pref.get("google", {}).get("services", {}).get("last_username", "")
                    if p_email:
                        is_valid_account = True
                except: pass
                
            # 只有确实有账号信息的实例才加入
            if is_valid_account and p_email and len(p_email) > 5:
                # [核心修复] 统一规范化路径，并处理 Default 后缀，作为唯一标识进行去重
                canonical_active_path = os.path.normpath(norm_ap).lower()
                if canonical_active_path.endswith("\\default"):
                    canonical_active_path = os.path.dirname(canonical_active_path)

                if canonical_active_path in processed_canonical_paths: continue
                processed_canonical_paths.add(canonical_active_path)
                
                p_id = os.path.basename(norm_ap)
                p_parent = os.path.basename(os.path.dirname(norm_ap))
                
                base_name = p_name_from_pref or p_email or p_id
                if p_id in ("Default", "") or p_id.startswith("Profile "):
                    p_name = f"[{p_parent}] {base_name}"
                else:
                    p_name = base_name
                
                found_profiles[canonical_active_path] = {
                    "name": p_name,
                    "email": p_email,
                    "remark": "当前运行中",
                    "id": p_id
                }

    # 5. 格式化输出
    for canonical_path, d in found_profiles.items():
        # 最终返回时，将规范化路径作为第一个元素
        profiles.append((canonical_path, d["name"], d["email"], d["remark"], d["id"]))
    
    profiles.sort(key=lambda x: x[1].lower())
    CHROME_PROFILES_CACHE["data"] = list(profiles)
    CHROME_PROFILES_CACHE["ts"] = time.time()
    CHROME_PROFILES_CACHE["scan_dirs"] = scan_dirs_snapshot
    return profiles

def force_activate_window(hwnd):
    """使用 Win32 API 强制激活窗口，确保不破坏最大化状态。"""
    if sys.platform != 'win32': return
    try:
        import ctypes
        from ctypes import wintypes
        user32 = ctypes.windll.user32
        # SW_SHOW=5, SW_RESTORE=9, SW_SHOWMAXIMIZED=3
        # 检查窗口状态
        class WINDOWPLACEMENT(ctypes.Structure):
            _fields_ = [
                ("length", wintypes.UINT),
                ("flags", wintypes.UINT),
                ("showCmd", wintypes.UINT),
                ("ptMinPosition", wintypes.POINT),
                ("ptMaxPosition", wintypes.POINT),
                ("rcNormalPosition", wintypes.RECT),
            ]
        placement = WINDOWPLACEMENT()
        placement.length = ctypes.sizeof(WINDOWPLACEMENT)
        user32.GetWindowPlacement(hwnd, ctypes.byref(placement))
        
        if placement.showCmd == 2: # SW_SHOWMINIMIZED
            user32.ShowWindow(hwnd, 9) # SW_RESTORE
        
        user32.SetForegroundWindow(hwnd)
    except:
        pass

def get_root_window_from_point(x, y):
    """按屏幕坐标获取顶层窗口句柄。"""
    if sys.platform != 'win32':
        return 0
    try:
        import ctypes
        from ctypes import wintypes
        user32 = ctypes.windll.user32
        user32.WindowFromPoint.argtypes = [wintypes.POINT]
        user32.WindowFromPoint.restype = wintypes.HWND
        user32.GetAncestor.argtypes = [wintypes.HWND, wintypes.UINT]
        user32.GetAncestor.restype = wintypes.HWND

        pt = wintypes.POINT(int(x), int(y))
        hwnd = user32.WindowFromPoint(pt)
        if not hwnd:
            return 0

        GA_ROOT = 2
        return user32.GetAncestor(hwnd, GA_ROOT) or hwnd
    except:
        return 0

def activate_window_from_point(x, y):
    """按目标坐标激活对应顶层窗口。"""
    hwnd = get_root_window_from_point(x, y)
    if not hwnd:
        raise RuntimeError(f"无法根据坐标定位窗口: ({x}, {y})")
    force_activate_window(hwnd)
    return hwnd

def get_window_text(hwnd):
    """安全读取窗口标题。"""
    if not hwnd or sys.platform != 'win32':
        return ""
    try:
        buf = ctypes.create_unicode_buffer(512)
        ctypes.windll.user32.GetWindowTextW(int(hwnd), buf, len(buf))
        return buf.value or ""
    except Exception:
        return ""

def _safe_guard_name(text):
    text = str(text or "").strip()
    text = re.sub(r'[\\/:*?"<>|\s]+', '_', text)
    text = text.strip("._")
    return text[:80] or "item"

def _capture_virtual_screen_region(left, top, width, height):
    """抓取虚拟桌面区域，兼容多屏与负坐标。"""
    if width <= 0 or height <= 0:
        raise ValueError("截图区域尺寸无效")
    try:
        from PIL import ImageGrab
        if sys.platform == 'win32':
            user32 = ctypes.windll.user32
            vx = int(user32.GetSystemMetrics(76))   # SM_XVIRTUALSCREEN
            vy = int(user32.GetSystemMetrics(77))   # SM_YVIRTUALSCREEN
            full_img = ImageGrab.grab(all_screens=True)
            crop_left = max(0, int(left) - vx)
            crop_top = max(0, int(top) - vy)
            crop_right = min(full_img.size[0], crop_left + int(width))
            crop_bottom = min(full_img.size[1], crop_top + int(height))
            if crop_right <= crop_left or crop_bottom <= crop_top:
                raise RuntimeError("目标区域超出虚拟桌面范围")
            return full_img.crop((crop_left, crop_top, crop_right, crop_bottom))
        return ImageGrab.grab(bbox=(int(left), int(top), int(left) + int(width), int(top) + int(height)))
    except Exception as e:
        raise RuntimeError(f"抓取界面守卫截图失败: {e}")

def create_action_guard_snapshot(task_name, step_name, x, y, width=160, height=96):
    """录制坐标时自动保存一张周边界面快照，供执行前校验。"""
    guard_dir = os.path.join(BASE_DIR, "step_guards")
    os.makedirs(guard_dir, exist_ok=True)
    left = int(round(x - width / 2))
    top = int(round(y - height / 2))
    img = _capture_virtual_screen_region(left, top, width, height)
    file_name = f"{_safe_guard_name(task_name)}__{_safe_guard_name(step_name)}__{int(x)}_{int(y)}.png"
    guard_path = os.path.join(guard_dir, file_name)
    img.save(guard_path)
    hwnd = get_root_window_from_point(x, y)
    return {
        "guard_image": guard_path,
        "guard_region": [left, top, img.size[0], img.size[1]],
        "guard_window_title": get_window_text(hwnd),
        "guard_window_class": get_window_class_name(hwnd),
        "guard_threshold": 0.72,
    }

def evaluate_guard_snapshot_similarity(guard_image_path, region):
    """比较当前界面与录制快照的相似度，返回 0~1。"""
    if not guard_image_path or not os.path.exists(guard_image_path):
        raise FileNotFoundError(f"守卫快照不存在: {guard_image_path}")
    if not isinstance(region, (list, tuple)) or len(region) < 4:
        raise ValueError("守卫区域参数无效")
    try:
        from PIL import Image, ImageChops, ImageStat
        left, top, width, height = [int(float(v)) for v in region[:4]]
        current = _capture_virtual_screen_region(left, top, width, height).convert("L")
        recorded = Image.open(guard_image_path).convert("L")
        if current.size != recorded.size:
            current = current.resize(recorded.size)
        diff = ImageChops.difference(recorded, current)
        rms = ImageStat.Stat(diff).rms[0]
        return max(0.0, min(1.0, 1.0 - (float(rms) / 255.0)))
    except Exception as e:
        raise RuntimeError(f"界面快照比对失败: {e}")

class StepGuardMismatchError(RuntimeError):
    """界面守卫未通过时抛出，调用方应跳过当前行而不是终止整个任务。"""
    pass

class ExecutionInterrupted(RuntimeError):
    """步骤在执行中被人工控制打断（停止/跳步/下一行/重试）时抛出。"""
    pass

def is_same_path(p1, p2):
    """[深度归一化版] 智能路径比对。
    核心逻辑：只要物理位置相同，或者规范化后的路径（处理了 Default 后缀差异）相同，即视为同一账号。"""
    if not p1 or not p2: return False
    try:
        # 1. 规范化路径（处理斜杠、大小写）
        n1 = os.path.normpath(os.path.abspath(p1)).lower()
        n2 = os.path.normpath(os.path.abspath(p2)).lower()
        
        # 2. 处理 Default 后缀的差异（助手模式 vs 标准模式）
        # 统一将以 \default 结尾的路径去掉该结尾进行比对
        # [修复] 确保在比较前，如果路径是 User Data 目录本身，则其对应的 Profile 路径应视为 Default
        c1 = n1
        if c1.endswith("\\user data") and not n1.endswith("\\default"): # 如果是 User Data 目录，但不是 Default Profile
            c1 = os.path.join(c1, "default") # 视为 Default Profile
        c1 = c1[:-8] if c1.endswith("\\default") else c1

        c2 = n2
        if c2.endswith("\\user data") and not n2.endswith("\\default"): # 如果是 User Data 目录，但不是 Default Profile
            c2 = os.path.join(c2, "default") # 视为 Default Profile
        c2 = c2[:-8] if c2.endswith("\\default") else c2

        if c1 == c2: return True
        
        # 3. 物理一致性校验（处理 subst 映射等）
        if os.path.exists(p1) and os.path.exists(p2):
            try:
                if os.path.samefile(p1, p2): return True
            except Exception as e:
                log_internal_issue(f"samefile 比较失败: {p1} <-> {p2}", e)
            
    except Exception as e:
        log_internal_issue(f"路径比较失败: {p1} <-> {p2}", e)
    return False

def extract_cmd_switch_value(cmdline, switch_name):
    """从命令行中提取参数值，兼容带空格但未加引号的写法。"""
    if not cmdline or not switch_name:
        return ""
    try:
        pattern = rf'--{re.escape(switch_name)}=(?:"([^"]+)"|(.+?))(?=\s+--[A-Za-z0-9_-]+(?:=|\b)|$)'
        m = re.search(pattern, cmdline)
        if not m:
            return ""
        val = m.group(1) if m.group(1) is not None else m.group(2)
        return (val or "").strip().strip('"').rstrip("\\")
    except Exception as e:
        log_internal_issue(f"解析命令行参数失败: --{switch_name}", e)
        return ""

def get_active_chrome_profiles():
    """[终极重构] 窗口中心版：返回所有正在运行的 Chrome 实例的路径标识。
    直接以进程命令行参数为准，有多少个不同的路径，就认多少个账号。"""
    active_paths = set()
    if sys.platform != 'win32': return active_paths
    
    try:
        import subprocess, re
        # 获取所有 chrome 进程的命令行
        cmd = 'wmic process where "name=\'chrome.exe\'" get commandline'
        output = subprocess.check_output(cmd, shell=True).decode('gbk', errors='ignore')
        
        for line in output.splitlines():
            line = line.strip()
            if not line or '--user-data-dir=' not in line: continue
            
            u_path = extract_cmd_switch_value(line, "user-data-dir")
            p_dir = extract_cmd_switch_value(line, "profile-directory")
            if u_path:
                
                # 核心逻辑：以完整的 UserData + Profile 作为唯一身份
                if p_dir:
                    full_p = os.path.normpath(os.path.join(u_path, p_dir))
                else:
                    full_p = os.path.normpath(u_path)
                
                active_paths.add(full_p)
    except Exception as e:
        log_internal_issue("获取活动 Chrome 账号失败", e)
    return active_paths

def get_profile_display_name(profile_id):
    """根据路径 ID 获取账户显示名称。
    
    [深度修复] 对 D 盘映射等非标准路径，不再盲目 basename 兜底（避免回退成误导性的 Default）。
    策略：
      1) 先尝试 get_chrome_profiles() 正常匹配；
      2) 若未命中，临时扫描该路径的父目录（user-data-dir），直接读 Preferences 提取账号信息；
      3) 最终兜底返回 '自定义路径账号 (parent/basename)' 而非纯 basename。
    """
    if not profile_id: return "未设置账户"
    # [增强] 允许在 open_url 的“账号/目标”位置填入窗口句柄（用于直接激活已打开的浏览器窗口）
    # 形如：xxx::hwnd=123456 或 ::hwnd=123456
    try:
        if isinstance(profile_id, str) and "::hwnd=" in profile_id:
            left = profile_id.split("::hwnd=", 1)[0].strip()
            return left if left else "已打开窗口"
    except Exception:
        pass
    # 如果是路径，尝试从路径中提取
    if "\\" in profile_id or "/" in profile_id:
        norm_id = os.path.normpath(profile_id)

        # 第一步：尝试已缓存的 get_chrome_profiles() 列表
        for prof_data in get_chrome_profiles():
            pid, pname = prof_data[0], prof_data[1]
            if os.path.normpath(pid) == norm_id:
                return pname

        # 第二步：临时扫描父目录（解决 D 盘映射路径不在自动探测 base_dirs 内的问题）
        parent_dir = os.path.dirname(norm_id)
        if parent_dir and os.path.exists(parent_dir):
            try:
                p_id = os.path.basename(norm_id)
                # 读 Local State 缓存备注
                temp_state_info = {}
                state_path = os.path.join(parent_dir, "Local State")
                if os.path.exists(state_path):
                    try:
                        with open(state_path, "r", encoding="utf-8", errors="ignore") as _sf:
                            _sd = json.load(_sf)
                        for _k, _v in _sd.get("profile", {}).get("info_cache", {}).items():
                            if _v.get("name"): temp_state_info[_k] = _v["name"]
                    except: pass
                p_name = temp_state_info.get(p_id, "")
                # 读 Preferences 提取邮箱/备注
                pref_name, email_info = "", ""
                pref_path = os.path.join(norm_id, "Preferences")
                if os.path.exists(pref_path):
                    try:
                        with open(pref_path, 'r', encoding='utf-8') as _pf:
                            _pref = json.load(_pf)
                        pref_name = _pref.get("profile", {}).get("name", "")
                        acc_info = _pref.get("account_info", [])
                        if isinstance(acc_info, list) and acc_info:
                            email_info = acc_info[0].get("email", "")
                        if not email_info:
                            email_info = _pref.get("google", {}).get("services", {}).get("last_username", "")
                    except: pass
                final_name = p_name if p_name else pref_name
                if not final_name or "Person " in final_name:
                    final_name = email_info if email_info else final_name
                if final_name and final_name != p_id:
                    parent_hint = os.path.basename(parent_dir)
                    return f"[{parent_hint}] {final_name}"
            except: pass

        # 第三步：最终兜底——返回有意义的路径描述，而非纯 basename
        p_basename = os.path.basename(norm_id)
        p_parent   = os.path.basename(os.path.dirname(norm_id))
        
        # [优化] 针对运行中的账号，如果 basename 太通用，尝试从路径中找特征
        if p_basename in ("Default", "") or p_basename.startswith("Profile "):
            # 检查父目录名是否像一个账号 ID（例如指纹浏览器的 ID）
            if len(p_parent) > 5:
                return f"账号 {p_parent}"
            return f"自定义路径账号 ({p_parent}/{p_basename})"
        return p_basename
    if profile_id == "Default": return "默认账户"
    return profile_id

WINDOW_CLASS_NAME_CACHE = {}
WINDOW_ACCOUNT_MARKER_CACHE = {}

def get_profile_tag(profile_path, profile_meta=None):
    """根据 profile 路径读取用户手动标签。"""
    if not profile_path:
        return ""
    profile_meta = profile_meta or {}
    if profile_path in profile_meta:
        return str(profile_meta.get(profile_path, {}).get("tag", "")).strip()
    for meta_path, meta in profile_meta.items():
        try:
            if is_same_path(profile_path, meta_path):
                return str(meta.get("tag", "")).strip()
        except Exception:
            continue
    return ""

def get_profile_brief_info(profile_id):
    """返回 profile 的基础资料，优先走缓存，缺失时再回退到本地文件。"""
    info = {"path": "", "name": "", "email": "", "remark": "", "id": ""}
    if not profile_id:
        return info

    profile_path = os.path.normpath(profile_id) if ("\\" in profile_id or "/" in profile_id) else profile_id
    info["path"] = profile_path
    info["id"] = os.path.basename(profile_path) if ("\\" in profile_id or "/" in profile_id) else profile_id

    if "\\" in profile_id or "/" in profile_id:
        for pid, pname, pemail, premark, prawid in get_chrome_profiles():
            try:
                if is_same_path(pid, profile_path):
                    return {
                        "path": pid,
                        "name": pname or "",
                        "email": pemail or "",
                        "remark": premark or "",
                        "id": prawid or os.path.basename(pid)
                    }
            except Exception:
                continue

        parent_dir = os.path.dirname(profile_path)
        p_id = os.path.basename(profile_path)
        temp_name, temp_email, temp_remark = "", "", ""
        if parent_dir and os.path.exists(parent_dir):
            state_path = os.path.join(parent_dir, "Local State")
            if os.path.exists(state_path):
                try:
                    with open(state_path, "r", encoding="utf-8", errors="ignore") as _sf:
                        _sd = json.load(_sf)
                    _state = _sd.get("profile", {}).get("info_cache", {}).get(p_id, {})
                    temp_name = _state.get("name", "") or temp_name
                    temp_email = _state.get("user_name", "") or _state.get("gaia_name", "") or temp_email
                    temp_remark = _state.get("shortcut_name", "") or temp_remark
                except Exception:
                    pass

        pref_path = os.path.join(profile_path, "Preferences")
        if os.path.exists(pref_path):
            try:
                with open(pref_path, 'r', encoding='utf-8', errors='ignore') as _pf:
                    _pref = json.load(_pf)
                temp_name = _pref.get("profile", {}).get("name", "") or temp_name
                acc_info = _pref.get("account_info", [])
                if isinstance(acc_info, list) and acc_info:
                    temp_email = acc_info[0].get("email", "") or temp_email
                if not temp_email:
                    temp_email = _pref.get("google", {}).get("services", {}).get("last_username", "") or temp_email
            except Exception:
                pass

        final_name = temp_name if temp_name and "Person " not in temp_name else (temp_email or temp_name)
        if final_name and parent_dir:
            parent_hint = os.path.basename(parent_dir)
            if parent_hint and "user data" not in parent_hint.lower() and parent_hint.lower() not in final_name.lower():
                final_name = f"[{parent_hint}] {final_name}"

        info.update({
            "name": final_name or "",
            "email": temp_email or "",
            "remark": temp_remark or "",
            "id": p_id or info["id"]
        })
    return info

def get_window_class_name(hwnd):
    """按 hwnd 获取窗口类名，用于快速判断是否为 Chrome 主窗口。"""
    if sys.platform != 'win32' or not hwnd:
        return ""
    try:
        hwnd = int(hwnd)
    except Exception:
        return ""
    if hwnd in WINDOW_CLASS_NAME_CACHE:
        return WINDOW_CLASS_NAME_CACHE[hwnd]
    class_name = ""
    try:
        buf = ctypes.create_unicode_buffer(256)
        ctypes.windll.user32.GetClassNameW(hwnd, buf, 255)
        class_name = buf.value or ""
    except Exception as e:
        log_internal_issue(f"读取窗口类名失败: hwnd={hwnd}", e)
    WINDOW_CLASS_NAME_CACHE[hwnd] = class_name
    return class_name

def get_window_profile_descriptor(hwnd):
    """按 hwnd 提取 Chrome 窗口对应的 profile 资料，并做缓存复用。"""
    if sys.platform != 'win32' or not hwnd:
        return {}
    try:
        hwnd = int(hwnd)
    except Exception:
        return {}
    if hwnd in WINDOW_PROFILE_INFO_CACHE:
        return dict(WINDOW_PROFILE_INFO_CACHE[hwnd])

    info = {}
    try:
        if get_window_class_name(hwnd) != "Chrome_WidgetWin_1":
            WINDOW_PROFILE_INFO_CACHE[hwnd] = {}
            return {}

        _pid = wintypes.DWORD()
        ctypes.windll.user32.GetWindowThreadProcessId(hwnd, ctypes.byref(_pid))
        pid = int(_pid.value)
        if not pid:
            WINDOW_PROFILE_INFO_CACHE[hwnd] = {}
            return {}

        cmd_out = subprocess.check_output(
            f'wmic process where processid={pid} get commandline',
            shell=True
        ).decode('gbk', errors='ignore')
        cmd_lower = cmd_out.lower()
        if "chrome.exe" not in cmd_lower and "chrome_proxy.exe" not in cmd_lower:
            WINDOW_PROFILE_INFO_CACHE[hwnd] = {}
            return {}

        u_path = extract_cmd_switch_value(cmd_out, "user-data-dir")
        p_dir = extract_cmd_switch_value(cmd_out, "profile-directory")
        if u_path:
            profile_path = os.path.normpath(os.path.join(u_path, p_dir)) if p_dir else os.path.normpath(u_path)
            info = get_profile_brief_info(profile_path)
            if not info.get("path"):
                info["path"] = profile_path
            if not info.get("id"):
                info["id"] = p_dir or os.path.basename(profile_path)
    except Exception as e:
        log_internal_issue(f"按 hwnd 识别 Chrome 账号失败: hwnd={hwnd}", e)

    WINDOW_PROFILE_INFO_CACHE[hwnd] = dict(info)
    return dict(info)

def build_compact_profile_text(tag="", remark="", display_name="", profile_id=""):
    """将标签/备注/名称压缩成更顺眼的紧凑文案。"""
    tag = (tag or "").strip()
    remark = (remark or "").strip()
    display_name = (display_name or "").strip()
    profile_id = (profile_id or "").strip()

    parts = []
    if tag:
        parts.append(f"🏷️{tag}")

    main_text = display_name
    if remark:
        if main_text:
            if remark.lower() not in main_text.lower():
                main_text = f"[{remark}] {main_text}"
        else:
            main_text = f"[{remark}]"

    if not main_text:
        if profile_id:
            if profile_id == "Default" or profile_id.startswith("Profile "):
                main_text = f"未命名账号 | {profile_id}"
            else:
                main_text = profile_id
        else:
            main_text = "未命名账号"

    parts.append(main_text)
    return " | ".join([p for p in parts if p])

def get_window_account_marker(hwnd, profile_meta=None):
    """根据窗口句柄识别 Chrome 账号标识，优先返回更有辨识度的文案。"""
    if sys.platform != 'win32' or not hwnd:
        return ""
    try:
        hwnd = int(hwnd)
    except Exception:
        return ""
    if hwnd in WINDOW_ACCOUNT_MARKER_CACHE:
        cached_marker = WINDOW_ACCOUNT_MARKER_CACHE[hwnd]
        if cached_marker:
            return cached_marker

    info = get_window_profile_descriptor(hwnd)
    tag = get_profile_tag(info.get("path", ""), profile_meta)
    remark = (info.get("remark") or "").strip()
    display_name = (info.get("name") or info.get("email") or "").strip()
    profile_id = (info.get("id") or "").strip()
    marker = build_compact_profile_text(tag, remark, display_name, profile_id) if info else ""
    WINDOW_ACCOUNT_MARKER_CACHE[hwnd] = marker or ""
    return WINDOW_ACCOUNT_MARKER_CACHE[hwnd]

def build_window_display_text(raw_title, hwnd=None, prefix="", profile_meta=None):
    """构造窗口显示文案：标题 + 紧凑账号标识 + 短 hwnd。"""
    title = (raw_title or "").strip()
    extras = []
    info = get_window_profile_descriptor(hwnd) if hwnd else {}
    tag = get_profile_tag(info.get("path", ""), profile_meta)
    remark = (info.get("remark") or "").strip()
    display_name = (info.get("name") or info.get("email") or "").strip()
    profile_id = (info.get("id") or "").strip()
    marker = build_compact_profile_text(tag, remark, display_name, profile_id) if info else ""
    if marker:
        extras.append(marker)
    if hwnd:
        extras.append(f"#{str(int(hwnd))[-4:]}")
    suffix = f" 〔{' | '.join(extras)}〕" if extras else ""
    return f"{prefix}{title}{suffix}"

def find_browser_window_hwnd_by_hint(target_hint):
    """按窗口标题/账号标识/展示文案匹配已打开浏览器窗口，返回 hwnd。"""
    if sys.platform != "win32" or not target_hint:
        return None
    hint = str(target_hint or "").strip()
    if not hint:
        return None
    hint_lower = hint.lower()
    seen_hwnds = set()
    try:
        from pywinauto import Desktop
        desktop = Desktop(backend="uia")
        for b_class in ["Chrome_WidgetWin_1", "MozillaWindowClass"]:
            for win in desktop.windows(class_name=b_class):
                try:
                    hwnd = int(getattr(win, "handle", 0) or 0)
                except Exception:
                    hwnd = 0
                if not hwnd or hwnd in seen_hwnds:
                    continue
                seen_hwnds.add(hwnd)

                title = str(get_window_text(hwnd) or "").strip()
                marker = str(get_window_account_marker(hwnd) or "").strip()
                display_text = str(build_window_display_text(title, hwnd) or "").strip()
                info = get_window_profile_descriptor(hwnd) or {}
                profile_path = str(info.get("path", "") or "").strip()
                profile_id = str(info.get("id", "") or "").strip()
                profile_name = str(info.get("name") or info.get("email") or "").strip()
                profile_display = get_profile_display_name(profile_path or profile_id) if (profile_path or profile_id) else ""

                candidates = [
                    title,
                    marker,
                    display_text,
                    profile_name,
                    profile_id,
                    profile_path,
                    profile_display,
                ]
                matched = any(
                    c and (hint_lower in c.lower() or c.lower() in hint_lower)
                    for c in candidates
                )
                if matched:
                    return hwnd

                try:
                    for tab in win.descendants(control_type="TabItem"):
                        tab_title = str(tab.window_text() or "").strip()
                        if tab_title and hint_lower in tab_title.lower():
                            return hwnd
                except Exception:
                    pass
    except Exception as e:
        log_internal_issue(f"按提示词匹配浏览器窗口失败: {hint}", e)
    return None

# --- Shutdown Countdown Dialog ---
class ShutdownDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent); self.setWindowTitle("⚠️ 自动关机提醒"); self.resize(350, 180)
        self.setWindowFlags(Qt.WindowStaysOnTopHint | Qt.Dialog)
        layout = QVBoxLayout(self)
        self.label = QLabel("任务已全部完成！\n电脑即将在 60 秒后关机..."); self.label.setAlignment(Qt.AlignCenter); self.label.setStyleSheet("font-size: 16px; font-weight: bold; color: #d32f2f;"); layout.addWidget(self.label)
        self.count = 60
        self.timer = QTimer(self); self.timer.timeout.connect(self._tick); self.timer.start(1000)
        btn = QPushButton("❌ 取消关机"); btn.setFixedHeight(40); btn.setStyleSheet("background-color: #757575; color: white; font-weight: bold;"); btn.clicked.connect(self.reject); layout.addWidget(btn)
    def _tick(self):
        self.count -= 1
        self.label.setText(f"任务已全部完成！\n电脑即将在 {self.count} 秒后关机...")
        if self.count <= 0: self.accept()

class ModernTaskManager(QDialog):
    def __init__(self, tasks, parent=None):
        super().__init__(parent); self.setWindowTitle("🚀 任务中心 - 快速管理与搜索"); self.resize(950, 650)
        self.tasks = list(tasks); self.current_filter = "全部"
        self.config = parent.config if parent else {}
        self.parent_obj = parent

        main_layout = QHBoxLayout(self)

        # 左侧管理区
        left_layout = QVBoxLayout()
        self.side_bar = QListWidget(); self.side_bar.setFixedWidth(180)
        self.side_bar.setAcceptDrops(True)
        self.side_bar.setContextMenuPolicy(Qt.CustomContextMenu)
        self.side_bar.customContextMenuRequested.connect(self._show_side_bar_context_menu)
        self.side_bar.setStyleSheet("""
            QListWidget { background-color: #f8f9fa; border: 1px solid #e0e0e0; border-radius: 8px; outline: none; }
            QListWidget::item { 
                height: 45px; padding-left: 15px; border-radius: 6px; margin: 2px 5px;
                color: #444;
            }
            QListWidget::item:selected { 
                background-color: #ffffff; color: #0078d4; font-weight: 600; 
                border-left: 4px solid #0078d4;
            }
            QListWidget::item:hover:not(:selected) { background-color: #e9ecef; }
        """)
        self.side_bar.currentTextChanged.connect(self._on_category_changed)
        self.side_bar.dragEnterEvent = self._side_bar_dragEnterEvent
        self.side_bar.dragMoveEvent = self._side_bar_dragMoveEvent
        self.side_bar.dropEvent = self._side_bar_dropEvent
        left_layout.addWidget(self.side_bar)

        btn_new_cat = QPushButton("➕ 新建分类"); btn_new_cat.setFixedHeight(38)
        btn_new_cat.setStyleSheet("background-color: #0078d4; color: white; border: none; font-weight: 600;")
        btn_new_cat.clicked.connect(self._new_category_only)
        left_layout.addWidget(btn_new_cat)

        btn_rename_cat = QPushButton("✏️ 重命名分类"); btn_rename_cat.setFixedHeight(38)
        btn_rename_cat.clicked.connect(self._rename_current_category)
        left_layout.addWidget(btn_rename_cat)

        btn_del_cat = QPushButton("🗑️ 删除分类"); btn_del_cat.setFixedHeight(38)
        btn_del_cat.setStyleSheet("background-color: #fee2e2; border: 1px solid #fecaca; font-weight: 600; color: #dc2626;")
        btn_del_cat.clicked.connect(self._delete_current_category)
        left_layout.addWidget(btn_del_cat)
        main_layout.addLayout(left_layout)

        # 右侧内容区
        content_layout = QVBoxLayout()
        self.search_edit = QLineEdit(); self.search_edit.setPlaceholderText("🔍 搜索任务名称、分类或关键字..."); self.search_edit.setFixedHeight(40)
        self.search_edit.setStyleSheet("""
            QLineEdit { 
                padding: 0 15px; border-radius: 20px; border: 1px solid #dcdcdc; 
                background: white; font-size: 14px;
            }
            QLineEdit:focus { border: 2px solid #0078d4; }
        """)
        self.search_edit.textChanged.connect(self._refresh_list)
        content_layout.addWidget(self.search_edit)

        self.task_list = QListWidget()
        self.task_list.setDragEnabled(True)
        self.task_list.setContextMenuPolicy(Qt.CustomContextMenu)
        self.task_list.customContextMenuRequested.connect(self._show_context_menu)
        self.task_list.setStyleSheet("""
            QListWidget { background-color: white; border: 1px solid #e0e0e0; border-radius: 8px; outline: none; }
            QListWidget::item { 
                height: 60px; padding: 10px 15px; border-bottom: 1px solid #f8f9fa; 
                border-radius: 6px; margin: 2px 5px; color: #333;
            }
            QListWidget::item:selected { background-color: #eef6fc; color: #0078d4; font-weight: 600; }
            QListWidget::item:hover:not(:selected) { background-color: #f8f9fa; }
        """)
        self.task_list.itemDoubleClicked.connect(self.accept)
        content_layout.addWidget(self.task_list)

        # 底部工具栏
        tool_h = QHBoxLayout()
        btn_move = QPushButton("📁 归类到..."); btn_move.clicked.connect(self._move_to_category); tool_h.addWidget(btn_move)
        btn_rename = QPushButton("✏️ 重命名"); btn_rename.clicked.connect(self._rename_selected_task); tool_h.addWidget(btn_rename)
        btn_clone = QPushButton("📋 克隆任务"); btn_clone.clicked.connect(self._clone_selected_task); tool_h.addWidget(btn_clone)
        tool_h.addStretch()
        btn_ok = QPushButton("✅ 进入选中任务"); btn_ok.setFixedWidth(150); btn_ok.setFixedHeight(40); btn_ok.setStyleSheet("background-color: #228be6; color: white; font-weight: bold; border-radius: 5px;"); btn_ok.clicked.connect(self.accept); tool_h.addWidget(btn_ok)
        content_layout.addLayout(tool_h)

        main_layout.addLayout(content_layout)
        self._init_categories()
        self._refresh_list()

    def _init_categories(self):
        old_cat = self.current_filter
        self.side_bar.clear(); self.side_bar.addItem("全部"); self.side_bar.addItem("未分类")
        cats = sorted(list(set([t.split("/", 1)[0] for t in self.tasks if "/" in t])))
        for c in cats:
            if c not in ["全部", "未分类"]: self.side_bar.addItem(c)
        items = self.side_bar.findItems(old_cat, Qt.MatchExactly)
        if items: self.side_bar.setCurrentItem(items[0])
        else: self.side_bar.setCurrentRow(0)

    def _new_category_only(self):
        name, ok = QInputDialog.getText(self, "新建分类", "请输入分类名称:")
        if ok and name.strip():
            new_cat = name.strip()
            if new_cat in ["全部", "未分类"]:
                QMessageBox.warning(self, "提示", "不能使用系统保留名称作为分类！")
                return
            exists = False
            for i in range(self.side_bar.count()):
                if self.side_bar.item(i).text() == new_cat:
                    exists = True; self.side_bar.setCurrentRow(i); break
            if not exists:
                self.side_bar.addItem(new_cat)
                self.side_bar.setCurrentItem(self.side_bar.item(self.side_bar.count() - 1))

    def _show_side_bar_context_menu(self, pos):
        item = self.side_bar.itemAt(pos)
        if not item: return
        cat = item.text()
        if cat in ["全部", "未分类"]: return
        menu = QMenu()
        rename_act = menu.addAction("✏️ 重命名当前分类")
        del_act = menu.addAction("🗑️ 删除当前分类")
        action = menu.exec_(self.side_bar.mapToGlobal(pos))
        if action == rename_act:
            self.side_bar.setCurrentItem(item)
            self._rename_current_category()
        elif action == del_act:
            self.side_bar.setCurrentItem(item)
            self._delete_current_category()

    def _rename_current_category(self):
        current_item = self.side_bar.currentItem()
        old_cat = current_item.text() if current_item else self.current_filter
        if old_cat in ["", "全部", "未分类"]:
            QMessageBox.information(self, "提示", "请选择一个可重命名的自定义分类。")
            return

        new_cat, ok = QInputDialog.getText(self, "重命名分类", f"请输入分类 '{old_cat}' 的新名称:", QLineEdit.Normal, old_cat)
        if not ok or not new_cat.strip() or new_cat == old_cat:
            return
        
        new_cat = new_cat.strip()
        if new_cat in ["全部", "未分类"]:
            QMessageBox.warning(self, "错误", "不能使用系统保留名称！")
            return
        
        # 检查新分类名是否已存在
        for i in range(self.side_bar.count()):
            if self.side_bar.item(i).text() == new_cat:
                QMessageBox.warning(self, "错误", f"分类 '{new_cat}' 已存在！")
                return

        affected_tasks = [t for t in self.tasks if t.startswith(f"{old_cat}/")]
        if not affected_tasks:
            # 只是空分类改名，直接修改侧边栏
            current_item.setText(new_cat)
            self.current_filter = new_cat
            return

        rename_map = {}
        for old_full_name in affected_tasks:
            pure_name = old_full_name.split("/", 1)[1]
            new_full_name = f"{new_cat}/{pure_name}"
            if new_full_name in self.config['tasks']:
                QMessageBox.warning(self, "重命名失败", f"重命名后任务 '{new_full_name}' 与现有任务冲突，请先处理重名任务。")
                return
            rename_map[old_full_name] = new_full_name

        self.current_filter = new_cat
        self._apply_rename_map(rename_map)

    def _delete_current_category(self):
        current_item = self.side_bar.currentItem()
        cat = current_item.text() if current_item else self.current_filter
        if cat in ["", "全部", "未分类"]:
            QMessageBox.information(self, "提示", "请选择一个可删除的自定义分类。")
            return

        affected_tasks = [t for t in self.tasks if t.startswith(f"{cat}/")]
        if not affected_tasks:
            row = self.side_bar.row(current_item) if current_item else -1
            if row >= 0: self.side_bar.takeItem(row)
            self.current_filter = "全部"
            self._refresh_list()
            return

        rename_map = {}
        conflicts = []
        affected_set = set(affected_tasks)
        for old_name in affected_tasks:
            pure_name = old_name.split("/", 1)[1]
            if pure_name in self.config['tasks'] and pure_name not in affected_set:
                conflicts.append(pure_name)
            else:
                rename_map[old_name] = pure_name

        if conflicts:
            QMessageBox.warning(self, "无法删除分类", "以下任务名称在根目录已存在，无法自动移出分类：\n\n" + "\n".join(conflicts))
            return

        reply = QMessageBox.question(
            self,
            "确认删除分类",
            f"确定删除分类 '{cat}' 吗？\n\n该分类下共 {len(affected_tasks)} 个任务，删除后会自动移动到根目录。",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return

        self.current_filter = "全部"
        self._apply_rename_map(rename_map)

    def _apply_rename_map(self, rename_map):
        if not rename_map: return
        if self.parent_obj and hasattr(self.parent_obj, '_rename_tasks_in_config'):
            self.parent_obj._rename_tasks_in_config(rename_map)
            preferred_task = rename_map.get(self.parent_obj.current_task, self.parent_obj.current_task)
            self.parent_obj._reload_task_combo_after_config_change(preferred_task)
            if hasattr(self.parent_obj, '_refresh_schedule_task_options'):
                self.parent_obj._refresh_schedule_task_options(rename_map=rename_map)
        else:
            self.config['tasks'] = {rename_map.get(k, k): v for k, v in self.config['tasks'].items()}
            self.config['task_data'] = {rename_map.get(k, k): v for k, v in self.config['task_data'].items()}
            save_config(self.config)
        self.tasks = list(self.config['tasks'].keys())
        self._init_categories()
        self._refresh_list()

    def _on_category_changed(self, cat):
        self.current_filter = cat
        self._refresh_list()

    def _refresh_list(self):
        self.task_list.clear(); search_text = self.search_edit.text().lower()
        for t in self.tasks:
            cat = t.split("/", 1)[0] if "/" in t else "未分类"
            display_name = t.split("/", 1)[1] if "/" in t else t

            if self.current_filter != "全部" and cat != self.current_filter: continue
            if search_text and search_text not in t.lower(): continue

            item = QListWidgetItem(f"📄 {display_name}")
            item.setToolTip(f"完整路径: {t}")
            item.setData(Qt.UserRole, t)
            self.task_list.addItem(item)

        if self.task_list.count() > 0: self.task_list.setCurrentRow(0)

    def _side_bar_dragEnterEvent(self, event):
        if event.mimeData().hasFormat("application/x-qabstractitemmodeldatalist"): event.accept()
        else: event.ignore()

    def _side_bar_dragMoveEvent(self, event):
        if event.mimeData().hasFormat("application/x-qabstractitemmodeldatalist"):
            item = self.side_bar.itemAt(event.pos())
            if item: self.side_bar.setCurrentItem(item)
            event.accept()
        else: event.ignore()

    def _side_bar_dropEvent(self, event):
        target_item = self.side_bar.itemAt(event.pos())
        if not target_item: return
        target_cat = target_item.text()

        source_item = self.task_list.currentItem()
        if not source_item: return
        old_full_name = source_item.data(Qt.UserRole)
        old_pure_name = old_full_name.split("/", 1)[1] if "/" in old_full_name else old_full_name

        new_full_name = old_pure_name if target_cat in ["全部", "未分类"] else f"{target_cat}/{old_pure_name}"
        if new_full_name != old_full_name:
            self._do_rename(old_full_name, new_full_name)
        event.accept()

    def _show_context_menu(self, pos):
        item = self.task_list.itemAt(pos)
        if not item: return
        menu = QMenu()
        move_act = menu.addAction("📁 移动到分类...")
        rename_act = menu.addAction("✏️ 重命名任务")
        clone_act = menu.addAction("📋 克隆任务")
        action = menu.exec_(self.task_list.mapToGlobal(pos))
        if action == move_act: self._move_to_category()
        elif action == rename_act: self._rename_selected_task()
        elif action == clone_act: self._clone_selected_task()

    def _move_to_category(self):
        item = self.task_list.currentItem()
        if not item: return
        old_full_name = item.data(Qt.UserRole)
        old_pure_name = old_full_name.split("/", 1)[1] if "/" in old_full_name else old_full_name

        cats = ["(根目录)"]
        for i in range(self.side_bar.count()):
            c = self.side_bar.item(i).text()
            if c not in ["全部", "未分类"] and c not in cats:
                cats.append(c)

        cat, ok = QInputDialog.getItem(self, "移动任务", f"将任务 '{old_pure_name}' 移动到:", cats, 0, False)
        if ok:
            new_full_name = old_pure_name if cat == "(根目录)" else f"{cat}/{old_pure_name}"
            if new_full_name != old_full_name:
                self._do_rename(old_full_name, new_full_name)

    def _new_category(self):
        item = self.task_list.currentItem()
        if not item: return
        old_full_name = item.data(Qt.UserRole)
        old_pure_name = old_full_name.split("/", 1)[1] if "/" in old_full_name else old_full_name

        cat, ok = QInputDialog.getText(self, "新建分类", "请输入新分类名称:")
        if ok and cat.strip():
            new_full_name = f"{cat.strip()}/{old_pure_name}"
            self._do_rename(old_full_name, new_full_name)

    def _clone_selected_task(self):
        item = self.task_list.currentItem()
        if not item: return
        old_full_name = item.data(Qt.UserRole)
        cat_prefix = f"{old_full_name.split('/', 1)[0]}/" if "/" in old_full_name else ""
        old_pure_name = old_full_name.split('/', 1)[1] if "/" in old_full_name else old_full_name

        new_name, ok = QInputDialog.getText(self, "克隆任务", f"请输入新任务名称 (基于 '{old_pure_name}' 克隆):", QLineEdit.Normal, f"{old_pure_name}_副本")
        if not ok or not new_name.strip():
            return
        
        new_full_name = f"{cat_prefix}{new_name.strip()}"

        if new_full_name in self.config['tasks']:
            QMessageBox.warning(self, "错误", "目标任务名称已存在！")
            return
        
        # 复制任务配置
        self.config['tasks'][new_full_name] = self.config['tasks'][old_full_name]
        # 复制任务数据（如果存在）
        if old_full_name in self.config['task_data']:
            self.config['task_data'][new_full_name] = self.config['task_data'][old_full_name]
        
        save_config(self.config)
        self.tasks = list(self.config['tasks'].keys())
        self._init_categories()
        self._refresh_list()
        
        # 选中新克隆的任务
        new_item = self.task_list.findItems(new_full_name, Qt.MatchExactly)
        if new_item: self.task_list.setCurrentItem(new_item[0])
        
    def _rename_selected_task(self):
        item = self.task_list.currentItem()
        if not item: return
        old_full_name = item.data(Qt.UserRole)
        cat_prefix = f"{old_full_name.split('/', 1)[0]}/" if "/" in old_full_name else ""
        old_pure_name = old_full_name.split("/", 1)[1] if "/" in old_full_name else old_full_name

        new_name, ok = QInputDialog.getText(self, "重命名", "请输入新任务名称:", QLineEdit.Normal, old_pure_name)
        if ok and new_name.strip():
            self._do_rename(old_full_name, f"{cat_prefix}{new_name.strip()}")

    def _do_rename(self, old_name, new_name):
        if not new_name or new_name == old_name: return
        if new_name in self.config['tasks']:
            QMessageBox.warning(self, "错误", "目标名称已存在！")
            return
        self._apply_rename_map({old_name: new_name})

    def get_selection(self):
        if self.exec_() == QDialog.Accepted:
            item = self.task_list.currentItem()
            if item: return item.data(Qt.UserRole)
        return None

# --- Chrome Profile Selector Dialog ---
class ChromeProfileSelector(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent); self.setWindowTitle("👤 选择 Chrome 账户"); self.resize(550, 650)
        layout = QVBoxLayout(self)
        
        # 顶部筛选区
        filter_h = QHBoxLayout()
        self.search_edit = QLineEdit(); self.search_edit.setPlaceholderText("🔍 搜索名称或 ID..."); self.search_edit.textChanged.connect(self._filter_list)
        filter_h.addWidget(self.search_edit)
        
        self.chk_only_active = QCheckBox("仅显示已打开的账户")
        self.chk_only_active.stateChanged.connect(lambda: self._filter_list(self.search_edit.text()))
        filter_h.addWidget(self.chk_only_active)
        layout.addLayout(filter_h)

        # [新增] 自定义目录管理区
        custom_h = QHBoxLayout()
        btn_add_dir = QPushButton("📂 添加自定义数据目录 (User Data)"); btn_add_dir.setFixedHeight(30)
        btn_add_dir.setStyleSheet("background-color: #f1f3f5; border: 1px solid #dee2e6; font-size: 12px;")
        btn_add_dir.clicked.connect(self._add_custom_dir)
        custom_h.addWidget(btn_add_dir)
        
        btn_clear_extra = QPushButton("🗑️ 清空手动添加"); btn_clear_extra.setFixedWidth(100); btn_clear_extra.setFixedHeight(30)
        btn_clear_extra.setStyleSheet("color: #c92a2a; font-size: 11px;")
        btn_clear_extra.clicked.connect(self._clear_extra_dirs)
        custom_h.addWidget(btn_clear_extra)
        layout.addLayout(custom_h)

        self.list_widget = QListWidget()
        self.list_widget.setStyleSheet("""
            QListWidget::item { height: 45px; padding: 5px; color: black; border-bottom: 1px solid #eee; }
            QListWidget::item:selected { background-color: #e3f2fd; color: #0078d7; }
        """)
        layout.addWidget(self.list_widget)
        self.lbl_selection_count = create_table_selection_label()
        layout.addWidget(self.lbl_selection_count)
        bind_item_view_selection_label(self.list_widget, self.lbl_selection_count, kind_text="个账户")
        
        self.profiles = get_chrome_profiles(force_refresh=True)
        self.active_pids = get_active_chrome_profiles()
        self._filter_list("")
        
        btn_h = QHBoxLayout()
        btn_refresh = QPushButton("🔄 刷新状态"); btn_refresh.clicked.connect(self._refresh_status)
        btn_h.addWidget(btn_refresh)
        btn_h.addStretch()
        self.btn_select = QPushButton("✅ 确定选择"); self.btn_select.setStyleSheet("background-color: #0078d7; color: white; font-weight: bold; width: 120px; height: 35px;"); self.btn_select.clicked.connect(self.accept)
        btn_h.addWidget(self.btn_select)
        layout.addLayout(btn_h)

    def _add_custom_dir(self):
        path = QFileDialog.getExistingDirectory(self, "选择 Chrome User Data 目录", "")
        if path:
            norm_path = os.path.normpath(path)
            if norm_path not in _extra_scan_dirs:
                _extra_scan_dirs.add(norm_path)
                # 尝试保存到配置文件（如果 parent 是主窗口且有 config）
                if self.parent() and hasattr(self.parent(), 'config'):
                    cfg = self.parent().config
                    if 'extra_scan_dirs' not in cfg: cfg['extra_scan_dirs'] = []
                    if norm_path not in cfg['extra_scan_dirs']:
                        cfg['extra_scan_dirs'].append(norm_path)
                        save_config(cfg)
                
                clear_chrome_profile_cache()
                self.profiles = get_chrome_profiles(force_refresh=True)
                self._filter_list(self.search_edit.text())
                QMessageBox.information(self, "成功", f"已添加目录: {norm_path}\n现在可以识别该目录下的账户了。")

    def _clear_extra_dirs(self):
        if not _extra_scan_dirs: return
        reply = QMessageBox.question(self, "确认", "确定要清空所有手动添加的扫描目录吗？", QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            _extra_scan_dirs.clear()
            if self.parent() and hasattr(self.parent(), 'config'):
                cfg = self.parent().config
                cfg['extra_scan_dirs'] = []
                save_config(cfg)
            clear_chrome_profile_cache()
            self.profiles = get_chrome_profiles(force_refresh=True)
            self._filter_list(self.search_edit.text())

    def _refresh_status(self):
        self.active_pids = get_active_chrome_profiles()
        clear_chrome_profile_cache()
        self.profiles = get_chrome_profiles(force_refresh=True) # 刷新时重新扫描物理目录
        self._filter_list(self.search_edit.text())

    def _filter_list(self, text):
        self.list_widget.clear()
        search_text = text.strip().lower()
        only_active = self.chk_only_active.isChecked()
        
        for prof_data in self.profiles:
            pid, pname = prof_data[0], prof_data[1]
            # [修复] 使用 is_same_path 进行路径容错比对，防止因盘符映射、大小写、反斜杠导致的漏判
            is_active = any(is_same_path(pid, ap) for ap in self.active_pids)
            if only_active and not is_active: continue
            
            if not search_text or search_text in pname.lower() or search_text in pid.lower():
                item = QListWidgetItem(pname)
                if is_active:
                    item.setText(f"🟢 {pname} [已打开]")
                    item.setForeground(QColor("#2e7d32"))
                else:
                    item.setText(f"⚪ {pname}")
                    item.setForeground(QColor("#757575"))
                
                item.setData(Qt.UserRole, pid)
                self.list_widget.addItem(item)
        
        if self.list_widget.count() > 0: self.list_widget.setCurrentRow(0)

    def get_selection(self):
        if self.exec_() == QDialog.Accepted:
            item = self.list_widget.currentItem()
            return item.data(Qt.UserRole) if item else ""
        return None

# --- Command Mapping ---
CMD_MAP = {
    "📷 图像识别点击": "image_click",
    "❓ 如果找图成功": "if_image",
    "🪟 如果窗口存在": "if_win",
    "左键点击": "click", 
    "双击": "double_click", 
    "右键点击": "right_click", 
    "移动鼠标": "move",
    "悬停点击": "hover_click",
    "输入文本": "input", 
    "清空并输入": "clear_input",
    "✨ 清空并输入(增强版)": "clear_input_plus",
    "上传文件": "upload", 
    "🖱️ 拖拽文件": "drag_file",
    "单按键": "press", 
    "组合键": "hotkey", 
    "滚轮滚动": "scroll",
    "激活窗口": "win_active", 
    "🌐 打开网址": "open_url",
    "运行程序": "run_app",
    "屏幕截图": "screenshot",
    "⏸️ 延后执行": "defer",
    "等待": "wait",
    "💻 CMD 指令": "cmd"
}

# --- Hotkey Recorder Widget ---
class KeyRecorder(QLineEdit):
    key_recorded = pyqtSignal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._saved_text = ""
        self._temporary_text = False
        self.setPlaceholderText("点击后按快捷键，例如 Ctrl+C；Esc/Backspace 可清空")
        self.setReadOnly(True)
        self.setFocusPolicy(Qt.StrongFocus)
        self.setToolTip("点击输入框后直接按快捷键。单独按 Ctrl/Shift/Alt/Win 不会保存，请继续按主键；按 Esc 或 Backspace 可清空后重新录制。")
        self.setStyleSheet("background-color: #fffde7; border: 1px solid #fbc02d; font-weight: bold; color: #f57f17;")

    def setText(self, text):
        super().setText(text)
        if not self._temporary_text:
            self._saved_text = text

    def focusInEvent(self, event):
        super().focusInEvent(event)
        self.selectAll()

    def focusOutEvent(self, event):
        if self._temporary_text:
            self._temporary_text = False
            super().setText(self._saved_text)
        super().focusOutEvent(event)

    def _set_temporary_text(self, text):
        self._temporary_text = True
        super().setText(text)

    def _save_recorded_text(self, text):
        self._temporary_text = False
        self._saved_text = text
        super().setText(text)
        self.key_recorded.emit(text)

    def keyPressEvent(self, event):
        key = event.key()
        modifiers = event.modifiers()

        if key in (Qt.Key_Escape, Qt.Key_Backspace, Qt.Key_Delete):
            self._save_recorded_text("")
            event.accept()
            return

        modifier_names = []
        if modifiers & Qt.ControlModifier: modifier_names.append("ctrl")
        if modifiers & Qt.ShiftModifier: modifier_names.append("shift")
        if modifiers & Qt.AltModifier: modifier_names.append("alt")
        if modifiers & Qt.MetaModifier: modifier_names.append("win")

        modifier_only_keys = {Qt.Key_Control, Qt.Key_Shift, Qt.Key_Alt, Qt.Key_Meta}
        if key in modifier_only_keys:
            if modifier_names:
                self._set_temporary_text("+".join(modifier_names) + "+")
            event.accept()
            return

        # Mapping Qt keys to pyautogui names
        key_map = {Qt.Key_Enter: "enter", Qt.Key_Return: "enter", Qt.Key_Tab: "tab", Qt.Key_Space: "space",
                   Qt.Key_Up: "up", Qt.Key_Down: "down", Qt.Key_Left: "left", Qt.Key_Right: "right",
                   Qt.Key_F1: "f1", Qt.Key_F2: "f2", Qt.Key_F3: "f3", Qt.Key_F4: "f4", Qt.Key_F5: "f5",
                   Qt.Key_F6: "f6", Qt.Key_F7: "f7", Qt.Key_F8: "f8", Qt.Key_F9: "f9", Qt.Key_F10: "f10",
                   Qt.Key_F11: "f11", Qt.Key_F12: "f12", Qt.Key_Home: "home", Qt.Key_End: "end",
                   Qt.Key_PageUp: "pageup", Qt.Key_PageDown: "pagedown", Qt.Key_Insert: "insert"}

        main_key = key_map.get(key)
        if not main_key:
            text = event.text()
            if 32 <= key <= 126:
                main_key = chr(key).lower()
            elif text and text.isprintable():
                main_key = text.lower()
            else:
                main_key = ""

        keys = modifier_names[:]
        if main_key and main_key not in keys:
            keys.append(main_key)

        final_keys = "+".join([k for k in keys if k])
        if final_keys:
            self._save_recorded_text(final_keys)
        event.accept()

class MultiLineTextEdit(QTextEdit):
    editingFinished = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptRichText(False)
        self.setWordWrapMode(QTextOption.WrapAnywhere)
        self.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self._last_committed_text = ""

    def text(self):
        return self.toPlainText()

    def setText(self, text):
        text = "" if text is None else str(text)
        self.blockSignals(True)
        self.setPlainText(text)
        self.blockSignals(False)
        self._last_committed_text = text

    def focusOutEvent(self, event):
        current_text = self.toPlainText()
        if current_text != self._last_committed_text:
            self._last_committed_text = current_text
            self.editingFinished.emit()
        super().focusOutEvent(event)

def _find_text_input(widget):
    if widget is None:
        return None
    if isinstance(widget, (QLineEdit, MultiLineTextEdit)):
        return widget
    editor = widget.findChild(MultiLineTextEdit)
    if editor:
        return editor
    editor = widget.findChild(QLineEdit)
    if editor:
        return editor
    return None

# --- Improved Window Selector Dialog ---
class WindowSelector(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent); self.setWindowTitle("🔍 窗口快速选择器"); self.resize(600, 700)
        layout = QVBoxLayout(self)
        search_h = QHBoxLayout(); search_h.addWidget(QLabel("搜索窗口:"))
        self.search_edit = QLineEdit(); self.search_edit.setPlaceholderText("输入关键字过滤窗口标题、账号标识或浏览器标签..."); self.search_edit.textChanged.connect(self._filter_list); search_h.addWidget(self.search_edit); layout.addLayout(search_h)
        self.list_widget = QListWidget()
        self.list_widget.setStyleSheet("""
            QListWidget::item { height: 45px; padding: 5px; color: black; border-bottom: 1px solid #eee; }
            QListWidget::item:selected { background-color: #e3f2fd; color: #0078d7; }
        """)
        layout.addWidget(self.list_widget)
        self.lbl_selection_count = create_table_selection_label()
        layout.addWidget(self.lbl_selection_count)
        bind_item_view_selection_label(self.list_widget, self.lbl_selection_count, kind_text="个窗口")
        btn_h = QHBoxLayout()
        btn_refresh = QPushButton("🔄 刷新列表"); btn_refresh.clicked.connect(self._refresh_list); btn_h.addWidget(btn_refresh)
        btn_pick = QPushButton("🎯 瞄准拾取 (3秒倒计时)"); btn_pick.setStyleSheet("background-color: #ff9800; color: white;"); btn_pick.clicked.connect(self._start_pick); btn_h.addWidget(btn_pick)
        self.btn_select = QPushButton("✅ 确定选择"); self.btn_select.setStyleSheet("background-color: #0078d7; color: white; font-weight: bold;"); self.btn_select.clicked.connect(self.accept); btn_h.addWidget(self.btn_select)
        layout.addLayout(btn_h)
        self.all_items = []
        self._refresh_list()

    def _get_profile_meta(self):
        parent = self.parent()
        if parent and hasattr(parent, "config") and isinstance(parent.config, dict):
            return parent.config.get("profile_meta", {}) or {}
        return {}

    def _get_browser_tabs(self):
        """
        [修复] 增强型浏览器标签页获取逻辑。
        返回 (display_text, raw_title, hwnd) 元组列表，保留每个标签页所属窗口的 hwnd 唯一标识。
        """
        tabs = []  # 存储 (display_text, raw_title, hwnd) 元组
        if sys.platform != 'win32': return tabs
        profile_meta = self._get_profile_meta()
        try:
            from pywinauto import Desktop
            desktop = Desktop(backend="uia")
            
            for b_class in ["Chrome_WidgetWin_1", "MozillaWindowClass"]:
                wins = desktop.windows(class_name=b_class)
                for win in wins:
                    try:
                        win_hwnd = win.handle  # 获取该浏览器窗口的 hwnd
                        tab_items = win.descendants(control_type="TabItem")
                        for tab in tab_items:
                            try:
                                name = tab.window_text()
                                if name and name.strip() and name != "新标签页":
                                    # [修复] 将 hwnd 一并存入元组，同名标签页可通过账号标识 + hwnd 区分所属窗口
                                    tabs.append((build_window_display_text(name, win_hwnd, "[浏览器标签] ", profile_meta), name, win_hwnd))
                            except: continue
                        
                        if not tabs:
                            btns = win.descendants(control_type="Button")
                            for btn in btns:
                                try:
                                    name = btn.window_text()
                                    if name and len(name) > 2 and "Google Chrome" not in name:
                                        tabs.append((build_window_display_text(name, win_hwnd, "[浏览器标签] ", profile_meta), name, win_hwnd))
                                except: continue
                    except: continue
        except: pass
        
        tabs.sort(key=lambda x: x[0])
        return tabs

    def _filter_list(self, text):
        """[修复] 适配 all_items 改为 (display_text, raw_title, hwnd) 元组后的渲染逻辑。"""
        self.list_widget.clear()
        search_text = text.strip().lower()
        for entry in self.all_items:
            if isinstance(entry, tuple) and len(entry) >= 3:
                display_text, raw_title, hwnd = entry
            elif isinstance(entry, tuple):
                display_text, hwnd = entry
                raw_title = display_text
            else:
                display_text, raw_title, hwnd = entry, entry, None
            if search_text and search_text not in display_text.lower():
                continue
            item = QListWidgetItem(display_text)
            item.setData(Qt.UserRole, raw_title)      # 原始标题
            item.setData(Qt.UserRole + 1, hwnd)       # hwnd 唯一标识
            self.list_widget.addItem(item)
        if self.list_widget.count() > 0:
            self.list_widget.setCurrentRow(0)

    def _refresh_list(self):
        """[修复] 将每个窗口信息存为 (display_text, raw_title, hwnd) 元组，保留同名窗口的唯一标识。"""
        # all_items 改为存储 (display_text, raw_title, hwnd_or_None) 元组
        clear_window_profile_caches()
        win_items = []
        profile_meta = self._get_profile_meta()
        for w in gw.getAllWindows():
            if w.title and w.title.strip() and w.width > 0 and w.height > 0:
                hwnd = getattr(w, '_hWnd', None)  # pygetwindow 在 Windows 上有 _hWnd
                win_items.append((build_window_display_text(w.title, hwnd, "[窗口] ", profile_meta), w.title, hwnd))
        win_items.sort(key=lambda x: x[0])
        browser_tabs = self._get_browser_tabs()  # 返回 (display_text, raw_title, hwnd) 列表
        # 不去重，保留所有同名窗口
        self.all_items = win_items + browser_tabs
        self._filter_list(self.search_edit.text())

    def _start_pick(self):
        self.btn_select.setText("请切换到目标窗口..."); self.btn_select.setEnabled(False)
        QTimer.singleShot(3000, self._do_pick)

    def _do_pick(self):
        active_win = gw.getActiveWindow()
        if active_win:
            title = active_win.title
            hwnd = getattr(active_win, '_hWnd', None)
            if title:
                self._refresh_list()
                self.search_edit.setText(title)
                # 优先按 hwnd 匹配，确保选中正确的那个窗口
                matched = None
                for i in range(self.list_widget.count()):
                    it = self.list_widget.item(i)
                    if hwnd and it.data(Qt.UserRole + 1) == hwnd:
                        matched = it; break
                if not matched:
                    items = self.list_widget.findItems(title, Qt.MatchContains)
                    if items: matched = items[0]
                if matched: self.list_widget.setCurrentItem(matched)
        self.btn_select.setText("✅ 确定选择"); self.btn_select.setEnabled(True)

    def get_selection(self):
        """[修复] 返回带 hwnd 唯一标识的字符串格式: “标题::hwnd=12345”。
        执行层解析时优先用 hwnd 直接定位窗口，彻底解决同名窗口无法区分的问题。"""
        if self.exec_() == QDialog.Accepted:
            item = self.list_widget.currentItem()
            if item:
                raw_title = item.data(Qt.UserRole) or item.text()
                # 取出存储在 UserRole+1 中的 hwnd
                hwnd = item.data(Qt.UserRole + 1)
                if hwnd:
                    # 带 hwnd 唯一标识，执行层可直接定位
                    return f"{raw_title}::hwnd={hwnd}"
                return raw_title
        return ""

# --- Failure/Subtask Manager Dialog ---
class FailureManagerDialog(QDialog):
    """可复用的“子任务管理器”对话框：
    - 默认用于“失败任务管理器”（传入失败项列表）
    - 也可用于对“勾选的数据行”执行子任务（传入勾选条目列表）
    """
    def __init__(self, parent=None, items=None, title="❌ 失败任务管理器", context_task_id=""):
        super().__init__(parent)
        self.setWindowTitle(str(title or "❌ 失败任务管理器"))
        self.resize(980, 720)
        self.items = list(items or [])
        self.context_task_id = str(context_task_id or "")
        self._runtime_entry_key = None
        self._runtime_parent_step_idx = None
        self._runtime_running = False

        layout = QVBoxLayout(self)

        top = QHBoxLayout()
        top.addWidget(QLabel("搜索:"))
        self.ed_search = QLineEdit()
        self.ed_search.setPlaceholderText("按任务名 / 行号 / 步骤名 / 错误 / 窗口标题过滤…")
        self.ed_search.textChanged.connect(self._refresh)
        top.addWidget(self.ed_search, 1)
        self.chk_browser_only = QCheckBox("仅浏览器窗口")
        self.chk_browser_only.stateChanged.connect(self._refresh)
        top.addWidget(self.chk_browser_only)
        self.btn_items_all = QPushButton("☑ 全选条目")
        self.btn_items_all.clicked.connect(lambda: self._set_filtered_entry_checks(True))
        top.addWidget(self.btn_items_all)
        self.btn_items_none = QPushButton("☐ 清空条目")
        self.btn_items_none.clicked.connect(lambda: self._set_filtered_entry_checks(False))
        top.addWidget(self.btn_items_none)
        layout.addLayout(top)

        self.grp_parent_steps = QGroupBox("前置步骤")
        parent_ly = QVBoxLayout(self.grp_parent_steps)
        self.lbl_parent_steps = QLabel("从当前父任务里勾选要复用的步骤。这些步骤会先于“子任务”执行，并且继续使用当前行自己的数据。")
        self.lbl_parent_steps.setWordWrap(True)
        parent_ly.addWidget(self.lbl_parent_steps)

        parent_ctrl = QHBoxLayout()
        self.btn_parent_steps_all = QPushButton("☑ 全选前置步骤")
        self.btn_parent_steps_all.clicked.connect(lambda: self._set_parent_step_checks(True))
        parent_ctrl.addWidget(self.btn_parent_steps_all)
        self.btn_parent_steps_none = QPushButton("☐ 清空前置步骤")
        self.btn_parent_steps_none.clicked.connect(lambda: self._set_parent_step_checks(False))
        parent_ctrl.addWidget(self.btn_parent_steps_none)
        parent_ctrl.addStretch()
        parent_ly.addLayout(parent_ctrl)

        self.list_parent_steps = QListWidget()
        self.list_parent_steps.setMaximumHeight(150)
        parent_ly.addWidget(self.list_parent_steps)

        self.chk_continue_parent = QCheckBox("执行完前置步骤/子任务后，继续父任务")
        self.chk_continue_parent.setToolTip("勾选后，执行顺序会变成：前置步骤 → 子任务 → 父任务原流程。")
        parent_ly.addWidget(self.chk_continue_parent)
        layout.addWidget(self.grp_parent_steps)

        self._load_parent_steps()

        self.table_widget = QTableWidget(0, 8)
        self.table_widget.setHorizontalHeaderLabels(["选择", "执行", "状态", "任务", "行号", "步骤", "错误", "窗口"])
        self.table_widget.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.table_widget.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table_widget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table_widget.verticalHeader().setDefaultSectionSize(30)
        self.table_widget.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table_widget.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.table_widget.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.table_widget.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.table_widget.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        self.table_widget.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeToContents)
        self.table_widget.horizontalHeader().setSectionResizeMode(6, QHeaderView.Stretch)
        self.table_widget.horizontalHeader().setSectionResizeMode(7, QHeaderView.Stretch)
        self.table_widget.itemChanged.connect(self._on_table_item_changed)
        layout.addWidget(self.table_widget, 1)

        bottom = QHBoxLayout()
        self.btn_activate = QPushButton("🪟 激活窗口")
        self.btn_activate.clicked.connect(self._activate_selected)
        bottom.addWidget(self.btn_activate)

        self.btn_jump = QPushButton("🎯 跳到任务/行")
        self.btn_jump.clicked.connect(self._jump_selected)
        bottom.addWidget(self.btn_jump)

        self.btn_copy = QPushButton("📋 复制窗口列表")
        self.btn_copy.clicked.connect(self._copy_browser_windows)
        bottom.addWidget(self.btn_copy)

        bottom.addStretch()

        self.chk_auto_activate = QCheckBox("自动激活窗口")
        self.chk_auto_activate.setToolTip("如果条目里捕获到了 hwnd，则在执行子任务前自动插入“激活窗口(::hwnd=xxx)”步骤。")
        self.chk_auto_activate.setChecked(True)
        bottom.addWidget(self.chk_auto_activate)

        self.chk_highlight_running = QCheckBox("执行中高亮当前项")
        self.chk_highlight_running.setToolTip("执行子任务时，在列表里高亮当前正在跑的条目，并同步标出当前前置步骤。")
        self.chk_highlight_running.setChecked(True)
        self.chk_highlight_running.stateChanged.connect(self._apply_runtime_marks)
        bottom.addWidget(self.chk_highlight_running)

        self.chk_scroll_running = QCheckBox("滚动到当前项")
        self.chk_scroll_running.setToolTip("执行子任务时，自动把列表滚动到当前正在执行的条目/前置步骤。")
        self.chk_scroll_running.setChecked(True)
        self.chk_scroll_running.stateChanged.connect(self._apply_runtime_marks)
        bottom.addWidget(self.chk_scroll_running)

        self.chk_only_selected = QCheckBox("仅执行勾选条目")
        self.chk_only_selected.setToolTip("勾选后，只对表格里已勾选的条目执行；不勾选则对当前列表(过滤后的全部)执行。")
        self.chk_only_selected.setChecked(False)
        bottom.addWidget(self.chk_only_selected)

        self.cb_subtask = QComboBox()
        self.cb_subtask.setMinimumWidth(260)
        self.cb_subtask.addItem("(不追加子任务)", "")
        try:
            p = self.parent()
            if p and hasattr(p, "_get_task_names"):
                for tid in p._get_task_names():
                    self.cb_subtask.addItem(p._get_task_display_text(tid, with_folder=True), tid)
        except Exception:
            pass
        bottom.addWidget(self.cb_subtask)

        self.btn_new_subtask = QPushButton("➕ 新建子任务(录制)")
        self.btn_new_subtask.setToolTip("基于当前选中的条目创建一个新的子任务，并尽量自动带上“激活窗口(hwnd)”作为第1步，方便你继续录制后续动作。")
        self.btn_new_subtask.clicked.connect(self._create_new_subtask_task)
        bottom.addWidget(self.btn_new_subtask)

        # 用图标代替文字，避免窗口较窄时按钮文字被截断看不见
        self.btn_run_subtask = QPushButton("")
        try:
            self.btn_run_subtask.setIcon(self.style().standardIcon(QStyle.SP_MediaPlay))
            self.btn_run_subtask.setIconSize(QSize(16, 16))
        except Exception:
            pass
        self.btn_run_subtask.setFixedWidth(34)
        self.btn_run_subtask.setToolTip("执行子任务（对当前列表/勾选条目批量执行）")
        self.btn_run_subtask.clicked.connect(self._run_subtask_selected_or_all)
        bottom.addWidget(self.btn_run_subtask)

        self.btn_close = QPushButton("关闭")
        self.btn_close.clicked.connect(self.accept)
        bottom.addWidget(self.btn_close)

        layout.addLayout(bottom)
        self._refresh()

    def _filtered_items(self):
        txt = (self.ed_search.text() or "").strip().lower()
        browser_only = self.chk_browser_only.isChecked()
        out = []
        for it in self.items:
            if not isinstance(it, dict):
                if isinstance(it, (list, tuple)):
                    try:
                        if len(it) >= 2 and isinstance(it[1], dict):
                            it = dict(it[1])
                        elif len(it) >= 1 and isinstance(it[0], dict):
                            it = dict(it[0])
                        else:
                            continue
                    except Exception:
                        continue
                else:
                    continue
            if browser_only and not bool(it.get("is_browser", False)):
                continue
            blob = " ".join([
                str(it.get("task_display", "")),
                str(int(it.get("row_index", 0)) + 1),
                str(it.get("step_name", "")),
                str(it.get("error", "")),
                str(it.get("window_title", "")),
            ]).lower()
            if txt and txt not in blob:
                continue
            out.append(it)
        return out

    def _refresh(self):
        items = self._filtered_items()
        self._refreshing_table = True
        try:
            self.table_widget.blockSignals(True)
            self.table_widget.clearContents()
            self.table_widget.setRowCount(len(items))
            for row_idx, it in enumerate(items):
                if not isinstance(it, dict):
                    continue
                checked = bool(it.get("_checked", True))
                status = str(it.get("status", "") or "").strip()
                task_text = str(it.get("task_display", ""))
                row_text = f"第{int(it.get('row_index', 0)) + 1}行"
                step_text = str(it.get("step_name", "") or f"步骤{int(it.get('step_index', 0))+1}")
                err_full = str(it.get("error", "") or "").strip()
                err_short = err_full if len(err_full) <= 120 else err_full[:120] + "…"
                win_text = str(it.get("window_title", "") or it.get("window_class", "") or "")
                if not win_text:
                    win_text = "(未捕获)"

                select_item = QTableWidgetItem()
                select_item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                select_item.setCheckState(Qt.Checked if checked else Qt.Unchecked)
                select_item.setTextAlignment(Qt.AlignCenter)
                select_item.setData(Qt.UserRole, it)
                self.table_widget.setItem(row_idx, 0, select_item)

                self.table_widget.removeCellWidget(row_idx, 1)
                # 用图标代替文字，避免窄列下文字不可见
                btn_run_row = QPushButton("")
                try:
                    btn_run_row.setIcon(self.style().standardIcon(QStyle.SP_MediaPlay))
                    btn_run_row.setIconSize(QSize(14, 14))
                except Exception:
                    pass
                btn_run_row.setToolTip("只执行这一条子任务。")
                btn_run_row.setFixedSize(30, 24)
                btn_run_row.clicked.connect(lambda _checked=False, entry=it: self._run_single_subtask_entry(entry))
                self.table_widget.setCellWidget(row_idx, 1, btn_run_row)

                status_item = QTableWidgetItem()
                status_item.setFlags(status_item.flags() & ~Qt.ItemIsEditable)
                status_item.setData(Qt.UserRole, it)
                self._apply_entry_status_style(status_item, status)
                self.table_widget.setItem(row_idx, 2, status_item)

                task_item = QTableWidgetItem(task_text)
                task_item.setFlags(task_item.flags() & ~Qt.ItemIsEditable)
                task_item.setData(Qt.UserRole, it)
                task_item.setToolTip(task_text)
                self.table_widget.setItem(row_idx, 3, task_item)

                row_item = QTableWidgetItem(row_text)
                row_item.setFlags(row_item.flags() & ~Qt.ItemIsEditable)
                row_item.setData(Qt.UserRole, it)
                row_item.setTextAlignment(Qt.AlignCenter)
                self.table_widget.setItem(row_idx, 4, row_item)

                step_item = QTableWidgetItem(step_text)
                step_item.setFlags(step_item.flags() & ~Qt.ItemIsEditable)
                step_item.setData(Qt.UserRole, it)
                step_item.setToolTip(step_text)
                self.table_widget.setItem(row_idx, 5, step_item)

                err_item = QTableWidgetItem(err_short if err_short else "(无错误详情)")
                err_item.setFlags(err_item.flags() & ~Qt.ItemIsEditable)
                err_item.setData(Qt.UserRole, it)
                err_item.setToolTip(err_full if err_full else "(无错误详情)")
                self.table_widget.setItem(row_idx, 6, err_item)

                win_item = QTableWidgetItem(win_text)
                win_item.setFlags(win_item.flags() & ~Qt.ItemIsEditable)
                win_item.setData(Qt.UserRole, it)
                win_item.setToolTip(win_text)
                self.table_widget.setItem(row_idx, 7, win_item)
        finally:
            self.table_widget.blockSignals(False)
            self._refreshing_table = False
        if self.table_widget.rowCount() > 0:
            self.table_widget.setCurrentCell(0, 0)
        self._apply_runtime_marks()

    def _load_parent_steps(self):
        self.list_parent_steps.clear()
        p = self.parent()
        task_id = self.context_task_id
        if not p or not task_id or not hasattr(p, "config"):
            self.grp_parent_steps.setVisible(False)
            return
        acts = (p.config.get("tasks", {}) or {}).get(task_id, []) or []
        if not acts:
            self.grp_parent_steps.setVisible(False)
            return
        self.grp_parent_steps.setVisible(True)
        for idx, act in enumerate(acts):
            if not isinstance(act, dict):
                continue
            raw_action = str(act.get("action", "") or "")
            step_name = str(act.get("name", f"步骤{idx+1}") or f"步骤{idx+1}")
            item = QListWidgetItem(f"{idx+1}. {step_name}  [{raw_action}]")
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Unchecked)
            item.setData(Qt.UserRole, idx)
            self.list_parent_steps.addItem(item)

    def _set_parent_step_checks(self, checked):
        state = Qt.Checked if checked else Qt.Unchecked
        for i in range(self.list_parent_steps.count()):
            item = self.list_parent_steps.item(i)
            if item:
                item.setCheckState(state)

    def _get_selected_parent_step_indices(self):
        rows = []
        for i in range(self.list_parent_steps.count()):
            item = self.list_parent_steps.item(i)
            if item and item.checkState() == Qt.Checked:
                try:
                    rows.append(int(item.data(Qt.UserRole)))
                except Exception:
                    pass
        return rows

    def _on_table_item_changed(self, item):
        if getattr(self, "_refreshing_table", False):
            return
        if not item or item.column() != 0:
            return
        data = item.data(Qt.UserRole)
        if isinstance(data, dict):
            data["_checked"] = (item.checkState() == Qt.Checked)

    def _set_filtered_entry_checks(self, checked):
        target_keys = {self._entry_key(it) for it in self._filtered_items() if isinstance(it, dict)}
        for it in self.items:
            if isinstance(it, dict) and self._entry_key(it) in target_keys:
                it["_checked"] = bool(checked)
        self._refresh()

    def _apply_entry_status_style(self, item, status):
        p = self.parent()
        if p and hasattr(p, "_apply_row_status_style"):
            p._apply_row_status_style(item, status)
        else:
            item.setText(str(status or ""))
            item.setTextAlignment(Qt.AlignCenter)

    def _get_entry_from_table_row(self, row_idx):
        if row_idx < 0 or row_idx >= self.table_widget.rowCount():
            return None
        for col in [0, 2, 3, 4, 5, 6, 7]:
            item = self.table_widget.item(row_idx, col)
            if item:
                data = item.data(Qt.UserRole)
                if isinstance(data, dict):
                    return data
        return None

    def _get_checked_items(self):
        out = []
        for row_idx in range(self.table_widget.rowCount()):
            item = self.table_widget.item(row_idx, 0)
            if item and item.checkState() == Qt.Checked:
                data = item.data(Qt.UserRole)
                if isinstance(data, dict):
                    out.append(data)
        return out

    def _set_runtime_table_row_style(self, row_idx, active=False):
        for col in [0, 3, 4, 5, 6, 7]:
            item = self.table_widget.item(row_idx, col)
            if item:
                self._set_runtime_item_style(item, active=active, bg="#fff176", fg="#1f2937")
        btn = self.table_widget.cellWidget(row_idx, 1)
        if isinstance(btn, QPushButton):
            btn.setStyleSheet("background:#fff59d; font-weight:bold;" if active else "")

    def update_entry_status(self, entry=None, status="", error=None):
        entry_key = self._entry_key(entry)
        if not entry_key:
            return
        target_entry = None
        for it in self.items:
            if isinstance(it, dict) and self._entry_key(it) == entry_key:
                target_entry = it
                break
        if target_entry is None and isinstance(entry, dict):
            target_entry = entry
        if not isinstance(target_entry, dict):
            return

        target_entry["status"] = str(status or "")
        if error is not None:
            target_entry["error"] = str(error or "")

        for row_idx in range(self.table_widget.rowCount()):
            row_entry = self._get_entry_from_table_row(row_idx)
            if self._entry_key(row_entry) != entry_key:
                continue
            status_item = self.table_widget.item(row_idx, 2)
            if status_item:
                status_item.setData(Qt.UserRole, target_entry)
                self._apply_entry_status_style(status_item, target_entry.get("status", ""))
            err_item = self.table_widget.item(row_idx, 6)
            if err_item is not None and error is not None:
                err_full = str(target_entry.get("error", "") or "").strip()
                err_short = err_full if len(err_full) <= 120 else err_full[:120] + "…"
                err_item.setText(err_short if err_short else "(无错误详情)")
                err_item.setToolTip(err_full if err_full else "(无错误详情)")
                err_item.setData(Qt.UserRole, target_entry)
            break

    def _get_selected_item(self):
        return self._get_entry_from_table_row(self.table_widget.currentRow())

    def _activate_selected(self):
        it = self._get_selected_item()
        if not it:
            return
        hwnd = it.get("hwnd")
        if hwnd:
            try:
                force_activate_window(int(hwnd))
            except Exception:
                pass

    def _jump_selected(self):
        it = self._get_selected_item()
        p = self.parent()
        if not it or not p:
            return
        if hasattr(p, "_jump_to_failure"):
            p._jump_to_failure(it)

    def _copy_browser_windows(self):
        items = [it for it in self._filtered_items() if it.get("hwnd") and it.get("is_browser")]
        text_lines = []
        for it in items:
            title = it.get("window_title") or ""
            hwnd = it.get("hwnd")
            text_lines.append(f"{title}::hwnd={hwnd}")
        txt = "\n".join(text_lines).strip()
        if not txt:
            QMessageBox.information(self, "提示", "当前没有捕获到可复制的窗口(hwnd)。")
            return
        try:
            pyperclip.copy(txt)
            QMessageBox.information(self, "已复制", f"已复制 {len(text_lines)} 个窗口到剪贴板。")
        except Exception:
            QMessageBox.information(self, "提示", txt)

    def _run_subtask_selected_or_all(self):
        p = self.parent()
        if not p:
            return
        subtask_task_id = self.cb_subtask.currentData()
        parent_step_indices = self._get_selected_parent_step_indices()
        continue_parent = self.chk_continue_parent.isChecked() if hasattr(self, "chk_continue_parent") and self.grp_parent_steps.isVisible() else False
        if not subtask_task_id and not parent_step_indices:
            QMessageBox.information(self, "提示", "请至少选择一个“前置步骤”或一个“子任务”。")
            return
        if not subtask_task_id:
            subtask_task_id = ""
        # 默认对列表(过滤后的全部)执行；仅在勾选“仅执行选中条目”时才执行选中项
        targets = []
        if hasattr(self, "chk_only_selected") and self.chk_only_selected.isChecked():
            targets = self._get_checked_items()
        else:
            targets = self._filtered_items()
        if not targets:
            QMessageBox.information(self, "提示", "当前没有可执行的条目。")
            return
        auto_activate = self.chk_auto_activate.isChecked() if hasattr(self, "chk_auto_activate") else True
        if hasattr(p, "_start_subtask_for_entries"):
            started = p._start_subtask_for_entries(
                targets,
                subtask_task_id,
                auto_activate_hwnd=auto_activate,
                parent_step_indices=parent_step_indices,
                continue_with_parent=continue_parent,
                context_task_id=self.context_task_id
            )
            if started:
                self.set_runtime_running(True)
        elif hasattr(p, "_start_repair_for_failures"):
            # 兼容旧版本：退化为“修复子任务”入口
            p._start_repair_for_failures(targets, subtask_task_id)
            self.accept()

    def _create_new_subtask_task(self):
        p = self.parent()
        if not p:
            return
        sel = self._get_selected_item()
        if not sel:
            QMessageBox.information(self, "提示", "请先在列表中选中一条条目。")
            return
        if hasattr(p, "_create_repair_task_from_failure"):
            p._create_repair_task_from_failure(sel)
        self.accept()

    def _run_single_subtask_entry(self, entry):
        if not isinstance(entry, dict):
            return
        p = self.parent()
        if not p:
            return
        subtask_task_id = self.cb_subtask.currentData()
        parent_step_indices = self._get_selected_parent_step_indices()
        continue_parent = self.chk_continue_parent.isChecked() if hasattr(self, "chk_continue_parent") and self.grp_parent_steps.isVisible() else False
        if not subtask_task_id and not parent_step_indices:
            QMessageBox.information(self, "提示", "请至少选择一个“前置步骤”或一个“子任务”。")
            return
        auto_activate = self.chk_auto_activate.isChecked() if hasattr(self, "chk_auto_activate") else True
        if hasattr(p, "_start_subtask_for_entries"):
            started = p._start_subtask_for_entries(
                [entry],
                subtask_task_id or "",
                auto_activate_hwnd=auto_activate,
                parent_step_indices=parent_step_indices,
                continue_with_parent=continue_parent,
                context_task_id=self.context_task_id
            )
            if started:
                self.set_runtime_running(True)

    def _entry_key(self, it):
        if not isinstance(it, dict):
            return None
        try:
            row_idx = int(it.get("row_index", -1))
        except Exception:
            row_idx = -1
        return (
            str(it.get("task_id") or ""),
            row_idx,
            str(it.get("step_name") or ""),
            str(it.get("window_title") or ""),
        )

    def _set_runtime_item_style(self, item, active=False, bg="#fff176", fg="#1f2937"):
        if not item:
            return
        font = item.font()
        font.setBold(bool(active))
        item.setFont(font)
        if active:
            item.setBackground(QColor(bg))
            item.setForeground(QColor(fg))
        else:
            item.setBackground(QBrush(Qt.NoBrush))
            item.setForeground(QBrush(Qt.NoBrush))

    def _apply_runtime_marks(self):
        highlight_enabled = self.chk_highlight_running.isChecked() if hasattr(self, "chk_highlight_running") else True
        scroll_enabled = self.chk_scroll_running.isChecked() if hasattr(self, "chk_scroll_running") else True

        active_entry_row = None
        active_entry_key = self._runtime_entry_key
        for i in range(self.table_widget.rowCount()):
            item_data = self._get_entry_from_table_row(i)
            matched = bool(active_entry_key and self._entry_key(item_data) == active_entry_key)
            self._set_runtime_table_row_style(i, active=bool(highlight_enabled and matched))
            if matched:
                active_entry_row = i
        if active_entry_row is not None and (highlight_enabled or scroll_enabled):
            self.table_widget.setCurrentCell(active_entry_row, 0)
            if scroll_enabled:
                self.table_widget.scrollToItem(self.table_widget.item(active_entry_row, 0), QAbstractItemView.PositionAtCenter)

        active_parent_item = None
        active_parent_step_idx = self._runtime_parent_step_idx
        for i in range(self.list_parent_steps.count()):
            item = self.list_parent_steps.item(i)
            matched = False
            if item is not None and active_parent_step_idx is not None:
                try:
                    matched = int(item.data(Qt.UserRole)) == int(active_parent_step_idx)
                except Exception:
                    matched = False
            self._set_runtime_item_style(item, active=bool(highlight_enabled and matched), bg="#dbeafe", fg="#1d4ed8")
            if matched:
                active_parent_item = item
        if active_parent_item and (highlight_enabled or scroll_enabled):
            self.list_parent_steps.setCurrentItem(active_parent_item)
            if scroll_enabled:
                self.list_parent_steps.scrollToItem(active_parent_item, QAbstractItemView.PositionAtCenter)

    def set_runtime_progress(self, entry=None, parent_step_idx=None):
        self._runtime_entry_key = self._entry_key(entry) if entry else None
        self._runtime_parent_step_idx = int(parent_step_idx) if parent_step_idx is not None and int(parent_step_idx) >= 0 else None
        self._apply_runtime_marks()

    def clear_runtime_progress(self):
        self._runtime_entry_key = None
        self._runtime_parent_step_idx = None
        self._apply_runtime_marks()

    def set_runtime_running(self, running):
        self._runtime_running = bool(running)
        if hasattr(self, "btn_run_subtask"):
            self.btn_run_subtask.setEnabled(not self._runtime_running)
            # 运行中也用图标反馈（按钮会被禁用，但图标仍可见）
            try:
                icon = QStyle.SP_MediaPause if self._runtime_running else QStyle.SP_MediaPlay
                self.btn_run_subtask.setIcon(self.style().standardIcon(icon))
            except Exception:
                pass
            self.btn_run_subtask.setText("")
        if hasattr(self, "table_widget"):
            for row_idx in range(self.table_widget.rowCount()):
                btn = self.table_widget.cellWidget(row_idx, 1)
                if isinstance(btn, QPushButton):
                    btn.setEnabled(not self._runtime_running)
        if hasattr(self, "btn_close"):
            self.btn_close.setText("关闭监看" if self._runtime_running else "关闭")
        if not self._runtime_running:
            self.clear_runtime_progress()


# --- Schedule Configuration Dialog ---
class ScheduleDialog(QDialog):
    def __init__(self, parent=None, current_config=None):
        super().__init__(parent)
        self.setWindowTitle("⏰ 计划时间")
        self.resize(400, 500)
        layout = QVBoxLayout(self)
        
        self.mode_group = QButtonGroup(self)
        
        # 1. Once Mode
        once_group = QGroupBox("📅 单次执行")
        once_layout = QVBoxLayout(once_group)
        self.radio_once = QRadioButton("指定日期和时间，跑完一次就结束")
        self.mode_group.addButton(self.radio_once, 0)
        once_layout.addWidget(self.radio_once)
        self.dt_once = QDateTimeEdit(QDateTime.currentDateTime().addSecs(3600))
        self.dt_once.setCalendarPopup(True)
        self.dt_once.setDisplayFormat("yyyy-MM-dd HH:mm")
        once_layout.addWidget(self.dt_once)
        layout.addWidget(once_group)
        
        # 2. Daily Mode
        daily_group = QGroupBox("🔄 每天重复")
        daily_layout = QVBoxLayout(daily_group)
        self.radio_daily = QRadioButton("每天到了这个点就自动开跑")
        self.mode_group.addButton(self.radio_daily, 1)
        daily_layout.addWidget(self.radio_daily)
        self.time_daily = QTimeEdit(QTime(10, 0))
        self.time_daily.setDisplayFormat("HH:mm")
        daily_layout.addWidget(self.time_daily)
        layout.addWidget(daily_group)
        
        # 3. Specific Date Mode (Shortcut for common usage)
        spec_group = QGroupBox("📍 指定日期 (日历选择)")
        spec_layout = QVBoxLayout(spec_group)
        self.radio_spec = QRadioButton("在特定日期执行")
        self.mode_group.addButton(self.radio_spec, 2)
        spec_layout.addWidget(self.radio_spec)
        self.calendar = QCalendarWidget()
        self.calendar.setGridVisible(True)
        spec_layout.addWidget(self.calendar)
        self.time_spec = QTimeEdit(QTime(10, 0))
        self.time_spec.setDisplayFormat("HH:mm")
        spec_layout.addWidget(self.time_spec)
        layout.addWidget(spec_group)
        
        # Load current config
        if current_config:
            mode = current_config.get("mode", 0)
            if mode == 0:
                self.radio_once.setChecked(True)
                self.dt_once.setDateTime(QDateTime.fromString(current_config.get("value"), "yyyy-MM-dd HH:mm"))
            elif mode == 1:
                self.radio_daily.setChecked(True)
                self.time_daily.setTime(QTime.fromString(current_config.get("value"), "HH:mm"))
            elif mode == 2:
                self.radio_spec.setChecked(True)
                dt = QDateTime.fromString(current_config.get("value"), "yyyy-MM-dd HH:mm")
                self.calendar.setSelectedDate(dt.date())
                self.time_spec.setTime(dt.time())
        else:
            self.radio_once.setChecked(True)
            
        btn_box = QHBoxLayout()
        btn_ok = QPushButton("✅ 开启调度"); btn_ok.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold; height: 35px;")
        btn_ok.clicked.connect(self.accept)
        btn_cancel = QPushButton("❌ 取消"); btn_cancel.setFixedHeight(35)
        btn_cancel.clicked.connect(self.reject)
        btn_box.addWidget(btn_ok); btn_box.addWidget(btn_cancel)
        layout.addLayout(btn_box)
        
    def get_config(self):
        mode = self.mode_group.checkedId()
        if mode == 0:
            val = self.dt_once.dateTime().toString("yyyy-MM-dd HH:mm")
        elif mode == 1:
            val = self.time_daily.time().toString("HH:mm")
        else:
            val = f"{self.calendar.selectedDate().toString('yyyy-MM-dd')} {self.time_spec.time().toString('HH:mm')}"
        return {"mode": mode, "value": val}

# --- Custom Table for Smart Paste & Editing ---
class MultiLineTextDelegate(QStyledItemDelegate):
    """让表格单元格支持更顺手的多行文本编辑。"""
    def createEditor(self, parent, option, index):
        editor = QTextEdit(parent)
        editor.setAcceptRichText(False)
        editor.setWordWrapMode(QTextOption.WrapAnywhere)
        editor.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        editor.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        editor.setMinimumHeight(96)
        return editor

    def setEditorData(self, editor, index):
        editor.setPlainText(index.data() or "")
        editor.selectAll()

    def setModelData(self, editor, model, index):
        model.setData(index, editor.toPlainText())

    def updateEditorGeometry(self, editor, option, index):
        rect = option.rect.adjusted(0, 0, 0, 40)
        editor.setGeometry(rect)

class PrefixPresetEditorDialog(QDialog):
    """用于完整查看和编辑前缀内容的弹窗。"""
    def __init__(self, parent=None, name="", prefix="", title="✏️ 编辑前缀"):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.resize(760, 560)

        layout = QVBoxLayout(self)
        tip = QLabel("这里会完整显示当前前缀内容，可直接编辑多行文本。")
        tip.setStyleSheet("color: #666; padding: 2px 0 4px 2px;")
        layout.addWidget(tip)

        form = QFormLayout()
        self.name_edit = QLineEdit(str(name))
        self.name_edit.setPlaceholderText("输入前缀名称...")
        form.addRow("名称：", self.name_edit)
        layout.addLayout(form)

        layout.addWidget(QLabel("前缀内容："))
        self.prefix_edit = QTextEdit()
        self.prefix_edit.setAcceptRichText(False)
        self.prefix_edit.setWordWrapMode(QTextOption.WrapAnywhere)
        self.prefix_edit.setPlainText(str(prefix))
        self.prefix_edit.setPlaceholderText("输入完整前缀内容...")
        self.prefix_edit.setMinimumHeight(360)
        layout.addWidget(self.prefix_edit, 1)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_ok = QPushButton("✅ 保存")
        btn_ok.setStyleSheet("background-color: #e8f5e9; font-weight: bold; min-width: 96px;")
        btn_cancel = QPushButton("❌ 取消")
        btn_cancel.setMinimumWidth(96)
        btn_ok.clicked.connect(self.accept)
        btn_cancel.clicked.connect(self.reject)
        btn_row.addWidget(btn_ok)
        btn_row.addWidget(btn_cancel)
        layout.addLayout(btn_row)

        if str(prefix).strip():
            self.prefix_edit.setFocus()
        else:
            self.name_edit.selectAll()
            self.name_edit.setFocus()

    def get_data(self):
        return {
            "name": self.name_edit.text().strip(),
            "prefix": self.prefix_edit.toPlainText()
        }

def parse_spreadsheet_clipboard_text(raw_text):
    """解析来自 Excel / Google 表格 的制表符文本，保留被引号包裹单元格内的换行。"""
    text = str(raw_text or "")
    if not text:
        return []
    text = text.replace('\r\n', '\n').replace('\r', '\n')
    if not text.strip():
        return []

    rows = []
    try:
        reader = csv.reader(StringIO(text), delimiter='\t', quotechar='"')
        rows = [list(row) for row in reader]
    except Exception:
        rows = [line.split('\t') for line in text.split('\n')]

    while rows and not any(str(cell).strip() for cell in rows[-1]):
        rows.pop()
    if not rows:
        return []

    max_cols = max((len(row) for row in rows), default=0)
    while max_cols > 0:
        col_has_value = False
        for row in rows:
            if max_cols - 1 < len(row) and str(row[max_cols - 1]).strip():
                col_has_value = True
                break
        if col_has_value:
            break
        max_cols -= 1

    normalized = []
    for row in rows:
        normalized.append([str(row[i]) if i < len(row) else "" for i in range(max_cols)])
    return normalized

def make_excel_column_name(idx):
    idx = int(idx)
    name = ""
    while idx >= 0:
        idx, rem = divmod(idx, 26)
        name = chr(ord('A') + rem) + name
        idx -= 1
    return name

def create_table_selection_label(empty_text="当前未选择"):
    lbl = QLabel(empty_text)
    lbl.setStyleSheet("color: #666; padding: 2px 0 2px 2px;")
    return lbl

def get_table_selection_summary(table):
    try:
        indexes = table.selectedIndexes()
    except Exception:
        return "当前未选择"

    if not indexes:
        return "当前未选择"

    unique_indexes = []
    seen = set()
    for idx in indexes:
        key = (idx.row(), idx.column())
        if key in seen:
            continue
        seen.add(key)
        unique_indexes.append(idx)

    rows = sorted({idx.row() for idx in unique_indexes})
    cols = sorted({idx.column() for idx in unique_indexes})
    behavior = table.selectionBehavior()

    if behavior == QAbstractItemView.SelectRows:
        return f"当前已选 {len(rows)} 行"
    if behavior == QAbstractItemView.SelectColumns:
        return f"当前已选 {len(cols)} 列"

    cell_count = len(unique_indexes)
    if cell_count <= 0:
        return "当前未选择"

    is_full_rect = bool(rows) and bool(cols) and cell_count == len(rows) * len(cols)
    if is_full_rect:
        return f"当前已选 {cell_count} 格 ({len(rows)} 行 × {len(cols)} 列)"
    return f"当前已选 {cell_count} 格 (涉及 {len(rows)} 行 / {len(cols)} 列)"

def bind_table_selection_label(table, label, prefix=""):
    def _update_label():
        summary = get_table_selection_summary(table)
        label.setText(f"{prefix}{summary}" if prefix else summary)

    table.itemSelectionChanged.connect(_update_label)
    _update_label()
    table._selection_counter_updater = _update_label
    return _update_label

def bind_item_view_selection_label(view, label, prefix="", kind_text="项"):
    def _update_label():
        count = 0
        try:
            if hasattr(view, "selectedItems"):
                count = len(view.selectedItems())
            else:
                indexes = view.selectedIndexes()
                seen = set()
                for idx in indexes:
                    seen.add((idx.row(), idx.column()))
                count = len(seen)
        except Exception:
            count = 0

        summary = f"当前已选 {count} 个{kind_text}" if count > 0 else "当前未选择"
        label.setText(f"{prefix}{summary}" if prefix else summary)

    if hasattr(view, "itemSelectionChanged"):
        view.itemSelectionChanged.connect(_update_label)
    _update_label()
    view._selection_counter_updater = _update_label
    return _update_label

class SpreadsheetPasteTable(QTableWidget):
    """简易表格式文本输入区，支持像表格软件一样粘贴/复制二维数据。"""

    def __init__(self, rows=12, cols=4, parent=None):
        super().__init__(rows, cols, parent)
        self.setAlternatingRowColors(True)
        self.setWordWrap(True)
        self.setShowGrid(True)
        self.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.SelectedClicked | QAbstractItemView.AnyKeyPressed)
        self.setItemDelegate(MultiLineTextDelegate(self))
        self.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.verticalHeader().setDefaultSectionSize(42)
        self.verticalHeader().setMinimumSectionSize(28)
        self._refresh_headers()
        self._ensure_all_items()

    def _refresh_headers(self):
        self.setHorizontalHeaderLabels([make_excel_column_name(i) for i in range(self.columnCount())])
        self.setVerticalHeaderLabels([str(i + 1) for i in range(self.rowCount())])

    def _ensure_all_items(self):
        for r in range(self.rowCount()):
            for c in range(self.columnCount()):
                if not self.item(r, c):
                    self.setItem(r, c, QTableWidgetItem(""))

    def ensure_size(self, min_rows, min_cols):
        changed = False
        if min_rows > self.rowCount():
            self.setRowCount(min_rows)
            changed = True
        if min_cols > self.columnCount():
            self.setColumnCount(min_cols)
            changed = True
        if changed:
            self._refresh_headers()
        self._ensure_all_items()

    def keyPressEvent(self, event):
        key = event.key()
        mod = event.modifiers()
        if mod == Qt.ControlModifier and key == Qt.Key_V:
            self.smart_paste()
        elif mod == Qt.ControlModifier and key == Qt.Key_C:
            self.smart_copy()
        elif key in (Qt.Key_Delete, Qt.Key_Backspace) and mod == Qt.NoModifier:
            self.clear_selection_contents()
        else:
            super().keyPressEvent(event)

    def clear_selection_contents(self):
        items = self.selectedItems()
        if not items:
            cur = self.currentItem()
            if cur:
                items = [cur]
        for it in items:
            it.setText("")

    def smart_copy(self):
        selected = self.selectedRanges()
        if not selected:
            return
        r = selected[0]
        lines = []
        for row in range(r.topRow(), r.bottomRow() + 1):
            cells = []
            for col in range(r.leftColumn(), r.rightColumn() + 1):
                it = self.item(row, col)
                cells.append(it.text() if it else "")
            lines.append("\t".join(cells))
        pyperclip.copy("\n".join(lines))

    def smart_paste(self, text=None):
        matrix = parse_spreadsheet_clipboard_text(text if text is not None else pyperclip.paste())
        if not matrix:
            return False
        start_row = max(0, self.currentRow())
        start_col = max(0, self.currentColumn())
        self.ensure_size(start_row + len(matrix), start_col + max(len(row) for row in matrix))
        self.blockSignals(True)
        for r_off, row in enumerate(matrix):
            for c_off, val in enumerate(row):
                it = self.item(start_row + r_off, start_col + c_off)
                if not it:
                    it = QTableWidgetItem("")
                    self.setItem(start_row + r_off, start_col + c_off, it)
                it.setText(val)
        self.blockSignals(False)
        for r_off in range(len(matrix)):
            self.resizeRowToContents(start_row + r_off)
            self.setRowHeight(start_row + r_off, min(max(self.rowHeight(start_row + r_off), 42), 140))
        return True

    def get_effective_matrix(self):
        max_row = -1
        max_col = -1
        for r in range(self.rowCount()):
            for c in range(self.columnCount()):
                it = self.item(r, c)
                if it and str(it.text()).strip():
                    max_row = max(max_row, r)
                    max_col = max(max_col, c)
        if max_row < 0 or max_col < 0:
            return []
        matrix = []
        for r in range(max_row + 1):
            row = []
            for c in range(max_col + 1):
                it = self.item(r, c)
                row.append(it.text() if it else "")
            matrix.append(row)
        return matrix

    def get_selected_matrix(self):
        selected = self.selectedRanges()
        if not selected:
            return []
        r = selected[0]
        matrix = []
        for row in range(r.topRow(), r.bottomRow() + 1):
            line = []
            for col in range(r.leftColumn(), r.rightColumn() + 1):
                it = self.item(row, col)
                line.append(it.text() if it else "")
            matrix.append(line)

        while matrix and not any(str(cell).strip() for cell in matrix[-1]):
            matrix.pop()
        if not matrix:
            return []

        max_cols = max((len(row) for row in matrix), default=0)
        while max_cols > 0:
            if any(str(row[max_cols - 1]).strip() for row in matrix if max_cols - 1 < len(row)):
                break
            max_cols -= 1

        return [row[:max_cols] for row in matrix]

    def get_active_matrix(self):
        selected_matrix = self.get_selected_matrix()
        return selected_matrix if selected_matrix else self.get_effective_matrix()

    def clear_all_contents(self):
        self.blockSignals(True)
        for r in range(self.rowCount()):
            for c in range(self.columnCount()):
                it = self.item(r, c)
                if it:
                    it.setText("")
                else:
                    self.setItem(r, c, QTableWidgetItem(""))
        self.blockSignals(False)

class DataEditorTable(QTableWidget):
    double_clicked_cell = pyqtSignal(int, int)
    rows_reordered = pyqtSignal()

    def is_editing(self):
        """判断当前是否有单元格正在编辑状态。"""
        return self.state() == QAbstractItemView.EditingState

    def __init__(self, *args):
        super().__init__(*args)
        self.itemDoubleClicked.connect(lambda item: self.double_clicked_cell.emit(item.row(), item.column()))
        self.setEditTriggers(QTableWidget.DoubleClicked | QTableWidget.SelectedClicked | QTableWidget.AnyKeyPressed)
        self.setAlternatingRowColors(True); self.setShowGrid(True)
        self.setStyleSheet("alternate-background-color: #f9f9f9; gridline-color: #ccc; color: black;")
        self.setContextMenuPolicy(Qt.CustomContextMenu)
        self.setAcceptDrops(True) # Keep for file drops
        self.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self._last_checkbox_row = None

    def _resolve_widget(self, widget):
        """If widget is a container (e.g. QWidget wrapping a ComboBox + refresh button),
        return the first meaningful child (QComboBox or text editor). Otherwise return as-is."""
        if widget is None: return None
        if isinstance(widget, (QComboBox, QLineEdit, MultiLineTextEdit)): return widget
        # Look for a QComboBox or text editor child
        cb = widget.findChild(QComboBox)
        if cb: return cb
        ed = _find_text_input(widget)
        if ed: return ed
        return widget

    def _find_manager(self):
        manager = self.parent()
        while manager and not hasattr(manager, '_handle_data_checkbox_shift_click'):
            manager = manager.parent()
        return manager

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            idx = self.indexAt(event.pos())
            if idx.isValid() and idx.column() == 0:
                item = self.item(idx.row(), 0)
                if item and (event.modifiers() & Qt.ShiftModifier):
                    manager = self._find_manager()
                    if manager:
                        target_checked = (item.checkState() != Qt.Checked)
                        handled = manager._handle_data_checkbox_shift_click(
                            idx.row(),
                            target_checked,
                            anchor_row=self._last_checkbox_row
                        )
                        if handled:
                            self._last_checkbox_row = idx.row()
                            event.accept()
                            return
                self._last_checkbox_row = idx.row()
        super().mousePressEvent(event)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls(): event.acceptProposedAction()

    def dragMoveEvent(self, event):
        if event.mimeData().hasUrls(): event.acceptProposedAction()

    def dropEvent(self, event):
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            files = [os.path.normpath(u.toLocalFile()) for u in urls if u.isLocalFile()]
            if not files: return
            pos = event.pos(); item = self.itemAt(pos)
            start_row = item.row() if item else max(0, self.rowCount() - 1)
            start_col = item.column() if item else 0
            manager = self.parent()
            while manager and not hasattr(manager, '_add_data_row'): manager = manager.parent()
            rows_needed = (start_row + len(files)) - self.rowCount()
            for _ in range(max(0, rows_needed)):
                if manager: manager._add_data_row()
                else: self.insertRow(self.rowCount())
            self.blockSignals(True)
            for i, f_path in enumerate(files):
                target_row = start_row + i
                widget = self.cellWidget(target_row, start_col); w = self._resolve_widget(widget)
                if w:
                    if isinstance(w, QComboBox): w.setCurrentText(f_path)
                    elif isinstance(w, (QLineEdit, MultiLineTextEdit)):
                        # [修复] 处理复合单元格（open_url/clear_input_plus）的拖放
                        if "|" in f_path:
                            parts = f_path.split("|", 1)
                            w.setText(parts[0])
                            # [修复] 同步更新 open_url 的 profile_id 按鈕属性，防止账号脱黄
                            if widget:
                                btn = widget.findChild(QPushButton)
                                if btn and "profile_id" in btn.dynamicPropertyNames():
                                    btn.setProperty("profile_id", parts[1])
                                    btn.setText(get_profile_display_name(parts[1]))
                        else:
                            w.setText(f_path)
                else:
                    target_item = self.item(target_row, start_col)
                    if target_item and (target_item.flags() & Qt.ItemIsEditable): target_item.setText(f_path)
                    elif not target_item: self.setItem(target_row, start_col, QTableWidgetItem(f_path))
            self.blockSignals(False)
            if manager: manager._save_data_table()
            event.acceptProposedAction()

    def keyPressEvent(self, event):
        key = event.key()
        mod = event.modifiers()

        if mod == Qt.ControlModifier and key == Qt.Key_V:
            self._smart_paste()
        elif mod == Qt.ControlModifier and key == Qt.Key_C:
            self._smart_copy()
        elif key in (Qt.Key_Delete, Qt.Key_Backspace) and mod == Qt.NoModifier:
            self._clear_selection()
        else:
            super().keyPressEvent(event)

    def _clear_selection(self):
        """Delete/Backspace 清空选中的所有可编辑单元格，就像 Excel 一样。"""
        manager = self.parent()
        while manager and not hasattr(manager, '_save_data_table'):
            manager = manager.parent()

        self.blockSignals(True)
        for item in self.selectedItems():
            if item.flags() & Qt.ItemIsEditable:
                item.setText("")
            # 同时清空该格对应的 cellWidget（如 ComboBox / KeyRecorder）
            widget = self.cellWidget(item.row(), item.column())
            w = self._resolve_widget(widget)
            if w:
                if isinstance(w, QComboBox):
                    w.setCurrentText("")
                elif isinstance(w, (QLineEdit, MultiLineTextEdit)):
                    w.setText("")
        self.blockSignals(False)

        if manager:
            manager._save_data_table()

    def _smart_copy(self):
        """Ctrl+C 把选中区域按 Excel 格式（Tab 分隔列、换行分隔行）复制到剪贴板。"""
        selected = self.selectedRanges()
        if not selected:
            return
        r = selected[0]
        rows = range(r.topRow(), r.bottomRow() + 1)
        cols = range(r.leftColumn(), r.rightColumn() + 1)
        lines = []
        for row in rows:
            cells = []
            for col in cols:
                widget = self.cellWidget(row, col)
                w = self._resolve_widget(widget)
                if w:
                    if isinstance(w, QComboBox):
                        cells.append(w.currentText())
                    elif isinstance(w, (QLineEdit, MultiLineTextEdit)):
                        cells.append(w.text())
                    else:
                        cells.append("")
                else:
                    item = self.item(row, col)
                    cells.append(item.text() if item else "")
            lines.append("\t".join(cells))
        pyperclip.copy("\n".join(lines))

    def _smart_paste(self):
        text = pyperclip.paste()
        if not text: return
        text = text.replace('\r\n', '\n').replace('\r', '\n')
        rows_data = text.split('\n')
        while rows_data and not rows_data[-1].strip():
            rows_data.pop()
        if not rows_data: return

        current_row = max(0, self.currentRow())
        current_col = max(0, self.currentColumn())

        # Find parent manager
        manager = self.parent()
        while manager and not hasattr(manager, '_add_data_row'):
            manager = manager.parent()

        # Step 1: pre-add rows BEFORE blockSignals so each row is properly initialised
        rows_needed = (current_row + len(rows_data)) - self.rowCount()
        for _ in range(max(0, rows_needed)):
            if manager:
                manager._add_data_row()
            else:
                self.insertRow(self.rowCount())

        # Step 2: fill cells — handle plain items AND cellWidgets (ComboBox / KeyRecorder)
        self.blockSignals(True)
        for i, row_text in enumerate(rows_data):
            target_row = current_row + i
            cols_data = row_text.split('\t')
            for j, col_text in enumerate(cols_data):
                target_col = current_col + j
                if target_col >= self.columnCount():
                    break
                widget = self.cellWidget(target_row, target_col)
                w = self._resolve_widget(widget)
                if w:
                    # [修复] 智能粘贴：针对复合单元格（open_url / clear_input_plus）进行特殊处理
                    # 如果粘贴的内容包含 |，说明是完整数据格式，需要拆分更新 UI
                    if isinstance(w, (QLineEdit, MultiLineTextEdit)) and "|" in col_text and widget:
                        parts = col_text.split("|", 1)
                        btn = widget.findChild(QPushButton)
                        lbl = widget.findChild(QLabel)
                        if btn and "profile_id" in btn.dynamicPropertyNames():
                            w.setText(parts[0]) # open_url 的网址
                            btn.setProperty("profile_id", parts[1])
                            btn.setText(get_profile_display_name(parts[1]))
                        elif lbl:
                            w.setText(parts[1]) # clear_input_plus 的内容
                            lbl.setText(f" {parts[0]}")
                        else:
                            w.setText(col_text)
                    elif isinstance(w, QComboBox):
                        w.setCurrentText(col_text)
                    elif isinstance(w, (QLineEdit, MultiLineTextEdit)):
                        w.setText(col_text)
                    
                    # 统一同步到 backing item
                    bk = self.item(target_row, target_col)
                    if bk: bk.setText(col_text)
                else:
                    item = self.item(target_row, target_col)
                    if item and (item.flags() & Qt.ItemIsEditable):
                        item.setText(col_text)
                    elif not item:
                        self.setItem(target_row, target_col, QTableWidgetItem(col_text))
        self.blockSignals(False)

        if manager:
            manager._save_data_table()

# --- Manual Width Header ---
class ManualWidthHeader(QHeaderView):
    """彻底禁用表头双击自动适应列宽，只允许手动拖动或外部控件调整。"""

    def mouseDoubleClickEvent(self, event):
        event.accept()

# --- Drag-Sort Action Table ---
class DragSortActionTable(QTableWidget):
    """Action table with drag-row reorder support."""
    rows_reordered = pyqtSignal()

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setDragEnabled(True)
        self.setAcceptDrops(True)
        self.setDropIndicatorShown(True)
        self.setDragDropMode(QAbstractItemView.InternalMove)
        self.setDragDropOverwriteMode(False)
        self.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._drag_source_row = -1

    def mousePressEvent(self, event):
        self._drag_source_row = self.rowAt(event.pos().y())
        super().mousePressEvent(event)

    def dropEvent(self, event):
        src = self._drag_source_row
        if src < 0:
            super().dropEvent(event)
            return

        selection_model = self.selectionModel()
        selected_rows = sorted({idx.row() for idx in selection_model.selectedRows()}) if selection_model else []
        drag_rows = selected_rows if src in selected_rows else [src]
        drag_rows = [r for r in drag_rows if 0 <= r < self.rowCount()]
        if not drag_rows:
            event.ignore()
            return

        target_row = self.rowAt(event.pos().y())
        if target_row < 0:
            # 拖到末尾空白区域，插入到最后一行之后
            target_row = self.rowCount() - 1
            insert_before = False
        else:
            # [修复] 根据鼠标在目标行的上/下半区决定插入位置
            # 上半区 → 插到目标行前面；下半区 → 插到目标行后面
            row_rect = self.visualRect(self.model().index(target_row, 0))
            mid_y = row_rect.top() + row_rect.height() / 2
            insert_before = event.pos().y() < mid_y

        # 计算最终插入行号（相对于移动前的行号）
        if insert_before:
            insert_at = target_row
        else:
            insert_at = target_row + 1

        # Emit signal; the manager will do the actual reorder in config
        self._pending_reorder = (drag_rows, insert_at)
        self.rows_reordered.emit()
        event.accept()



# ============================================================
# 自定义任务树控件
# - 支持将任务拖入文件夹（分类）
# - 支持精确的项间排序（拖到边缘）
# - 防止非法的嵌套（文件夹套文件夹、任务套任务）
# ============================================================
class TaskTreeWidget(QTreeWidget):
    structure_changed = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self._drag_items = []
        self._origin_role = Qt.UserRole + 1
        self._last_move_info = []
        self._drag_in_progress = False
        self._last_drag_hover_signature = None
        self.setDragDropMode(QAbstractItemView.DragDrop)
        self.setDefaultDropAction(Qt.MoveAction)

    def _debug_item_label(self, item):
        if not item:
            return "<root>"
        role = item.data(0, Qt.UserRole)
        item_type = "folder" if self._is_folder_role(role) else "task"
        return f"{item_type}(text={item.text(0)!r}, role={role!r}, path={self._item_path_from_parent(item)!r})"

    def _debug_items_label(self, items):
        if not items:
            return "[]"
        return "[" + ", ".join(self._debug_item_label(it) for it in items if it) + "]"

    def _debug_log(self, message):
        write_drag_debug(f"[TaskTree] {message}")

    def startDrag(self, supportedActions):
        self._drag_items = self.selectedItems()
        if not self._drag_items:
            self._debug_log("startDrag: 未选中任何节点，直接返回")
            return
        self._drag_in_progress = True
        self._last_move_info = []
        self._last_drag_hover_signature = None
        for item in self._drag_items:
            self._remember_origin_paths(item)
        self._debug_log(
            f"startDrag: supportedActions={int(supportedActions)} selected={self._debug_items_label(self._drag_items)}"
        )
        try:
            mime_data = self.model().mimeData(self.selectedIndexes())
            if not mime_data:
                self._drag_in_progress = False
                self._debug_log("startDrag: mimeData 为空，取消拖拽")
                return
            drag = QDrag(self)
            drag.setMimeData(mime_data)
            result = drag.exec_(Qt.MoveAction)
            self._debug_log(f"startDrag: custom QDrag finished result={int(result)}")
        finally:
            self._debug_log(
                f"startDrag:end drag_in_progress={self._drag_in_progress} last_move_info={self._last_move_info}"
            )

    def _is_folder_role(self, role):
        return isinstance(role, str) and role.startswith("[FOLDER]")

    def _folder_role(self, folder_path=""):
        return f"[FOLDER]::{folder_path}" if folder_path else "[FOLDER]"

    def _remember_origin_paths(self, item):
        if not item:
            return
        role = item.data(0, Qt.UserRole)
        if self._is_folder_role(role):
            item.setData(0, self._origin_role, self._folder_role(self._folder_path_from_item(item)))
            for i in range(item.childCount()):
                self._remember_origin_paths(item.child(i))
        else:
            item.setData(0, self._origin_role, role or self._item_path_from_parent(item))

    def _item_path_from_parent(self, item):
        parts = []
        cursor = item
        while cursor:
            parts.append((cursor.text(0) or "").strip().strip("/"))
            cursor = cursor.parent()
        parts = [p for p in reversed(parts) if p]
        return "/".join(parts)

    def _folder_path_from_item(self, item):
        role = item.data(0, Qt.UserRole)
        if self._is_folder_role(role) and "::" in role:
            return role.split("::", 1)[1]
        return self._item_path_from_parent(item)

    def _collect_move_info(self, item, results):
        if not item:
            return
        current_path = self._item_path_from_parent(item)
        origin_path = item.data(0, self._origin_role)
        role = item.data(0, Qt.UserRole)
        if self._is_folder_role(role):
            for i in range(item.childCount()):
                self._collect_move_info(item.child(i), results)
            return
        if isinstance(origin_path, str) and origin_path and not self._is_folder_role(origin_path):
            results.append((origin_path, current_path))

    def dragEnterEvent(self, event):
        self._debug_log(f"dragEnterEvent: source_is_self={event.source() == self}")
        if event.source() == self:
            event.acceptProposedAction()
        else:
            super().dragEnterEvent(event)

    def _get_drag_items(self):
        items = [it for it in self._drag_items if it and not self._is_folder_role(it.data(0, Qt.UserRole))]
        folders = [it for it in self._drag_items if it and self._is_folder_role(it.data(0, Qt.UserRole))]
        return items, folders

    def _hover_ratio(self, event, target_item):
        rect = self.visualItemRect(target_item)
        if rect.height() <= 0:
            return 0.5
        return (event.pos().y() - rect.top()) / rect.height()

    def dragMoveEvent(self, event):
        if event.source() != self:
            self._debug_log("dragMoveEvent: 外部来源，交给父类处理")
            super().dragMoveEvent(event)
            return

        drag_tasks, drag_folders = self._get_drag_items()
        target_item = self.itemAt(event.pos())
        ratio = self._hover_ratio(event, target_item) if target_item else None
        decision = "accept_default"
        if not drag_tasks and not drag_folders:
            decision = "ignore_no_drag_items"
            signature = (decision, self._debug_item_label(target_item))
            if signature != self._last_drag_hover_signature:
                self._last_drag_hover_signature = signature
                self._debug_log(f"dragMoveEvent: decision={decision} target={self._debug_item_label(target_item)}")
            event.ignore()
            return

        if target_item and target_item in self._drag_items:
            decision = "ignore_target_in_drag_items"
            signature = (decision, self._debug_item_label(target_item))
            if signature != self._last_drag_hover_signature:
                self._last_drag_hover_signature = signature
                self._debug_log(f"dragMoveEvent: decision={decision} target={self._debug_item_label(target_item)}")
            event.ignore()
            return

        if target_item:
            target_role = target_item.data(0, Qt.UserRole)
            is_hovering_center = 0.15 < ratio < 0.85

            if self._is_folder_role(target_role):
                if (drag_tasks or drag_folders) and is_hovering_center:
                    decision = "accept_into_folder_center"
                    signature = (
                        decision,
                        self._debug_item_label(target_item),
                        len(drag_tasks),
                        len(drag_folders),
                    )
                    if signature != self._last_drag_hover_signature:
                        self._last_drag_hover_signature = signature
                        self._debug_log(
                            f"dragMoveEvent: decision={decision} ratio={ratio:.3f} target={self._debug_item_label(target_item)} "
                            f"tasks={self._debug_items_label(drag_tasks)} folders={self._debug_items_label(drag_folders)}"
                        )
                    event.acceptProposedAction()
                    self.setDropIndicatorShown(False)
                    return
            elif drag_tasks and is_hovering_center:
                decision = "ignore_task_center"
                signature = (decision, self._debug_item_label(target_item), len(drag_tasks))
                if signature != self._last_drag_hover_signature:
                    self._last_drag_hover_signature = signature
                    self._debug_log(
                        f"dragMoveEvent: decision={decision} ratio={ratio:.3f} target={self._debug_item_label(target_item)}"
                    )
                event.ignore()
                return

        signature = (
            decision,
            self._debug_item_label(target_item),
            None if ratio is None else round(ratio, 3),
            len(drag_tasks),
            len(drag_folders),
        )
        if signature != self._last_drag_hover_signature:
            self._last_drag_hover_signature = signature
            self._debug_log(
                f"dragMoveEvent: decision={decision} ratio={'n/a' if ratio is None else f'{ratio:.3f}'} "
                f"target={self._debug_item_label(target_item)} tasks={self._debug_items_label(drag_tasks)} "
                f"folders={self._debug_items_label(drag_folders)}"
            )
        self.setDropIndicatorShown(True)
        event.acceptProposedAction()

    def dropEvent(self, event):
        if event.source() != self:
            self._debug_log("dropEvent: 外部来源，交给父类处理")
            super().dropEvent(event)
            return

        try:
            drag_tasks, drag_folders = self._get_drag_items()
            if not drag_tasks and not drag_folders:
                self._debug_log("dropEvent: 没有可拖动节点，ignore")
                event.ignore()
                return

            target_item = self.itemAt(event.pos())
            ratio = self._hover_ratio(event, target_item) if target_item else None
            self._debug_log(
                f"dropEvent:start ratio={'n/a' if ratio is None else f'{ratio:.3f}'} target={self._debug_item_label(target_item)} "
                f"tasks={self._debug_items_label(drag_tasks)} folders={self._debug_items_label(drag_folders)}"
            )
            moved = False
            moved_items = []
            conflict_names = []

            self.blockSignals(True)
            try:
                if not target_item:
                    for folder in drag_folders:
                        clone = self._move_item(folder, None, self.topLevelItemCount())
                        if clone:
                            moved_items.append(clone)
                            moved = True
                    for task in drag_tasks:
                        clone = self._move_item(task, None, self.topLevelItemCount())
                        if clone:
                            moved_items.append(clone)
                            moved = True
                        else:
                            conflict_names.append(task.text(0))
                else:
                    target_role = target_item.data(0, Qt.UserRole)

                    if self._is_folder_role(target_role) and 0.15 < ratio < 0.85:
                        for folder in drag_folders:
                            if folder is target_item or self._is_ancestor(folder, target_item):
                                continue
                            clone = self._move_item(folder, target_item, target_item.childCount())
                            if clone:
                                moved_items.append(clone)
                                moved = True
                            else:
                                conflict_names.append(folder.text(0))
                        for task in drag_tasks:
                            if task is target_item or self._is_ancestor(task, target_item):
                                continue
                            clone = self._move_item(task, target_item, target_item.childCount())
                            if clone:
                                moved_items.append(clone)
                                moved = True
                            else:
                                conflict_names.append(task.text(0))
                        target_item.setExpanded(True)
                    else:
                        insert_before = ratio < 0.5
                        target_parent = target_item.parent()
                        if target_parent:
                            target_index = target_parent.indexOfChild(target_item)
                        else:
                            target_index = self.indexOfTopLevelItem(target_item)
                        if not insert_before:
                            target_index += 1

                        for folder in drag_folders:
                            if folder is target_item or self._is_ancestor(folder, target_item):
                                continue
                            clone = self._move_item(folder, target_parent, target_index)
                            if clone:
                                moved_items.append(clone)
                                target_index += 1
                                moved = True
                            else:
                                conflict_names.append(folder.text(0))

                        if drag_tasks:
                            for task in drag_tasks:
                                if task is target_item:
                                    continue
                                clone = self._move_item(task, target_parent, target_index)
                                if clone:
                                    moved_items.append(clone)
                                    target_index += 1
                                    moved = True
                                else:
                                    conflict_names.append(task.text(0))
            finally:
                self.blockSignals(False)

            self._drag_items = []
            if moved_items:
                move_info = []
                for moved_item in moved_items:
                    self._collect_move_info(moved_item, move_info)
                self._last_move_info = move_info
                self._debug_log(f"dropEvent:moved move_info={move_info}")
                self.blockSignals(True)
                try:
                    self.clearSelection()
                    for moved_item in moved_items:
                        moved_item.setSelected(True)
                    current_item = next(
                        (it for it in reversed(moved_items) if not self._is_folder_role(it.data(0, Qt.UserRole))),
                        moved_items[-1],
                    )
                    self.setCurrentItem(current_item)
                finally:
                    self.blockSignals(False)

            if conflict_names:
                conflict_text = "、".join(sorted(set(conflict_names)))
                self._debug_log(f"dropEvent: conflict_names={sorted(set(conflict_names))}")
                QMessageBox.warning(self, "无法移动任务", f"目标位置已存在同名任务：{conflict_text}")

            if moved:
                event.acceptProposedAction()
                self._debug_log(
                    f"dropEvent: accepted, 50ms 后发出 structure_changed；tree={format_tree_snapshot(self, self._is_folder_role)}"
                )
                QTimer.singleShot(50, self.structure_changed.emit)
            else:
                self._drag_in_progress = False
                self._debug_log("dropEvent: 未发生移动，ignore")
                event.ignore()
        except Exception as e:
            self._drag_items = []
            self._drag_in_progress = False
            self._debug_log(f"dropEvent: 异常 {e}")
            write_drag_debug(traceback.format_exc())
            raise

    def _is_ancestor(self, ancestor, item):
        parent = item.parent()
        while parent:
            if parent is ancestor:
                return True
            parent = parent.parent()
        return False

    def _has_name_conflict(self, item, new_parent):
        """文件夹仍禁止同级同名；任务允许重名，由 task_id 区分。"""
        if not item:
            return False

        item_role = item.data(0, Qt.UserRole)
        item_name = item.text(0)
        if not self._is_folder_role(item_role):
            return False

        if new_parent:
            for i in range(new_parent.childCount()):
                sibling = new_parent.child(i)
                if not sibling or sibling is item:
                    continue
                if self._is_folder_role(sibling.data(0, Qt.UserRole)) and sibling.text(0) == item_name:
                    return True
            return False

        for i in range(self.topLevelItemCount()):
            sibling = self.topLevelItem(i)
            if not sibling or sibling is item:
                continue
            if self._is_folder_role(sibling.data(0, Qt.UserRole)) and sibling.text(0) == item_name:
                return True
        return False

    def _move_item(self, item, new_parent, index):
        """统一的节点移动逻辑"""
        if not item:
            self._debug_log("_move_item: item 为空")
            return None

        if self._has_name_conflict(item, new_parent):
            self._debug_log(
                f"_move_item: 命名冲突 item={self._debug_item_label(item)} new_parent={self._debug_item_label(new_parent)}"
            )
            return None

        old_path = self._item_path_from_parent(item)
        old_parent = item.parent()
        if old_parent:
            old_idx = old_parent.indexOfChild(item)
        else:
            old_idx = self.indexOfTopLevelItem(item)

        if old_idx == -1:
            self._debug_log(f"_move_item: old_idx=-1 item={self._debug_item_label(item)}")
            return None

        if old_parent:
            clone = old_parent.takeChild(old_idx)
        else:
            clone = self.takeTopLevelItem(old_idx)
            if new_parent is None and old_idx < index:
                index = max(0, index - 1)

        if not clone:
            self._debug_log(f"_move_item: takeChild/takeTopLevelItem 失败 item={self._debug_item_label(item)}")
            return None

        if new_parent:
            max_idx = new_parent.childCount()
            if old_parent == new_parent and old_idx < index:
                index = max(0, index - 1)
            index = min(index, max_idx)
            new_parent.insertChild(index, clone)
        else:
            max_idx = self.topLevelItemCount()
            index = min(index, max_idx)
            self.insertTopLevelItem(index, clone)

        self._debug_log(
            f"_move_item: moved old_path={old_path!r} old_idx={old_idx} new_parent={self._debug_item_label(new_parent)} new_index={index} "
            f"clone={self._debug_item_label(clone)}"
        )
        return clone

CONFIG_PATH = os.path.join(BASE_DIR, "app_config.json")
CONFIG_LOAD_ERROR = ""
CONFIG_LOAD_BACKUP_PATH = ""

def _empty_config():
    return _migrate_config_schema({"tasks": {}, "task_data": {}, "extra_scan_dirs": []})

def _split_legacy_task_path(path):
    path = (path or "").strip().strip("/")
    if not path:
        return "", ""
    if "/" not in path:
        return "", path
    return path.rsplit("/", 1)

def _new_task_id(used_ids=None):
    used_ids = used_ids or set()
    while True:
        task_id = f"task_{uuid.uuid4().hex[:12]}"
        if task_id not in used_ids:
            used_ids.add(task_id)
            return task_id

def _normalize_task_meta(task_id, meta):
    meta = meta if isinstance(meta, dict) else {}
    name = (meta.get("name") or "").strip()
    folder = (meta.get("folder") or "").strip().replace("\\", "/").strip("/")
    if not name:
        _, legacy_name = _split_legacy_task_path(task_id)
        name = legacy_name or task_id
    return {
        "name": name,
        "folder": folder,
        "created_at": meta.get("created_at") or datetime.now().isoformat(timespec="seconds")
    }

def _migrate_config_schema(config):
    config = config if isinstance(config, dict) else {}
    config.setdefault("tasks", {})
    config.setdefault("task_data", {})
    config.setdefault("extra_scan_dirs", [])
    config.setdefault("folders", [])
    config.setdefault("layout", {})
    config.setdefault("tasks_layout", {})
    config.setdefault("schedule_bundles", {})

    tasks = config.get("tasks", {})
    task_data = config.get("task_data", {})
    task_meta = config.get("task_meta", {})
    tasks_layout = config.get("tasks_layout", {})
    layout = config.get("layout", {})

    used_ids = set()
    new_tasks = {}
    new_task_data = {}
    new_task_meta = {}
    new_tasks_layout = {}
    key_to_id = {}

    already_new_schema = bool(tasks) and all(
        isinstance(task_meta.get(task_id), dict) and task_meta[task_id].get("name")
        for task_id in tasks.keys()
    )

    if already_new_schema:
        for task_id, actions in tasks.items():
            used_ids.add(task_id)
            new_tasks[task_id] = actions
            new_task_meta[task_id] = _normalize_task_meta(task_id, task_meta.get(task_id))
            new_task_data[task_id] = task_data.get(task_id, [])
            if task_id in tasks_layout:
                new_tasks_layout[task_id] = tasks_layout.get(task_id, [])
            key_to_id[task_id] = task_id
    else:
        for legacy_path, actions in tasks.items():
            task_id = _new_task_id(used_ids)
            folder, name = _split_legacy_task_path(legacy_path)
            new_tasks[task_id] = actions
            new_task_meta[task_id] = _normalize_task_meta(task_id, {"name": name, "folder": folder})
            new_task_data[task_id] = task_data.get(legacy_path, [])
            if legacy_path in tasks_layout:
                new_tasks_layout[task_id] = tasks_layout.get(legacy_path, [])
            key_to_id[legacy_path] = task_id

    for task_id in list(new_tasks.keys()):
        new_task_data.setdefault(task_id, [])
        new_task_meta[task_id] = _normalize_task_meta(task_id, new_task_meta.get(task_id))

    layout_order = []
    for item in layout.get("task_order", []):
        mapped = key_to_id.get(item, item if item in new_tasks else "")
        if mapped and mapped in new_tasks and mapped not in layout_order:
            layout_order.append(mapped)
    for task_id in new_tasks.keys():
        if task_id not in layout_order:
            layout_order.append(task_id)

    last_task = layout.get("last_task", "")
    last_task = key_to_id.get(last_task, last_task if last_task in new_tasks else "")
    if not last_task and layout_order:
        last_task = layout_order[0]

    normalized_folders = set()
    for folder_path in config.get("folders", []):
        folder_path = (folder_path or "").strip().replace("\\", "/").strip("/")
        if not folder_path:
            continue
        parts = folder_path.split("/")
        for i in range(1, len(parts) + 1):
            normalized_folders.add("/".join(parts[:i]))
    for meta in new_task_meta.values():
        folder_path = meta.get("folder", "")
        if not folder_path:
            continue
        parts = folder_path.split("/")
        for i in range(1, len(parts) + 1):
            normalized_folders.add("/".join(parts[:i]))

    config["tasks"] = new_tasks
    config["task_data"] = new_task_data
    config["task_meta"] = new_task_meta
    config["tasks_layout"] = new_tasks_layout
    config["folders"] = sorted(normalized_folders, key=lambda p: (p.count("/"), p.lower()))
    config["layout"] = layout
    config["layout"]["task_order"] = layout_order
    config["layout"]["last_task"] = last_task
    schedule_bundles = config.get("schedule_bundles", {})
    if not isinstance(schedule_bundles, dict):
        schedule_bundles = {}
    normalized_bundles = {}
    for bundle_name, bundle in schedule_bundles.items():
        if not str(bundle_name).strip():
            continue
        bundle = bundle if isinstance(bundle, dict) else {}
        items = bundle.get("items", [])
        normalized_items = []
        if isinstance(items, list):
            for item in items:
                if isinstance(item, dict):
                    task_id = item.get("task_id", "")
                    enabled = bool(item.get("enabled", True))
                else:
                    task_id = item
                    enabled = True
                if task_id in new_tasks:
                    normalized_items.append({"task_id": task_id, "enabled": enabled})
        timer_cfg = bundle.get("timer_config")
        if not isinstance(timer_cfg, dict):
            timer_cfg = None
        normalized_bundles[str(bundle_name).strip()] = {
            "items": normalized_items,
            "timer_config": timer_cfg,
            "timer_enabled": bool(bundle.get("timer_enabled", False))
        }
    if not normalized_bundles:
        normalized_bundles["默认排程包"] = {"items": [], "timer_config": None, "timer_enabled": False}
    current_bundle = layout.get("current_schedule_bundle", "")
    if current_bundle not in normalized_bundles:
        current_bundle = next(iter(normalized_bundles.keys()))
    config["schedule_bundles"] = normalized_bundles
    config["layout"]["current_schedule_bundle"] = current_bundle
    return config

def load_config():
    global CONFIG_LOAD_ERROR, CONFIG_LOAD_BACKUP_PATH
    CONFIG_LOAD_ERROR = ""
    CONFIG_LOAD_BACKUP_PATH = ""
    if os.path.exists(CONFIG_PATH):
        try:
            with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
                d = _migrate_config_schema(json.load(f))
                # 同步到全局变量
                for p in d["extra_scan_dirs"]:
                    if os.path.exists(p): _extra_scan_dirs.add(os.path.normpath(p))
                return d
        except Exception as e:
            CONFIG_LOAD_ERROR = f"{type(e).__name__}: {e}"
            try:
                stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_path = os.path.join(BASE_DIR, f"app_config.broken_{stamp}.json")
                shutil.copy2(CONFIG_PATH, backup_path)
                CONFIG_LOAD_BACKUP_PATH = backup_path
            except Exception:
                CONFIG_LOAD_BACKUP_PATH = ""
            return _empty_config()
    return _empty_config()
def save_config(config):
    config = _migrate_config_schema(config)
    with open(CONFIG_PATH, 'w', encoding='utf-8') as f: json.dump(config, f, indent=4, ensure_ascii=False)

# --- Floating OSD Progress Window ---
class FloatingProgressWindow(QWidget):
    """始终置顶的半透明微型控制台，支持滚动日志显示。"""
    request_pause      = pyqtSignal()
    request_stop       = pyqtSignal()
    request_skip_step  = pyqtSignal()   # 跳过当前步骤，执行下一步
    request_next_row   = pyqtSignal()   # 放弃当前行剩余步骤，跳到下一行
    request_retry_step = pyqtSignal()   # 重试当前步骤

    def __init__(self, parent=None):
        super().__init__(parent)
        self._pinned = True   # 是否始终置顶
        self._collapsed = False  # 是否折叠日志区
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint | Qt.Tool | Qt.X11BypassWindowManagerHint)
        self.setAttribute(Qt.WA_TranslucentBackground, True)
        self.setAttribute(Qt.WA_ShowWithoutActivating, True)

        self.resize(520, 210)  # 增大宽度以容纳更多按钮

        # 默认位置：屏幕上方正中央
        self._move_to_top_center()

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        # 背景容器
        self.container = QFrame()
        self.container.setStyleSheet("""
            QFrame {
                background-color: rgba(20, 20, 20, 220);
                border: 1px solid #555;
                border-radius: 8px;
            }
            QLabel { font-weight: 400; color: #222; color: #e0e0e0; font-family: 'Consolas', '微软雅黑'; font-size: 11px; }
        """)
        container_layout = QVBoxLayout(self.container)
        container_layout.setContentsMargins(10, 8, 10, 8)
        container_layout.setSpacing(5)

        # --- 第一行：任务信息 ---
        self.lbl_info = QLabel("等待开始...")
        self.lbl_info.setStyleSheet("font-weight: bold; color: #82b1ff;")
        container_layout.addWidget(self.lbl_info)

        # --- 第二行：进度条 ---
        prog_row = QHBoxLayout()
        self.bar = QProgressBar()
        self.bar.setFixedHeight(10)
        self.bar.setTextVisible(False)
        self.bar.setStyleSheet("""
            QProgressBar { background-color: #333; border: none; border-radius: 5px; }
            QProgressBar::chunk { background-color: #4CAF50; border-radius: 5px; }
        """)
        prog_row.addWidget(self.bar)
        self.lbl_pct = QLabel("0%")
        self.lbl_pct.setFixedWidth(35)
        prog_row.addWidget(self.lbl_pct)
        container_layout.addLayout(prog_row)

        # --- 第三行：实时动作/延时描述 ---
        self.lbl_detail = QLabel("准备就绪")
        self.lbl_detail.setStyleSheet("color: #69f0ae; font-weight: bold; font-size: 12px; padding: 2px 0;")
        container_layout.addWidget(self.lbl_detail)

        # --- 第四行：主控制按钮（暂停/跳步/下一行/重试/停止）---
        _btn_style_base = "QPushButton { color: white; border-radius: 3px; font-size: 11px; font-weight: bold; padding: 1px 4px; } QPushButton:hover { opacity: 0.85; }"

        ctrl_row = QHBoxLayout()
        ctrl_row.setSpacing(4)

        self.btn_pause = QPushButton("⏸ 暂停")
        self.btn_pause.setFixedHeight(24)
        self.btn_pause.setStyleSheet(_btn_style_base + "QPushButton { background-color: #ff9800; }")
        self.btn_pause.setToolTip("暂停 / 继续执行")
        self.btn_pause.clicked.connect(self.request_pause.emit)
        ctrl_row.addWidget(self.btn_pause)

        self.btn_skip_step = QPushButton("⏭ 跳步")
        self.btn_skip_step.setFixedHeight(24)
        self.btn_skip_step.setStyleSheet(_btn_style_base + "QPushButton { background-color: #2196F3; }")
        self.btn_skip_step.setToolTip("跳过当前步骤，直接执行下一步")
        self.btn_skip_step.clicked.connect(self.request_skip_step.emit)
        ctrl_row.addWidget(self.btn_skip_step)

        self.btn_next_row = QPushButton("⏩ 下一行")
        self.btn_next_row.setFixedHeight(24)
        self.btn_next_row.setStyleSheet(_btn_style_base + "QPushButton { background-color: #9C27B0; }")
        self.btn_next_row.setToolTip("放弃当前行剩余步骤，跳到下一行开始执行")
        self.btn_next_row.clicked.connect(self.request_next_row.emit)
        ctrl_row.addWidget(self.btn_next_row)

        self.btn_retry = QPushButton("🔁 重试")
        self.btn_retry.setFixedHeight(24)
        self.btn_retry.setStyleSheet(_btn_style_base + "QPushButton { background-color: #009688; }")
        self.btn_retry.setToolTip("重新执行当前步骤（不跳过，重来一次）")
        self.btn_retry.clicked.connect(self.request_retry_step.emit)
        ctrl_row.addWidget(self.btn_retry)

        self.btn_stop = QPushButton("🛑 停止")
        self.btn_stop.setFixedHeight(24)
        self.btn_stop.setStyleSheet(_btn_style_base + "QPushButton { background-color: #f44336; }")
        self.btn_stop.setToolTip("终止整个任务")
        self.btn_stop.clicked.connect(self.request_stop.emit)
        ctrl_row.addWidget(self.btn_stop)

        container_layout.addLayout(ctrl_row)

        # --- 第五行：辅助功能按钮（复制日志/置顶/折叠）---
        _aux_style = "QPushButton { color: #ccc; border-radius: 3px; font-size: 10px; padding: 1px 6px; background-color: rgba(80,80,80,180); border: 1px solid #666; } QPushButton:hover { background-color: rgba(110,110,110,200); }"

        aux_row = QHBoxLayout()
        aux_row.setSpacing(4)

        self.btn_copy_log = QPushButton("📋 复制日志")
        self.btn_copy_log.setFixedHeight(20)
        self.btn_copy_log.setStyleSheet(_aux_style)
        self.btn_copy_log.setToolTip("将小窗口日志复制到剪贴板")
        self.btn_copy_log.clicked.connect(self._copy_log)
        aux_row.addWidget(self.btn_copy_log)

        self.btn_pin = QPushButton("📌 取消置顶")
        self.btn_pin.setFixedHeight(20)
        self.btn_pin.setStyleSheet(_aux_style)
        self.btn_pin.setToolTip("切换窗口是否始终在最前面")
        self.btn_pin.clicked.connect(self._toggle_pin)
        aux_row.addWidget(self.btn_pin)

        self.btn_collapse = QPushButton("🔽 折叠")
        self.btn_collapse.setFixedHeight(20)
        self.btn_collapse.setStyleSheet(_aux_style)
        self.btn_collapse.setToolTip("折叠/展开日志区，减少遮挡")
        self.btn_collapse.clicked.connect(self._toggle_collapse)
        aux_row.addWidget(self.btn_collapse)

        aux_row.addStretch()
        container_layout.addLayout(aux_row)

        # --- 第六行：微型滚动日志区 ---
        self.log_area = QTextEdit()
        self.log_area.setReadOnly(True)
        self.log_area.setFrameStyle(QFrame.NoFrame)
        self.log_area.setFixedHeight(65)
        self.log_area.setStyleSheet("""
            QTextEdit {
                background-color: rgba(0, 0, 0, 100);
                color: #aaa;
                font-family: 'Consolas', 'Courier New';
                font-size: 10px;
                border-radius: 4px;
            }
        """)
        self.log_area.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        container_layout.addWidget(self.log_area)

        layout.addWidget(self.container)

        self._dragging = False
        self._drag_pos = QPoint()

        # 强制 Windows 置顶
        if sys.platform == 'win32':
            self._force_topmost(-1)

        self._dragging = False
        self._drag_pos = QPoint()

    def _move_to_top_center(self, top_margin=20):
        """把悬浮窗放到当前主屏幕的上方正中央。"""
        screen = QApplication.primaryScreen()
        if not screen:
            return
        screen_geo = screen.availableGeometry()
        x = screen_geo.x() + max(0, (screen_geo.width() - self.width()) // 2)
        y = screen_geo.y() + max(0, int(top_margin))
        self.move(x, y)

    def update_progress(self, task_name, loop, group, total_groups, step, total_steps, step_name, percent):
        info = f"<b>{task_name}</b> | 组 {group}/{total_groups} | 步 {step}/{total_steps}: <font color='#82b1ff'>{step_name}</font>"
        self.lbl_info.setText(info)
        self.bar.setValue(percent)
        self.lbl_pct.setText(f"{percent}%")
        # 切换步骤时清空之前的延时详情
        self.lbl_detail.setText("")
        if not self.isVisible():
            self.show()
            # 再次强制置顶（针对某些全屏应用抢占置顶的情况）
            if sys.platform == 'win32':
                self._force_topmost(-1)

    def mousePressEvent(self, event):
        child = self.childAt(event.pos())
        if isinstance(child, (QPushButton, QTextEdit)):
            event.ignore()
            return super().mousePressEvent(event)
        if event.button() == Qt.LeftButton:
            self._dragging = True
            self._drag_pos = event.globalPos() - self.frameGeometry().topLeft()
            event.accept()

    def mouseMoveEvent(self, event):
        if not self._dragging:
            return super().mouseMoveEvent(event)
        if self._dragging and event.buttons() & Qt.LeftButton:
            self.move(event.globalPos() - self._drag_pos)
            event.accept()

    def mouseReleaseEvent(self, event):
        self._dragging = False
        super().mouseReleaseEvent(event)

    def update_detail(self, text):
        """更新具体的动作详情或倒计时描述。"""
        self.lbl_detail.setText(text)

    def add_log(self, msg, color="white"):
        """向微型控制台添加一条日志。"""
        time_str = datetime.now().strftime('%H:%M:%S')
        color_map = {
            "red": "#ff5252", "green": "#69f0ae", "blue": "#82b1ff",
            "orange": "#ffab40", "purple": "#ea80fc", "gray": "#888888",
            "black": "#cccccc", "white": "#cccccc"
        }
        c = color_map.get(color, "#cccccc")
        html = f'<div style="color:{c}">[{time_str}] {msg}</div>'
        self.log_area.append(html)
        # 自动滚动到底部
        self.log_area.verticalScrollBar().setValue(self.log_area.verticalScrollBar().maximum())
        # 限制日志行数，防止内存占用过高
        if self.log_area.document().blockCount() > 50:
            cursor = self.log_area.textCursor()
            cursor.movePosition(cursor.Start)
            cursor.select(cursor.BlockUnderCursor)
            cursor.removeSelectedText()
            cursor.deleteChar()

    def _copy_log(self):
        """将日志区内容复制到剪贴板。"""
        plain_text = self.log_area.toPlainText()
        if plain_text.strip():
            pyperclip.copy(plain_text)
            # 短暂改变按钮文字作为反馈
            self.btn_copy_log.setText("✅ 已复制")
            QTimer.singleShot(1500, lambda: self.btn_copy_log.setText("📋 复制日志"))
        else:
            self.btn_copy_log.setText("⚠️ 无日志")
            QTimer.singleShot(1500, lambda: self.btn_copy_log.setText("📋 复制日志"))

    def _toggle_pin(self):
        """切换窗口是否始终置顶。"""
        self._pinned = not self._pinned
        if self._pinned:
            self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)
            self.btn_pin.setText("📌 取消置顶")
            if sys.platform == 'win32':
                self._force_topmost(-1)
        else:
            flags = self.windowFlags() & ~Qt.WindowStaysOnTopHint
            self.setWindowFlags(flags)
            self.btn_pin.setText("📌 恢复置顶")
            if sys.platform == 'win32':
                self._force_topmost(-2)
        self.show()  # 重新显示以应用新的窗口标志

    def _force_topmost(self, z_order):
        try:
            ctypes.windll.user32.SetWindowPos(int(self.winId()), z_order, 0, 0, 0, 0, 0x0001 | 0x0002 | 0x0040)
        except Exception as e:
            log_internal_issue("设置 OSD 窗口置顶状态失败", e)

    def _toggle_collapse(self):
        """折叠/展开日志区，减少对操作区域的遮挡。"""
        self._collapsed = not self._collapsed
        if self._collapsed:
            self.log_area.hide()
            self.btn_collapse.setText("🔼 展开")
            self.adjustSize()
        else:
            self.log_area.show()
            self.btn_collapse.setText("🔽 折叠")
            self.adjustSize()
        self._move_to_top_center()

# --- Engine ---
# Row status constants
ROW_STATUS_OK    = "✅"
ROW_STATUS_FAIL  = "❌"
ROW_STATUS_SKIP  = "⏭️"
ROW_STATUS_DEFER = "⏸️"
ROW_STATUS_MANUAL = "⚠️"

class AutoEngine(QThread):
    log_sig        = pyqtSignal(str, str)
    prog_sig       = pyqtSignal(int)
    done_sig       = pyqtSignal()
    pause_sig      = pyqtSignal(int, int, int)
    # (row_index, status_emoji)
    row_status_sig = pyqtSignal(int, str)
    # (row_index, step_index) — highlight current cell; -1/-1 clears
    highlight_sig  = pyqtSignal(int, int)
    # emitted when engine pauses itself mid-step (hotkey pause)
    hotkey_paused_sig = pyqtSignal(int, int, int)
    detail_sig        = pyqtSignal(str) # [新增] 详细描述信号（用于 OSD 倒计时等）
    deferred_queue_sig = pyqtSignal(object)
    # (dict) per-row structured result, used for failure aggregation / window extraction in UI.
    row_result_sig = pyqtSignal(object)

    def __init__(self, actions, data_list, loop_delay, start_t=0, start_s=0, loops=1, start_l=0,
                 retry_count=0, on_error="stop", dry_run=False, ignore_data=False, standardize_window=False):
        super().__init__()
        self.actions     = actions
        self.data_list   = data_list if data_list else [{}]
        self.ignore_data = ignore_data # 核心修复：是否忽略数据表（流程测试模式）
        self.loop_delay  = loop_delay
        self.start_t     = start_t
        self.start_s     = start_s
        self.loops       = loops
        self.start_l     = start_l
        self.retry_count = retry_count   # 0 = no retry
        # on_error:
        # - "stop": stop whole task immediately
        # - "skip": skip current step and continue (may still end up row OK)
        # - "fail_row": mark current row failed and continue next row
        self.on_error    = on_error
        self.dry_run     = dry_run       # True = only print, no actual execution
        self.standardize_window = standardize_window # [新增] 是否强制标准化窗口大小位置
        self._stop       = False
        self._paused     = False
        self._skip_step  = False  # [v3] 跳过当前步骤
        self._next_row   = False  # [v3] 放弃当前行剩余步骤
        self._retry_step = False  # [v3] 重试当前步骤
        self._cur_l      = 0
        self._cur_t      = 0
        self._cur_s      = 0
        self._last_percent = 0
        self._final_status = "idle"
        self._stop_reason = ""
        self._last_error = ""
        self._deferred_queue = []
        self._defer_seq = 0
        self._row_fail_ctx = None  # per-row failure context snapshot
        self._row_runtime_ctx = None  # per-row runtime context snapshot (for successful rows too)

    def _extract_hwnd_from_value(self, v):
        """从类似 'xxx::hwnd=123' 的字符串中解析 hwnd。"""
        try:
            if not v:
                return None
            s = str(v)
            if "::hwnd=" not in s:
                return None
            tail = s.split("::hwnd=", 1)[1].strip()
            digits = "".join(ch for ch in tail if ch.isdigit())
            return int(digits) if digits else None
        except Exception:
            return None

    def _snapshot_foreground_window(self):
        """抓取当前前台窗口信息，用于失败定位。"""
        if sys.platform != "win32":
            return {"hwnd": None, "title": "", "class": ""}
        try:
            hwnd = ctypes.windll.user32.GetForegroundWindow()
            hwnd = int(hwnd) if hwnd else None
            return {
                "hwnd": hwnd,
                "title": get_window_text(hwnd) if hwnd else "",
                "class": get_window_class_name(hwnd) if hwnd else ""
            }
        except Exception:
            return {"hwnd": None, "title": "", "class": ""}

    def _set_row_fail_ctx(self, l_idx, t_idx, s_idx, step_name, act_type, raw_action, val, err):
        """记录一次行级失败信息（不影响主流程控制）。"""
        try:
            fg = self._snapshot_foreground_window()
            self._row_fail_ctx = {
                "loop_index": int(l_idx),
                "row_index": int(t_idx),
                "step_index": int(s_idx),
                "step_name": str(step_name or ""),
                "act_type": str(act_type or ""),
                "action": str(raw_action or ""),
                "value": str(val or ""),
                "error": str(err or ""),
                "target_hwnd": self._extract_hwnd_from_value(val),
                "foreground": fg,
            }
        except Exception:
            self._row_fail_ctx = None

    def _set_row_runtime_ctx(self, l_idx, t_idx, s_idx, step_name, act_type, raw_action, val):
        """记录当前行最近一次成功执行时的窗口/步骤上下文，便于子任务管理器复用。"""
        try:
            fg = self._snapshot_foreground_window()
            self._row_runtime_ctx = {
                "loop_index": int(l_idx),
                "row_index": int(t_idx),
                "step_index": int(s_idx),
                "step_name": str(step_name or ""),
                "act_type": str(act_type or ""),
                "action": str(raw_action or ""),
                "value": str(val or ""),
                "error": "",
                "target_hwnd": self._extract_hwnd_from_value(val),
                "foreground": fg,
            }
        except Exception:
            self._row_runtime_ctx = None

    def stop(self):
        self._stop_reason = "stopped"
        self._stop = True
        self._paused = False  # unblock spin if paused

    def pause(self):
        self._paused = True

    def resume(self):
        self._paused = False

    def skip_step(self):
        """[v3] 跳过当前步骤，在步骤边界或延时等待中生效。"""
        self._skip_step = True
        self._paused = False  # 如果当前处于暂停状态，同时解除暂停

    def next_row(self):
        """[v3] 放弃当前行剩余步骤，跳到下一行开始。"""
        self._next_row = True
        self._paused = False

    def retry_step(self):
        """[v3] 重试当前步骤：将 s_idx 回退一步，让主循环重新执行当前步。"""
        self._retry_step = True
        self._paused = False

    def _spin_while_paused(self):
        """Block engine thread while paused; stop breaks out."""
        while self._paused and not self._stop:
            time.sleep(0.05)

    def _cooperative_abort(self):
        """在长步骤中主动检查人工控制信号，让步骤尽快让出控制权。"""
        if self._stop:
            raise ExecutionInterrupted("任务已停止")
        if self._next_row or self._skip_step or self._retry_step:
            raise ExecutionInterrupted("当前步骤已被人工控制打断")

    def _interruptible_sleep(self, seconds):
        """Sleep in small increments so stop/pause can break out early."""
        if seconds <= 0:
            return
        end = time.time() + float(seconds)
        while time.time() < end:
            if self._stop:
                return
            if self._paused:
                pause_started = time.time()
                self._spin_while_paused()
                if self._stop:
                    return
                end += time.time() - pause_started
            time.sleep(min(0.05, end - time.time()))

    def _cooperative_sleep(self, seconds):
        self._interruptible_sleep(seconds)
        self._cooperative_abort()

    def _abort_subprocess(self, process):
        if not process:
            return
        try:
            if process.poll() is not None:
                return
        except Exception:
            pass
        try:
            process.terminate()
        except Exception:
            pass
        try:
            process.wait(timeout=1.0)
        except Exception:
            try:
                process.kill()
            except Exception:
                pass

    def _wait_process_interruptibly(self, process, poll_interval=0.1):
        """等待子进程结束，同时响应暂停/停止/跳步等人工控制。"""
        while True:
            self._cooperative_abort()
            if self._paused:
                self._spin_while_paused()
                self._cooperative_abort()
            try:
                ret = process.poll()
            except Exception:
                ret = None
            if ret is not None:
                return ret
            time.sleep(max(0.02, float(poll_interval)))

    def _validate_step_guard(self, act, act_type):
        """执行前界面守卫：优先比对录制快照，退化到窗口标题/类名校验。"""
        guarded_types = {
            "click", "double_click", "right_click", "move", "hover_click",
            "scroll", "input", "clear_input", "clear_input_plus", "upload", "drag_file"
        }
        if self.dry_run or act_type not in guarded_types:
            return
        if not bool(act.get("guard_enabled", True)):
            return
        x, y = act.get('x'), act.get('y')
        if x is None or y is None:
            return
        x, y = int(x), int(y)
        if x == 0 and y == 0:
            return

        guard_image = str(act.get("guard_image", "") or "").strip()
        guard_region = act.get("guard_region")
        guard_threshold = float(act.get("guard_threshold", 0.72) or 0.72)
        if guard_image and os.path.exists(guard_image) and isinstance(guard_region, (list, tuple)) and len(guard_region) >= 4:
            score = evaluate_guard_snapshot_similarity(guard_image, guard_region)
            if score < guard_threshold:
                raise StepGuardMismatchError(
                    f"界面守卫未通过：当前画面与录制时差异较大（相似度 {score:.0%} < {guard_threshold:.0%}）"
                )
            self.log_sig.emit(f"🧭 界面守卫通过：相似度 {score:.0%}", "gray")
            return

        guard_class = str(act.get("guard_window_class", "") or "").strip()
        guard_title = str(act.get("guard_window_title", "") or "").strip()
        if not guard_class and not guard_title:
            return

        actual_hwnd = get_root_window_from_point(x, y)
        if not actual_hwnd:
            raise StepGuardMismatchError(f"界面守卫未通过：无法定位坐标 ({x}, {y}) 所在窗口")
        actual_class = str(get_window_class_name(actual_hwnd) or "").strip()
        actual_title = str(get_window_text(actual_hwnd) or "").strip()

        if guard_class and actual_class != guard_class:
            raise StepGuardMismatchError(f"界面守卫未通过：窗口类名不匹配（当前={actual_class or '空'}，录制={guard_class}）")
        if guard_title:
            title_ok = (
                actual_title == guard_title or
                (guard_title and guard_title in actual_title) or
                (actual_title and actual_title in guard_title)
            )
            if not title_ok:
                raise StepGuardMismatchError(f"界面守卫未通过：窗口标题不匹配（当前={actual_title or '空'}，录制={guard_title}）")
        self.log_sig.emit(f"🧭 窗口守卫通过：{actual_title or actual_class or '已匹配'}", "gray")

    def _write_progress_status(self, percent, cur_loop, cur_group, total_groups, cur_step, total_steps, step_name="", task_name="", status="running"):
        """将当前执行进度实时写入 progress_status.json，供无界面监控工具读取。"""
        try:
            self._last_percent = percent
            now = datetime.now()
            elapsed = (now - self._start_time).total_seconds() if hasattr(self, '_start_time') else 0
            eta_str = ""
            if percent > 0 and elapsed > 0:
                total_est = elapsed / (percent / 100.0)
                remaining = max(0, total_est - elapsed)
                h, m, s_r = int(remaining // 3600), int((remaining % 3600) // 60), int(remaining % 60)
                eta_str = f"{h:02d}:{m:02d}:{s_r:02d}" if h > 0 else f"{m:02d}:{s_r:02d}"
            elapsed_str = self._format_time_display(int(elapsed))
            status_data = {
                "task_name":    task_name,
                "status":       status,
                "percent":      percent,
                "cur_loop":     cur_loop,
                "total_loops":  self.loops,
                "cur_group":    cur_group,
                "total_groups": total_groups,
                "cur_step":     cur_step,
                "total_steps":  total_steps,
                "step_name":    step_name,
                "elapsed":      elapsed_str,
                "eta":          eta_str,
                "updated_at":   now.strftime("%Y-%m-%d %H:%M:%S")
            }
            status_path = os.path.join(BASE_DIR, "progress_status.json")
            with open(status_path, "w", encoding="utf-8") as f:
                json.dump(status_data, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _wait_with_countdown(self, seconds, prefix="⏳ 等待中"):
        if seconds <= 0:
            return
        end = time.time() + float(seconds)
        last_display = None
        while True:
            if self._stop or self._skip_step or self._next_row:
                break
            if self._paused:
                pause_started = time.time()
                self._spin_while_paused()
                if self._stop or self._skip_step or self._next_row:
                    break
                end += time.time() - pause_started

            remaining = end - time.time()
            if remaining <= 0:
                break

            remaining_text = self._format_time_display(max(1, int(math.ceil(remaining))))
            msg = f"{prefix}: 剩余 {remaining_text}"
            if msg != last_display:
                self.log_sig.emit(msg, "gray")
                self.detail_sig.emit(msg)
                last_display = msg

            time.sleep(min(0.1, max(0.01, remaining)))

        self.detail_sig.emit("") # 等待结束清空详情

    def _format_time_display(self, s):
        if s < 60: return f"{s}s"
        if s < 3600: return f"{s//60:02d}:{s%60:02d}"
        return f"{s//3600:02d}:{(s%3600)//60:02d}:{s%60:02d}"

    def _replace_vars(self, text, data_row):
        if not isinstance(text, str): return text
        
        # 核心修复：解析序号变量 {{n}}
        for i in range(len(self.actions)):
            step_num = i + 1
            act = self.actions[i]
            step_name = act.get('name', f'步骤{step_num}')
            
            # 取值优先级：
            # 如果是 ignore_data 模式（流程测试），强制只取 act['value']，不看 data_row
            # 如果是批量执行模式，data_row 优先，act['value'] 为兜底
            if getattr(self, 'ignore_data', False):
                val = str(act.get('value', ''))
            else:
                val = str(data_row.get(step_name, act.get('value', '')))

            # 增强匹配：支持 步骤n, 第n步, step n
            for pattern in [f"{{{{{step_num}}}}}", f"{{{{步骤{step_num}}}}}", f"{{{{第{step_num}步}}}}", f"{{{{step{step_num}}}}}"]:
                if pattern in text:
                    text = text.replace(pattern, val)
                    self.log_sig.emit(f"📝 [解析成功] {pattern} → {val}", "purple")

        # 核心修复：解析名称变量 {{步骤名}}
        for act in self.actions:
            name = act.get('name', '').strip()
            if name and f"{{{{{name}}}}}" in text:
                if getattr(self, 'ignore_data', False):
                    val = str(act.get('value', ''))
                else:
                    val = str(data_row.get(name, act.get('value', '')))
                text = text.replace(f"{{{{{name}}}}}", val)
                self.log_sig.emit(f"📝 [解析成功] {{ {name} }} → {val}", "purple")

        # 3. 替换数据表变量 (CSV/Excel)
        if data_row:
            for k, v in data_row.items():
                if f"{{{{{k}}}}}" in text:
                    text = text.replace(f"{{{{{k}}}}}", str(v))
                    self.log_sig.emit(f"📝 [解析成功] 数据列 [{k}] → {str(v)}", "purple")
        
        # 4. 替换系统变量
        now = datetime.now()
        vars_map = {"日期": now.strftime("%Y-%m-%d"), "时间": now.strftime("%H:%M:%S"), "行号": f"{self._cur_t + 1:03d}", "循环号": str(self._cur_l + 1)}
        for k, v in vars_map.items():
            text = text.replace(f"{{{{{k}}}}}", v)

        return text

    def _standardize_browser_window(self):
        """[核心功能] 强制将当前活动的浏览器窗口标准化：移动到 (0,0) 并设为全屏。"""
        if self.dry_run: return
        try:
            active_win = gw.getActiveWindow()
            if not active_win: return
            self.log_sig.emit("📏 正在标准化窗口位置与大小...", "blue")
            self.detail_sig.emit("📏 标准化窗口中...")
            if active_win.isMaximized:
                active_win.restore()
                time.sleep(0.2)
            active_win.moveTo(0, 0)
            time.sleep(0.2)
            active_win.maximize()
            time.sleep(0.5)
            self.log_sig.emit("✅ 窗口已标准化 (0,0 + 全屏)", "green")
        except Exception as e:
            self.log_sig.emit(f"⚠️ 窗口标准化失败: {e}", "orange")


    def _resolve_step_index(self, target_name):
        target_name = str(target_name or "").strip()
        if not target_name:
            return None
        for ji, ja in enumerate(self.actions):
            act_name = str(ja.get('name', '')).strip()
            if act_name == target_name:
                return ji
        target_name_lower = target_name.lower()
        for ji, ja in enumerate(self.actions):
            act_name = str(ja.get('name', '')).strip()
            if act_name and act_name.lower() == target_name_lower:
                return ji
        clean_target = target_name.split(' ')[-1]
        for ji, ja in enumerate(self.actions):
            act_name = ja.get('name', '').strip()
            if act_name == target_name or target_name in act_name or (clean_target and clean_target in act_name):
                return ji
        return None

    def _parse_defer_value(self, raw_val, current_step_idx):
        raw_text = str(raw_val or "").replace('\r', '').replace('\n', ' ').strip()
        parts = [p.strip() for p in raw_text.split('|')]
        delay_text = parts[0] if parts else ""
        try:
            seconds = float(delay_text)
        except Exception:
            raise RuntimeError("延后执行参数格式错误：请填写“秒数 | 下一步”或“秒数 | 指定步骤 | 步骤名”")
        if seconds < 0:
            raise RuntimeError("延后执行秒数不能小于 0")

        known_policies = {
            "到时优先恢复": "due_first",
            "到时间优先恢复": "due_first",
            "优先恢复": "due_first",
            "整轮后恢复": "after_round",
            "整轮后再恢复": "after_round",
            "一轮后恢复": "after_round"
        }
        mode = parts[1] if len(parts) > 1 else "下一步"
        target_text = parts[2] if len(parts) > 2 else ""
        policy_text = parts[3] if len(parts) > 3 else ""
        mode_norm = str(mode).strip().lower()

        if not policy_text:
            if len(parts) == 2 and mode in known_policies:
                policy_text = mode
                mode = "下一步"
                mode_norm = "next"
            elif len(parts) >= 3 and target_text in known_policies and mode in ("", "下一步"):
                policy_text = target_text
                target_text = ""
            else:
                policy_text = "整轮后再恢复"

        policy_key = known_policies.get(policy_text, "")
        if not policy_key:
            policy_key = "after_round"
            policy_text = "整轮后再恢复"

        if mode in ("", "下一步") or mode_norm in ("next", "continue", "顺序", "顺序执行"):
            resume_step = current_step_idx + 1
            if resume_step >= len(self.actions):
                raise RuntimeError("延后执行后已经没有后续步骤，请把它放在需要稍后继续的位置之前")
            resume_label = self.actions[resume_step].get('name', f'步骤{resume_step+1}')
            resume_mode = "下一步"
        else:
            if mode in ("指定步骤", "指定", "跳转到", "恢复到") or mode_norm in ("step", "goto", "resume"):
                target_name = target_text.strip()
            else:
                target_name = str(mode).strip()
            if not target_name:
                raise RuntimeError("延后执行缺少恢复步骤名")
            resume_step = self._resolve_step_index(target_name)
            if resume_step is None:
                raise RuntimeError(f"延后执行无法匹配恢复步骤: {target_name}")
            resume_label = self.actions[resume_step].get('name', f'步骤{resume_step+1}')
            resume_mode = "指定步骤"

        return {
            "seconds": seconds,
            "resume_step": resume_step,
            "resume_label": resume_label,
            "resume_mode": resume_mode,
            "policy": policy_key,
            "policy_label": "到时优先恢复" if policy_key == "due_first" else "整轮后再恢复",
            "raw_text": raw_text
        }

    def _get_deferred_queue_snapshot(self):
        now = time.time()
        out = []
        for entry in sorted(self._deferred_queue, key=lambda e: (e.get("due_at", 0), e.get("order", 0))):
            due_at = float(entry.get("due_at", now))
            remaining = max(0.0, due_at - now)
            out.append({
                "loop_index": entry.get("loop_index", 0),
                "row_index": entry.get("row_index", 0),
                "origin_step": entry.get("origin_step", 0),
                "origin_name": entry.get("origin_name", ""),
                "resume_step": entry.get("resume_step", 0),
                "resume_label": entry.get("resume_label", ""),
                "resume_mode": entry.get("resume_mode", "下一步"),
                "seconds": entry.get("seconds", 0),
                "due_at": due_at,
                "remaining_seconds": remaining,
                "remaining_text": self._format_time_display(max(1, int(math.ceil(remaining)))) if remaining > 0 else "已到时",
                "policy": entry.get("policy", "after_round"),
                "policy_label": entry.get("policy_label", "整轮后再恢复"),
            })
        return out

    def _emit_deferred_queue(self):
        try:
            self.deferred_queue_sig.emit(self._get_deferred_queue_snapshot())
        except Exception:
            pass

    def _enqueue_deferred_row(self, l_idx, t_idx, current_step_idx, defer_info):
        wait_seconds = max(0.0, float(defer_info.get("seconds", 0) or 0))
        due_at = time.time() if self.dry_run else (time.time() + wait_seconds)
        self._defer_seq += 1
        target_loop_idx = l_idx if defer_info.get("policy", "after_round") == "due_first" else (l_idx + 1)
        entry = {
            "loop_index": target_loop_idx,
            "row_index": t_idx,
            "origin_step": current_step_idx,
            "origin_name": self.actions[current_step_idx].get('name', f'步骤{current_step_idx+1}'),
            "resume_step": defer_info["resume_step"],
            "resume_label": defer_info["resume_label"],
            "resume_mode": defer_info.get("resume_mode", "下一步"),
            "seconds": wait_seconds,
            "due_at": due_at,
            "policy": defer_info.get("policy", "after_round"),
            "policy_label": defer_info.get("policy_label", "整轮后再恢复"),
            "order": self._defer_seq,
        }
        self._deferred_queue.append(entry)
        wait_text = self._format_time_display(max(1, int(math.ceil(wait_seconds)))) if wait_seconds > 0 else "0s"
        self.log_sig.emit(
            f"⏸️ 第 {t_idx+1} 组已挂起：等待 {wait_text} 后，从 [{entry['resume_label']}] 恢复（{entry['policy_label']}）",
            "blue"
        )
        self.detail_sig.emit(f"⏸️ 已挂起，稍后从 [{entry['resume_label']}] 恢复")
        self._emit_deferred_queue()
        return entry

    def _run_row(self, l_idx, t_idx, s_start, total, task_name="", resumed=False, resume_entry=None):
        self._cur_t = t_idx
        # reset per-row failure snapshot
        self._row_fail_ctx = None
        self._row_runtime_ctx = None
        data = self.data_list[t_idx]
        if not self.ignore_data and not data.get("_选中", True):
            self.log_sig.emit(f"⏭️ 跳过第 {t_idx+1} 组 (未勾选执行)", "gray")
            self.row_status_sig.emit(t_idx, ROW_STATUS_SKIP)
            return {"state": "skipped", "row_ok": True}

        if resumed and resume_entry:
            waited = max(0.0, time.time() - float(resume_entry.get("due_at", time.time())))
            waited_text = self._format_time_display(int(waited)) if waited >= 1 else "已到时"
            self.log_sig.emit(
                f"↩️ 恢复第 {t_idx+1}/{total} 组：从 [{resume_entry['resume_label']}] 继续执行 ({waited_text})",
                "blue"
            )
        else:
            self.log_sig.emit(f"--- 执行第 {t_idx+1}/{total} 组 ---", "blue")

        row_vars = {
            "日期": datetime.now().strftime("%Y-%m-%d"),
            "时间": datetime.now().strftime("%H:%M:%S"),
            "行号": f"{t_idx+1:03d}",
            "循环号": str(l_idx+1)
        }

        row_ok = True
        row_state = "done"
        manual_reason = ""
        s_idx = s_start
        while s_idx < len(self.actions):
            if self._stop:
                break
            self._cur_s = s_idx

            if self._next_row:
                self._next_row = False
                self.log_sig.emit(f"⏩ [手动] 放弃第 {t_idx+1} 行剩余步骤，跳到下一行", "orange")
                row_ok = True
                row_state = "manual"
                manual_reason = "人工切下一行"
                break

            if self._skip_step:
                self._skip_step = False
                self.log_sig.emit(f"⏭ [手动] 跳过步骤: [{self.actions[s_idx].get('name', f'步骤{s_idx+1}')}]", "orange")
                if row_state == "done":
                    row_state = "manual"
                    manual_reason = "人工跳步"
                s_idx += 1
                continue

            if self._retry_step:
                self._retry_step = False
                self.log_sig.emit(f"🔁 [手动] 重试步骤: [{self.actions[s_idx].get('name', f'步骤{s_idx+1}')}]", "blue")
                continue

            if self._paused:
                self._cur_s = s_idx
                self.log_sig.emit("⏸️ 已暂停 (全局热键)，等待恢复...", "orange")
                self.hotkey_paused_sig.emit(l_idx, t_idx, s_idx)
                self._spin_while_paused()
                if self._stop:
                    break
                self.log_sig.emit("▶️ 已恢复执行", "green")

            act = self.actions[s_idx]
            name = act.get('name', 'Step')
            if not self.ignore_data and data.get(f"{name}_跳过", False):
                self.log_sig.emit(f"⏭️ 跳过步骤: [{name}] (开关已关闭)", "gray")
                s_idx += 1
                continue

            act_type = CMD_MAP.get(act.get('action'), "click")

            if self.ignore_data:
                val = str(act.get('value', ''))
                row_delay = ""
            else:
                val_raw = data.get(name, None)
                if val_raw is not None:
                    if val_raw == "[SKIP_ROW]" or val_raw == "":
                        self.log_sig.emit(f"⏭️ [步骤 {name}] 数据表内容为空或触发词，正在结束当前行...", "orange")
                        row_ok = True
                        row_state = "skipped"
                        break

                val = str(val_raw) if val_raw is not None else ""
                delay_key = f"{name}_延时"
                row_delay = data.get(delay_key, "")
                if val_raw is None:
                    val = str(act.get('value', ''))

                for k, v in row_vars.items():
                    val = val.replace(f"{{{{{k}}}}}", v)
                if not self.ignore_data:
                    for k, v in data.items():
                        val = val.replace(f"{{{{{k}}}}}", str(v))

                if not self.ignore_data:
                    val_parts = [p.strip() for p in str(val).split('|')] if '|' in str(val) else [str(val).strip()]
                    if "[SKIP_ROW]" in val_parts:
                        self.log_sig.emit(f"⏭️ [步骤 {name}] 变量解析为跳过触发词，正在结束当前行...", "orange")
                        row_ok = True
                        row_state = "skipped"
                        break

                    if act_type in ["input", "clear_input", "clear_input_plus", "upload", "drag_file", "open_url", "cmd"]:
                        if not val.strip():
                            self.log_sig.emit(f"⏭️ [步骤 {name}] 输入内容解析为空，正在结束当前行...", "orange")
                            row_ok = True
                            row_state = "skipped"
                            break

            self.highlight_sig.emit(t_idx, s_idx)

            if self.dry_run:
                msg = f"🧪 [模拟] {name} ({act.get('action')})"
                self.log_sig.emit(f"{msg}  值=「{val}」  坐标=({act.get('x',0)},{act.get('y',0)})", "purple")
                self.detail_sig.emit(msg)
            else:
                msg = f"👉 执行: {name} ({act.get('action')})"
                self.log_sig.emit(msg, "black")
                self.detail_sig.emit(msg)

            step_ok = False
            attempts = 0
            max_attempts = max(1, self.retry_count + 1)
            jump_target = None
            guard_skip_row = False

            while attempts < max_attempts:
                if self._stop:
                    break
                try:
                    result = self._execute_step(act, val, act_type, data)
                    self._set_row_runtime_ctx(
                        l_idx, t_idx, s_idx,
                        name, act_type, act.get("action", ""), val
                    )
                    if result and result[0] == "jump_if":
                        jump_target = result[1]
                    elif result and result[0] == "defer":
                        self._enqueue_deferred_row(l_idx, t_idx, s_idx, result[1])
                        row_state = "deferred"
                    step_ok = True
                    break
                except ExecutionInterrupted:
                    break
                except StepGuardMismatchError as e:
                    error_detail = str(e) if str(e) else "界面守卫未通过"
                    self._last_error = error_detail
                    self.log_sig.emit(f"⚠️ {error_detail}，已跳过当前行并继续下一行", "orange")
                    row_ok = True
                    row_state = "manual"
                    manual_reason = "界面守卫不匹配，已跳过当前行"
                    guard_skip_row = True
                    break
                except Exception as e:
                    error_detail = str(e) if str(e) else f"未知错误类型: {type(e).__name__}"
                    self._last_error = error_detail
                    attempts += 1
                    if attempts < max_attempts:
                        self.log_sig.emit(f"⚠️ 第{attempts}次重试: {error_detail}", "orange")
                        self._interruptible_sleep(1)
                    else:
                        self.log_sig.emit(f"❌ 错误: {error_detail}", "red")
                        # 记录失败上下文，便于执行结束后快速定位失败窗口
                        self._set_row_fail_ctx(
                            l_idx, t_idx, s_idx,
                            name, act_type, act.get("action", ""), val,
                            error_detail
                        )
                        if self.on_error == "skip":
                            self.log_sig.emit(f"⏭️ 已跳过步骤: {name}", "orange")
                            step_ok = True
                        elif self.on_error == "fail_row":
                            # 只标记当前行失败，继续后续行（不终止整个任务）
                            row_ok = False
                            row_state = "failed"
                            break
                        elif self.on_error == "stop":
                            self._stop_reason = "failed"
                            self._stop = True
                            row_ok = False
                        else:
                            row_ok = False

            if row_state == "deferred":
                break
            if guard_skip_row:
                break
            if row_state == "failed":
                break

            if jump_target:
                if jump_target != "(顺序执行)":
                    jumped = False
                    clean_target = jump_target.strip().split(' ')[-1]
                    for ji, ja in enumerate(self.actions):
                        act_name = ja.get('name', '').strip()
                        if act_name == jump_target or jump_target in act_name or (clean_target and clean_target in act_name):
                            s_idx = ji - 1
                            jumped = True
                            self.log_sig.emit(f"🔀 成功定位跳转目标: {act_name}", "blue")
                            break
                    if not jumped:
                        self.log_sig.emit(f"⚠️ 无法匹配跳转目标 '{jump_target}'，将按顺序执行下一步", "orange")

            if self._stop:
                break

            if self._next_row:
                self._next_row = False
                self.log_sig.emit(f"⏩ [手动] 放弃第 {t_idx+1} 行剩余步骤，跳到下一行", "orange")
                row_ok = True
                row_state = "manual"
                manual_reason = "人工切下一行"
                break
            if self._skip_step:
                self._skip_step = False
                self.log_sig.emit(f"⏭ [手动] 跳过当前步骤延时，直接进入下一步", "orange")
                if row_state == "done":
                    row_state = "manual"
                    manual_reason = "人工跳步"
                s_idx += 1
                continue

            if self._retry_step:
                self._retry_step = False
                self.log_sig.emit(f"🔁 [手动] 立即重执行当前步骤: [{name}]", "blue")
                if row_state == "done":
                    row_state = "manual"
                    manual_reason = "人工重试"
                continue

            s_delay = float(row_delay) if (row_delay and str(row_delay).replace('.','',1).isdigit()) else float(act.get('delay', 1))
            if self.dry_run:
                self._interruptible_sleep(min(0.1, s_delay))
            elif s_delay > 0:
                self._wait_with_countdown(s_delay, f"⏱️ [{name}] 步后延时")

            if self._next_row:
                self._next_row = False
                self.log_sig.emit(f"⏩ [手动] 延时中触发，跳到下一行", "orange")
                row_ok = True
                row_state = "manual"
                manual_reason = "人工切下一行"
                break
            if self._skip_step:
                self._skip_step = False
                self.log_sig.emit(f"⏭ [手动] 延时中触发，跳入下一步", "orange")
                if row_state == "done":
                    row_state = "manual"
                    manual_reason = "人工跳步"
                s_idx += 1
                continue

            s_idx += 1
            _pct = int((t_idx * len(self.actions) + s_idx) / (total * len(self.actions)) * 100) if total and self.actions else 0
            self.prog_sig.emit(_pct)
            _cur_act = self.actions[s_idx - 1] if s_idx > 0 else {}
            _status_str = "paused" if self._paused else "running"
            self._write_progress_status(
                _pct, l_idx, t_idx + 1, total, s_idx, len(self.actions),
                step_name=_cur_act.get('name', ''),
                task_name=task_name, status=_status_str
            )

        self.highlight_sig.emit(-1, -1)
        if self._stop and self._stop_reason in ("stopped", "failed"):
            row_ok = False

        status_emoji = ""
        if row_state == "deferred":
            status_emoji = ROW_STATUS_DEFER
            self.row_status_sig.emit(t_idx, status_emoji)
            row_result_text = "⏸️ 挂起"
        elif row_state == "skipped":
            status_emoji = ROW_STATUS_SKIP
            self.row_status_sig.emit(t_idx, status_emoji)
            row_result_text = "⏭️ 跳过"
        elif not row_ok:
            status_emoji = ROW_STATUS_FAIL
            self.row_status_sig.emit(t_idx, status_emoji)
            row_result_text = "❌ 失败"
        elif row_state == "manual":
            status_emoji = ROW_STATUS_MANUAL
            self.row_status_sig.emit(t_idx, status_emoji)
            row_result_text = f"⚠️ {manual_reason or '人工介入'}"
        else:
            status_emoji = ROW_STATUS_OK
            self.row_status_sig.emit(t_idx, status_emoji)
            row_result_text = "✅ 成功"

        self.log_sig.emit(
            f"═══ 组 {t_idx+1}/{total} 完成 | 结果: {row_result_text} | "
            f"已耐时: {self._format_time_display(int((datetime.now()-self._start_time).total_seconds()))} ═══",
            "blue"
        )
        # 结构化结果：供 UI 在“全部任务结束后”汇总失败项与失败窗口
        try:
            self.row_result_sig.emit({
                "task_id": getattr(self, "_task_id", ""),
                "task_name": task_name,
                "loop_index": int(l_idx),
                "row_index": int(t_idx),
                "status": status_emoji,
                "row_state": row_state,
                "row_ok": bool(row_ok),
                "last_error": str(self._last_error or ""),
                "fail_ctx": self._row_fail_ctx,
                "row_ctx": self._row_runtime_ctx,
            })
        except Exception:
            pass
        return {"state": row_state, "row_ok": row_ok}

    def _drain_deferred_queue(self, l_idx, total, task_name="", priority_only=False, wait_for_due=True):
        while not self._stop:
            pending = [e for e in self._deferred_queue if e.get("loop_index", 0) <= l_idx]
            if not pending:
                return

            pending.sort(key=lambda e: (e.get("due_at", 0), e.get("order", 0)))
            now = time.time()
            due_items = [e for e in pending if e.get("due_at", now) <= now]
            if priority_only:
                due_items = [e for e in due_items if e.get("policy") == "due_first"]
            else:
                due_first_items = [e for e in due_items if e.get("policy") == "due_first"]
                after_round_items = [e for e in due_items if e.get("policy") != "due_first"]
                due_items = due_first_items + (after_round_items[:1] if after_round_items else [])

            if not due_items:
                if priority_only or not wait_for_due:
                    return
                next_entry = pending[0]
                wait_seconds = max(0.0, float(next_entry.get("due_at", now)) - now)
                if wait_seconds > 0:
                    wait_text = self._format_time_display(max(1, int(math.ceil(wait_seconds))))
                    self.log_sig.emit(
                        f"🕒 挂起队列等待中：最近一项将在 {wait_text} 后恢复第 {next_entry['row_index']+1} 组",
                        "gray"
                    )
                    if self.dry_run:
                        time.sleep(min(0.1, wait_seconds))
                    else:
                        self._wait_with_countdown(wait_seconds, "🕒 等待挂起项到时")
                continue

            for entry in due_items:
                if self._stop:
                    return
                if entry in self._deferred_queue:
                    self._deferred_queue.remove(entry)
                    self._emit_deferred_queue()
                self._run_row(
                    l_idx,
                    entry["row_index"],
                    entry["resume_step"],
                    total,
                    task_name=task_name,
                    resumed=True,
                    resume_entry=entry
                )
                if self._stop:
                    return
                if self.loop_delay and not self.dry_run and not priority_only:
                    still_pending = [e for e in self._deferred_queue if e.get("loop_index") == l_idx]
                    if still_pending:
                        self._wait_with_countdown(self.loop_delay, "⏳ 组间等待")
            return

    def _execute_step(self, act, val, act_type, data):
        """Execute one action step; raises on failure. In dry_run mode just logs."""
        self._cooperative_abort()
        if self.dry_run:
            x, y = act.get('x'), act.get('y')
            if x is not None and y is not None and act_type not in ["run_app", "wait", "screenshot", "open_url", "defer"]:
                try:
                    pyautogui.moveTo(x, y, duration=0.5)
                    pyautogui.moveRel(10, 0, duration=0.1); pyautogui.moveRel(-20, 0, duration=0.1); pyautogui.moveRel(10, 0, duration=0.1)
                except: pass
            if act_type == "image_click" or act_type == "if_image":
                img_path = val.split('|')[0].strip() if '|' in val else val
                if img_path and os.path.exists(img_path):
                    self.log_sig.emit(f"🧪 [模拟] 正在屏幕搜索图片: {os.path.basename(img_path)}", "purple")
                    try:
                        res = pyautogui.locateCenterOnScreen(img_path, confidence=0.8)
                        if res: pyautogui.moveTo(res, duration=0.5)
                    except: pass
            if act_type == "win_active": self.log_sig.emit(f"🧪 [模拟] 准备激活窗口: {val}", "purple")
            if act_type == "open_url": self.log_sig.emit(f"🧪 [模拟] 准备打开网址: {val}", "purple")
            if act_type == "defer":
                defer_raw = self._replace_vars(val, data)
                defer_info = self._parse_defer_value(defer_raw, self._cur_s)
                self.log_sig.emit(
                    f"🧪 [模拟] 延后执行 {defer_info['seconds']} 秒，稍后从 [{defer_info['resume_label']}] 恢复",
                    "purple"
                )
                return ("defer", defer_info)
            return None
        
        # [新增] 如果开启了标准化，且当前是打开网页或激活窗口动作，则在操作后执行标准化
        # 我们在这里记录是否需要标准化，等动作执行完后再操作
        should_std = self.standardize_window and act_type in ["open_url", "win_active"]
        self._validate_step_guard(act, act_type)
        if act_type == "image_click":
            # 解析跳转参数：图片路径|成功跳转步骤|失败跳转步骤
            # 增强：过滤掉可能存在的换行符和干扰字符
            raw_val = val.replace('\r', '').replace('\n', ' ').strip()
            parts = [p.strip() for p in raw_val.split('|')]
            img_raw_path = parts[0]
            then_step = parts[1] if len(parts) > 1 else ""
            else_step = parts[2] if len(parts) > 2 else ""

            img_path = os.path.normpath(os.path.abspath(img_raw_path))
            if not os.path.exists(img_path):
                alt_path = os.path.join(BASE_DIR, img_raw_path)
                if os.path.exists(alt_path): img_path = alt_path
                else: raise RuntimeError(f"图片文件不存在: {img_path}")
            
            self.log_sig.emit(f"📷 正在搜索并尝试点击: {os.path.basename(img_path)}", "gray")
            res = None
            try:
                from PIL import Image
                with Image.open(img_path) as img_obj:
                    start_time = time.time()
                    locate_err = None
                    while time.time() - start_time < 3.0:
                        self._cooperative_abort()
                        try:
                            res = pyautogui.locateCenterOnScreen(img_obj, confidence=0.8)
                            if res: break
                        except Exception as e:
                            locate_err = e
                            res = None
                        self._cooperative_sleep(0.5)
                    if locate_err is not None and res is None:
                        log_internal_issue(f"图像点击识别异常: {img_path}", locate_err)
                
                if res: 
                    pyautogui.click(res)
                    self.log_sig.emit(f"✅ 已找到并点击图片: {os.path.basename(img_path)}", "green")
                    return ("jump_if", then_step)
                else: 
                    self.log_sig.emit(f"🔍 未能点到图片 (正常跳过): {os.path.basename(img_path)}", "gray")
                    return ("jump_if", else_step)
            except Exception as e:
                if "Failed to read" in str(e) or "OSError" in str(e):
                    raise RuntimeError(f"无法读取图片文件: {img_path}")
                raise e
        elif act_type == "if_image":
            # 增强：过滤掉可能存在的换行符和干扰字符，防止路径解析错误
            raw_val = val.replace('\r', '').replace('\n', ' ').strip()
            parts = [p.strip() for p in raw_val.split('|')]
            
            img_raw_path = parts[0]
            then_step = parts[1] if len(parts) > 1 else ""
            else_step = parts[2] if len(parts) > 2 else ""

            img_path = os.path.normpath(os.path.abspath(img_raw_path))
            if not os.path.exists(img_path):
                alt_path = os.path.join(BASE_DIR, img_raw_path)
                if os.path.exists(alt_path): img_path = alt_path
            res = None
            try: 
                from PIL import Image
                with Image.open(img_path) as img_obj:
                    try:
                        res = pyautogui.locateCenterOnScreen(img_obj, confidence=0.8)
                    except Exception as e:
                        log_internal_issue(f"图像判断识别异常: {img_path}", e)
                        res = None
                
                if res: self.log_sig.emit(f"🔍 图像判断：已找到目标 {os.path.basename(img_path)}", "gray")
                else: self.log_sig.emit(f"🔍 图像判断：未找到目标 {os.path.basename(img_path)}", "gray")
            except Exception as e:
                self.log_sig.emit(f"⚠️ 图像判断出错: {str(e)}", "orange")
            return ("jump_if", then_step if res else else_step)
        elif act_type == "if_win":
            parts = [p.strip() for p in val.split('|')]
            target_win = parts[0]
            then_step  = parts[1] if len(parts) > 1 else ""
            else_step  = parts[2] if len(parts) > 2 else ""
            found = any(target_win in t for t in gw.getAllTitles())
            return ("jump_if", then_step if found else else_step)
        elif act_type == "click":          pyautogui.click(act.get('x',0), act.get('y',0))
        elif act_type == "double_click":   pyautogui.doubleClick(act.get('x',0), act.get('y',0))
        elif act_type == "right_click":    pyautogui.rightClick(act.get('x',0), act.get('y',0))
        elif act_type == "move":           pyautogui.moveTo(act.get('x',0), act.get('y',0))
        elif act_type == "hover_click":    pyautogui.moveTo(act.get('x',0), act.get('y',0)); self._cooperative_sleep(0.5); pyautogui.click()
        elif act_type == "input":
            final_val = self._replace_vars(val, data)
            self.log_sig.emit(f"✍️ 正在输入文本: {final_val[:20]}{'...' if len(final_val)>20 else ''}", "gray")
            # 如果有坐标，先点击再输入
            x, y = act.get('x', 0), act.get('y', 0)
            if x or y:
                pyautogui.click(x, y); self._interruptible_sleep(0.3)
            pyperclip.copy(final_val); self._interruptible_sleep(0.2)
            pyautogui.hotkey('ctrl', 'v')
        elif act_type == "clear_input":
            final_val = self._replace_vars(val, data)
            self.log_sig.emit(f"✍️ 正在清空并输入: {final_val[:20]}{'...' if len(final_val)>20 else ''}", "gray")
            pyautogui.click(act.get('x',0), act.get('y',0)); self._interruptible_sleep(0.2)
            pyautogui.hotkey('ctrl', 'a'); pyautogui.press('backspace'); self._interruptible_sleep(0.2)
            pyperclip.copy(final_val); self._interruptible_sleep(0.2)
            pyautogui.hotkey('ctrl', 'v')
        elif act_type == "clear_input_plus":
            # 增强版：支持 val 格式为 "前缀|内容"
            raw_val = self._replace_vars(val, data)
            prefix = ""
            content = raw_val
            if "|" in raw_val:
                prefix, content = raw_val.split("|", 1)
            
            final_val = f"{prefix}{content}"
            self.log_sig.emit(f"✨ 增强输入(前缀:{prefix}): {content[:20]}...", "gray")
            
            pyautogui.click(act.get('x',0), act.get('y',0)); self._interruptible_sleep(0.2)
            pyautogui.hotkey('ctrl', 'a'); pyautogui.press('backspace'); self._interruptible_sleep(0.2)
            pyperclip.copy(final_val); self._interruptible_sleep(0.2)
            pyautogui.hotkey('ctrl', 'v')
        elif act_type == "upload":
            if not val: self.log_sig.emit("⚠️ 上传文件路径为空，跳过该步骤", "orange"); return None
            abs_path = os.path.normpath(os.path.abspath(val))
            self.log_sig.emit(f"📋 准备上传: {abs_path}", "gray")
            x, y = act.get('x', 0), act.get('y', 0)
            if x or y:
                pyautogui.click(x, y)
                self._interruptible_sleep(0.8)
            pyperclip.copy(abs_path); self._interruptible_sleep(0.5)
            pyautogui.hotkey('ctrl', 'v'); self._cooperative_sleep(0.8); pyautogui.press('enter')
            self.log_sig.emit("✅ 已尝试粘贴路径并回车", "green")
        elif act_type == "drag_file":
            if not val: self.log_sig.emit("⚠️ 拖拽文件路径为空，跳过", "orange"); return None
            if sys.platform != 'win32': self.log_sig.emit("⚠️ 拖拽功能仅支持 Windows 系统", "orange"); return None
            
            abs_path = os.path.normpath(os.path.abspath(val))
            x, y = act.get('x', 0), act.get('y', 0)
            if not x and not y:
                self.log_sig.emit("⚠️ 拖拽指令未设置目标坐标，跳过", "orange"); return None
            drag_target_hwnd = 0
            if act.get('activate_target_before_drag', False):
                try:
                    drag_target_hwnd = get_root_window_from_point(x, y)
                    pyautogui.click(x, y)
                    if drag_target_hwnd:
                        self.log_sig.emit(f"🪟 拖拽前已点击目标坐标: ({x}, {y}) hwnd={drag_target_hwnd}", "gray")
                    else:
                        self.log_sig.emit(f"🪟 拖拽前已点击目标坐标: ({x}, {y})", "gray")
                    self._interruptible_sleep(0.25)
                except Exception as e:
                    self.log_sig.emit(f"⚠️ 拖拽前点击目标坐标失败: {str(e)}", "orange")
            
            self.log_sig.emit(f"🖱️ 正在执行真实拖拽上传: {os.path.basename(abs_path)}", "gray")
            try:
                perform_explorer_assisted_drag(
                    abs_path,
                    x,
                    y,
                    log_func=lambda msg: self.log_sig.emit(f"🧩 {msg}", "gray"),
                    target_hwnd=drag_target_hwnd
                )
                self.log_sig.emit("✅ 资源管理器真实拖拽已完成", "green")
            except Exception as explorer_err:
                self.log_sig.emit(f"⚠️ 资源管理器真实拖拽失败，回退 OLE 合成拖拽: {str(explorer_err)}", "orange")
                try:
                    perform_native_file_drag(
                        [abs_path],
                        x,
                        y,
                        log_func=lambda msg: self.log_sig.emit(f"🧩 {msg}", "gray")
                    )
                    self.log_sig.emit("✅ OLE 合成拖拽已完成", "green")
                except Exception as native_err:
                    self.log_sig.emit(f"⚠️ OLE 合成拖拽失败，回退窗口级拖放: {str(native_err)}", "orange")
                    try:
                        # 1. 获取目标坐标处的窗口句柄
                        hwnd = get_root_window_from_point(x, y)
                        if not hwnd: raise RuntimeError("无法定位目标窗口")
                        
                        # 2. 构造 DROPFILES 结构体
                        class DROPFILES(ctypes.Structure):
                            _fields_ = [("pFiles", wintypes.DWORD),
                                        ("pt", wintypes.POINT),
                                        ("fNC", wintypes.BOOL),
                                        ("fWide", wintypes.BOOL)]

                        files = abs_path + '\0\0'
                        files_u16 = files.encode('utf-16le')
                        offset = ctypes.sizeof(DROPFILES)
                        size = offset + len(files_u16)
                        
                        GHND = 0x0042
                        hGlobal = ctypes.windll.kernel32.GlobalAlloc(GHND, size)
                        if not hGlobal: raise RuntimeError("全局内存分配失败")
                        
                        pGlobal = ctypes.windll.kernel32.GlobalLock(hGlobal)
                        if not pGlobal:
                            ctypes.windll.kernel32.GlobalFree(hGlobal)
                            raise RuntimeError("GlobalLock 失败，无法写入拖放数据")
                        df = DROPFILES()
                        df.pFiles = offset
                        df.pt = wintypes.POINT(x, y)
                        df.fNC = False
                        df.fWide = True
                        
                        ctypes.memmove(pGlobal, ctypes.addressof(df), offset)
                        ctypes.memmove(pGlobal + offset, files_u16, len(files_u16))
                        ctypes.windll.kernel32.GlobalUnlock(hGlobal)
                        
                        WM_DROPFILES = 0x0233
                        ctypes.windll.user32.PostMessageW(hwnd, WM_DROPFILES, hGlobal, 0)
                        pyautogui.moveTo(x, y, duration=0.2)
                        self.log_sig.emit("✅ 已发送窗口级拖放消息", "green")
                    except Exception as legacy_err:
                        raise RuntimeError(f"资源管理器拖拽、OLE 拖拽与窗口级拖放均失败: {legacy_err}") from legacy_err
        elif act_type == "press":    pyautogui.press(val)
        elif act_type == "hotkey":   pyautogui.hotkey(*val.split('+'))
        elif act_type == "scroll":
            # [修复] 增加坐标支持：如果设置了坐标，先移动鼠标到该位置，确保滚动作用于正确区域
            x, y = act.get('x'), act.get('y')
            if x is not None and y is not None and (x != 0 or y != 0):
                pyautogui.moveTo(x, y)
            
            clicks = int(val) if val else 0
            if clicks != 0:
                # [修复] 解决大数值滚动只动一下的问题：分段执行滚动
                # 许多应用程序（如浏览器或 Excel）在接收到单次超大滚动消息时会限流或只处理为一格
                abs_clicks = abs(clicks)
                direction = 1 if clicks > 0 else -1
                step = 120  # Windows 标准滚动增量
                
                done = 0
                while done < abs_clicks:
                    self._cooperative_abort()
                    curr = min(step, abs_clicks - done)
                    pyautogui.scroll(curr * direction)
                    done += curr
                    # 极短的延迟（5ms）让系统消息队列有时间处理滚动事件，提高成功率
                    time.sleep(0.005)
                self.log_sig.emit(f"🖱️ 滚轮已滚动 {clicks} 单位 ({'向上' if clicks > 0 else '向下'})", "gray")
        elif act_type == "cmd":
            # 强制解析变量，确保 {{2}} 等序号引用生效
            cmd_str = self._replace_vars(val, data)
            self.log_sig.emit(f"💻 正在执行 CMD 指令: {cmd_str}", "gray")
            try:
                # 使用 shell=True 以支持复杂的命令和管道
                # 优化：对于 start 命令，我们不需要等待它结束，否则会阻塞 UI
                if "start " in cmd_str.lower():
                    subprocess.Popen(cmd_str, shell=True)
                    self.log_sig.emit(f"🚀 已发起启动指令", "green")
                else:
                    # 解决 echo 等内置命令在 Popen 中的编码问题，确保中文内容正确写入
                    process = subprocess.Popen(cmd_str, shell=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, encoding='gbk' if sys.platform == 'win32' else 'utf-8', errors='replace')
                    try:
                        self._wait_process_interruptibly(process)
                        stdout, _ = process.communicate()
                    except ExecutionInterrupted:
                        self._abort_subprocess(process)
                        raise
                    if stdout:
                        self.log_sig.emit(f"📝 CMD 输出: {stdout.strip()}", "gray")
                    if process.returncode == 0:
                        self.log_sig.emit(f"✅ CMD 执行成功", "green")
                    else:
                        self.log_sig.emit(f"⚠️ CMD 执行返回错误码: {process.returncode}", "orange")
            except Exception as e:
                self.log_sig.emit(f"❌ CMD 执行失败: {str(e)}", "red")
        elif act_type == "win_active":
            raw_val = self._replace_vars(val, data)
            _profile_path_for_win = None
            display_name = raw_val  # 默认显示名

            # ── [修复] 解析 hwnd 标识字段 ───────────────────────────────────
            # 格式: "标题::hwnd=12345" (新格式) 或 "纯标题" (旧格式)
            _target_hwnd = None
            target_title = raw_val
            if "::hwnd=" in raw_val:
                parts = raw_val.split("::hwnd=", 1)
                target_title = parts[0].strip()
                try:
                    _target_hwnd = int(parts[1].strip())
                except Exception as e:
                    log_internal_issue(f"解析 hwnd 失败: {raw_val}", e)

            # ── 判断是否为 Chrome 账号路径 ────────────────────────────────
            if "\\" in target_title or "/" in target_title:
                _profile_path_for_win = target_title
                display_name = get_profile_display_name(target_title)
                self.log_sig.emit(f"🔍 正在尝试切换到账号窗口: {display_name}", "gray")
                _is_generic = (
                    display_name.startswith("自定义路径账号")
                    or display_name in ("默认账户", "Default", "未设置账户")
                )
                if _is_generic:
                    target_title = ""
                else:
                    target_hint = display_name.split('] ')[-1] if '] ' in display_name else display_name
                    target_title = target_hint.split(' (')[0]
            else:
                self.log_sig.emit(f"🔍 正在尝试激活窗口: {target_title}", "gray")

            found = False
            for _ in range(5):
                self._cooperative_abort()
                # ── [最高优先级] 直接用 hwnd 激活，完全不依赖标题匹配 ────────────
                if _target_hwnd and sys.platform == 'win32':
                    try:
                        # 先验证窗口句柄仍有效
                        if ctypes.windll.user32.IsWindow(_target_hwnd):
                            force_activate_window(_target_hwnd)
                            self.log_sig.emit(f"✅ 已通过 hwnd 直接激活窗口: {target_title} (hwnd={_target_hwnd})", "green")
                            found = True; break
                        else:
                            self.log_sig.emit(f"⚠️ hwnd={_target_hwnd} 已失效，回退到标题匹配模式", "orange")
                            _target_hwnd = None  # hwnd 失效后不再重试
                    except Exception as e:
                        log_internal_issue(f"通过 hwnd 激活窗口失败: {_target_hwnd}", e)

                # ── 常规标题匹配 ────────────────────────────────────────────────────
                if not found and target_title:
                    wins = gw.getWindowsWithTitle(target_title)
                    if wins:
                        try:
                            if sys.platform == 'win32':
                                force_activate_window(wins[0]._hWnd)
                            else:
                                wins[0].activate()
                            self.log_sig.emit(f"✅ 已激活窗口: {target_title}", "green")
                            found = True; break
                        except Exception as e:
                            log_internal_issue(f"按标题激活窗口失败: {target_title}", e)

                # ── Chrome 账号路径溢源匹配 ───────────────────────────────────
                if not found and _profile_path_for_win:
                    try:
                        from pywinauto import Desktop
                        _norm_p = os.path.normpath(_profile_path_for_win)
                        _desktop = Desktop(backend="uia")
                        for _b_class in ["Chrome_WidgetWin_1", "MozillaWindowClass"]:
                            _wins_uia = _desktop.windows(class_name=_b_class)
                            for _w in _wins_uia:
                                try:
                                    _pid = _w.process_id()
                                    _cmd_out = subprocess.check_output(
                                        f'wmic process where processid={_pid} get commandline',
                                        shell=True
                                    ).decode('gbk', errors='ignore')
                                    _u_path = extract_cmd_switch_value(_cmd_out, "user-data-dir")
                                    _p_dir  = extract_cmd_switch_value(_cmd_out, "profile-directory") or "Default"
                                    if _u_path:
                                        _full_p = os.path.normpath(os.path.join(_u_path, _p_dir))
                                        if is_same_path(_full_p, _norm_p):
                                            force_activate_window(_w.handle)
                                            self.log_sig.emit(f"✅ 已通过路径溢源激活账号窗口: {display_name}", "green")
                                            found = True; break
                                except Exception as e:
                                    log_internal_issue(f"路径溢源匹配窗口失败: {display_name}", e)
                                    continue
                            if found: break
                    except Exception as e:
                        log_internal_issue(f"路径溢源激活账号窗口失败: {display_name}", e)

                # ── UIA 标签页匹配 ──────────────────────────────────────────────────
                if not found and target_title:
                    try:
                        from pywinauto import Desktop
                        for b_class in ["Chrome_WidgetWin_1", "MozillaWindowClass"]:
                            wins_uia = Desktop(backend="uia").windows(class_name=b_class)
                            for w in wins_uia:
                                tab_items = w.descendants(control_type="TabItem")
                                for tab in tab_items:
                                    if target_title in tab.window_text():
                                        tab.select(); w.set_focus(); found = True; break
                                if found: break
                            if found: break
                    except Exception as e:
                        log_internal_issue(f"UIA 标签匹配失败: {target_title}", e)
                if found: break
                self._cooperative_sleep(1.0)

            if not found:
                _hint = target_title or display_name
                self.log_sig.emit(f"⚠️ 未能找到或激活窗口: {_hint}", "orange")
        elif act_type == "open_url":
            # [修复 v3] val 已经由引擎 run() 按「数据表优先、步骤默认兜底」组合好：
            #   批量模式：val = data[name]（包含批量数据行的 url|profile）
            #   测试模式：val = act['value']（步骤默认值）
            # 因此直接解析 val，不再重复读 act['value'] 覆盖批量数据行的账号。
            val_parts = [p.strip() for p in str(val).split('|')]
            url     = val_parts[0]
            profile = val_parts[1] if len(val_parts) > 1 else ""
            options = val_parts[2] if len(val_parts) > 2 else ""

            # 如果 val 中 url/profile 为空，才回退到步骤默认值（act['value']）
            step_val = str(act.get('value', ''))
            step_parts = [p.strip() for p in step_val.split('|')]
            if not url or url == "AUTO_KEEP_URL":
                url = step_parts[0]
            if not profile:
                profile = step_parts[1] if len(step_parts) > 1 else ""
            if not options:
                options = step_parts[2] if len(step_parts) > 2 else ""

            url = self._replace_vars(url, data)
            profile = self._replace_vars(profile, data)
            if "|" in profile:
                parts = profile.split("|")
                if not url or url == "AUTO_KEEP_URL": url = parts[0]
                profile = parts[1]
            
            if not url.startswith(('http://', 'https://', 'file://')):
                url = "https://" + url

            # [增强] 支持把“账号(profile)”位置填成已打开窗口（::hwnd=xxx）。
            # 当检测到 hwnd 时，优先激活该窗口并在其地址栏打开 url（不依赖账号识别）。
            target_hwnd = None
            try:
                target_hwnd = self._extract_hwnd_from_value(profile) or self._extract_hwnd_from_value(options)
            except Exception:
                target_hwnd = None
            if not target_hwnd and profile and sys.platform == "win32":
                try:
                    target_hwnd = find_browser_window_hwnd_by_hint(profile)
                except Exception:
                    target_hwnd = None
            if target_hwnd and sys.platform == "win32":
                try:
                    self.log_sig.emit(f"🌐 使用已打开窗口打开网址: {get_profile_display_name(profile)}", "gray")
                    force_activate_window(int(target_hwnd))
                    self._cooperative_sleep(0.2)
                    try:
                        pyperclip.copy(url)
                    except Exception:
                        pass
                    pyautogui.hotkey('ctrl', 'l')
                    self._cooperative_sleep(0.05)
                    # 粘贴优先，失败则退回输入
                    try:
                        pyautogui.hotkey('ctrl', 'v')
                    except Exception:
                        pyautogui.typewrite(url, interval=0.01)
                    self._cooperative_sleep(0.05)
                    pyautogui.press('enter')
                    return None
                except Exception as e:
                    raise RuntimeError(f"使用已打开窗口打开网址失败: {e}")

            # [终极修复版] 解决多账户切换、映射路径识别与书签栏坐标统一。
            if profile and sys.platform == 'win32':
                display_name = get_profile_display_name(profile)
                self.log_sig.emit(f"🌐 正在请求打开账号: {display_name}", "gray")
                
                # 1. 路径解析 (回归 Chrome 原生 Profile 目录结构)
                # Chrome 的多账号本质上是：一个总的数据目录 (User Data) 下有多个 Profile 文件夹
                # 我们的目标是准确拆分出这两部分。
                norm_p = os.path.normpath(profile)
                u_dir, p_dir = "", "Default"
                
                # 查找 "User Data" 在路径中的位置
                # 比如：C:\Users\Admin\AppData\Local\Google\Chrome\User Data\Profile 1
                ud_marker = "user data"
                lower_p = norm_p.lower()
                if ud_marker in lower_p:
                    idx = lower_p.find(ud_marker)
                    # u_dir 必须包含 "User Data" 这一级
                    u_dir = norm_p[:idx + len(ud_marker)]
                    # p_dir 是 "User Data" 之后的那一级文件夹名
                    rest = norm_p[idx + len(ud_marker):].lstrip(os.sep)
                    if rest:
                        p_dir = rest.split(os.sep)[0]
                    else:
                        p_dir = "Default"
                else:
                    # 如果路径里根本没有 "User Data"（比如自定义的隔离路径）
                    # 则将父目录作为数据目录，当前目录名作为 Profile 名
                    u_dir = os.path.dirname(norm_p)
                    p_dir = os.path.basename(norm_p)
                
                # 特殊情况：如果用户选的是 User Data 目录本身，Profile 应为 Default
                if p_dir.lower() == "user data": p_dir = "Default"
                
                # 2. 寻找浏览器路径
                chrome_path = None
                for p in [r"C:\Program Files\Google\Chrome\Application\chrome.exe", 
                          r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
                          os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe")]:
                    if os.path.exists(p): chrome_path = p; break
                
                if chrome_path:
                    # [调整] 为了支持多账号并行多开，我们不再强制清理进程
                    # 而是通过确保 --user-data-dir 和 --profile-directory 的组合唯一性来实现多开
                    
                    # 3. [增强] 启动前预探测：如果账号已打开，先强制最大化对齐，再打开网址
                    def _force_std_window(hwnd):
                        try:
                            _user32 = ctypes.windll.user32
                            _user32.ShowWindow(hwnd, 9) # Restore
                            self._interruptible_sleep(0.1)
                            _user32.SetForegroundWindow(hwnd)
                            self._interruptible_sleep(0.1)
                            _user32.ShowWindow(hwnd, 3) # Maximize
                            # 触发重新渲染
                            try:
                                import pygetwindow as pgw
                                w = pgw.Window(hwnd)
                                rect = w._rect
                                _user32.MoveWindow(hwnd, rect.left, rect.top, rect.width - 1, rect.height - 1, True)
                                self._interruptible_sleep(0.05)
                                _user32.ShowWindow(hwnd, 3)
                            except Exception as e:
                                log_internal_issue(f"窗口重绘微调失败: hwnd={hwnd}", e)
                        except Exception as e:
                            log_internal_issue(f"强制标准化浏览器窗口失败: hwnd={hwnd}", e)

                    # 扫描当前是否已有该账号窗口
                    try:
                        import pygetwindow as pgw
                        from pywinauto import Application
                        all_chromes = [w for w in pgw.getWindowsWithTitle("") if "Chrome" in w.title or "RunningHub" in w.title]
                        for cw in all_chromes:
                            try:
                                _hwnd = cw._hWnd
                                _app = Application().connect(handle=_hwnd)
                                _pid = _app.window(handle=_hwnd).process_id()
                                _c_line = subprocess.check_output(f'wmic process where processid={_pid} get commandline', shell=True).decode('gbk', errors='ignore')
                                _u_m = extract_cmd_switch_value(_c_line, "user-data-dir")
                                _p_m = extract_cmd_switch_value(_c_line, "profile-directory")
                                
                                if is_same_path(u_dir, _u_m) and (not p_dir or p_dir == _p_m):
                                    self.log_sig.emit(f"📐 账号 [{p_dir}] 已在运行，正在强制对齐坐标...", "blue")
                                    _force_std_window(_hwnd)
                                    break
                            except Exception as e:
                                log_internal_issue(f"预探测账号窗口失败: {p_dir}", e)
                                continue
                    except Exception as e:
                        log_internal_issue(f"扫描已有账号窗口失败: {p_dir}", e)

                    # 4. 命令行启动：加入 --start-maximized 参数，让 Chrome 启动时自带最大化属性
                    cmd_args = [
                        f'"{chrome_path}"', 
                        f'--user-data-dir="{u_dir}"', 
                        f'--profile-directory="{p_dir}"', 
                        "--new-window",
                        "--start-maximized" # [新增] 浏览器原生参数，最稳的最大化方式
                    ]
                    if "AUTO_KEEP_URL" not in url: cmd_args.append(f'"{url}"')
                    
                    full_cmd = " ".join(cmd_args)
                    self.log_sig.emit(f"🚀 执行启动命令: {full_cmd}", "gray")
                    subprocess.Popen(full_cmd, shell=True)
                    
                    # 5. [极速诊断版] 窗口状态监控与对齐逻辑
                    def _async_force_standardize():
                        try:
                            user32 = ctypes.windll.user32
                            kernel32 = ctypes.windll.kernel32
                            
                            _success = False
                            target_profile_key = f'--profile-directory="{p_dir}"'.lower()
                            # 容错：有些命令行可能不带引号
                            target_profile_key_alt = f'--profile-directory={p_dir}'.lower()
                            
                            self.log_sig.emit(f"🔍 监控启动: 目标账号标识 [{p_dir}]", "gray")
                            
                            for _attempt in range(12):
                                if _success: break
                                time.sleep(1.0)
                                
                                import pygetwindow as pgw
                                for w in pgw.getWindowsWithTitle(""):
                                    try:
                                        _hwnd = w._hWnd
                                        _class = ctypes.create_unicode_buffer(256)
                                        user32.GetClassNameW(_hwnd, _class, 256)
                                        if _class.value != "Chrome_WidgetWin_1": continue
                                        
                                        # 极速获取 PID
                                        _pid = ctypes.c_ulong()
                                        user32.GetWindowThreadProcessId(_hwnd, ctypes.byref(_pid))
                                        _pid = _pid.value
                                        
                                        # 获取命令行
                                        try:
                                            _c_line = subprocess.check_output(f'wmic process where processid={_pid} get commandline', shell=True).decode('gbk', errors='ignore').lower()
                                        except: continue
                                        
                                        # 模糊匹配 Profile 参数
                                        if target_profile_key in _c_line or target_profile_key_alt in _c_line:
                                            # 命中目标！
                                            is_max = user32.IsZoomed(_hwnd)
                                            if not is_max:
                                                self.log_sig.emit(f"📐 发现目标窗口 (PID:{_pid})，当前非全屏，正在强制对齐...", "orange")
                                                user32.ShowWindow(_hwnd, 9) # Restore
                                                time.sleep(0.1)
                                                user32.SetForegroundWindow(_hwnd)
                                                time.sleep(0.1)
                                                user32.ShowWindow(_hwnd, 3) # Maximize
                                                # 补一个快捷键 Win+Up 确保万无一失
                                                pyautogui.hotkey('win', 'up')
                                                time.sleep(0.5)
                                                pyautogui.hotkey('ctrl', 'shift', 'b')
                                                self.log_sig.emit(f"✅ 账号 [{p_dir}] 窗口已强行铺满全屏", "green")
                                            else:
                                                user32.SetForegroundWindow(_hwnd)
                                                self.log_sig.emit(f"✨ 账号 [{p_dir}] 已是全屏状态，已置顶唤醒", "gray")
                                            
                                            _success = True
                                            break
                                    except: continue
                            if not _success:
                                self.log_sig.emit(f"ℹ️ 监控结束，未捕获到账号 [{p_dir}] 的窗口，请检查账号配置路径是否正确", "gray")
                        except Exception as e:
                            self.log_sig.emit(f"❌ 监控异常: {str(e)}", "red")
                    
                    import threading
                    threading.Thread(target=_async_force_standardize, daemon=True).start()
                else:
                    import webbrowser; webbrowser.open(url)
                return None
            else:
                import webbrowser; webbrowser.open(url)
        elif act_type == "run_app":  os.startfile(val) if sys.platform == 'win32' else os.system(val)
        elif act_type == "screenshot":
            shot_dir = os.path.join(BASE_DIR, "screenshots")
            if not os.path.exists(shot_dir): os.makedirs(shot_dir)
            pyautogui.screenshot(os.path.join(shot_dir, f"shot_{datetime.now().strftime('%H%M%S')}.png"))
        elif act_type == "defer":
            defer_raw = self._replace_vars(val, data)
            return ("defer", self._parse_defer_value(defer_raw, self._cur_s))
        elif act_type == "wait": self._wait_with_countdown(float(val) if val else 1, "⏳ 等待中")
        
        # 如果标记了需要标准化，在动作执行完、窗口出现后执行
        if should_std:
            # 额外等待 1 秒确保窗口已渲染
            self._cooperative_sleep(1)
            self._standardize_browser_window()
            
        return None


    def run(self):
        self._start_time = datetime.now()  # 记录任务开始时间，用于计算已耗时间和预估剩余时间
        self._final_status = "running"
        self._stop_reason = ""
        self._last_error = ""
        self._deferred_queue = []
        self._defer_seq = 0
        self._emit_deferred_queue()
        _task_name = getattr(self, '_task_name', '')  # 任务名称（由 AutoManager 在启动前注入）
        if self.dry_run:
            self.log_sig.emit("🧪 ===== 试运行模式：不会执行任何真实操作 =====", "purple")
        total = len(self.data_list)
        self._write_progress_status(0, self.start_l, self.start_t, total, self.start_s, len(self.actions),
                                    task_name=_task_name, status="running")  # 初始化进度状态

        l_idx = self.start_l
        while not self._stop:
            has_pending_deferred = bool(self._deferred_queue)
            is_normal_round = l_idx < self.loops
            if not is_normal_round and not has_pending_deferred:
                break

            if self._stop:
                break
            self._cur_l = l_idx
            if is_normal_round:
                if self.loops > 1 or l_idx > 0:
                    self.log_sig.emit(f"🔄 === 第 {l_idx+1}/{self.loops} 轮循环 ===", "purple")
            else:
                self.log_sig.emit(f"🔄 === 延后恢复轮 {l_idx+1} ===", "purple")

            if is_normal_round:
                t_start = self.start_t if l_idx == self.start_l else 0
                for t_idx in range(t_start, total):
                    if self._stop:
                        break
                    self._drain_deferred_queue(l_idx, total, task_name=_task_name, priority_only=True, wait_for_due=False)
                    if self._stop:
                        break
                    s_start = self.start_s if (l_idx == self.start_l and t_idx == self.start_t) else 0
                    result = self._run_row(l_idx, t_idx, s_start, total, task_name=_task_name, resumed=False)
                    if self._stop:
                        break
                    self._drain_deferred_queue(l_idx, total, task_name=_task_name, priority_only=True, wait_for_due=False)
                    if self._stop:
                        break
                    if result.get("state") != "skipped" and t_idx < total - 1 and not self.dry_run:
                        self._wait_with_countdown(self.loop_delay, "⏳ 组间等待")

            if self._stop:
                break

            self._drain_deferred_queue(l_idx, total, task_name=_task_name)

            l_idx += 1

        if self.dry_run:
            self.log_sig.emit("🧪 ===== 试运行完成 =====", "purple")
        final_status = "done"
        final_percent = 100
        if self._stop_reason == "stopped":
            final_status = "stopped"
            final_percent = max(0, getattr(self, '_last_percent', 0))
            self.log_sig.emit("🛑 任务已按用户请求停止", "orange")
        elif self._stop_reason == "failed":
            final_status = "failed"
            final_percent = max(0, getattr(self, '_last_percent', 0))
            if self._last_error:
                self.log_sig.emit(f"❌ 任务因错误终止: {self._last_error}", "red")
        self._final_status = final_status
        self._write_progress_status(
            final_percent, self._cur_l, total if total else 0, total, self._cur_s + 1, len(self.actions),
            step_name=self.actions[self._cur_s].get('name', '') if self.actions and 0 <= self._cur_s < len(self.actions) else "",
            task_name=_task_name, status=final_status
        )
        self._emit_deferred_queue()
        self.done_sig.emit()


# --- Screenshot Overlay for Region Capture (Feature 9) ---
class ScreenshotOverlay(QWidget):
    """Full-screen translucent overlay; user drags to select a region and saves it as PNG."""
    captured = pyqtSignal(str)   # emits the saved file path

    def __init__(self, save_dir, callback=None):
        super().__init__()
        self._save_dir = save_dir
        self._callback = callback
        self._origin   = QPoint()
        self._rubber   = QRubberBand(QRubberBand.Rectangle, self)

        # Freeze the desktop BEFORE going fullscreen — capture ALL screens for dual monitor support
        screen = QApplication.primaryScreen()
        virtual_geo = QRect()
        for s in QApplication.screens():
            virtual_geo = virtual_geo.united(s.geometry())
        self._virtual_offset = virtual_geo.topLeft()
        self._bg = QApplication.primaryScreen().grabWindow(
            0,
            virtual_geo.x(), virtual_geo.y(),
            virtual_geo.width(), virtual_geo.height()
        )

        self.setWindowFlags(Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint | Qt.Tool)
        self.setAttribute(Qt.WA_TranslucentBackground, False)
        self.setCursor(QCursor(Qt.CrossCursor))
        # Cover the full virtual desktop (all monitors)
        virtual_geo = QRect()
        for s in QApplication.screens():
            virtual_geo = virtual_geo.united(s.geometry())
        self.setGeometry(virtual_geo)

    def paintEvent(self, event):
        p = QPainter(self)
        p.drawPixmap(0, 0, self._bg)
        p.fillRect(self.rect(), QColor(0, 0, 0, 80))

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self._origin = event.pos()
            self._rubber.setGeometry(QRect(self._origin, QSize()))
            self._rubber.show()

    def mouseMoveEvent(self, event):
        if not self._origin.isNull():
            self._rubber.setGeometry(QRect(self._origin, event.pos()).normalized())

    def mouseReleaseEvent(self, event):
        if event.button() == Qt.LeftButton and not self._origin.isNull():
            rect = QRect(self._origin, event.pos()).normalized()
            self._rubber.hide()
            self.hide()
            if rect.width() > 5 and rect.height() > 5:
                cropped = self._bg.copy(rect)
                if not os.path.exists(self._save_dir):
                    os.makedirs(self._save_dir)
                fname = os.path.join(self._save_dir,
                                     f"capture_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png")
                cropped.save(fname, "PNG")
                self.captured.emit(fname)
                if self._callback:
                    self._callback(fname)
            self.close()

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Escape:
            self.close()

# --- Reliable Coord Picker Overlay (New) ---
class CoordPickerOverlay(QWidget):
    """全屏透明遮罩，直接捕获鼠标点击坐标，解决 pynput 钩子被拦截的问题。"""
    picked = pyqtSignal(int, int)

    def __init__(self):
        super().__init__()
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint | Qt.Tool)
        self.setAttribute(Qt.WA_TranslucentBackground, True)
        self.setCursor(QCursor(Qt.CrossCursor))
        
        # 覆盖所有显示器
        virtual_geo = QRect()
        for s in QApplication.screens():
            virtual_geo = virtual_geo.united(s.geometry())
        self.setGeometry(virtual_geo)
        
        # 记录虚拟桌面偏移，用于计算真实坐标
        self._offset_x = virtual_geo.x()
        self._offset_y = virtual_geo.y()

    def paintEvent(self, event):
        p = QPainter(self)
        # 淡淡的半透明层，提示用户正在拾取模式
        p.fillRect(self.rect(), QColor(0, 0, 0, 1)) 

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            # 获取相对于虚拟桌面的全局坐标
            global_pos = event.globalPos()
            self.picked.emit(global_pos.x(), global_pos.y())
            self.close()
        elif event.button() == Qt.RightButton:
            self.close()

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Escape:
            self.close()


class CommandPresetDialog(QDialog):
    """CMD 预设中心窗口 - 表格布局版"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("📋 CMD 指令预设中心")
        self.resize(700, 500)
        
        layout = QVBoxLayout(self)
        
        # 搜索框
        search_h = QHBoxLayout()
        search_h.addWidget(QLabel("🔍 搜索预设:"))
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("输入名称或指令关键词...")
        self.search_edit.textChanged.connect(self._filter_table)
        search_h.addWidget(self.search_edit)
        layout.addLayout(search_h)
        
        # 预设表格
        self.table = QTableWidget()
        self.table.setColumnCount(2)
        self.table.setHorizontalHeaderLabels(["预设名称", "CMD 指令内容"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Interactive)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.setColumnWidth(0, 150)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.doubleClicked.connect(self.accept)
        layout.addWidget(self.table)
        self.lbl_selection_count = create_table_selection_label()
        layout.addWidget(self.lbl_selection_count)
        bind_table_selection_label(self.table, self.lbl_selection_count)
        
        self.presets = [
            # --- 基础文件操作 ---
            {"name": "追加文本到文件", "cmd": 'echo {{内容}} >> "D:\\log.txt"'},
            {"name": "追加文本并打开", "cmd": 'echo {{内容}} >> "D:\\log.txt" && start "" "D:\\log.txt"'},
            {"name": "创建空文件", "cmd": 'type nul > "D:\\new.txt"'},
            {"name": "清空文件内容", "cmd": 'break > "D:\\target.txt"'},
            {"name": "移动文件并改名", "cmd": 'move "C:\\old.txt" "D:\\new.txt"'},
            {"name": "复制文件夹(含子目录)", "cmd": 'xcopy "C:\\src" "D:\\dst" /e /h /y /c'},
            {"name": "强制删除文件夹", "cmd": 'rd /s /q "D:\\old_folder"'},
            {"name": "隐藏文件/文件夹", "cmd": 'attrib +h +s "D:\\private"'},
            {"name": "取消隐藏文件", "cmd": 'attrib -h -s "D:\\private"'},
            {"name": "查找包含关键词的文件", "cmd": 'findstr /s /i "关键词" *.*'},

            # --- 批量与高级文件管理 ---
            {"name": "批量修改后缀(如txt转bak)", "cmd": 'ren *.txt *.bak'},
            {"name": "合并多个文本文件", "cmd": 'copy /b *.txt combined.txt'},
            {"name": "提取当前目录文件名列表", "cmd": 'dir /b > file_list.txt'},
            {"name": "提取完整路径列表", "cmd": 'dir /s /b /a-d > paths.txt'},
            {"name": "镜像同步文件夹(Robocopy)", "cmd": 'robocopy "C:\\src" "D:\\dst" /mir /mt:16'},
            {"name": "仅复制新增/修改的文件", "cmd": 'robocopy "C:\\src" "D:\\dst" /xo /s'},
            {"name": "批量设置只读属性", "cmd": 'attrib +r /s /d *.*'},
            {"name": "生成详细目录结构树", "cmd": "tree /f /a > tree.txt"},
            {"name": "按大小列出前10个文件", "cmd": 'powershell "ls | sort length -descending | select -first 10"'},
            {"name": "快速生成1GB测试文件", "cmd": 'fsutil file createnew test_1gb.dat 1073741824'}
        ]
        
        # 加载用户自定义预设
        self.config_path = os.path.join(BASE_DIR, "command_presets.json")
        if os.path.exists(self.config_path):
            try:
                with open(self.config_path, "r", encoding="utf-8") as f:
                    user_presets = json.load(f)
                    self.presets.extend(user_presets)
            except Exception as e:
                log_internal_issue(f"加载 CMD 用户预设失败: {self.config_path}", e)
            
        self._refresh_table()
        
        # 按钮区
        btn_layout = QHBoxLayout()
        btn_use = QPushButton("✅ 使用选中指令"); btn_use.setStyleSheet("background-color: #e8f5e9; height: 35px; font-weight: bold;")
        btn_use.clicked.connect(self.accept)
        btn_add = QPushButton("➕ 添加新预设"); btn_add.setFixedHeight(35)
        btn_add.clicked.connect(self._add_preset)
        btn_del = QPushButton("❌ 删除预设"); btn_del.setFixedHeight(35)
        btn_del.clicked.connect(self._del_preset)
        btn_layout.addWidget(btn_use); btn_layout.addWidget(btn_add); btn_layout.addWidget(btn_del)
        layout.addLayout(btn_layout)
        
    def _refresh_table(self):
        self.table.setRowCount(0)
        for p in self.presets:
            row = self.table.rowCount()
            self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(p['name']))
            self.table.setItem(row, 1, QTableWidgetItem(p['cmd']))
            
    def _filter_table(self, text):
        text = text.lower()
        for i in range(self.table.rowCount()):
            name = self.table.item(i, 0).text().lower()
            cmd = self.table.item(i, 1).text().lower()
            self.table.setRowHidden(i, text not in name and text not in cmd)
            
    def _add_preset(self):
        name, ok1 = QInputDialog.getText(self, "添加预设", "输入预设名称:")
        if ok1 and name:
            cmd, ok2 = QInputDialog.getText(self, "添加预设", f"输入 [{name}] 的 CMD 指令:")
            if ok2 and cmd:
                new_p = {"name": name, "cmd": cmd}
                self.presets.append(new_p)
                self._refresh_table()
                self._save_user_presets()
                
    def _del_preset(self):
        row = self.table.currentRow()
        if row >= 0:
            if row < 20: # 保护内置预设 (20个)
                QMessageBox.warning(self, "提示", "内置预设无法删除")
                return
            self.presets.pop(row)
            self._refresh_table()
            self._save_user_presets()
            
    def _save_user_presets(self):
        user_only = self.presets[7:] # 只保存内置以外的
        try:
            with open(self.config_path, "w", encoding="utf-8") as f:
                json.dump(user_only, f, ensure_ascii=False, indent=4)
        except Exception as e:
            log_internal_issue(f"保存 CMD 用户预设失败: {self.config_path}", e)
            QMessageBox.warning(self, "保存失败", f"CMD 预设保存失败：\n{e}")
        
    def get_command(self):
        row = self.table.currentRow()
        return self.table.item(row, 1).text() if row >= 0 else None

class ClearInputPrefixPresetDialog(QDialog):
    """清空并输入(增强版)前缀库"""
    DEFAULT_PRESETS = []

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("📚 常用前缀库")
        self.resize(520, 430)
        layout = QVBoxLayout(self)

        search_h = QHBoxLayout()
        search_h.addWidget(QLabel("🔍 搜索前缀:"))
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("输入名称或前缀内容...")
        self.search_edit.textChanged.connect(self._filter_table)
        search_h.addWidget(self.search_edit)
        layout.addLayout(search_h)

        self.table = QTableWidget()
        self.table.setColumnCount(2)
        self.table.setHorizontalHeaderLabels(["名称", "前缀内容"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Interactive)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.setColumnWidth(0, 150)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setWordWrap(True)
        self.table.verticalHeader().setDefaultSectionSize(34)
        self.table.setItemDelegateForColumn(1, MultiLineTextDelegate(self.table))
        layout.addWidget(self.table)
        self.lbl_selection_count = create_table_selection_label()
        layout.addWidget(self.lbl_selection_count)
        bind_table_selection_label(self.table, self.lbl_selection_count)
        self._table_updating = False
        tip = QLabel("提示：双击行或点“编辑前缀”可弹出完整编辑窗口；支持“上移/下移”调整顺序。")
        tip.setStyleSheet("color: #666; padding: 2px 0 4px 2px;")
        layout.addWidget(tip)

        self.config_path = os.path.join(BASE_DIR, "input_prefixes.json")
        self.presets = []
        if os.path.exists(self.config_path):
            try:
                with open(self.config_path, "r", encoding="utf-8") as f:
                    user_presets = json.load(f)
                    if isinstance(user_presets, list):
                        self.presets = [p for p in user_presets if isinstance(p, dict)]
            except Exception as e:
                log_internal_issue(f"加载增强输入前缀库失败: {self.config_path}", e)
        self._refresh_table()
        self.table.itemChanged.connect(self._on_table_item_changed)
        self.table.itemDoubleClicked.connect(lambda *_: self._edit_preset())

        btn_layout = QHBoxLayout()
        btn_use = QPushButton("✅ 使用选中前缀")
        btn_use.setStyleSheet("background-color: #e8f5e9; height: 35px; font-weight: bold;")
        btn_use.clicked.connect(self.accept)
        btn_add = QPushButton("➕ 添加前缀")
        btn_add.clicked.connect(self._add_preset)
        btn_edit = QPushButton("✏️ 编辑前缀")
        btn_edit.clicked.connect(self._edit_preset)
        btn_up = QPushButton("⬆️ 上移")
        btn_up.clicked.connect(lambda: self._move_preset(-1))
        btn_down = QPushButton("⬇️ 下移")
        btn_down.clicked.connect(lambda: self._move_preset(1))
        btn_import = QPushButton("📥 导入")
        btn_import.clicked.connect(self._import_presets)
        btn_export = QPushButton("📤 导出")
        btn_export.clicked.connect(self._export_presets)
        btn_del = QPushButton("❌ 删除前缀")
        btn_del.clicked.connect(self._del_preset)
        btn_layout.addWidget(btn_use)
        btn_layout.addWidget(btn_add)
        btn_layout.addWidget(btn_edit)
        btn_layout.addWidget(btn_up)
        btn_layout.addWidget(btn_down)
        btn_layout.addWidget(btn_import)
        btn_layout.addWidget(btn_export)
        btn_layout.addWidget(btn_del)
        layout.addLayout(btn_layout)

    def _refresh_table(self, selected_row=None):
        self._table_updating = True
        self.table.setRowCount(0)
        for p in self.presets:
            row = self.table.rowCount()
            self.table.insertRow(row)
            it_name = QTableWidgetItem(str(p.get("name", "")))
            it_name.setData(Qt.UserRole, row)
            it_prefix = QTableWidgetItem(str(p.get("prefix", "")))
            it_prefix.setData(Qt.UserRole, row)
            self.table.setItem(row, 0, it_name)
            self.table.setItem(row, 1, it_prefix)
        self._table_updating = False
        for row in range(self.table.rowCount()):
            self.table.resizeRowToContents(row)
            self.table.setRowHeight(row, min(max(self.table.rowHeight(row), 34), 120))
        self._filter_table(self.search_edit.text())
        if selected_row is None:
            selected_row = self.table.currentRow()
        if 0 <= selected_row < self.table.rowCount():
            self.table.setCurrentCell(selected_row, 0)
            self.table.selectRow(selected_row)

    def _on_table_item_changed(self, item):
        if self._table_updating or not item:
            return
        row = item.row()
        if row < 0 or row >= len(self.presets):
            return
        name_item = self.table.item(row, 0)
        prefix_item = self.table.item(row, 1)
        name = name_item.text().strip() if name_item else ""
        prefix = prefix_item.text() if prefix_item else ""
        if not name:
            name = f"前缀{row + 1}"
            self._table_updating = True
            if name_item:
                name_item.setText(name)
            else:
                it_name = QTableWidgetItem(name)
                it_name.setData(Qt.UserRole, row)
                self.table.setItem(row, 0, it_name)
            self._table_updating = False
        self.presets[row] = {"name": name, "prefix": prefix}
        self._save_user_presets()
        self.table.resizeRowToContents(row)
        self.table.setRowHeight(row, min(max(self.table.rowHeight(row), 34), 120))

    def _filter_table(self, text):
        text = text.lower()
        for i in range(self.table.rowCount()):
            name = self.table.item(i, 0).text().lower()
            prefix = self.table.item(i, 1).text().lower()
            self.table.setRowHidden(i, text not in name and text not in prefix)

    def _save_user_presets(self):
        try:
            with open(self.config_path, "w", encoding="utf-8") as f:
                json.dump(self.presets, f, ensure_ascii=False, indent=4)
        except Exception as e:
            log_internal_issue(f"保存增强输入前缀库失败: {self.config_path}", e)
            QMessageBox.warning(self, "保存失败", f"前缀库保存失败：\n{e}")

    def _parse_import_text(self, raw_text):
        text = str(raw_text or "").lstrip("\ufeff").strip()
        if not text:
            return []

        imported = []
        try:
            obj = json.loads(text)
            if isinstance(obj, list):
                for idx, p in enumerate(obj):
                    if not isinstance(p, dict):
                        continue
                    name = str(p.get("name", "")).strip() or f"导入{idx + 1}"
                    prefix = str(p.get("prefix", ""))
                    imported.append({"name": name, "prefix": prefix})
                if imported:
                    return imported
        except Exception:
            pass

        for line in text.splitlines():
            raw_line = line.rstrip("\r\n")
            s = raw_line.strip()
            if not s or s.startswith("#"):
                continue

            compact = s.lower().replace(" ", "")
            if compact in {
                "name|prefix", "名称|前缀", "名称|前缀内容",
                "name,prefix", "名称,前缀", "名称,前缀内容",
                "name\tprefix", "名称\t前缀", "名称\t前缀内容"
            }:
                continue

            if "\t" in raw_line:
                name, prefix = raw_line.split("\t", 1)
            elif "|" in raw_line:
                name, prefix = raw_line.split("|", 1)
            elif "," in raw_line:
                name, prefix = raw_line.split(",", 1)
            else:
                name, prefix = "", raw_line

            name = str(name).strip() or f"导入{len(imported) + 1}"
            prefix = str(prefix)
            if name or prefix.strip():
                imported.append({"name": name, "prefix": prefix})
        return imported

    def _import_presets(self):
        path, selected_filter = QFileDialog.getOpenFileName(
            self,
            "选择前缀数据文件",
            "",
            "Data Files (*.json *.csv *.txt);;All Files (*.*)"
        )
        if not path:
            return

        try:
            items = []
            lower_path = path.lower()
            if lower_path.endswith(".csv") or "CSV" in selected_filter:
                with open(path, "r", encoding="utf-8-sig", errors="ignore", newline="") as f:
                    reader = csv.reader(f)
                    for idx, row in enumerate(reader):
                        if not row:
                            continue
                        header0 = str(row[0]).strip().lower() if len(row) > 0 else ""
                        header1 = str(row[1]).strip().lower() if len(row) > 1 else ""
                        if idx == 0 and header0 in {"name", "名称"} and header1 in {"prefix", "前缀", "前缀内容"}:
                            continue
                        name = str(row[0]).strip() if len(row) > 0 else ""
                        prefix = str(row[1]) if len(row) > 1 else ""
                        if name or prefix.strip():
                            items.append({"name": name or f"导入{len(items) + 1}", "prefix": prefix})
            else:
                with open(path, "r", encoding="utf-8-sig", errors="ignore") as f:
                    items = self._parse_import_text(f.read())

            if not items:
                QMessageBox.warning(self, "导入失败", "没有解析到可用的前缀数据。")
                return

            mode, ok = QInputDialog.getItem(
                self,
                "导入模式",
                "请选择导入方式：",
                ["同名覆盖", "仅追加", "先清空再导入"],
                0,
                False
            )
            if not ok:
                return

            if mode == "先清空再导入":
                self.presets = []

            name_to_idx = {str(p.get("name", "")): i for i, p in enumerate(self.presets)}
            added = 0
            updated = 0
            for p in items:
                name = str(p.get("name", "")).strip() or f"导入{len(self.presets) + 1}"
                prefix = str(p.get("prefix", ""))
                if mode == "仅追加":
                    base_name = name
                    suffix = 2
                    while name in name_to_idx:
                        name = f"{base_name}_{suffix}"
                        suffix += 1
                    name_to_idx[name] = len(self.presets)
                    self.presets.append({"name": name, "prefix": prefix})
                    added += 1
                elif name in name_to_idx:
                    self.presets[name_to_idx[name]] = {"name": name, "prefix": prefix}
                    updated += 1
                else:
                    name_to_idx[name] = len(self.presets)
                    self.presets.append({"name": name, "prefix": prefix})
                    added += 1

            self._refresh_table()
            self._save_user_presets()
            QMessageBox.information(self, "导入完成", f"已导入完成：新增 {added} 条，更新 {updated} 条。")
        except Exception as e:
            QMessageBox.warning(self, "导入失败", f"读取或导入前缀数据失败：\n{e}")

    def _export_presets(self):
        sel_model = self.table.selectionModel()
        sel_rows_model = sel_model.selectedRows() if sel_model else []
        export_items = []
        for m in sel_rows_model:
            row = m.row()
            if 0 <= row < len(self.presets):
                export_items.append({
                    "name": str(self.presets[row].get("name", "")),
                    "prefix": str(self.presets[row].get("prefix", ""))
                })
        if not export_items:
            export_items = [
                {"name": str(p.get("name", "")), "prefix": str(p.get("prefix", ""))}
                for p in self.presets
            ]
        if not export_items:
            QMessageBox.warning(self, "提示", "前缀库为空，暂无可导出的内容。")
            return

        path, selected_filter = QFileDialog.getSaveFileName(
            self,
            "导出前缀库",
            "prefix_presets.json",
            "JSON Files (*.json);;CSV Files (*.csv);;Text Files (*.txt)"
        )
        if not path:
            return

        try:
            if path.lower().endswith(".csv") or "CSV" in selected_filter:
                if not path.lower().endswith(".csv"):
                    path += ".csv"
                with open(path, "w", encoding="utf-8-sig", newline="") as f:
                    writer = csv.writer(f)
                    writer.writerow(["name", "prefix"])
                    for p in export_items:
                        writer.writerow([p["name"], p["prefix"]])
            elif path.lower().endswith(".txt") or "Text" in selected_filter:
                if not path.lower().endswith(".txt"):
                    path += ".txt"
                with open(path, "w", encoding="utf-8") as f:
                    for p in export_items:
                        f.write(f"{p['name']}|{p['prefix']}\n")
            else:
                if not path.lower().endswith(".json"):
                    path += ".json"
                with open(path, "w", encoding="utf-8") as f:
                    json.dump(export_items, f, ensure_ascii=False, indent=4)
            QMessageBox.information(self, "导出完成", f"已导出 {len(export_items)} 条前缀。")
        except Exception as e:
            QMessageBox.warning(self, "导出失败", f"导出前缀库失败：\n{e}")

    def _get_current_row(self):
        row = self.table.currentRow()
        return row if 0 <= row < len(self.presets) else -1

    def _add_preset(self):
        default_name = f"前缀{len(self.presets) + 1}"
        dlg = PrefixPresetEditorDialog(self, name=default_name, prefix="", title="➕ 添加前缀")
        if dlg.exec_() != QDialog.Accepted:
            return
        data = dlg.get_data()
        self.presets.append({
            "name": data["name"] or default_name,
            "prefix": data["prefix"]
        })
        row = len(self.presets) - 1
        self._refresh_table(selected_row=row)
        self._save_user_presets()

    def _edit_preset(self):
        row = self._get_current_row()
        if row < 0:
            return
        current = self.presets[row]
        dlg = PrefixPresetEditorDialog(
            self,
            name=current.get("name", ""),
            prefix=current.get("prefix", ""),
            title=f"✏️ 编辑前缀 - {current.get('name', '') or f'前缀{row + 1}'}"
        )
        if dlg.exec_() != QDialog.Accepted:
            return
        data = dlg.get_data()
        self.presets[row] = {
            "name": data["name"] or f"前缀{row + 1}",
            "prefix": data["prefix"]
        }
        self._refresh_table(selected_row=row)
        self._save_user_presets()

    def _move_preset(self, direction):
        row = self._get_current_row()
        if row < 0:
            return
        target_row = row + direction
        if target_row < 0 or target_row >= len(self.presets):
            return
        self.presets[row], self.presets[target_row] = self.presets[target_row], self.presets[row]
        self._refresh_table(selected_row=target_row)
        self._save_user_presets()

    def _del_preset(self):
        row = self._get_current_row()
        if row < 0:
            return
        self.presets.pop(row)
        self._refresh_table(selected_row=min(row, len(self.presets) - 1))
        self._save_user_presets()

    def get_prefix(self):
        row = self._get_current_row()
        return self.table.item(row, 1).text() if row >= 0 and self.table.item(row, 1) else None

class StepPostCheckDialog(QDialog):
    """步骤执行后校验配置。"""

    TYPE_OPTIONS = [
        ("不校验", "none"),
        ("图片出现", "image_exists"),
        ("图片消失", "image_not_exists"),
        ("窗口出现", "window_exists"),
        ("窗口消失", "window_not_exists"),
    ]

    def __init__(self, current_config=None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("🧪 执行后结果判断")
        self.resize(560, 260)

        cfg = current_config if isinstance(current_config, dict) else {}
        layout = QVBoxLayout(self)

        tip = QLabel("动作执行完成后，会按这里的条件做验收。验收失败时，会自动进入当前步骤的重试次数和出错策略。")
        tip.setWordWrap(True)
        tip.setStyleSheet("color: #555;")
        layout.addWidget(tip)

        form = QFormLayout()
        self.type_combo = QComboBox()
        for label, value in self.TYPE_OPTIONS:
            self.type_combo.addItem(label, value)
        current_type = str(cfg.get("type", "none") or "none")
        idx = max(0, self.type_combo.findData(current_type))
        self.type_combo.setCurrentIndex(idx)
        form.addRow("校验方式:", self.type_combo)

        target_row = QHBoxLayout()
        self.target_edit = QLineEdit(str(cfg.get("target", "") or ""))
        self.target_edit.setPlaceholderText("图片路径，或窗口标题 / 选择结果...")
        self.btn_pick_target = QPushButton("选择")
        self.btn_pick_target.setFixedWidth(64)
        self.btn_pick_target.clicked.connect(self._pick_target)
        target_row.addWidget(self.target_edit, 1)
        target_row.addWidget(self.btn_pick_target)
        form.addRow("校验目标:", target_row)

        self.timeout_spin = QSpinBox()
        self.timeout_spin.setRange(0, 60)
        self.timeout_spin.setSuffix(" 秒")
        self.timeout_spin.setValue(int(cfg.get("timeout", 3) or 3))
        form.addRow("最长等待:", self.timeout_spin)

        self.confidence_spin = QSpinBox()
        self.confidence_spin.setRange(50, 99)
        self.confidence_spin.setSuffix(" %")
        self.confidence_spin.setValue(int(round(float(cfg.get("confidence", 0.8) or 0.8) * 100)))
        form.addRow("找图相似度:", self.confidence_spin)

        layout.addLayout(form)

        self.lbl_hint = QLabel("")
        self.lbl_hint.setWordWrap(True)
        self.lbl_hint.setStyleSheet("color: #666;")
        layout.addWidget(self.lbl_hint)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_ok = QPushButton("✅ 保存")
        btn_cancel = QPushButton("❌ 取消")
        btn_ok.clicked.connect(self._save_and_accept)
        btn_cancel.clicked.connect(self.reject)
        btn_row.addWidget(btn_ok)
        btn_row.addWidget(btn_cancel)
        layout.addLayout(btn_row)

        self.type_combo.currentIndexChanged.connect(self._refresh_ui)
        self._refresh_ui()

    def _current_type(self):
        return self.type_combo.currentData() or "none"

    def _refresh_ui(self):
        mode = self._current_type()
        is_image = mode in ("image_exists", "image_not_exists")
        needs_target = mode != "none"
        self.target_edit.setEnabled(needs_target)
        self.btn_pick_target.setEnabled(needs_target)
        self.confidence_spin.setEnabled(is_image)
        self.btn_pick_target.setText("选图" if is_image else ("选窗" if needs_target else "选择"))
        if mode == "none":
            self.lbl_hint.setText("关闭后，该步骤只按“有没有报错”判断。")
        elif is_image:
            self.lbl_hint.setText("适合按钮出现、成功弹窗出现、上传完成标记出现等场景。")
        else:
            self.lbl_hint.setText("适合页面切换、弹窗出现、指定窗口关闭等场景。")

    def _pick_target(self):
        mode = self._current_type()
        if mode in ("image_exists", "image_not_exists"):
            file_path, _ = QFileDialog.getOpenFileName(self, "选择校验图片", "", "Images (*.png *.jpg *.jpeg *.bmp *.gif *.webp)")
            if file_path:
                self.target_edit.setText(os.path.normpath(file_path))
        elif mode in ("window_exists", "window_not_exists"):
            selection = WindowSelector(self).get_selection()
            if selection:
                self.target_edit.setText(selection)

    def _save_and_accept(self):
        mode = self._current_type()
        target = self.target_edit.text().strip()
        if mode != "none" and not target:
            QMessageBox.warning(self, "提示", "已开启执行后校验，请先填写校验目标。")
            return
        self.accept()

    def get_config(self):
        return {
            "type": self._current_type(),
            "target": self.target_edit.text().strip(),
            "timeout": int(self.timeout_spin.value()),
            "confidence": round(self.confidence_spin.value() / 100.0, 2),
        }

class AutoManager(QMainWindow):
    def __init__(self):
        super().__init__(); self.setWindowTitle("SimuOps Auto Manager Pro"); self.resize(1200, 850)
        resolved_icon = None
        icon_candidates = [
            os.path.join(BASE_DIR, "app_icon.ico"),
            os.path.join(BASE_DIR, "app.ico"),
            os.path.join(BASE_DIR, "app_icon.png"),
        ]
        for icon_path in icon_candidates:
            if os.path.exists(icon_path):
                resolved_icon = QIcon(icon_path)
                self.setWindowIcon(resolved_icon)
                break
        self.config = load_config(); self.current_task = ""  # will be set correctly when task_combo is built
        self._config_load_error = CONFIG_LOAD_ERROR
        self._config_load_backup_path = CONFIG_LOAD_BACKUP_PATH
        self.recording_idx = -1; self.resume_point = (0, 0, 0); self._timer_config = None
        self._timer_enabled = False
        self._loading_schedule_ui = False
        self._row_statuses = {}  # task -> {row_idx: status_emoji}
        # 本次“批量执行”产生的失败记录（跨任务汇总）
        self._last_run_row_results = []  # raw row_result_sig payload list
        self._last_run_failures = []     # filtered failure list for quick navigation
        self._hotkey_hooks = []  # keyboard hook handles
        self._task_queue = []  # pending tasks for sequential run
        self._is_initializing = True # [新增] 初始化标志位，防止提前触发 UI 逻辑导致崩溃
        self.osd = FloatingProgressWindow() # [新增] 置顶悬浮进度窗
        if resolved_icon is not None:
            self.osd.setWindowIcon(resolved_icon)
        self.osd.request_pause.connect(self._toggle_pause)
        self.osd.request_stop.connect(self._stop_execution)
        self.osd.request_skip_step.connect(self._osd_skip_step)   # [v3] 跳步
        self.osd.request_next_row.connect(self._osd_next_row)     # [v3] 下一行
        self.osd.request_retry_step.connect(self._osd_retry_step) # [v3] 重试
        
        # Restore layout from config
        layout_cfg = self.config.get("layout", {})
        if layout_cfg.get("size"): self.resize(QSize(*layout_cfg["size"]))
        if layout_cfg.get("pos"): self.move(QPoint(*layout_cfg["pos"]))
        
        # Styles - Modern Windows 10/11 Theme
        self.setStyleSheet("""
            QMainWindow { background-color: #f9f9f9; }
            QWidget { font-family: "Segoe UI", "Microsoft YaHei UI", "Microsoft YaHei"; color: #333; }
            
            /* Splitter Handle */
            QSplitter::handle:horizontal { background-color: #e0e0e0; width: 6px; }
            QSplitter::handle:vertical { background-color: #e0e0e0; height: 6px; }
            QSplitter::handle:hover { background-color: #0078d4; }
            
            /* --- Title & Labels --- */
            #TitleLabel { font-size: 18px; font-weight: 600; color: #1a1a1a; margin: 12px 0; }
            QLabel { font-weight: 400; color: #222; color: #444; }
            
            /* --- Buttons --- */
            QPushButton { 
                padding: 8px 16px; border-radius: 6px; border: 1px solid #e0e0e0; 
                background-color: #ffffff; color: #333; font-weight: 400;
            }
            QPushButton:hover { background-color: #f3f3f3; border-color: #d0d0d0; }
            QPushButton:pressed { background-color: #ebebeb; }
            QPushButton:disabled { background-color: #f5f5f5; color: #aaa; border-color: #eee; }
            
            /* Primary Action Button */
            QPushButton#RunBtn { 
                background-color: #0078d4; color: white; border: none; 
                font-weight: 600; font-size: 14px; 
            }
            QPushButton#RunBtn:hover { background-color: #106ebe; }
            QPushButton#RunBtn:pressed { background-color: #005a9e; }
            
            QPushButton#ResumeBtn { 
                background-color: #28a745; color: white; border: none; font-weight: 600; 
            }
            QPushButton#ResumeBtn:hover { background-color: #218838; }
            
            QPushButton#DryRunBtn { 
                background-color: #8e44ad; color: white; border: none; font-weight: 600; 
            }
            QPushButton#DryRunBtn:hover { background-color: #7d3c98; }
            
            /* --- Input Fields --- */
            QLineEdit, QSpinBox, QComboBox { 
                padding: 2px 8px; border-radius: 4px; border: 1px solid #dcdcdc; 
                background-color: white; selection-background-color: #0078d4;
                min-height: 24px;
            }
            QLineEdit:focus, QSpinBox:focus, QComboBox:focus { border: 1px solid #0078d4; background-color: #fff; }
            
            /* --- Tables --- */
            QTableWidget { 
                background-color: white; border: 1px solid #e0e0e0; border-radius: 8px; 
                gridline-color: #f0f0f0; outline: none;
            }
            QTableWidget::item { padding: 2px; }
            QTableWidget::item:selected { background-color: #e5f1fb; color: #005a9e; }
            
            QHeaderView::section { font-weight: bold; color: #222; background-color: #f8f8f8; 
                background-color: #fcfcfc; padding: 6px; border: none; 
                border-bottom: 1px solid #e0e0e0; border-right: 1px solid #f0f0f0;
                font-weight: 600; color: #666;
            }
            
            /* --- Tabs --- */
            QTabWidget::pane { border: 1px solid #e0e0e0; border-radius: 8px; top: -1px; background: white; }
            QTabBar::tab { 
                font-size: 14px; font-weight: 400; padding: 12px 24px; 
                background: transparent; color: #666; border-bottom: 3px solid transparent;
            }
            QTabBar::tab:selected { color: #0078d4; border-bottom: 3px solid #0078d4; font-weight: 600; }
            QTabBar::tab:hover:not(:selected) { color: #333; background: #f0f0f0; border-radius: 4px 4px 0 0; }
            
            /* --- GroupBox --- */
            QGroupBox { font-weight: 600; color: #222; font-size: 12px; 
                font-weight: 600; border: 1px solid #e0e0e0; border-radius: 8px; 
                margin-top: 15px; padding-top: 15px; background-color: #ffffff;
            }
            QGroupBox::title { subcontrol-origin: margin; left: 15px; padding: 0 5px; color: #0078d4; }
            
            /* --- ScrollBars --- */
            QScrollBar:vertical { border: none; background: transparent; width: 10px; margin: 0px; }
            QScrollBar::handle:vertical { background: #cdcdcd; min-height: 20px; border-radius: 5px; }
            QScrollBar::handle:vertical:hover { background: #a6a6a6; }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0px; }
        """)

        central = QWidget(); self.setCentralWidget(central)
        main_h_layout = QHBoxLayout(central); main_h_layout.setContentsMargins(0, 0, 0, 0); main_h_layout.setSpacing(0)
        self.main_splitter = QSplitter(Qt.Horizontal)
        main_h_layout.addWidget(self.main_splitter)
        
        # --- 左侧任务导航栏 ---
        side_panel = QWidget(); side_panel.setMinimumWidth(96); side_panel.setObjectName("SidePanel")
        side_panel.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Preferred)
        self.side_panel = side_panel
        side_panel.setStyleSheet("""
            #SidePanel { background-color: #f3f3f3; border-right: 1px solid #e0e0e0; }
            QLabel#SideTitle { font-weight: 600; font-size: 16px; color: #222; padding: 10px 5px; }
            QLineEdit#SearchEdit { border-radius: 15px; padding: 5px 12px; background: white; border: 1px solid #e0e0e0; margin: 5px; }
            QLineEdit#SearchEdit:focus { border: 2px solid #0078d4; }
        """)
        side_ly = QVBoxLayout(side_panel); side_ly.setContentsMargins(10, 10, 10, 10); side_ly.setSpacing(8)
        
        side_ly.addWidget(QLabel("📂 任务管理", objectName="SideTitle"))
        self.task_search = QLineEdit(); self.task_search.setMinimumWidth(0); self.task_search.setObjectName("SearchEdit"); self.task_search.setPlaceholderText("🔍 搜索任务或分类..."); side_ly.addWidget(self.task_search)
        
        # [重构] 升级为 QTreeWidget 以支持文件夹管理
        self.task_tree = TaskTreeWidget()
        self.task_tree.setMinimumWidth(0)
        self.task_tree.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Expanding)
        self.task_tree.setHeaderHidden(True)
        self.task_tree.setIndentation(15)
        self.task_tree.setAnimated(False)
        self.task_tree.setUniformRowHeights(True)
        self.task_tree.setDragEnabled(True); self.task_tree.setAcceptDrops(True); self.task_tree.setDragDropMode(QAbstractItemView.DragDrop); self.task_tree.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self._task_tree_sync_timer = QTimer(self)
        self._task_tree_sync_timer.setSingleShot(True)
        self._task_tree_sync_timer.setInterval(180)
        self._task_tree_sync_timer.timeout.connect(self._sync_tree_structure_to_config)
        self.task_tree.structure_changed.connect(self._schedule_task_tree_sync)
        self.task_tree.setStyleSheet("""
            QTreeWidget { background: transparent; border: none; outline: none; }
            QTreeWidget::item { 
                padding: 8px 5px; border-radius: 6px; margin: 1px 0;
                color: #222222; border-left: 4px solid transparent;
                font-weight: 400; /* 默认加粗 */
            }
            QTreeWidget::item:hover { background: #e0e0e0; color: #222; }
            QTreeWidget::item:selected { 
                background: #ffffff; color: #005a9e; font-weight: 600; /* 选中时更粗 */
                border-left: 4px solid #0078d4;
            }
        """)
        self.task_tree.itemSelectionChanged.connect(self._on_tree_selection_changed)
        self.task_tree.itemDoubleClicked.connect(self._on_tree_item_double_clicked)
        self.task_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.task_tree.customContextMenuRequested.connect(self._show_task_list_menu)
        side_ly.addWidget(self.task_tree)
        
        self.task_search.textChanged.connect(self._filter_task_tree)

        # 任务操作按钮 (侧边栏底部)
        task_btn_grid = QGridLayout()
        btn_open_task = QPushButton("📖 打开"); btn_open_task.clicked.connect(self._open_task_in_editor)
        btn_add_task = QPushButton("➕ 新建任务"); btn_add_task.clicked.connect(self._add_task)
        btn_add_folder = QPushButton("📁 新建文件夹"); btn_add_folder.clicked.connect(self._add_folder)
        btn_move_task = QPushButton("📂 移动"); btn_move_task.clicked.connect(self._move_selected_tasks_to_folder)
        btn_rename_task = QPushButton("✏️ 改名"); btn_rename_task.clicked.connect(self._rename_task)
        btn_del_task = QPushButton("🗑️ 删除"); btn_del_task.clicked.connect(self._delete_task)
        btn_clone_task = QPushButton("📋 克隆"); btn_clone_task.clicked.connect(self._clone_task)
        btn_import_task = QPushButton("📥 导入任务"); btn_import_task.clicked.connect(self._import_task_templates)
        btn_export_task = QPushButton("📤 导出任务"); btn_export_task.clicked.connect(self._export_task_templates)
        for btn in (btn_open_task, btn_add_task, btn_add_folder, btn_move_task, btn_rename_task, btn_del_task, btn_clone_task, btn_import_task, btn_export_task):
            btn.setMinimumWidth(0)
            btn.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Preferred)
        
        task_btn_grid.addWidget(btn_open_task, 0, 0, 1, 2)
        task_btn_grid.addWidget(btn_add_task, 1, 0); task_btn_grid.addWidget(btn_add_folder, 1, 1)
        task_btn_grid.addWidget(btn_move_task, 2, 0); task_btn_grid.addWidget(btn_rename_task, 2, 1)
        task_btn_grid.addWidget(btn_del_task, 3, 0); task_btn_grid.addWidget(btn_clone_task, 3, 1)
        task_btn_grid.addWidget(btn_import_task, 4, 0); task_btn_grid.addWidget(btn_export_task, 4, 1)
        side_ly.addLayout(task_btn_grid)
        
        self.main_splitter.addWidget(side_panel)

        # --- 右侧主工作区 (使用纵向分栏) ---
        self.right_splitter = QSplitter(Qt.Vertical)
        self.main_splitter.addWidget(self.right_splitter)
        self.main_splitter.setHandleWidth(6)
        self.main_splitter.setChildrenCollapsible(False)
        self.main_splitter.setCollapsible(0, False)
        self.main_splitter.setCollapsible(1, False)
        self.main_splitter.setStretchFactor(1, 1)
        
        # 编辑区容器
        edit_container = QWidget(); edit_ly = QVBoxLayout(edit_container)
        edit_container.setMinimumWidth(0)
        edit_container.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.edit_container = edit_container
        edit_ly.setContentsMargins(10, 10, 10, 5)
        self.right_splitter.addWidget(edit_container)
        
        # --- Tab Layout ---
        self.tabs = QTabWidget()
        self.tabs.setMinimumWidth(0)
        self.tabs.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Preferred)
        
        self.tabs.setStyleSheet("""
            QTabWidget::pane { border: 1px solid #e0e0e0; top: -1px; background: white; }
            QTabBar::tab {
                background: #f8f8f8; color: #222222; font-weight: 600; font-size: 12px;
                padding: 10px 25px; border: 1px solid #e0e0e0;
                border-bottom: none; border-top-left-radius: 6px; border-top-right-radius: 6px;
                margin-right: 2px;
            }
            QTabBar::tab:selected { background: white; color: #005a9e; border-bottom: 2px solid #0078d4; font-weight: 700; }
            QTabBar::tab:hover { background: #eeeeee; }
        """)

        edit_ly.addWidget(self.tabs)

        # 恢复左右分栏宽度
        if layout_cfg.get("splitter_sizes"):
            self.main_splitter.setSizes(self._normalize_splitter_sizes(layout_cfg["splitter_sizes"]))
        else:
            self.main_splitter.setSizes([240, 960])

        # 初始化树形列表
        self._reload_task_tree()
        
        # [修复] 记忆上次任务功能
        last_task = self.config.get("layout", {}).get("last_task", "")
        if last_task:
            self._select_task_in_tree(last_task)
            self._on_task_changed(last_task)
        else:
            # 默认选中第一个任务（非文件夹）
            self._select_first_task()
            first_selected = self._get_primary_selected_task_path(fallback_current=False)
            if first_selected:
                self._on_task_changed(first_selected)
        
        # Tab 1: 流程编排
        self.tab_actions = QWidget(); act_ly = QVBoxLayout(self.tab_actions)
        act_ly.setContentsMargins(8, 10, 8, 8)
        act_ly.setSpacing(8)
        self.tabs.addTab(self.tab_actions, "🛠️ 流程编排")
        
        act_ctrl = QHBoxLayout()
        btn_add_act = QPushButton("➕ 添加步骤"); btn_add_act.clicked.connect(self._add_action); act_ctrl.addWidget(btn_add_act)
        btn_del_act = QPushButton("❌ 删除步骤"); btn_del_act.clicked.connect(self._del_action); act_ctrl.addWidget(btn_del_act)
        btn_copy_sel_act = QPushButton("📋 复制选中步骤"); btn_copy_sel_act.setToolTip("支持 Ctrl / Shift 多选后一起复制"); btn_copy_sel_act.clicked.connect(lambda _checked=False: self._copy_selected_actions()); act_ctrl.addWidget(btn_copy_sel_act)
        btn_move_sel_act = QPushButton("↕️ 批量挪到..."); btn_move_sel_act.setToolTip("将选中的步骤整体挪到指定位置"); btn_move_sel_act.clicked.connect(lambda _checked=False: self._move_selected_actions_to()); act_ctrl.addWidget(btn_move_sel_act)
        act_ctrl.addStretch()
        
        self.btn_test_flow = QPushButton("🧪 单次测试流程"); 
        self.btn_test_flow.setStyleSheet("background-color: #e3f2fd; border: 1px solid #2196F3; font-weight: bold; color: #1976d2;")
        self.btn_test_flow.setToolTip("仅执行一次当前编排的步骤（忽略批量数据表）")
        self.btn_test_flow.clicked.connect(lambda: self._execute(0, 0, 0, is_test=True))
        act_ctrl.addWidget(self.btn_test_flow)
        
        act_ly.addLayout(act_ctrl)
        
        self.action_table = DragSortActionTable(0, 5); self.action_table.setObjectName("ActionTable"); self.action_table.setHorizontalHeaderLabels(["步骤说明", "指令类型", "坐标/窗口", "默认参数/变量", "延时"])
        self.action_table.setHorizontalHeader(ManualWidthHeader(Qt.Horizontal, self.action_table))
        self.action_table.setAlternatingRowColors(True)
        self.action_table.setShowGrid(False)
        self.action_table.setWordWrap(False)
        self.action_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.action_table.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.action_table.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.action_table.verticalHeader().setMinimumSectionSize(34)
        self.action_table.verticalHeader().setDefaultSectionSize(38)
        self.action_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Interactive)
        self.action_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Fixed)
        self.action_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Interactive)
        self.action_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        self.action_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.Fixed)
        self.action_table.horizontalHeader().setStretchLastSection(False)
        self.action_table.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)
        self.action_table.setColumnWidth(1, 132)
        self.action_table.setColumnWidth(4, 92)
        self.action_table.horizontalHeader().sectionResized.connect(self._on_action_column_resized)
        self.action_table.verticalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        self.action_table.verticalHeader().customContextMenuRequested.connect(self._show_action_header_menu)
        # Fix: connect right-click context menu on action table rows
        self.action_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.action_table.customContextMenuRequested.connect(self._show_action_menu)
        # Fix: connect itemChanged so step name edits are saved immediately
        self.action_table.itemChanged.connect(self._on_action_item_changed)
        # Fix BUG 2: 绑定 rows_reordered 信号，将拖拽后的新顺序保存到配置中
        self.action_table.rows_reordered.connect(self._on_action_table_reordered)
        self.action_table.setStyleSheet("""
            QTableWidget#ActionTable {
                border: 1px solid #dbe4ee;
                border-radius: 10px;
                background: #ffffff;
                alternate-background-color: #f8fbff;
                gridline-color: transparent;
            }
            QTableWidget#ActionTable::item {
                padding: 6px 8px;
                border-bottom: 1px solid #edf2f7;
            }
            QTableWidget#ActionTable::item:selected {
                background: #e8f2ff;
                color: #0f4c81;
            }
            QTableWidget#ActionTable QLineEdit,
            QTableWidget#ActionTable QComboBox,
            QTableWidget#ActionTable QSpinBox,
            QTableWidget#ActionTable QTextEdit {
                min-height: 30px;
                border: 1px solid #d7dee7;
                border-radius: 7px;
                padding: 4px 8px;
                background: #ffffff;
                selection-background-color: #d8eafe;
            }
            QTableWidget#ActionTable QLineEdit:focus,
            QTableWidget#ActionTable QComboBox:focus,
            QTableWidget#ActionTable QSpinBox:focus,
            QTableWidget#ActionTable QTextEdit:focus {
                border: 1px solid #7aaef7;
                background: #ffffff;
            }
            QTableWidget#ActionTable QPushButton {
                min-height: 30px;
                padding: 4px 10px;
                border-radius: 7px;
                border: 1px solid #d7dee7;
                background: #f8fafc;
                color: #25364a;
                font-weight: 500;
            }
            QTableWidget#ActionTable QPushButton:hover {
                background: #eef5ff;
                border-color: #b9d2f7;
            }
            QTableWidget#ActionTable QPushButton:disabled {
                background: #f5f7fa;
                color: #9aa5b1;
                border-color: #e6ebf1;
            }
            QTableWidget#ActionTable QHeaderView::section {
                background: #f7f9fc;
                color: #5b6777;
                border: none;
                border-bottom: 1px solid #dbe4ee;
                border-right: 1px solid #edf2f7;
                padding: 8px 6px;
                font-weight: 600;
            }
        """)
        act_ly.addWidget(self.action_table)
        self.lbl_action_selection_count = create_table_selection_label()
        act_ly.addWidget(self.lbl_action_selection_count)
        bind_table_selection_label(self.action_table, self.lbl_action_selection_count)
        
        # Tab 2: 批量数据
        self.tab_data = QWidget(); data_ly = QVBoxLayout(self.tab_data)
        self.tabs.addTab(self.tab_data, "📊 批量数据")
        
        data_ctrl = QHBoxLayout()
        data_ctrl.setContentsMargins(0, 0, 0, 0)
        data_ctrl.setSpacing(6)
        btn_sync = QPushButton("🔄 同步表头"); btn_sync.setToolTip("根据流程编排同步批量数据的列结构：新增缺失列、移除无效列，不覆盖已填写内容。")
        btn_sync.clicked.connect(self._sync_data_headers); data_ctrl.addWidget(btn_sync)
        btn_reset = QPushButton("🧹 重置预设"); btn_reset.setStyleSheet("background-color: #ffebee; border: 1px solid #ef9a9a;")
        btn_reset.setToolTip("将当前任务的所有批量数据（可编辑步骤列）强制重置为流程编排里的默认参数/变量。不会影响“选择”勾选。")
        btn_reset.clicked.connect(self._reset_data_to_presets); data_ctrl.addWidget(btn_reset)
        btn_add_data = QPushButton("➕ 添加数据"); btn_add_data.clicked.connect(self._add_data_row); data_ctrl.addWidget(btn_add_data)
        btn_del_data = QPushButton("❌ 删除数据"); btn_del_data.clicked.connect(self._del_data_row); data_ctrl.addWidget(btn_del_data)
        btn_del_done_data = QPushButton("🧹 删除已完成"); btn_del_done_data.setToolTip("只删除当前任务里“已勾选”且状态为“完成”的数据行。"); btn_del_done_data.clicked.connect(self._del_completed_data_rows); data_ctrl.addWidget(btn_del_done_data)
        btn_batch_fill = QPushButton("🛠️ 批量填充"); btn_batch_fill.setStyleSheet("background-color: #f3e5f5; font-weight: bold;"); btn_batch_fill.setToolTip("批量填充账户、文件路径、输入文本等，支持文件夹扫描"); btn_batch_fill.clicked.connect(self._batch_assign_profiles); data_ctrl.addWidget(btn_batch_fill)
        btn_prefix_library = QPushButton("📚 前缀库"); btn_prefix_library.setStyleSheet("background-color: #fff8e1; border: 1px solid #ffd54f; font-weight: bold;"); btn_prefix_library.setToolTip("直接在主界面打开常用前缀库，无需先进入批量填充中心。"); btn_prefix_library.clicked.connect(self._open_prefix_library); data_ctrl.addWidget(btn_prefix_library)

        self.btn_select_all = QPushButton("☑️ 全选"); self.btn_select_all.clicked.connect(lambda: self._set_all_row_check_state(True)); data_ctrl.addWidget(self.btn_select_all)
        self.btn_deselect_all = QPushButton("☐ 取消全选"); self.btn_deselect_all.clicked.connect(lambda: self._set_all_row_check_state(False)); data_ctrl.addWidget(self.btn_deselect_all)
        self.btn_invert_select = QPushButton("🔁 反选"); self.btn_invert_select.clicked.connect(self._invert_row_check_state); data_ctrl.addWidget(self.btn_invert_select)
        self.btn_select_unsuccessful = QPushButton("⚠️ 选未成功"); self.btn_select_unsuccessful.setToolTip("一键勾选本次执行里未成功的行，例如失败、跳过、挂起或需人工介入的行。"); self.btn_select_unsuccessful.clicked.connect(self._select_unsuccessful_rows); data_ctrl.addWidget(self.btn_select_unsuccessful)

        # 子任务管理器：对“勾选的行”批量执行某个子任务（不限失败项）
        self.btn_subtask_mgr = QPushButton("🧩 子任务管理")
        self.btn_subtask_mgr.setMinimumWidth(122)
        self.btn_subtask_mgr.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Fixed)
        self.btn_subtask_mgr.setToolTip("打开子任务管理器：对当前任务里“已勾选”的数据行，批量执行你选择的子任务（可选自动激活窗口）。")
        self.btn_subtask_mgr.clicked.connect(self._open_subtask_manager)
        data_ctrl.addWidget(self.btn_subtask_mgr)

        self.btn_more_data_ops = QPushButton("⚙️ 更多")
        self.btn_more_data_ops.setToolTip("打开批量数据的更多操作和开关。")
        self.btn_more_data_ops.clicked.connect(self._show_data_toolbar_menu)
        data_ctrl.addWidget(self.btn_more_data_ops)
        data_ctrl.addStretch()

        self.btn_toggle_delay = QPushButton("⏱️ 显示延时"); self.btn_toggle_delay.setCheckable(True); self.btn_toggle_delay.clicked.connect(self._toggle_delay_cols)
        self.btn_toggle_noneditable = QPushButton()
        self.btn_toggle_noneditable.setCheckable(True)
        self.btn_toggle_noneditable.setChecked(layout_cfg.get("show_noneditable_steps", False))
        self.btn_toggle_noneditable.clicked.connect(self._toggle_noneditable_cols)
        self.data_row_height_spin = QSpinBox()
        self.data_row_height_spin.setRange(24, 96)
        self.data_row_height_spin.setSuffix(" px")
        self.data_row_height_spin.setValue(int(layout_cfg.get("data_row_height", 28)))
        self.data_row_height_spin.setToolTip("统一调整批量数据区所有行的高度；也可直接拖动左侧行号边界来一起调整。")
        self.data_row_height_spin.valueChanged.connect(self._set_data_row_height)
        self.data_col_width_label = QLabel("整体行高: 28 px")
        self.data_col_width_label.setToolTip("外部滑杆统一调整批量数据区所有行的高度。")
        self.data_col_width_slider = QSlider(Qt.Horizontal)
        self.data_col_width_slider.setRange(24, 96)
        self.data_col_width_slider.setSingleStep(1)
        self.data_col_width_slider.setPageStep(4)
        self.data_col_width_slider.setFixedWidth(180)
        self.data_col_width_slider.setValue(int(layout_cfg.get("data_row_height", 28)))
        self.data_col_width_slider.setToolTip("统一调整批量数据区所有行的高度。")
        self.data_col_width_slider.valueChanged.connect(self._set_data_row_height)
        self.chk_only_current = QCheckBox("仅执行当前任务"); self.chk_only_current.setChecked(True)
        self.chk_continue_on_fail = QCheckBox("❌ 失败也继续"); self.chk_continue_on_fail.setToolTip("当某个任务因“停止策略”失败时，是否继续执行后续任务；失败项会在最后汇总。"); self.chk_continue_on_fail.setChecked(True)
        self.chk_auto_shutdown = QCheckBox("🏁 任务完关机")
        self.chk_show_osd = QCheckBox("🖥️ 显示置顶进度条"); self.chk_show_osd.setChecked(True)
        self.chk_std_win = QCheckBox("📏 自动标准化窗口"); self.chk_std_win.setToolTip("开启后，每次打开网址或激活窗口，都会自动将窗口移动到(0,0)并全屏，确保点击位置准确。"); self.chk_std_win.setChecked(True)
        self.chk_multi_open = QCheckBox("👯 多账号并行模式"); self.chk_multi_open.setToolTip("开启后，不同账号的浏览器窗口可以同时并存（需确保账号路径唯一）。"); self.chk_multi_open.setChecked(True)
        data_ctrl.addWidget(self.data_col_width_label)
        data_ctrl.addWidget(self.data_col_width_slider)
        data_ly.addLayout(data_ctrl)
        
        self.data_table = DataEditorTable(0, 0); self.data_table.setColumnCount(3); self.data_table.setHorizontalHeaderLabels(["选择", "执行", "状态"])
        self.data_table.setHorizontalHeader(ManualWidthHeader(Qt.Horizontal, self.data_table))
        self.data_table.verticalHeader().setDefaultSectionSize(int(layout_cfg.get("data_row_height", 28))) # 批量数据区默认更紧凑，同时允许统一调整
        self.data_table.verticalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.data_table.verticalHeader().setMinimumSectionSize(24)
        self.data_table.verticalHeader().sectionResized.connect(self._on_data_row_resized)
        self.data_table.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)
        self.data_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.data_table.horizontalHeader().sectionResized.connect(self._on_data_column_resized)
        self.data_table.setWordWrap(False)
        self.data_table.setTextElideMode(Qt.ElideRight)
        self.data_table.verticalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        self.data_table.verticalHeader().customContextMenuRequested.connect(self._show_data_header_menu)
        # Fix: connect right-click on data cells
        self.data_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.data_table.customContextMenuRequested.connect(self._show_data_menu)
        self.data_table.horizontalHeader().sectionClicked.connect(self._on_data_header_clicked)
        self.data_table.itemChanged.connect(self._on_data_item_changed)
        # [新增] 点击单元格时如果是图片路径则自动弹出预览
        self.data_table.cellClicked.connect(self._on_data_cell_clicked_preview)
        self._config_flush_timer = QTimer(self)
        self._config_flush_timer.setSingleShot(True)
        self._config_flush_timer.setInterval(260)
        self._config_flush_timer.timeout.connect(lambda: save_config(self.config))
        self._update_noneditable_toggle_text(self.btn_toggle_noneditable.isChecked())
        self._syncing_data_row_height = False
        self._syncing_data_column_width = False
        self._syncing_action_column_width = False
        self._set_data_row_height(int(layout_cfg.get("data_row_height", 28)), save=False)
        data_ly.addWidget(self.data_table)
        self.lbl_data_selection_count = create_table_selection_label()
        data_ly.addWidget(self.lbl_data_selection_count)
        bind_table_selection_label(self.data_table, self.lbl_data_selection_count)

        # Tab 3: 任务排程
        self.tab_schedule = QWidget(); sched_ly = QVBoxLayout(self.tab_schedule)
        self.tabs.addTab(self.tab_schedule, "📅 任务排程")

        sched_bundle_bar = QHBoxLayout()
        sched_bundle_bar.addWidget(QLabel("排程包:"))
        self.sched_bundle_combo = QComboBox()
        self.sched_bundle_combo.setToolTip("一个排程包就是一组可复用的任务序列和定时设置")
        self.sched_bundle_combo.currentTextChanged.connect(self._on_schedule_bundle_changed)
        sched_bundle_bar.addWidget(self.sched_bundle_combo, 1)
        btn_add_bundle = QPushButton("➕ 新建排程包"); btn_add_bundle.clicked.connect(self._add_schedule_bundle); sched_bundle_bar.addWidget(btn_add_bundle)
        btn_rename_bundle = QPushButton("✏️ 重命名"); btn_rename_bundle.clicked.connect(self._rename_schedule_bundle); sched_bundle_bar.addWidget(btn_rename_bundle)
        btn_del_bundle = QPushButton("🗑️ 删除"); btn_del_bundle.clicked.connect(self._delete_schedule_bundle); sched_bundle_bar.addWidget(btn_del_bundle)
        sched_ly.addLayout(sched_bundle_bar)

        sched_ctrl = QHBoxLayout()
        btn_add_sched = QPushButton("➕ 添加到排程"); btn_add_sched.clicked.connect(self._add_to_schedule); sched_ctrl.addWidget(btn_add_sched)
        btn_del_sched = QPushButton("❌ 移除排程"); btn_del_sched.clicked.connect(self._del_from_schedule); sched_ctrl.addWidget(btn_del_sched)
        btn_move_sched_up = QPushButton("⬆️ 上移"); btn_move_sched_up.clicked.connect(self._move_selected_sched_up); sched_ctrl.addWidget(btn_move_sched_up)
        btn_move_sched_down = QPushButton("⬇️ 下移"); btn_move_sched_down.clicked.connect(self._move_selected_sched_down); sched_ctrl.addWidget(btn_move_sched_down)
        sched_ctrl.addStretch()
        self.chk_sched_shutdown = QCheckBox("🏁 任务完关机"); sched_ctrl.addWidget(self.chk_sched_shutdown)
        self.run_sched_btn = QPushButton("🔥 启动序列化排程"); self.run_sched_btn.setStyleSheet("background-color: #ffe0b2; font-weight: bold;"); self.run_sched_btn.clicked.connect(self._run_schedule); sched_ctrl.addWidget(self.run_sched_btn)
        sched_ly.addLayout(sched_ctrl)

        self.sched_table = QTableWidget(0, 3); self.sched_table.setHorizontalHeaderLabels(["执行", "任务选择", "状态"])
        self.sched_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.sched_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.sched_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.sched_table.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)
        self.sched_table.verticalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        self.sched_table.verticalHeader().customContextMenuRequested.connect(self._show_sched_header_menu)
        sched_ly.addWidget(self.sched_table)
        self.lbl_sched_selection_count = create_table_selection_label()
        sched_ly.addWidget(self.lbl_sched_selection_count)
        bind_table_selection_label(self.sched_table, self.lbl_sched_selection_count)

        console_panel = QWidget(); console_layout = QVBoxLayout(console_panel); console_layout.setContentsMargins(10, 5, 10, 10)
        self.right_splitter.addWidget(console_panel)
        self.right_splitter.setStretchFactor(0, 3) # 默认编辑区占大头
        self.right_splitter.setStretchFactor(1, 1)
        
        # 恢复上下分栏高度
        if layout_cfg.get("v_splitter_sizes"):
            self.right_splitter.setSizes(layout_cfg["v_splitter_sizes"])

        console_layout.addWidget(QLabel("🎮 3. 运行控制台", objectName="TitleLabel"))
        h_ctrl = QHBoxLayout()
        label_interval = QLabel("轮次间隔:")
        label_interval.setToolTip("执行完表格的一行数据后，等待多久再开始下一行")
        h_ctrl.addWidget(label_interval)
        self.delay_spin = QSpinBox(); self.delay_spin.setSuffix(" 秒"); self.delay_spin.setValue(3); h_ctrl.addWidget(self.delay_spin)
        h_ctrl.addWidget(QLabel("循环执行:"))
        self.loop_spin = QSpinBox(); self.loop_spin.setSuffix(" 次"); self.loop_spin.setMinimum(1); self.loop_spin.setValue(1); h_ctrl.addWidget(self.loop_spin)

        h_ctrl.addWidget(QLabel("出错重试:"))
        self.retry_spin = QSpinBox(); self.retry_spin.setSuffix(" 次"); self.retry_spin.setRange(0, 10); self.retry_spin.setValue(0)
        self.retry_spin.setToolTip("步骤出错时自动重试次数，0=不重试"); h_ctrl.addWidget(self.retry_spin)
        h_ctrl.addWidget(QLabel("出错策略:"))
        self.error_combo = QComboBox()
        self.error_combo.addItems(["🛑 停止", "⏭️ 跳过步骤", "❌ 标记失败并继续"])
        self.error_combo.setToolTip("出错时：停止整个任务 / 跳过当前步骤 / 标记当前行失败并继续下一行")
        h_ctrl.addWidget(self.error_combo)
        
        self.btn_timer = QPushButton("⏰ 计划时间"); self.btn_timer.setCheckable(True); self.btn_timer.setFixedHeight(45); self.btn_timer.clicked.connect(self._toggle_timer); h_ctrl.addWidget(self.btn_timer)

        self.btn_dry_run = QPushButton("🧪 试运行"); self.btn_dry_run.setObjectName("DryRunBtn")
        self.btn_dry_run.setFixedHeight(45)
        self.btn_dry_run.setStyleSheet("background-color: #9C27B0; color: white; border: none; font-weight: bold;")
        self.btn_dry_run.setToolTip("试运行：遍历所有步骤并打印将要执行的操作，但不移动鼠标、不敲键盘")
        self.btn_dry_run.clicked.connect(self._run_dry)
        h_ctrl.addWidget(self.btn_dry_run)
        
        self.btn_run = QPushButton("🚀 启动批量任务 (按数据表)"); self.btn_run.setObjectName("RunBtn"); self.btn_run.setFixedHeight(45); self.btn_run.clicked.connect(self._run_all); h_ctrl.addWidget(self.btn_run, 1)
        self.btn_pause = QPushButton("⏸️ 暂停"); self.btn_pause.setFixedHeight(45); self.btn_pause.setStyleSheet("background-color: #ff9800; color: white; font-weight: bold;"); self.btn_pause.setEnabled(False); self.btn_pause.clicked.connect(self._toggle_pause); h_ctrl.addWidget(self.btn_pause, 1)
        self.btn_resume = QPushButton("⏯️ 恢复执行"); self.btn_resume.setObjectName("ResumeBtn"); self.btn_resume.setFixedHeight(45); self.btn_resume.setEnabled(False); self.btn_resume.clicked.connect(self._resume_execution); h_ctrl.addWidget(self.btn_resume, 1)
        self.btn_stop = QPushButton("🛑 停止执行"); self.btn_stop.setFixedHeight(45); self.btn_stop.setStyleSheet("background-color: #f44336; color: white; font-weight: bold;"); self.btn_stop.setEnabled(False); self.btn_stop.clicked.connect(self._stop_execution); h_ctrl.addWidget(self.btn_stop, 1)
        console_layout.addLayout(h_ctrl)

        # Hotkey status bar
        hk_row_widget = QWidget()
        hk_row_widget.setMinimumHeight(45) # 显式设置热键栏最小高度，防止遮挡
        hk_row = QHBoxLayout(hk_row_widget)
        hk_row.setContentsMargins(0, 5, 0, 5)
        self.lbl_hotkey = QLabel()
        self.lbl_hotkey.setStyleSheet("color: #555; font-size: 12px;")
        hk_row.addWidget(self.lbl_hotkey)
        hk_row.addStretch()
        btn_hk_settings = QPushButton("⌨️ 全局热键设置")
        btn_hk_settings.setFixedHeight(32) # 增加按钮高度
        btn_hk_settings.clicked.connect(self._show_hotkey_settings)
        hk_row.addWidget(btn_hk_settings)
        # Screenshot button
        btn_cap = QPushButton("📸 框选截图 (图像识别用)")
        btn_cap.setFixedHeight(32) # 增加按钮高度
        btn_cap.setStyleSheet("background-color: #0288d1; color: white;")
        btn_cap.setToolTip("框选屏幕区域并保存为 PNG，可直接用作图像识别步骤的目标图片")
        btn_cap.clicked.connect(self._start_region_capture)
        hk_row.addWidget(btn_cap)
        console_layout.addWidget(hk_row_widget)
        defer_group = QGroupBox("⏸️ 挂起队列")
        defer_group.setStyleSheet("QGroupBox { font-weight: 600; color: #334155; }")
        defer_ly = QVBoxLayout(defer_group)
        defer_ly.setContentsMargins(8, 8, 8, 8)
        defer_ly.setSpacing(6)
        self.lbl_deferred_summary = QLabel("当前无挂起项")
        self.lbl_deferred_summary.setStyleSheet("color: #64748b;")
        defer_ly.addWidget(self.lbl_deferred_summary)
        self.defer_queue_table = QTableWidget(0, 6)
        self.defer_queue_table.setHorizontalHeaderLabels(["轮 / 组", "挂起步骤", "恢复步骤", "恢复策略", "剩余", "状态"])
        self.defer_queue_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.defer_queue_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.defer_queue_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.defer_queue_table.verticalHeader().setVisible(False)
        self.defer_queue_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.defer_queue_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.defer_queue_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.defer_queue_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.defer_queue_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        self.defer_queue_table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeToContents)
        self.defer_queue_table.setMinimumHeight(136)
        defer_ly.addWidget(self.defer_queue_table)
        console_layout.addWidget(defer_group)
        self._deferred_queue_data = []
        self._deferred_panel_timer = QTimer(self)
        self._deferred_panel_timer.setInterval(1000)
        self._deferred_panel_timer.timeout.connect(self._refresh_deferred_queue_panel)
        self.progress = QProgressBar(); console_layout.addWidget(self.progress); self.log_area = QTextEdit(); self.log_area.setReadOnly(True); self.log_area.setStyleSheet("background-color: #1e1e1e; color: #d4d4d4; font-family: 'Consolas', '微软雅黑'; font-size: 12px;"); self.log_area.setMinimumHeight(90); console_layout.addWidget(self.log_area, 1)
        # content_ly.addWidget(console_panel) # 已经加入到 right_splitter 中了
        self._refresh_actions()
        self._refresh_data_table()
        self._populate_schedule_bundle_combo()
        self._load_schedule_bundle(self._get_current_schedule_bundle_name())
        self._register_hotkeys()
        if self._config_load_error:
            QTimer.singleShot(0, self._show_config_load_warning)
        
        # [修复] 所有组件创建完成后，完成初始化状态切换
        self._is_initializing = False

    def _on_action_table_reordered(self):
        if not hasattr(self.action_table, '_pending_reorder') or not self.current_task: return
        rows, insert_at = self.action_table._pending_reorder
        if isinstance(rows, int):
            rows = [rows]
        self._reorder_action_rows(rows, insert_at, log_text="↕️ 步骤已通过拖拽重新排序")

    def _on_action_item_changed(self, item):
        # Guard: ignore changes triggered during _refresh_actions (blockSignals covers table-level signals
        # but setCellWidget/setItem on col 0 can still fire in edge cases; double-check signals state)
        if self.action_table.signalsBlocked(): return
        if item.column() == 0 and self.current_task:
            idx = item.row()
            if idx < len(self.config['tasks'][self.current_task]):
                old_name = self.config['tasks'][self.current_task][idx]['name']
                new_name = item.text()
                if old_name != new_name:
                    unique_name = self._make_unique_action_name(new_name, skip_idx=idx)
                    self.config['tasks'][self.current_task][idx]['name'] = unique_name
                    if unique_name != new_name:
                        self.action_table.blockSignals(True)
                        item.setText(unique_name)
                        self.action_table.blockSignals(False)
                    # Migrate data for all rows from old_name to new_name
                    if self.current_task in self.config['task_data']:
                        for row in self.config['task_data'][self.current_task]:
                            if not isinstance(row, dict):
                                continue
                            rename_pairs = [
                                (old_name, unique_name),
                                (f"{old_name}_延时", f"{unique_name}_延时"),
                                (f"{old_name}_跳过", f"{unique_name}_跳过"),
                            ]
                            for old_key, new_key in rename_pairs:
                                if old_key in row:
                                    row[new_key] = row.pop(old_key)
                    self._rename_defer_target_references(old_name, unique_name)
                    save_config(self.config)
                    self._refresh_defer_target_options()
                    self._refresh_data_table()

    def _format_seconds(self, s):
        if s < 60: return f"{s}秒"
        if s < 3600: return f"{s//60}分{s%60}秒"
        return f"{s//3600}时{(s%3600)//60}分{s%60}秒"

    def _on_task_changed(self, name):
        if not name or self._is_initializing: return
        # 检查是否真的发生了任务切换，避免重复点击导致的卡顿
        if hasattr(self, 'current_task') and self.current_task == name:
            return

        # [优化] 使用定时器延迟保存宽度，避免点击时同步 IO 导致卡顿
        if hasattr(self, 'current_task') and self.current_task:
            QTimer.singleShot(500, self._save_column_widths)
            
        self.current_task = name
        
        # [优化] 任务切换时不再立即写入磁盘，改为内存记忆，程序关闭时统一保存
        if "layout" not in self.config: self.config["layout"] = {}
        self.config["layout"]["last_task"] = name
        
        # [优化] 渲染前停止刷新，渲染后再恢复，减少重绘开销
        self.action_table.setUpdatesEnabled(False)
        self.data_table.setUpdatesEnabled(False)
        try:
            self._refresh_actions()
            self._refresh_data_table()
        finally:
            self.action_table.setUpdatesEnabled(True)
            self.data_table.setUpdatesEnabled(True)
        
        # 不再做任何自动列宽适应，只恢复用户自己调过的宽度
        QTimer.singleShot(100, self._restore_action_column_widths)
        QTimer.singleShot(180, self._restore_column_widths)
        QTimer.singleShot(240, lambda: self._set_data_row_height(self.config.get("layout", {}).get("data_row_height", 28), save=False))

    def _save_column_widths(self):
        """保存当前任务的列宽到配置中。"""
        # [修复] 增加对 data_table 的防御性检查
        if not self.current_task or not hasattr(self, 'data_table'): return
        widths = []
        for i in range(self.data_table.columnCount()):
            widths.append(self.data_table.columnWidth(i))
        
        if "tasks_layout" not in self.config:
            self.config["tasks_layout"] = {}
        self.config["tasks_layout"][self.current_task] = widths
        self._schedule_config_flush(400)

    def _restore_column_widths(self):
        """从配置中恢复当前任务的列宽。"""
        if not self.current_task or not hasattr(self, "data_table"):
            return
        widths = self.config.get("tasks_layout", {}).get(self.current_task, [])
        acts = self.config.get('tasks', {}).get(self.current_task, [])
        show_delay = self.btn_toggle_delay.isChecked() if hasattr(self, "btn_toggle_delay") else False
        default_widths = [42, 64, 54]
        for action in acts:
            default_widths.append(self._get_data_column_width(action))
            if show_delay:
                default_widths.append(self._get_data_column_width(is_delay=True))

        expected_old_count = 2 + len(acts) * (2 if show_delay else 1)
        expected_new_count = 3 + len(acts) * (2 if show_delay else 1)
        if len(widths) == expected_old_count and expected_new_count == expected_old_count + 1:
            widths = [widths[0], 64, widths[1], *widths[2:]]

        target_count = self.data_table.columnCount()
        if target_count <= 0:
            return

        merged_widths = []
        for i in range(target_count):
            if i < len(widths):
                merged_widths.append(min(max(42, int(widths[i])), 420))
            elif i < len(default_widths):
                merged_widths.append(min(max(42, int(default_widths[i])), 420))
            else:
                merged_widths.append(120)

        self._syncing_data_column_width = True
        try:
            for i, w in enumerate(merged_widths):
                self.data_table.setColumnWidth(i, w)
        finally:
            self._syncing_data_column_width = False
        self.data_table.horizontalHeader().setMinimumSectionSize(42)

    def _save_action_column_widths(self):
        if not hasattr(self, "action_table"):
            return
        hdr = self.action_table.horizontalHeader()
        self.config.setdefault("layout", {})
        self.config["layout"]["action_col_widths"] = [hdr.sectionSize(i) for i in range(hdr.count())]
        self._schedule_config_flush(400)

    def _restore_action_column_widths(self):
        if not hasattr(self, "action_table"):
            return
        widths = self.config.get("layout", {}).get("action_col_widths", [190, 132, 150, 340, 92])
        hdr = self.action_table.horizontalHeader()
        self._syncing_action_column_width = True
        try:
            for i in range(min(hdr.count(), len(widths))):
                width = int(widths[i])
                if i == 1:
                    width = min(max(120, width), 170)
                elif i == 3:
                    width = min(max(240, width), 640)
                elif i == 4:
                    width = min(max(88, width), 108)
                else:
                    width = min(max(96, width), 520)
                self.action_table.setColumnWidth(i, width)
        finally:
            self._syncing_action_column_width = False
        hdr.setMinimumSectionSize(88)

    def _on_action_column_resized(self, logical_index, old_size, new_size):
        if getattr(self, "_syncing_action_column_width", False):
            return
        self._save_action_column_widths()

    def _on_data_column_resized(self, logical_index, old_size, new_size):
        if getattr(self, "_syncing_data_column_width", False):
            return
        self._save_column_widths()

    def _get_task_meta(self, task_id, create_missing=False):
        self.config.setdefault("task_meta", {})
        meta = self.config["task_meta"].get(task_id)
        if meta is None and create_missing and task_id:
            meta = _normalize_task_meta(task_id, {})
            self.config["task_meta"][task_id] = meta
        return meta or {}

    def _get_task_name_only(self, task_id):
        return self._get_task_meta(task_id, create_missing=True).get("name", "")

    def _get_task_folder(self, task_id):
        return self._get_task_meta(task_id, create_missing=True).get("folder", "")

    def _get_task_path(self, task_id):
        folder = self._get_task_folder(task_id)
        name = self._get_task_name_only(task_id)
        return self._join_tree_path(folder, name)

    def _set_task_location(self, task_id, folder=None, name=None):
        meta = self._get_task_meta(task_id, create_missing=True)
        if folder is not None:
            meta["folder"] = (folder or "").strip().replace("\\", "/").strip("/")
        if name is not None:
            meta["name"] = (name or "").strip()
        self.config["task_meta"][task_id] = _normalize_task_meta(task_id, meta)

    def _build_duplicate_task_labels(self):
        labels = {}
        groups = {}
        for task_id in self._get_task_names():
            key = (self._get_task_folder(task_id), self._get_task_name_only(task_id))
            groups.setdefault(key, []).append(task_id)
        for (_, name), ids in groups.items():
            if len(ids) <= 1:
                labels[ids[0]] = name
                continue
            for idx, task_id in enumerate(ids, start=1):
                labels[task_id] = f"{name} 〔{idx}〕"
        return labels

    def _get_task_tree_label(self, task_id, label_map=None):
        label_map = label_map or self._build_duplicate_task_labels()
        return label_map.get(task_id, self._get_task_name_only(task_id))

    def _get_task_display_text(self, task_id, with_folder=False):
        if not task_id:
            return ""
        name = self._get_task_tree_label(task_id)
        folder = self._get_task_folder(task_id)
        if with_folder and folder:
            return f"{folder}/{name}"
        return name

    def _activate_task_by_id(self, task_id, open_editor=True):
        if not task_id or task_id not in self.config.get("tasks", {}):
            return
        self._select_task_in_tree(task_id)
        if open_editor:
            self._on_task_changed(task_id)

    def _get_task_names(self):
        # [优化] 支持自定义排序顺序；内部统一返回 task_id 列表
        all_tasks = list(self.config.get('tasks', {}).keys())
        order = self.config.get('layout', {}).get('task_order', [])
        if order:
            # 按照保存的顺序排序，新任务追加到末尾
            sorted_tasks = [t for t in order if t in all_tasks]
            sorted_tasks += [t for t in all_tasks if t not in sorted_tasks]
            return sorted_tasks
        return sorted(all_tasks)

    def _on_task_reordered(self):
        """当用户手动拖拽任务排序后触发，保存新顺序。"""
        # [优化] 使用定时器稍微延迟保存，防止频繁拖拽导致的卡顿
        QTimer.singleShot(500, self._save_task_order)

    def _schedule_task_tree_sync(self):
        """拖拽完成后稍作合并再落盘，减少连续操作时的卡顿。"""
        write_drag_debug(
            f"[AutoManager] _schedule_task_tree_sync: initializing={getattr(self, '_is_initializing', False)} "
            f"drag_in_progress={getattr(self.task_tree, '_drag_in_progress', False)} "
            f"last_move_info={getattr(self.task_tree, '_last_move_info', [])}"
        )
        if getattr(self, "_is_initializing", False):
            return
        if hasattr(self, "_task_tree_sync_timer"):
            write_drag_debug("[AutoManager] _schedule_task_tree_sync: start timer(180ms)")
            self._task_tree_sync_timer.start()
        else:
            write_drag_debug("[AutoManager] _schedule_task_tree_sync: timer missing, direct sync")
            self._sync_tree_structure_to_config()

    def _save_task_order(self):
        # [适配重构] 遍历树形结构获取所有任务的新顺序
        new_order = []
        iterator = QTreeWidgetItemIterator(self.task_tree)
        while iterator.value():
            task_path = iterator.value().data(0, Qt.UserRole)
            if task_path and not self._is_folder_role(task_path):
                new_order.append(task_path)
            iterator += 1
            
        if "layout" not in self.config: self.config["layout"] = {}
        self.config["layout"]["task_order"] = new_order
        write_drag_debug(f"[AutoManager] _save_task_order: new_order={new_order}")
        save_config(self.config)
        self._log("📍 任务列表顺序已更新并保存", "gray")

    def _is_folder_role(self, role):
        return isinstance(role, str) and role.startswith("[FOLDER]")

    def _make_folder_role(self, folder_path=""):
        return f"[FOLDER]::{folder_path}" if folder_path else "[FOLDER]"

    def _folder_path_from_role(self, role):
        if not self._is_folder_role(role):
            return ""
        return role.split("::", 1)[1] if "::" in role else ""

    def _join_tree_path(self, parent_path, name):
        name = (name or "").strip().strip("/")
        parent_path = (parent_path or "").strip().strip("/")
        if not parent_path:
            return name
        if not name:
            return parent_path
        return f"{parent_path}/{name}"

    def _split_tree_path(self, path):
        path = (path or "").strip().strip("/")
        if not path:
            return "", ""
        if "/" not in path:
            return "", path
        return path.rsplit("/", 1)

    def _tree_item_path(self, item):
        if not item:
            return ""
        role = item.data(0, Qt.UserRole)
        if self._is_folder_role(role):
            return self._folder_path_from_role(role)
        return role or item.text(0)

    def _get_all_folder_paths(self):
        folder_paths = set()
        for folder_path in self.config.get("folders", []):
            folder_path = (folder_path or "").strip().strip("/")
            if not folder_path:
                continue
            parts = folder_path.split("/")
            for i in range(1, len(parts) + 1):
                folder_paths.add("/".join(parts[:i]))
        for task_id in self.config.get("tasks", {}):
            folder_path = self._get_task_folder(task_id)
            if not folder_path:
                continue
            parts = folder_path.split("/")
            for i in range(1, len(parts) + 1):
                folder_paths.add("/".join(parts[:i]))
        return sorted(folder_paths, key=lambda p: (p.count("/"), p.lower()))

    def _refresh_tree_item_roles(self):
        def _walk(parent_item=None, parent_path=""):
            count = parent_item.childCount() if parent_item else self.task_tree.topLevelItemCount()
            getter = parent_item.child if parent_item else self.task_tree.topLevelItem
            for i in range(count):
                item = getter(i)
                item_path = self._join_tree_path(parent_path, item.text(0))
                role = item.data(0, Qt.UserRole)
                if self._is_folder_role(role):
                    item.setData(0, Qt.UserRole, self._make_folder_role(item_path))
                    _walk(item, item_path)
        _walk()

    def _clear_tree_origin_roles(self):
        origin_role = Qt.UserRole + 1
        iterator = QTreeWidgetItemIterator(self.task_tree)
        while iterator.value():
            item = iterator.value()
            item.setData(0, origin_role, None)
            iterator += 1

    def _normalize_splitter_sizes(self, sizes, minimum_left=96, minimum_right=360):
        if not sizes or len(sizes) < 2:
            return [240, 960]
        left = max(int(sizes[0]), minimum_left)
        right = max(int(sizes[1]), minimum_right)
        total = left + right
        if total <= 0:
            return [240, 960]
        available = max(self.width(), minimum_left + minimum_right + 120)
        if total > available:
            scale = available / total
            left = max(minimum_left, int(left * scale))
            right = max(minimum_right, available - left)
            if left + right > available:
                left = max(minimum_left, available - right)
        return [left, right]

    def _sync_tree_structure_to_config(self):
        """将树形结构的拖拽变动同步到 task_id 元信息，避免因路径变化而丢任务。"""
        write_drag_debug(
            f"[AutoManager] _sync_tree_structure_to_config:start tree={format_tree_snapshot(self.task_tree, self._is_folder_role)}"
        )
        if hasattr(self, "_task_tree_sync_timer") and self._task_tree_sync_timer.isActive():
            self._task_tree_sync_timer.stop()
            write_drag_debug("[AutoManager] _sync_tree_structure_to_config: stopped active timer before syncing")

        new_folders_set = set()
        new_task_order = []
        moved_count = 0

        def _collect_paths(item, parent_path=""):
            nonlocal moved_count
            if hasattr(item, "topLevelItemCount"):
                count = item.topLevelItemCount()
                get_child = item.topLevelItem
            else:
                count = item.childCount()
                get_child = item.child

            for i in range(count):
                child = get_child(i)
                if not child:
                    continue

                name = child.text(0)
                role = child.data(0, Qt.UserRole)
                new_path = self._join_tree_path(parent_path, name)

                if self._is_folder_role(role):
                    new_folders_set.add(new_path)
                    _collect_paths(child, new_path)
                else:
                    task_id = role
                    if not task_id or task_id not in self.config.get("tasks", {}):
                        continue
                    raw_name = child.data(0, Qt.UserRole + 2) or self._get_task_name_only(task_id)
                    old_folder = self._get_task_folder(task_id)
                    if old_folder != parent_path:
                        moved_count += 1
                    self._set_task_location(task_id, folder=parent_path, name=raw_name)
                    new_task_order.append(task_id)

        _collect_paths(self.task_tree)
        write_drag_debug(
            f"[AutoManager] _sync_tree_structure_to_config: collected new_folders={sorted(new_folders_set)} "
            f"new_task_order={new_task_order} moved_count={moved_count} last_move_info={getattr(self.task_tree, '_last_move_info', [])}"
        )

        original_task_count = len(self.config.get("tasks", {}))
        tree_task_count = len(new_task_order)
        if original_task_count > 0 and tree_task_count == 0:
            write_drag_debug(
                f"[AutoManager] _sync_tree_structure_to_config: protection triggered original_task_count={original_task_count} tree_task_count=0"
            )
            self._log("⚠️ 同步保护：检测到任务遍历异常，已拦截自动保存以防止数据丢失。", "red")
            self._reload_task_tree()
            if getattr(self, "current_task", ""):
                self._select_task_in_tree(self.current_task)
            self.task_tree._last_move_info = []
            self.task_tree._drag_in_progress = False
            return

        existing_tasks = set(self.config.get("tasks", {}).keys())
        normalized_order = [task_id for task_id in new_task_order if task_id in existing_tasks]
        normalized_order += [task_id for task_id in self._get_task_names() if task_id not in normalized_order]

        self.config["folders"] = sorted(list(new_folders_set), key=lambda p: (p.count("/"), p.lower()))
        if "layout" not in self.config:
            self.config["layout"] = {}
        self.config["layout"]["task_order"] = normalized_order

        if getattr(self, "current_task", "") in self.config.get("tasks", {}):
            self.config["layout"]["last_task"] = self.current_task
        elif normalized_order:
            self.config["layout"]["last_task"] = normalized_order[0]

        write_drag_debug(
            f"[AutoManager] _sync_tree_structure_to_config: normalized_order={normalized_order} folders={self.config['folders']} "
            f"last_task={self.config.get('layout', {}).get('last_task')}"
        )
        save_config(self.config)
        write_drag_debug("[AutoManager] _sync_tree_structure_to_config: save_config 完成")

        self._refresh_tree_item_roles()
        self._clear_tree_origin_roles()
        write_drag_debug(
            f"[AutoManager] _sync_tree_structure_to_config:end tree={format_tree_snapshot(self.task_tree, self._is_folder_role)}"
        )

        if moved_count:
            self._log(f"🔄 已同步 {moved_count} 个任务的位置变更", "blue")
            self._refresh_schedule_task_options()

        self.task_tree._last_move_info = []
        self.task_tree._drag_in_progress = False

    def _get_expanded_folder_paths(self):
        expanded = set()
        iterator = QTreeWidgetItemIterator(self.task_tree)
        while iterator.value():
            item = iterator.value()
            role = item.data(0, Qt.UserRole)
            if self._is_folder_role(role) and item.isExpanded():
                expanded.add(self._folder_path_from_role(role))
            iterator += 1
        return expanded

    def _restore_expanded_folder_paths(self, expanded_paths):
        if expanded_paths is None:
            return
        iterator = QTreeWidgetItemIterator(self.task_tree)
        while iterator.value():
            item = iterator.value()
            role = item.data(0, Qt.UserRole)
            if self._is_folder_role(role):
                item.setExpanded(self._folder_path_from_role(role) in expanded_paths)
            iterator += 1

    def _ensure_tree_structure_synced_if_needed(self):
        need_sync = False
        if hasattr(self, "_task_tree_sync_timer") and self._task_tree_sync_timer.isActive():
            need_sync = True
        if getattr(self.task_tree, "_drag_in_progress", False):
            need_sync = True
        if getattr(self.task_tree, "_last_move_info", []):
            need_sync = True
        if need_sync:
            write_drag_debug("[AutoManager] _ensure_tree_structure_synced_if_needed: execute sync")
            self._sync_tree_structure_to_config()
        else:
            write_drag_debug("[AutoManager] _ensure_tree_structure_synced_if_needed: skip sync")

    def _pick_preferred_task_after_delete(self, deleted_task_ids):
        deleted_set = set(deleted_task_ids or [])
        if self.current_task and self.current_task not in deleted_set and self.current_task in self.config.get("tasks", {}):
            return self.current_task
        selected_paths = [p for p in self._get_selected_task_paths() if p not in deleted_set and p in self.config.get("tasks", {})]
        if selected_paths:
            return selected_paths[0]
        for task_id in self._get_task_names():
            if task_id not in deleted_set and task_id in self.config.get("tasks", {}):
                return task_id
        return ""



    def _reload_task_tree(self):
        """[核心重构] 将任务列表渲染为树形结构。"""
        expanded_paths = self._get_expanded_folder_paths() if hasattr(self, "task_tree") else set()
        write_drag_debug(
            f"[AutoManager] _reload_task_tree:start drag_in_progress={getattr(self.task_tree, '_drag_in_progress', False)} "
            f"current_task={self.current_task!r} expanded={sorted(expanded_paths)}"
        )
        self.task_tree.setUpdatesEnabled(False)
        self.task_tree.blockSignals(True)
        self.task_tree.clear()
        
        all_tasks = self._get_task_names()
        label_map = self._build_duplicate_task_labels()
        folders = {} # { folder_path: QTreeWidgetItem }

        def ensure_folder(folder_path):
            folder_path = (folder_path or "").strip().strip("/")
            if not folder_path:
                return None
            if folder_path in folders:
                return folders[folder_path]
            parent_path, folder_name = self._split_tree_path(folder_path)
            parent_item = ensure_folder(parent_path)
            f_item = QTreeWidgetItem(parent_item or self.task_tree)
            f_item.setText(0, folder_name)
            f_item.setIcon(0, self.style().standardIcon(QStyle.SP_DirIcon))
            f_item.setData(0, Qt.UserRole, self._make_folder_role(folder_path))
            f_item.setExpanded(folder_path in expanded_paths)
            folders[folder_path] = f_item
            return f_item

        for folder_path in self._get_all_folder_paths():
            ensure_folder(folder_path)

        for task_id in all_tasks:
            parent_path = self._get_task_folder(task_id)
            parent_item = ensure_folder(parent_path)
            t_item = QTreeWidgetItem(parent_item or self.task_tree)
            t_item.setText(0, self._get_task_tree_label(task_id, label_map))
            t_item.setIcon(0, self.style().standardIcon(QStyle.SP_FileIcon))
            t_item.setData(0, Qt.UserRole, task_id)
            t_item.setData(0, Qt.UserRole + 2, self._get_task_name_only(task_id))
            t_item.setToolTip(0, f"任务ID: {task_id}\n位置: {self._get_task_path(task_id) or self._get_task_name_only(task_id)}")
        
        self.task_tree.blockSignals(False)
        self._restore_expanded_folder_paths(expanded_paths)
        self.task_tree.setUpdatesEnabled(True)
        write_drag_debug(
            f"[AutoManager] _reload_task_tree:end total_tasks={len(all_tasks)} tree={format_tree_snapshot(self.task_tree, self._is_folder_role)}"
        )

    def _on_tree_selection_changed(self):
        if getattr(self.task_tree, "_drag_in_progress", False):
            return
        item = self.task_tree.currentItem()
        if not item:
            return

    def _on_tree_item_double_clicked(self, item, column):
        if not item:
            return
        task_path = item.data(0, Qt.UserRole)
        if self._is_folder_role(task_path):
            item.setExpanded(not item.isExpanded())
            return
        self._on_task_changed(task_path)

    def _filter_task_tree(self, text):
        text = text.lower()
        def _filter_item(item):
            match = text in item.text(0).lower()
            child_match = False
            for i in range(item.childCount()):
                if _filter_item(item.child(i)):
                    child_match = True
            
            show = match or child_match
            item.setHidden(not show)
            if show and child_match:
                item.setExpanded(True)
            return show

        for i in range(self.task_tree.topLevelItemCount()):
            _filter_item(self.task_tree.topLevelItem(i))

    def _select_task_in_tree(self, task_path):
        """在树中寻找并选中指定路径的任务。"""
        iterator = QTreeWidgetItemIterator(self.task_tree)
        while iterator.value():
            item = iterator.value()
            if item.data(0, Qt.UserRole) == task_path:
                self.task_tree.setCurrentItem(item)
                # 确保父节点展开
                p = item.parent()
                while p:
                    p.setExpanded(True)
                    p = p.parent()
                break
            iterator += 1

    def _select_folder_in_tree(self, folder_path):
        iterator = QTreeWidgetItemIterator(self.task_tree)
        target_role = self._make_folder_role(folder_path)
        while iterator.value():
            item = iterator.value()
            if item.data(0, Qt.UserRole) == target_role:
                self.task_tree.setCurrentItem(item)
                p = item.parent()
                while p:
                    p.setExpanded(True)
                    p = p.parent()
                break
            iterator += 1

    def _select_first_task(self):
        iterator = QTreeWidgetItemIterator(self.task_tree)
        while iterator.value():
            item = iterator.value()
            if not self._is_folder_role(item.data(0, Qt.UserRole)):
                self.task_tree.setCurrentItem(item)
                break
            iterator += 1

    def _reload_task_combo_after_config_change(self, preferred_task=None):
        self._reload_task_tree()
        if preferred_task:
            self._select_task_in_tree(preferred_task)
            self._on_task_changed(preferred_task)
        else:
            self._select_first_task()
            first_task = self._get_primary_selected_task_path(fallback_current=False)
            if first_task:
                self._on_task_changed(first_task)

    def _get_primary_selected_task_path(self, fallback_current=True):
        paths = self._get_selected_task_paths()
        if paths:
            return paths[0]
        return self.current_task if fallback_current else ""

    def _get_selected_task_paths(self):
        paths = []
        for item in self.task_tree.selectedItems():
            task_path = item.data(0, Qt.UserRole)
            if task_path and not self._is_folder_role(task_path) and task_path not in paths:
                paths.append(task_path)
        if not paths and getattr(self, "current_task", ""):
            paths = [self.current_task]
        return paths

    def _make_unique_task_name(self, base_name, folder="", exclude_task_id=None):
        base_name = (base_name or "").strip() or "导入任务"
        folder = (folder or "").strip().replace("\\", "/").strip("/")
        existing_names = set()
        for task_id in self.config.get("tasks", {}).keys():
            if exclude_task_id and task_id == exclude_task_id:
                continue
            if self._get_task_folder(task_id) == folder:
                existing_names.add(self._get_task_name_only(task_id))
        if base_name not in existing_names:
            return base_name
        index = 2
        while True:
            candidate = f"{base_name}_{index}"
            if candidate not in existing_names:
                return candidate
            index += 1

    def _clear_exported_step_coordinates(self, step):
        return dict(step if isinstance(step, dict) else {})

    def _build_task_template_payload(self, task_ids):
        import copy
        exported_tasks = []
        for task_id in task_ids:
            if task_id not in self.config.get("tasks", {}):
                continue
            raw_actions = self.config.get("tasks", {}).get(task_id, [])
            actions = []
            coord_step_names = []
            for idx, step in enumerate(raw_actions):
                cleaned_step = self._clear_exported_step_coordinates(step)
                actions.append(cleaned_step)
                if isinstance(step, dict) and (step.get("x") is not None or step.get("y") is not None):
                    coord_step_names.append(step.get("name", f"步骤{idx + 1}"))
            meta = copy.deepcopy(self._get_task_meta(task_id, create_missing=True))
            meta.pop("created_at", None)
            exported_tasks.append({
                "name": self._get_task_name_only(task_id),
                "folder": self._get_task_folder(task_id),
                "meta": meta,
                "actions": actions,
                "task_data": copy.deepcopy(self.config.get("task_data", {}).get(task_id, [])),
                "tasks_layout": copy.deepcopy(self.config.get("tasks_layout", {}).get(task_id, [])),
                "coordinate_step_names": coord_step_names,
            })
        return {
            "format": "task_template_bundle",
            "version": 1,
            "exported_at": datetime.now().isoformat(timespec="seconds"),
            "coordinates_removed": False,
            "note": "导出的任务模板会完整保留坐标和界面守卫，导入后也会原样保留，后续可按需手动修改。",
            "tasks": exported_tasks,
        }

    def _normalize_imported_task_templates(self, payload):
        if isinstance(payload, dict) and payload.get("format") == "task_template_bundle":
            tasks = payload.get("tasks", [])
            return tasks if isinstance(tasks, list) else []
        if isinstance(payload, dict) and ("actions" in payload or "task_data" in payload):
            return [payload]
        if isinstance(payload, list):
            return [item for item in payload if isinstance(item, dict)]
        return []

    def _prompt_target_folder(self, initial_folder=""):
        folders = ["(根目录)"] + self._get_all_folder_paths()
        current_text = initial_folder if initial_folder else "(根目录)"
        current_index = folders.index(current_text) if current_text in folders else 0
        folder_text, ok = QInputDialog.getItem(
            self,
            "移动到文件夹",
            "选择目标文件夹，可直接输入新路径：",
            folders,
            current_index,
            True,
        )
        if not ok:
            return None
        folder_text = (folder_text or "").strip().replace("\\", "/").strip("/")
        if not folder_text or folder_text == "(根目录)":
            return ""
        if "//" in folder_text:
            QMessageBox.warning(self, "错误", "文件夹路径格式不正确。")
            return None
        return folder_text

    def _ensure_folder_path_exists(self, folder_path):
        folder_path = (folder_path or "").strip().strip("/")
        if not folder_path:
            return
        self.config.setdefault("folders", [])
        existing = set(self.config.get("folders", []))
        parts = folder_path.split("/")
        for i in range(1, len(parts) + 1):
            current = "/".join(parts[:i]).strip("/")
            if current and current not in existing:
                self.config["folders"].append(current)
                existing.add(current)

    def _move_selected_tasks_to_folder(self, task_paths=None):
        task_paths = task_paths or self._get_selected_task_paths()
        if not task_paths:
            QMessageBox.warning(self, "提示", "请选择要移动的任务。")
            return

        first_parent = self._get_task_folder(task_paths[0])
        target_folder = self._prompt_target_folder(first_parent)
        if target_folder is None:
            return

        changed_ids = []
        for task_id in task_paths:
            if self._get_task_folder(task_id) == target_folder:
                continue
            self._set_task_location(task_id, folder=target_folder)
            changed_ids.append(task_id)

        if not changed_ids:
            return

        self._ensure_folder_path_exists(target_folder)
        self.config["folders"] = sorted(set(self.config.get("folders", [])), key=lambda p: (p.count("/"), p.lower()))
        preferred_task = self.current_task if self.current_task in changed_ids else changed_ids[0]
        save_config(self.config)
        self._reload_task_combo_after_config_change(preferred_task if preferred_task in self.config.get("tasks", {}) else None)
        if target_folder:
            self._select_folder_in_tree(target_folder)
            self._select_task_in_tree(preferred_task)
        self._refresh_schedule_task_options()
        self._log(f"📂 已移动 {len(changed_ids)} 个任务到 '{target_folder or '根目录'}'", "blue")

    def _open_task_in_editor(self, task_path=None):
        task_path = task_path or self._get_primary_selected_task_path(fallback_current=False)
        if not task_path:
            return
        self._on_task_changed(task_path)

    def _rename_tasks_in_config(self, rename_map):
        if not rename_map: return
        if hasattr(self, 'data_table') and self.current_task:
            try: self._save_data_table()
            except: pass
        for old_key, new_key in rename_map.items():
            task_id = old_key if old_key in self.config.get("tasks", {}) else None
            if task_id is None:
                for candidate_id in self.config.get("tasks", {}).keys():
                    if self._get_task_path(candidate_id) == old_key:
                        task_id = candidate_id
                        break
            if not task_id or task_id not in self.config.get("tasks", {}):
                continue
            folder, name = self._split_tree_path(new_key)
            self._set_task_location(task_id, folder=folder, name=name)
        save_config(self.config)

    def _create_schedule_task_combo(self, current_text=""):
        cb = QComboBox()
        labels = self._build_duplicate_task_labels()
        task_ids = self._get_task_names()
        for task_id in task_ids:
            label = labels.get(task_id, self._get_task_name_only(task_id))
            folder = self._get_task_folder(task_id)
            if folder:
                label = f"{label}  ·  {folder}"
            cb.addItem(label, task_id)
        index = cb.findData(current_text)
        if index >= 0:
            cb.setCurrentIndex(index)
        elif task_ids:
            cb.setCurrentIndex(0)
        cb.currentIndexChanged.connect(lambda *_: self._save_current_schedule_bundle())
        return cb

    def _create_schedule_enabled_checkbox(self, checked=True):
        chk = QCheckBox()
        chk.setChecked(bool(checked))
        chk.stateChanged.connect(lambda *_: self._save_current_schedule_bundle())
        return chk

    def _ensure_schedule_bundles(self):
        bundles = self.config.setdefault("schedule_bundles", {})
        if not isinstance(bundles, dict):
            bundles = {}
            self.config["schedule_bundles"] = bundles
        if not bundles:
            bundles["默认排程包"] = {"items": [], "timer_config": None, "timer_enabled": False}
        self.config.setdefault("layout", {})
        current = self.config["layout"].get("current_schedule_bundle", "")
        if current not in bundles:
            self.config["layout"]["current_schedule_bundle"] = next(iter(bundles.keys()))
        return bundles

    def _get_current_schedule_bundle_name(self):
        self._ensure_schedule_bundles()
        current = self.config.get("layout", {}).get("current_schedule_bundle", "")
        if current in self.config["schedule_bundles"]:
            return current
        combo_text = self.sched_bundle_combo.currentText().strip() if hasattr(self, "sched_bundle_combo") else ""
        if combo_text in self.config["schedule_bundles"]:
            return combo_text
        return next(iter(self.config["schedule_bundles"].keys()))

    def _get_schedule_bundle(self, bundle_name=None):
        bundles = self._ensure_schedule_bundles()
        bundle_name = bundle_name or self._get_current_schedule_bundle_name()
        if bundle_name not in bundles:
            bundle_name = next(iter(bundles.keys()))
        return bundle_name, bundles[bundle_name]

    def _populate_schedule_bundle_combo(self):
        bundles = self._ensure_schedule_bundles()
        current = self.config.get("layout", {}).get("current_schedule_bundle", "")
        self.sched_bundle_combo.blockSignals(True)
        self.sched_bundle_combo.clear()
        for name in bundles.keys():
            self.sched_bundle_combo.addItem(name)
        if current in bundles:
            self.sched_bundle_combo.setCurrentText(current)
        self.sched_bundle_combo.blockSignals(False)

    def _insert_schedule_row(self, task_id="", enabled=True, status="等待中"):
        row = self.sched_table.rowCount()
        self.sched_table.insertRow(row)
        self.sched_table.setCellWidget(row, 0, self._create_schedule_enabled_checkbox(enabled))
        self.sched_table.setCellWidget(row, 1, self._create_schedule_task_combo(task_id))
        self._set_sched_status(row, status)
        return row

    def _load_schedule_bundle(self, bundle_name=None):
        bundle_name, bundle = self._get_schedule_bundle(bundle_name)
        self._loading_schedule_ui = True
        self.sched_table.setRowCount(0)
        for item in bundle.get("items", []):
            if isinstance(item, dict):
                task_id = item.get("task_id", "")
                enabled = item.get("enabled", True)
            else:
                task_id = item
                enabled = True
            self._insert_schedule_row(task_id, enabled=enabled, status="等待中")
        self._timer_config = bundle.get("timer_config") if isinstance(bundle.get("timer_config"), dict) else None
        self._timer_enabled = bool(bundle.get("timer_enabled", False))
        self._last_schedule_trigger_key = None
        self._loading_schedule_ui = False
        self._apply_loaded_schedule_timer()

    def _save_current_schedule_bundle(self):
        if getattr(self, "_loading_schedule_ui", False) or not hasattr(self, "sched_table"):
            return
        bundle_name, bundle = self._get_schedule_bundle()
        items = []
        for row in range(self.sched_table.rowCount()):
            task_id = self._get_sched_task_name(row)
            if not task_id:
                continue
            items.append({"task_id": task_id, "enabled": self._is_sched_enabled(row)})
        bundle["items"] = items
        bundle["timer_config"] = self._timer_config if isinstance(self._timer_config, dict) else None
        bundle["timer_enabled"] = bool(getattr(self, "_timer_enabled", False))
        self.config.setdefault("layout", {})
        self.config["layout"]["current_schedule_bundle"] = bundle_name
        self._schedule_config_flush()

    def _on_schedule_bundle_changed(self, bundle_name):
        if not bundle_name:
            return
        self._save_current_schedule_bundle()
        self.config.setdefault("layout", {})
        self.config["layout"]["current_schedule_bundle"] = bundle_name
        self._load_schedule_bundle(bundle_name)
        self._schedule_config_flush()

    def _add_schedule_bundle(self):
        self._save_current_schedule_bundle()
        name, ok = QInputDialog.getText(self, "新建排程包", "请输入排程包名称:")
        name = (name or "").strip()
        if not (ok and name):
            return
        bundles = self._ensure_schedule_bundles()
        if name in bundles:
            QMessageBox.warning(self, "提示", "排程包名称已存在。")
            return
        bundles[name] = {"items": [], "timer_config": None, "timer_enabled": False}
        self.config["layout"]["current_schedule_bundle"] = name
        self._populate_schedule_bundle_combo()
        self.sched_bundle_combo.setCurrentText(name)
        self._load_schedule_bundle(name)
        self._schedule_config_flush()

    def _rename_schedule_bundle(self):
        old_name = self._get_current_schedule_bundle_name()
        bundles = self._ensure_schedule_bundles()
        new_name, ok = QInputDialog.getText(self, "重命名排程包", "请输入新的排程包名称:", text=old_name)
        new_name = (new_name or "").strip()
        if not (ok and new_name) or new_name == old_name:
            return
        if new_name in bundles:
            QMessageBox.warning(self, "提示", "排程包名称已存在。")
            return
        self._save_current_schedule_bundle()
        bundles[new_name] = bundles.pop(old_name)
        self.config["layout"]["current_schedule_bundle"] = new_name
        self._populate_schedule_bundle_combo()
        self.sched_bundle_combo.setCurrentText(new_name)
        self._schedule_config_flush()

    def _delete_schedule_bundle(self):
        bundles = self._ensure_schedule_bundles()
        if len(bundles) <= 1:
            QMessageBox.warning(self, "提示", "至少保留一个排程包。")
            return
        cur = self._get_current_schedule_bundle_name()
        if QMessageBox.question(self, "删除排程包", f"确定删除排程包「{cur}」吗？") != QMessageBox.Yes:
            return
        bundles.pop(cur, None)
        new_cur = next(iter(bundles.keys()))
        self.config["layout"]["current_schedule_bundle"] = new_cur
        self._populate_schedule_bundle_combo()
        self.sched_bundle_combo.setCurrentText(new_cur)
        self._load_schedule_bundle(new_cur)
        self._schedule_config_flush()

    def _get_sched_task_name(self, row):
        widget = self.sched_table.cellWidget(row, 1)
        if isinstance(widget, QComboBox):
            return widget.currentData() or ""
        item = self.sched_table.item(row, 1)
        return item.data(Qt.UserRole) if item else ""

    def _is_sched_enabled(self, row):
        widget = self.sched_table.cellWidget(row, 0)
        if isinstance(widget, QCheckBox):
            return widget.isChecked()
        item = self.sched_table.item(row, 0)
        return item.checkState() == Qt.Checked if item else True

    def _set_sched_status(self, row, status):
        item = self.sched_table.item(row, 2)
        if item is None:
            self.sched_table.setItem(row, 2, QTableWidgetItem(status))
        else:
            item.setText(status)

    def _refresh_schedule_task_options(self, rename_map=None, deleted_name=None):
        rename_map = rename_map or {}
        task_names = self._get_task_names()
        for row in range(self.sched_table.rowCount()):
            current = rename_map.get(self._get_sched_task_name(row), self._get_sched_task_name(row))
            if deleted_name and current == deleted_name:
                current = task_names[0] if task_names else ""
            cb = self.sched_table.cellWidget(row, 1)
            if not isinstance(cb, QComboBox):
                cb = self._create_schedule_task_combo(current)
                self.sched_table.setCellWidget(row, 1, cb)
            else:
                cb.blockSignals(True)
                cb.clear()
                labels = self._build_duplicate_task_labels()
                for task_id in task_names:
                    label = labels.get(task_id, self._get_task_name_only(task_id))
                    folder = self._get_task_folder(task_id)
                    if folder:
                        label = f"{label}  ·  {folder}"
                    cb.addItem(label, task_id)
                index = cb.findData(current)
                if index >= 0:
                    cb.setCurrentIndex(index)
                elif task_names:
                    cb.setCurrentIndex(0)
                cb.blockSignals(False)
        self._save_current_schedule_bundle()

    def _move_selected_sched_up(self):
        row = self.sched_table.currentRow()
        if row > 0: self._move_sched_row(row, row - 1)

    def _move_selected_sched_down(self):
        row = self.sched_table.currentRow()
        if 0 <= row < self.sched_table.rowCount() - 1: self._move_sched_row(row, row + 1)

    def _add_task(self, folder=None):
        self._sync_tree_structure_to_config()
        prompt = f"在 [{folder}] 中新建任务名称:" if folder else "请输入任务名称:"
        name, ok = QInputDialog.getText(self, "新建任务", prompt)
        if not ok:
            return
        name = (name or "").strip()
        if not name:
            return
        if "/" in name:
            QMessageBox.warning(self, "错误", "任务名称不能包含 '/'")
            return
        task_id = _new_task_id(set(self.config.get("tasks", {}).keys()))
        self.config['tasks'][task_id] = []
        self.config['task_data'][task_id] = []
        self._set_task_location(task_id, folder=folder or "", name=name)
        save_config(self.config)
        self._reload_task_combo_after_config_change(task_id)
        self._refresh_schedule_task_options()

    def _add_folder(self, parent_folder=None):
        self._sync_tree_structure_to_config()
        prompt = f"在 [{parent_folder}] 中新建文件夹名称:" if parent_folder else "请输入文件夹名称:"
        name, ok = QInputDialog.getText(self, "新建文件夹", prompt)
        if not ok:
            return
        name = (name or "").strip()
        if not name:
            return
        if "/" in name:
            QMessageBox.warning(self, "错误", "文件夹名称不能包含 '/'")
            return
        folder_path = self._join_tree_path(parent_folder, name)
        existing_folders = set(self._get_all_folder_paths())
        if folder_path in existing_folders:
            QMessageBox.warning(self, "提示", f"文件夹 '{folder_path}' 已存在。")
            return
        self.config.setdefault('folders', [])
        if folder_path not in self.config['folders']:
            self.config['folders'].append(folder_path)
        save_config(self.config)
        self._reload_task_tree()
        self._select_folder_in_tree(folder_path)

    def _rename_folder(self, item):
        self._sync_tree_structure_to_config()
        old_folder = self._tree_item_path(item)
        if not old_folder:
            return
        parent_path, old_name = self._split_tree_path(old_folder)
        new_name, ok = QInputDialog.getText(self, "重命名文件夹", "请输入新名称:", text=old_name)
        if not ok:
            return
        new_name = (new_name or "").strip()
        if not new_name or "/" in new_name:
            QMessageBox.warning(self, "错误", "文件夹名称不能为空或包含 '/'")
            return
        new_folder = self._join_tree_path(parent_path, new_name)
        if new_folder == old_folder:
            return
        existing_folders = set(self._get_all_folder_paths())
        if new_folder in existing_folders:
            QMessageBox.warning(self, "错误", f"目标文件夹 '{new_folder}' 已存在。")
            return

        affected_task_ids = []
        for task_id in list(self.config['tasks'].keys()):
            folder_path = self._get_task_folder(task_id)
            if folder_path == old_folder or folder_path.startswith(f"{old_folder}/"):
                new_task_folder = f"{new_folder}{folder_path[len(old_folder):]}"
                self._set_task_location(task_id, folder=new_task_folder)
                affected_task_ids.append(task_id)

        updated_folders = []
        for folder_path in self._get_all_folder_paths():
            if folder_path == old_folder or folder_path.startswith(f"{old_folder}/"):
                updated_folders.append(f"{new_folder}{folder_path[len(old_folder):]}")
            else:
                updated_folders.append(folder_path)
        self.config['folders'] = sorted(set(updated_folders), key=lambda p: (p.count('/'), p.lower()))

        preferred_task = self.current_task if self.current_task in self.config.get('tasks', {}) else ""
        save_config(self.config)
        self._reload_task_combo_after_config_change(preferred_task if preferred_task in self.config.get('tasks', {}) else None)
        self._select_folder_in_tree(new_folder)
        self._refresh_schedule_task_options()

    def _delete_folder(self, item):
        self._ensure_tree_structure_synced_if_needed()
        folder = self._tree_item_path(item)
        if not folder:
            return

        affected_tasks = [
            task_id for task_id in self.config['tasks'].keys()
            if self._get_task_folder(task_id) == folder or self._get_task_folder(task_id).startswith(f"{folder}/")
        ]
        affected_folders = [f for f in self._get_all_folder_paths() if f == folder or f.startswith(f"{folder}/")]

        if not affected_tasks:
            self.config['folders'] = [f for f in self.config.get('folders', []) if f not in affected_folders]
            save_config(self.config)
            self._reload_task_tree()
            if self.current_task and self.current_task in self.config.get("tasks", {}):
                self._select_task_in_tree(self.current_task)
            return

        parent_path, _ = self._split_tree_path(folder)
        target_label = parent_path if parent_path else "根目录"

        msg = QMessageBox(self)
        msg.setWindowTitle("删除文件夹")
        msg.setText(f"文件夹 '{folder}' 内含有 {len(affected_tasks)} 个任务。\n\n您想如何处理这些任务？")
        btn_move = msg.addButton(f"📁 移至 {target_label}", QMessageBox.ActionRole)
        btn_del_all = msg.addButton("🗑️ 全部彻底删除", QMessageBox.DestructiveRole)
        btn_cancel = msg.addButton("❌ 取消", QMessageBox.RejectRole)
        msg.setDefaultButton(btn_move)
        msg.exec_()

        if msg.clickedButton() == btn_cancel:
            return

        if msg.clickedButton() == btn_move:
            for task_id in affected_tasks:
                old_task_folder = self._get_task_folder(task_id)
                relative_path = old_task_folder[len(folder):].lstrip("/")
                new_task_folder = self._join_tree_path(parent_path, relative_path)
                self._set_task_location(task_id, folder=new_task_folder)

            moved_folders = []
            for folder_path in affected_folders:
                if folder_path == folder:
                    continue
                suffix = folder_path[len(folder) + 1:]
                moved_folders.append(self._join_tree_path(parent_path, suffix))
            self.config['folders'] = sorted(set(
                [f for f in self.config.get('folders', []) if f not in affected_folders] + moved_folders
            ), key=lambda p: (p.count('/'), p.lower()))
            preferred_task = self._pick_preferred_task_after_delete([])
            save_config(self.config)
            self._reload_task_combo_after_config_change(preferred_task if preferred_task in self.config.get('tasks', {}) else None)
            self._refresh_schedule_task_options()
            self._log(f"📂 文件夹 '{folder}' 已移除，内容已迁移到 '{target_label}'", "blue")
        else:
            for task_id in affected_tasks:
                self.config['tasks'].pop(task_id, None)
                self.config['task_data'].pop(task_id, None)
                self.config.get('task_meta', {}).pop(task_id, None)
                if 'tasks_layout' in self.config:
                    self.config['tasks_layout'].pop(task_id, None)
                if hasattr(self, '_row_statuses'):
                    self._row_statuses.pop(task_id, None)
            self.config['folders'] = [f for f in self.config.get('folders', []) if f not in affected_folders]
            preferred_task = self._pick_preferred_task_after_delete(affected_tasks)
            save_config(self.config)
            self._reload_task_combo_after_config_change(preferred_task if preferred_task in self.config.get('tasks', {}) else None)
            self._refresh_schedule_task_options(deleted_name=self.current_task if self.current_task in affected_tasks else None)
            self._log(f"🗑️ 已删除文件夹 '{folder}' 及其内部所有任务", "orange")
    def _rename_task(self):
        self._sync_tree_structure_to_config()
        old_name = self._get_primary_selected_task_path()
        if not old_name: return
        pure_name = self._get_task_name_only(old_name)
        name, ok = QInputDialog.getText(self, "重命名任务", "请输入新名称:", text=pure_name)
        if not ok:
            return
        name = (name or "").strip()
        if not name:
            return
        if "/" in name:
            QMessageBox.warning(self, "错误", "任务名称不能包含 '/'")
            return
        if name == pure_name:
            return
        old_display = self._get_task_path(old_name)
        self._set_task_location(old_name, name=name)
        save_config(self.config)
        self._reload_task_combo_after_config_change(old_name)
        self._refresh_schedule_task_options()
        self._log(f"✅ 任务已从 '{old_display}' 重命名为 '{self._get_task_path(old_name)}'", "blue")

    def _delete_task(self):
        self._ensure_tree_structure_synced_if_needed()
        selected_items = self.task_tree.selectedItems()
        tasks_to_del = []
        for it in selected_items:
            path = it.data(0, Qt.UserRole)
            if path and not self._is_folder_role(path):
                tasks_to_del.append(path)
        
        if not tasks_to_del:
            # 如果没选中，默认删除当前正在编辑的任务
            if self.current_task:
                tasks_to_del = [self.current_task]
            else:
                QMessageBox.warning(self, "提示", "请选择要删除的任务。")
                return

        if len(self.config['tasks']) <= len(tasks_to_del):
            QMessageBox.warning(self, "错误", "至少保留一个任务！")
            return

        msg = f"确定要删除选中的 {len(tasks_to_del)} 个任务吗？此操作不可撤销！" if len(tasks_to_del) > 1 else f"确定要删除任务 '{self._get_task_path(tasks_to_del[0])}' 吗？"
        reply = QMessageBox.question(self, '确认删除', msg, QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            preferred_task = self._pick_preferred_task_after_delete(tasks_to_del)
            for name in tasks_to_del:
                self.config['tasks'].pop(name, None)
                self.config['task_data'].pop(name, None)
                self.config.get('task_meta', {}).pop(name, None)
                if 'tasks_layout' in self.config: self.config['tasks_layout'].pop(name, None)
                self._row_statuses.pop(name, None)
            
            save_config(self.config)
            self._reload_task_combo_after_config_change(preferred_task if preferred_task in self.config.get('tasks', {}) else None)
            self._refresh_schedule_task_options()
            self._log(f"🗑️ 已批量删除 {len(tasks_to_del)} 个任务", "orange")

    def _clone_task(self):
        old_name = self._get_primary_selected_task_path()
        if not old_name: return
        pure_name = self._get_task_name_only(old_name)
        cat_prefix = self._get_task_folder(old_name)
        
        name, ok = QInputDialog.getText(self, "克隆任务", f"请输入新任务名称 (基于 '{pure_name}' 克隆):", text=f"{pure_name}_副本")
        if ok and name:
            new_task_id = _new_task_id(set(self.config.get("tasks", {}).keys()))
            import copy
            self.config['tasks'][new_task_id] = copy.deepcopy(self.config['tasks'][old_name])
            self.config['task_data'][new_task_id] = copy.deepcopy(self.config['task_data'].get(old_name, []))
            self._set_task_location(new_task_id, folder=cat_prefix, name=name.strip())
            
            save_config(self.config)
            self._reload_task_combo_after_config_change(new_task_id)
            self._refresh_schedule_task_options()
            self._log(f"📋 已克隆任务 '{self._get_task_path(old_name)}' 为 '{self._get_task_path(new_task_id)}'", "blue")

    def _show_task_list_menu(self, pos):
        item = self.task_tree.itemAt(pos)
        if not item:
            menu = QMenu(self)
            menu.addAction("➕ 新建任务").triggered.connect(self._add_task)
            menu.addAction("📁 新建文件夹").triggered.connect(self._add_folder)
            menu.addSeparator()
            menu.addAction("📥 导入任务模板").triggered.connect(self._import_task_templates)
            menu.exec_(self.task_tree.viewport().mapToGlobal(pos))
            return

        if item not in self.task_tree.selectedItems():
            self.task_tree.blockSignals(True)
            self.task_tree.clearSelection()
            item.setSelected(True)
            self.task_tree.setCurrentItem(item)
            self.task_tree.blockSignals(False)
        task_path = item.data(0, Qt.UserRole)
        menu = QMenu(self)
        
        if self._is_folder_role(task_path):
            folder_name = self._tree_item_path(item)
            menu.addAction("➕ 在此文件夹新建任务").triggered.connect(lambda: self._add_task(folder=folder_name))
            menu.addAction("📁 在此文件夹新建子文件夹").triggered.connect(lambda: self._add_folder(parent_folder=folder_name))
            menu.addAction("📥 导入任务模板").triggered.connect(self._import_task_templates)
            menu.addSeparator()
            menu.addAction("📂 展开/收起").triggered.connect(lambda: item.setExpanded(not item.isExpanded()))
            menu.addAction("✏️ 重命名文件夹").triggered.connect(lambda: self._rename_folder(item))
            menu.addAction("🗑️ 删除文件夹(及任务)").triggered.connect(lambda: self._delete_folder(item))
        else:
            menu.addAction("📖 打开任务").triggered.connect(lambda checked=False, p=task_path: self._open_task_in_editor(p))
            menu.addAction("📂 移动到文件夹...").triggered.connect(self._move_selected_tasks_to_folder)
            menu.addAction("📋 克隆此任务").triggered.connect(self._clone_task)
            menu.addAction("📤 导出任务模板").triggered.connect(self._export_task_templates)
            menu.addAction("✏️ 重命名任务").triggered.connect(self._rename_task)
            menu.addAction("🗑️ 删除任务").triggered.connect(self._delete_task)
            
        menu.exec_(self.task_tree.viewport().mapToGlobal(pos))

    def _search_task(self):
        # 侧边栏已经有搜索框，此功能可作为快速跳转
        display_map = {self._get_task_display_text(task_id, with_folder=True): task_id for task_id in self._get_task_names()}
        sel = ModernTaskManager(list(display_map.keys()), self)
        name = sel.get_selection()
        if name:
            self._activate_task_by_id(display_map.get(name, ""))

    def _add_to_schedule(self):
        task_name = self.current_task
        if not task_name: return
        row = self._insert_schedule_row(task_name, enabled=True, status="等待中")
        self.sched_table.setCurrentCell(row, 1)
        self._save_current_schedule_bundle()

    def _del_from_schedule(self):
        row = self.sched_table.currentRow()
        if row >= 0:
            self.sched_table.removeRow(row)
            self._save_current_schedule_bundle()

    def _show_sched_header_menu(self, pos):
        row = self.sched_table.verticalHeader().visualIndexAt(pos.y())
        if row < 0: return
        menu = QMenu()
        up = menu.addAction("⬆️ 向上移动")
        down = menu.addAction("⬇️ 向下移动")
        act = menu.exec_(self.sched_table.verticalHeader().mapToGlobal(pos))
        if act == up and row > 0: self._move_sched_row(row, row - 1)
        elif act == down and row < self.sched_table.rowCount() - 1: self._move_sched_row(row, row + 1)

    def _move_sched_row(self, old, new):
        old_task = self._get_sched_task_name(old)
        old_status = self.sched_table.item(old, 2).text() if self.sched_table.item(old, 2) else ""
        old_enabled = self._is_sched_enabled(old)
        new_task = self._get_sched_task_name(new)
        new_status = self.sched_table.item(new, 2).text() if self.sched_table.item(new, 2) else ""
        new_enabled = self._is_sched_enabled(new)
        self.sched_table.setCellWidget(old, 0, self._create_schedule_enabled_checkbox(new_enabled))
        self.sched_table.setCellWidget(new, 0, self._create_schedule_enabled_checkbox(old_enabled))
        self.sched_table.setCellWidget(old, 1, self._create_schedule_task_combo(new_task))
        self.sched_table.setCellWidget(new, 1, self._create_schedule_task_combo(old_task))
        self._set_sched_status(old, new_status)
        self._set_sched_status(new, old_status)
        self.sched_table.setCurrentCell(new, 1)
        self._save_current_schedule_bundle()

    def _run_schedule(self):
        if hasattr(self, '_engine') and self._engine.isRunning(): return
        tasks_to_run = []
        for i in range(self.sched_table.rowCount()):
            if not self._is_sched_enabled(i):
                self._set_sched_status(i, "已禁用")
                continue
            task_name = self._get_sched_task_name(i)
            if not task_name: continue
            tasks_to_run.append({'row': i, 'task': task_name})
            self._set_sched_status(i, "等待中")

        if not tasks_to_run: return
        self._log("🚀 启动序列化排程运行...", "purple")
        self._run_next_scheduled_task(tasks_to_run, 0)

    def _run_next_scheduled_task(self, task_list, index):
        if index > 0:
            prev_row = task_list[index - 1]['row']
            self._set_sched_status(prev_row, "已完成")

        if index >= len(task_list):
            self._log("✅ 所有排程任务已完成", "green")
            if self.chk_sched_shutdown.isChecked():
                self._log("🏁 所有排程任务已完成，即将按设置进入自动关机倒计时", "green")
                ShutdownDialog(self).exec_()
            return

        task_name = task_list[index]['task']
        row = task_list[index]['row']
        # 切换侧边栏选中项
        self._activate_task_by_id(task_name)
        self._set_sched_status(row, "正在执行...")
        self.sched_table.setCurrentCell(row, 1)

        # 启动执行，执行完后回调执行下一个
        self._run_all(on_finished=lambda: self._run_next_scheduled_task(task_list, index + 1))

    def _move_task_up(self):
        """将当前任务在列表中上移一位。"""
        if not self.current_task: return
        keys = self._get_task_names()
        idx = keys.index(self.current_task)
        if idx <= 0: return
        keys[idx], keys[idx - 1] = keys[idx - 1], keys[idx]
        self.config.setdefault("layout", {})
        self.config['layout']['task_order'] = keys
        save_config(self.config)
        self._reload_task_combo_after_config_change(self.current_task)
        self._log(f"⬆️ 任务 [{self._get_task_display_text(self.current_task, with_folder=True)}] 已上移", "blue")

    def _move_task_down(self):
        """将当前任务在列表中下移一位。"""
        if not self.current_task: return
        keys = self._get_task_names()
        idx = keys.index(self.current_task)
        if idx >= len(keys) - 1: return
        keys[idx], keys[idx + 1] = keys[idx + 1], keys[idx]
        self.config.setdefault("layout", {})
        self.config['layout']['task_order'] = keys
        save_config(self.config)
        self._reload_task_combo_after_config_change(self.current_task)
        self._log(f"⬇️ 任务 [{self._get_task_display_text(self.current_task, with_folder=True)}] 已下移", "blue")

    def _import_task_templates(self):
        p, _ = QFileDialog.getOpenFileName(self, "导入任务模板", "", "JSON Files (*.json)")
        if not p:
            return
        try:
            import copy
            with open(p, 'r', encoding='utf-8') as f:
                payload = json.load(f)
            items = self._normalize_imported_task_templates(payload)
            if not items:
                QMessageBox.warning(self, "导入失败", "没有解析到可导入的任务模板。")
                return

            added_task_ids = []
            total_coord_steps = 0
            for item in items:
                if not isinstance(item, dict):
                    continue
                name = str(item.get("name", "")).strip() or "导入任务"
                folder = str(item.get("folder", "")).strip().replace("\\", "/").strip("/")
                actions = item.get("actions", [])
                actions = actions if isinstance(actions, list) else []
                task_data = item.get("task_data", [])
                task_data = copy.deepcopy(task_data if isinstance(task_data, list) else [])
                task_layout = item.get("tasks_layout", [])
                task_layout = copy.deepcopy(task_layout if isinstance(task_layout, list) else [])
                coord_step_names = item.get("coordinate_step_names", [])
                if not isinstance(coord_step_names, list):
                    coord_step_names = []

                cleaned_actions = []
                for idx, step in enumerate(actions):
                    cleaned_actions.append(self._clear_exported_step_coordinates(step))
                    if idx < len(coord_step_names):
                        total_coord_steps += 1
                    elif isinstance(step, dict) and (step.get("x") is not None or step.get("y") is not None):
                        total_coord_steps += 1

                unique_name = self._make_unique_task_name(name, folder=folder)
                self._ensure_folder_path_exists(folder)
                task_id = _new_task_id(set(self.config.get("tasks", {}).keys()))
                self.config['tasks'][task_id] = cleaned_actions
                self.config['task_data'][task_id] = task_data
                self.config.setdefault('tasks_layout', {})[task_id] = task_layout
                self._set_task_location(task_id, folder=folder, name=unique_name)

                raw_meta = item.get("meta", {})
                if isinstance(raw_meta, dict):
                    meta = self._get_task_meta(task_id, create_missing=True)
                    for key, value in raw_meta.items():
                        if key in ("name", "folder", "created_at"):
                            continue
                        meta[key] = copy.deepcopy(value)
                    self.config["task_meta"][task_id] = _normalize_task_meta(task_id, meta)

                added_task_ids.append(task_id)

            if not added_task_ids:
                QMessageBox.warning(self, "导入失败", "文件里没有有效任务。")
                return

            self.config['folders'] = sorted(set(self.config.get('folders', [])), key=lambda p: (p.count('/'), p.lower()))
            save_config(self.config)
            preferred_task = added_task_ids[0]
            self._reload_task_combo_after_config_change(preferred_task)
            self._refresh_schedule_task_options()
            self._log(f"📥 已导入 {len(added_task_ids)} 个任务模板", "green")
            import_hint = "\n文件中的步骤坐标和界面守卫已一并保留，可按需手动修改。" if total_coord_steps else ""
            QMessageBox.information(
                self,
                "导入完成",
                f"已导入 {len(added_task_ids)} 个任务。\n\n"
                f"这些模板会保留步骤、批量数据、坐标和界面守卫。"
                f"{import_hint}"
            )
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导入任务模板失败: {str(e)}")

    def _export_task_templates(self):
        task_ids = self._get_selected_task_paths()
        if not task_ids:
            QMessageBox.warning(self, "提示", "请先选择要导出的任务。")
            return

        try:
            if self.current_task and self.current_task in task_ids:
                self._force_sync_action_widgets()
                self._save_data_table()

            payload = self._build_task_template_payload(task_ids)
            if not payload.get("tasks"):
                QMessageBox.warning(self, "提示", "没有可导出的任务内容。")
                return

            if len(task_ids) == 1:
                default_name = f"{self._get_task_name_only(task_ids[0])}_模板.json"
            else:
                default_name = f"任务模板包_{datetime.now().strftime('%Y%m%d')}.json"
            p, _ = QFileDialog.getSaveFileName(self, "导出任务模板", default_name, "JSON Files (*.json)")
            if not p:
                return

            with open(p, 'w', encoding='utf-8') as f:
                json.dump(payload, f, indent=4, ensure_ascii=False)

            coord_count = sum(len(task.get("coordinate_step_names", [])) for task in payload.get("tasks", []))
            self._log(f"📤 已导出 {len(payload.get('tasks', []))} 个任务模板", "blue")
            export_hint = "\n导出文件已完整保留相关步骤坐标和界面守卫，后续可按需手动修改。" if coord_count else ""
            QMessageBox.information(
                self,
                "导出完成",
                f"已导出 {len(payload.get('tasks', []))} 个任务模板。\n\n"
                "导出文件会完整保留坐标和界面守卫，适合备份，也可发给别人后再手动调整。"
                f"{export_hint}"
            )
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出任务模板失败: {str(e)}")

    def _import_full_config(self):
        p, _ = QFileDialog.getOpenFileName(self, "导入完整配置", "", "JSON Files (*.json)")
        if p:
            try:
                with open(p, 'r', encoding='utf-8') as f:
                    new_cfg = json.load(f)
                    if 'tasks' in new_cfg:
                        self.config = _migrate_config_schema(new_cfg)
                        save_config(self.config)
                        self._reload_task_tree()
                        last_task = self.config.get("layout", {}).get("last_task", "")
                        if last_task:
                            self._activate_task_by_id(last_task)
                        else:
                            self._select_first_task()
                            self._on_task_changed(self._get_primary_selected_task_path(fallback_current=False))
                        self._refresh_schedule_task_options()
                        QMessageBox.information(self, "成功", "配置已成功导入！")
            except Exception as e:
                QMessageBox.critical(self, "错误", f"导入失败: {str(e)}")

    def _export_full_config(self):
        p, _ = QFileDialog.getSaveFileName(self, "导出完整配置", f"config_backup_{datetime.now().strftime('%Y%m%d')}.json", "JSON Files (*.json)")
        if p:
            try:
                with open(p, 'w', encoding='utf-8') as f:
                    json.dump(self.config, f, indent=4, ensure_ascii=False)
                QMessageBox.information(self, "成功", "配置已成功导出！")
            except Exception as e:
                QMessageBox.critical(self, "错误", f"导出失败: {str(e)}")

    def _capture_image(self, idx):
        self.setWindowOpacity(0.2)
        QTimer.singleShot(500, lambda: self._do_capture(idx))

    def _capture_region_for_step(self, idx):
        """Launch region capture overlay; on completion fill the step's value."""
        self.showMinimized()
        QTimer.singleShot(300, lambda: self._start_region_capture(target_step_idx=idx))

    def _do_capture(self, idx):
        f, _ = QFileDialog.getOpenFileName(self, "选择目标图片", "", "Image Files (*.png *.jpg)")
        self.setWindowOpacity(1.0)
        if f:
            self.config['tasks'][self.current_task][idx]['value'] = f
            save_config(self.config); self._refresh_actions()

    def _refresh_actions(self):
        if not self.current_task: return
        
        # 定义通用的变量提示
        var_tooltip = (
            "💡 变量引用技巧：\n"
            "1. 系统变量: {{日期}}, {{时间}}, {{行号}}\n"
            "2. 步骤编号: {{step1}}, {{step2}} (获取对应步骤的内容)\n"
            "3. 步骤名称: {{步骤标题}} (推荐！如 {{搜索}}, 获取该标题步骤的内容)\n"
            "4. 数据表变量: {{列名}} (从 CSV/Excel 中读取)"
        )

        acts = self.config['tasks'].get(self.current_task, []); self.action_table.blockSignals(True); self.action_table.setRowCount(len(acts))
        for i, a in enumerate(acts):
            item_name = QTableWidgetItem(a.get('name', f'步骤{i+1}'))
            item_name.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
            self.action_table.setItem(i, 0, item_name)
            cb = QComboBox(); cb.addItems(list(CMD_MAP.keys())); cb.blockSignals(True); cb.setCurrentText(a.get('action', '左键点击')); cb.blockSignals(False); cb.currentTextChanged.connect(lambda v, idx=i: self._update_config(idx, 'action', v)); self.action_table.setCellWidget(i, 1, cb)
            ds = QSpinBox(); ds.setRange(0, 99999); ds.blockSignals(True); ds.setValue(int(a.get('delay', 1))); ds.blockSignals(False); ds.setSuffix(" 秒")
            ds.setToolTip(self._format_seconds(ds.value()))
            def _apply_delay_style(s, v, default=1):
                if v != default:
                    s.setStyleSheet("QSpinBox { background-color: #fff7ed; color: #c2410c; font-weight: 600; border: 1px solid #fdba74; border-radius: 7px; padding: 4px 8px; }")
                    s.setToolTip(f"⚠️ 已自定义: {self._format_seconds(v)}（默认: {self._format_seconds(default)}）")
                else:
                    s.setStyleSheet("QSpinBox { background-color: #ffffff; color: #25364a; font-weight: 500; border: 1px solid #d7dee7; border-radius: 7px; padding: 4px 8px; }")
                    s.setToolTip(self._format_seconds(v))
            _apply_delay_style(ds, ds.value())
            ds.valueChanged.connect(lambda v, idx=i, s=ds: [self._update_config(idx, 'delay', v), _apply_delay_style(s, v)])
            self.action_table.setCellWidget(i, 4, ds)

            act_name = a.get('action')
            if act_name == "激活窗口":
                # [修复] 按鈕只显示纯标题，去掉 ::hwnd=... 后缀
                _raw_val = a.get('value', '') or ''
                _btn_label = _raw_val.split('::hwnd=')[0] if '::hwnd=' in _raw_val else _raw_val
                btn = QPushButton(_btn_label if _btn_label else "选择窗口")
                btn.setToolTip(_raw_val)  # 悬浮显示完整标识信息
                btn.clicked.connect(lambda chk, idx=i: self._select_window(idx))
                self.action_table.setCellWidget(i, 2, btn)
            elif "图像识别点击" in act_name:
                w_cap = QWidget(); l_cap = QHBoxLayout(w_cap); l_cap.setContentsMargins(0,0,0,0)
                btn_region = QPushButton("✂️ 框选"); btn_region.setFixedWidth(52)
                btn_region.setToolTip("框选屏幕区域截图并作为识别目标")
                btn_region.clicked.connect(lambda chk, idx=i: self._capture_region_for_step(idx))
                btn_file   = QPushButton("📁 选图"); btn_file.setFixedWidth(52)
                btn_file.clicked.connect(lambda chk, idx=i: self._capture_image(idx))
                l_cap.addWidget(btn_region); l_cap.addWidget(btn_file)
                self.action_table.setCellWidget(i, 2, w_cap)
            elif any(x in act_name for x in ["运行程序", "屏幕截图", "等待", "延后执行", "单按键", "组合键", "如果找图成功", "如果窗口存在", "打开网址"]):
                btn = QPushButton("- 无需坐标 -"); btn.setEnabled(False)
                self.action_table.setCellWidget(i, 2, btn)
            else:
                w_pos = QWidget()
                l_pos = QHBoxLayout(w_pos)
                l_pos.setContentsMargins(0, 0, 0, 0)
                l_pos.setSpacing(6)
                chk_guard = QCheckBox("匹配")
                chk_guard.setChecked(bool(a.get('guard_enabled', True)))
                chk_guard.setToolTip("开启时，执行前会校验当前界面是否与录制时一致；关闭后将直接按坐标执行，不做界面卫士匹配。")
                chk_guard.stateChanged.connect(lambda state, idx=i: self._update_config(idx, 'guard_enabled', state == Qt.Checked))
                btn = QPushButton(f"{a.get('x',0)}, {a.get('y',0)}")
                btn.setToolTip("点击重新录制坐标")
                btn.clicked.connect(lambda chk, idx=i: self._start_record(idx))
                l_pos.addWidget(chk_guard)
                l_pos.addWidget(btn, 1)
                self.action_table.setCellWidget(i, 2, w_pos)
            if a.get('action') == "上传文件":
                w = QWidget(); l = QHBoxLayout(w); l.setContentsMargins(0,0,0,0)
                le = QLineEdit(str(a.get('value', ''))); le.editingFinished.connect(lambda idx=i, target=None: self._update_config(idx, 'value', self.action_table.cellWidget(idx, 3).findChild(QLineEdit).text()))
                btn = QPushButton("📁"); btn.setFixedWidth(30); btn.clicked.connect(lambda chk, idx=i, target=le: self._pick_default_file(idx, target))
                l.addWidget(le); l.addWidget(btn); self.action_table.setCellWidget(i, 3, w)
            elif a.get('action') == "🖱️ 拖拽文件":
                w = QWidget(); l = QHBoxLayout(w); l.setContentsMargins(0,0,0,0); l.setSpacing(6)
                le = QLineEdit(str(a.get('value', ''))); le.setPlaceholderText("选择要拖拽的文件...")
                le.editingFinished.connect(lambda idx=i: self._update_drag_file_config(idx))
                btn = QPushButton("📁"); btn.setFixedWidth(30)
                btn.clicked.connect(lambda chk, idx=i, target=le: self._pick_default_file(idx, target, self._update_drag_file_config))
                chk = QCheckBox("先点目标坐标")
                chk.setChecked(bool(a.get('activate_target_before_drag', False)))
                chk.setToolTip("开启后，拖拽开始前会先点击一次目标坐标，便于你自己提前处理窗口激活。")
                chk.stateChanged.connect(lambda _state, idx=i: self._update_drag_file_config(idx))
                l.addWidget(le, 1); l.addWidget(btn); l.addWidget(chk)
                self.action_table.setCellWidget(i, 3, w)
            elif a.get('action') in ["单按键", "组合键"]:
                le = KeyRecorder(); le.setText(str(a.get('value', ''))); le.key_recorded.connect(lambda v, idx=i: self._update_config(idx, 'value', v)); self.action_table.setCellWidget(i, 3, le)
            elif a.get('action') == "🌐 打开网址":
                w = QWidget(); l = QHBoxLayout(w); l.setContentsMargins(0,0,0,0)
                val_parts = str(a.get('value', '')).split('|')
                url_val = val_parts[0] if val_parts else ""; prof_val = val_parts[1] if len(val_parts) > 1 else ""
                le = QLineEdit(url_val); le.setPlaceholderText("输入网址..."); le.editingFinished.connect(lambda idx=i: self._update_url_config(idx))
                btn_prof = QPushButton(get_profile_display_name(prof_val)); btn_prof.setToolTip("点击选择账号或已打开窗口")
                btn_prof.setProperty("profile_id", prof_val)  # Fix: must store the raw profile ID so _update_url_config can read it
                btn_prof.clicked.connect(lambda chk, idx=i, b=btn_prof: self._pick_profile_for_action(idx, b))
                l.addWidget(le, 2); l.addWidget(btn_prof, 1); self.action_table.setCellWidget(i, 3, w)
            elif a.get('action') == "⏸️ 延后执行":
                self._build_defer_action_widget(i, a)
            elif a.get('action') == "屏幕截图": le = QLineEdit("- 自动保存到 screenshots 目录 -"); le.setEnabled(False); self.action_table.setCellWidget(i, 3, le)
            elif a.get('action') == "移动鼠标": le = QLineEdit("- 仅移动，不点击 -"); le.setEnabled(False); self.action_table.setCellWidget(i, 3, le)
            elif a.get('action') == "运行程序" or a.get('action') == "💻 CMD 指令":
                w = QWidget(); l = QHBoxLayout(w); l.setContentsMargins(0,0,0,0)
                le = QLineEdit(str(a.get('value', '')))
                le.setPlaceholderText("输入命令或使用右侧预设..." if a.get('action') == "💻 CMD 指令" else "输入 exe 路径或命令...")
                le.setToolTip(var_tooltip)
                # 修复保存逻辑：使用更稳定的闭包方式绑定 le
                le.editingFinished.connect(lambda idx=i, edit=le: self._update_config(idx, 'value', edit.text()))
                btn_pre = QPushButton("📋"); btn_pre.setFixedWidth(30); btn_pre.setToolTip("CMD 预设中心")
                btn_pre.clicked.connect(lambda chk, idx=i, target=le: self._show_cmd_presets(idx, target))
                l.addWidget(le); l.addWidget(btn_pre); self.action_table.setCellWidget(i, 3, w)
            elif a.get('action') == "滚轮滚动": le = QLineEdit(str(a.get('value', ''))); le.setPlaceholderText("正数向上，负数向下 (如 -500)"); le.editingFinished.connect(lambda idx=i, edit=le: self._update_config(idx, 'value', edit.text())); self.action_table.setCellWidget(i, 3, le)
            elif a.get('action') in ["输入文本", "清空输入"]:
                ed = MultiLineTextEdit()
                ed.setText(str(a.get('value', '')))
                ed.setPlaceholderText("输入多行内容...")
                ed.setToolTip(var_tooltip)
                ed.setMinimumHeight(76)
                ed.editingFinished.connect(lambda idx=i, edit=ed: self._update_config(idx, 'value', edit.text()))
                self.action_table.setCellWidget(i, 3, ed)
                self.action_table.setRowHeight(i, max(self.action_table.rowHeight(i), 82))
            elif a.get('action') == "✨ 清空并输入(增强版)":
                w = QWidget(); l = QHBoxLayout(w); l.setContentsMargins(0,0,0,0); l.setSpacing(2)
                val_parts = str(a.get('value', '')).split('|')
                prefix_val = val_parts[0] if val_parts else ""
                content_val = val_parts[1] if len(val_parts) > 1 else ""
                le_prefix = QLineEdit(prefix_val); le_prefix.setPlaceholderText("预置前缀..."); le_prefix.setFixedWidth(80)
                le_content = MultiLineTextEdit(); le_content.setText(content_val); le_content.setPlaceholderText("输入多行内容..."); le_content.setToolTip(var_tooltip); le_content.setMinimumHeight(76)
                btn_prefix = QPushButton("📚"); btn_prefix.setFixedWidth(30); btn_prefix.setToolTip("打开常用前缀库")
                def _save_plus(idx=i, lp=le_prefix, lc=le_content):
                    self._update_config(idx, 'value', f"{lp.text()}|{lc.text()}")
                le_prefix.editingFinished.connect(_save_plus); le_content.editingFinished.connect(_save_plus)
                btn_prefix.clicked.connect(lambda chk, idx=i, target=le_prefix, saver=_save_plus: self._show_clear_input_prefix_presets(idx, target, saver))
                l.addWidget(le_prefix); l.addWidget(btn_prefix); l.addWidget(QLabel("+")); l.addWidget(le_content)
                self.action_table.setCellWidget(i, 3, w)
                self.action_table.setRowHeight(i, max(self.action_table.rowHeight(i), 82))
            elif "图像识别点击" in act_name: le = QLineEdit(str(a.get('value', ''))); le.setPlaceholderText("图片路径..."); le.editingFinished.connect(lambda idx=i, edit=le: self._update_config(idx, 'value', edit.text())); self.action_table.setCellWidget(i, 3, le)
            elif any(x in act_name for x in ["如果找图成功", "如果窗口存在"]):
                w = QWidget(); l = QHBoxLayout(w); l.setContentsMargins(0,0,0,0); l.setSpacing(2)
                val_parts = [p.strip() for p in str(a.get('value', '')).split('|')]
                target_val = val_parts[0] if val_parts else ""
                ok_val = val_parts[1] if len(val_parts) > 1 else ""
                fail_val = val_parts[2] if len(val_parts) > 2 else ""
                
                # 1. 目标输入框 (带选择按钮)
                le_target = QLineEdit(target_val); le_target.setPlaceholderText("图片路径" if "找图" in act_name else "窗口标题")
                le_target.setStyleSheet("background-color: #faf7ff; border: 1px solid #ddd6fe; color: #3f3f46;")
                btn_pick = QPushButton("📁" if "找图" in act_name else "🔍"); btn_pick.setFixedWidth(30); btn_pick.setStyleSheet("background: #f1edff; border: 1px solid #ddd6fe;")
                if "找图" in act_name:
                    btn_pick.clicked.connect(lambda chk, idx=i, target=le_target: target.setText(os.path.normpath(QFileDialog.getOpenFileName(self, "选择图片", "", "Images (*.png *.jpg *.bmp)")[0])) or self._update_if_config(idx))
                else:
                    btn_pick.clicked.connect(lambda chk, idx=i, target=le_target: target.setText(WindowSelector(self).get_selection()) or self._update_if_config(idx))
                le_target.editingFinished.connect(lambda idx=i: self._update_if_config(idx))
                
                # 2. 成功跳转
                step_names = ["(顺序)"] + [act.get('name', f'步骤{idx+1}') for idx, act in enumerate(acts)]
                cb_ok = QComboBox(); cb_ok.addItems(step_names); cb_ok.setCurrentText(ok_val if ok_val else "(顺序)")
                cb_ok.setToolTip("✅ 成功跳转到"); cb_ok.setFixedWidth(86); cb_ok.setStyleSheet("background-color: #eefbf3; border: 1px solid #b7e4c7; color: #1f6f43;")
                cb_ok.currentTextChanged.connect(lambda v, idx=i: self._update_if_config(idx))
                
                # 3. 失败跳转
                cb_fail = QComboBox(); cb_fail.addItems(step_names); cb_fail.setCurrentText(fail_val if fail_val else "(顺序)")
                cb_fail.setToolTip("❌ 失败跳转到"); cb_fail.setFixedWidth(86); cb_fail.setStyleSheet("background-color: #fff1f2; border: 1px solid #fecdd3; color: #b42318;")
                cb_fail.currentTextChanged.connect(lambda v, idx=i: self._update_if_config(idx))
                
                l.addWidget(le_target, 2); l.addWidget(btn_pick); l.addWidget(cb_ok, 1); l.addWidget(cb_fail, 1)
                self.action_table.setCellWidget(i, 3, w)
            else:
                le = QLineEdit(str(a.get('value', ''))); le.setPlaceholderText("输入内容..."); le.setToolTip(var_tooltip); le.editingFinished.connect(lambda idx=i, edit=le: self._update_config(idx, 'value', edit.text())); self.action_table.setCellWidget(i, 3, le)
        self.action_table.blockSignals(False)
        self._restore_action_column_widths()

    def _normalize_defer_policy_label(self, text):
        return "到时优先恢复" if str(text).strip() in ("到时优先恢复", "到时间优先恢复", "优先恢复") else "整轮后再恢复"

    def _list_action_step_names(self):
        acts = self.config['tasks'].get(self.current_task, []) if self.current_task else []
        return [act.get('name', f'步骤{idx+1}') for idx, act in enumerate(acts)]

    def _get_default_defer_target(self, step_idx, step_names=None):
        step_names = step_names if step_names is not None else self._list_action_step_names()
        if step_idx + 1 < len(step_names):
            return step_names[step_idx + 1]
        if step_names:
            return step_names[-1]
        return ""

    def _make_unique_action_name(self, base_name, skip_idx=None):
        if not self.current_task:
            return str(base_name or "").strip() or "步骤"
        acts = self.config['tasks'].get(self.current_task, [])
        existing = {
            str(act.get('name', '')).strip()
            for idx, act in enumerate(acts)
            if idx != skip_idx and str(act.get('name', '')).strip()
        }
        base_name = str(base_name or "").strip() or "步骤"
        if base_name not in existing:
            return base_name
        seed = re.sub(r'_(副本|复\d+|\d+)$', '', base_name).strip() or base_name
        n = 2
        candidate = f"{seed}_{n}"
        while candidate in existing:
            n += 1
            candidate = f"{seed}_{n}"
        return candidate

    def _rename_defer_target_references(self, old_name, new_name):
        if not self.current_task:
            return False
        old_name = str(old_name or "").strip()
        new_name = str(new_name or "").strip()
        if not old_name or not new_name or old_name == new_name:
            return False

        acts = self.config['tasks'].get(self.current_task, [])
        changed = False
        for idx, act in enumerate(acts):
            if act.get('action') != "⏸️ 延后执行":
                continue
            cfg = self._decode_defer_config(act.get('value', ''), idx)
            if cfg["resume_mode"] == "指定步骤" and cfg["target"] == old_name:
                act['value'] = self._encode_defer_config(cfg["seconds"], cfg["resume_mode"], new_name, cfg["policy"])
                changed = True
        return changed

    def _refresh_defer_target_options(self, removed_names=None, persist_changes=False):
        if not self.current_task:
            return False

        acts = self.config['tasks'].get(self.current_task, [])
        step_names = self._list_action_step_names()
        removed_names = {str(name).strip() for name in (removed_names or []) if str(name).strip()}
        changed = False

        for idx, act in enumerate(acts):
            if act.get('action') != "⏸️ 延后执行":
                continue

            cfg = self._decode_defer_config(act.get('value', ''), idx)
            target = str(cfg.get("target", "")).strip()
            if cfg["resume_mode"] == "指定步骤":
                target_missing = bool(target) and target not in step_names
                target_removed = bool(target) and target in removed_names
                if target_missing or target_removed:
                    fallback = self._get_default_defer_target(idx, step_names)
                    if fallback:
                        cfg["target"] = fallback
                    else:
                        cfg["resume_mode"] = "下一步"
                        cfg["target"] = ""

            new_val = self._encode_defer_config(cfg["seconds"], cfg["resume_mode"], cfg["target"], cfg["policy"])
            if act.get('value', '') != new_val:
                act['value'] = new_val
                changed = True

            if not hasattr(self, "action_table"):
                continue
            w = self.action_table.cellWidget(idx, 3)
            if not w:
                continue
            cb_resume = w.findChild(QComboBox, "defer_resume_mode")
            cb_target = w.findChild(QComboBox, "defer_target_step")
            cb_policy = w.findChild(QComboBox, "defer_policy")
            sp_seconds = w.findChild(QSpinBox, "defer_seconds")
            if not cb_resume or not cb_target or not cb_policy or not sp_seconds:
                continue

            cb_resume.blockSignals(True)
            cb_target.blockSignals(True)
            cb_policy.blockSignals(True)
            sp_seconds.blockSignals(True)
            try:
                if cb_resume.currentText() != cfg["resume_mode"]:
                    cb_resume.setCurrentText(cfg["resume_mode"])
                if cb_policy.currentText() != cfg["policy"]:
                    cb_policy.setCurrentText(cfg["policy"])
                if sp_seconds.value() != int(cfg["seconds"]):
                    sp_seconds.setValue(int(cfg["seconds"]))
                current_target = str(cfg.get("target", "")).strip()
                cb_target.clear()
                if step_names:
                    cb_target.addItems(step_names)
                if current_target and cb_target.findText(current_target) < 0:
                    cb_target.addItem(current_target)
                if current_target:
                    cb_target.setCurrentText(current_target)
            finally:
                sp_seconds.blockSignals(False)
                cb_policy.blockSignals(False)
                cb_target.blockSignals(False)
                cb_resume.blockSignals(False)
            self._update_defer_target_visibility(idx)

        if changed and persist_changes:
            save_config(self.config)
        return changed

    def _decode_defer_config(self, raw_val, step_idx=0):
        step_names = self._list_action_step_names()
        raw_text = str(raw_val or "").replace('\r', '').replace('\n', ' ').strip()
        parts = [p.strip() for p in raw_text.split('|')]
        seconds = 0
        if parts and parts[0]:
            try:
                seconds = int(float(parts[0]))
            except Exception:
                seconds = 0

        policy = "整轮后再恢复"
        resume_mode = "下一步"
        target = ""

        if len(parts) > 1:
            token = parts[1]
            if token in ("", "下一步", "顺序执行"):
                resume_mode = "下一步"
            elif token in ("指定步骤", "指定", "恢复到", "跳转到"):
                resume_mode = "指定步骤"
            elif token in ("到时优先恢复", "到时间优先恢复", "优先恢复", "整轮后恢复", "整轮后再恢复"):
                policy = self._normalize_defer_policy_label(token)
            else:
                resume_mode = "指定步骤"
                target = token

        if len(parts) > 2:
            token = parts[2]
            if token in ("到时优先恢复", "到时间优先恢复", "优先恢复", "整轮后恢复", "整轮后再恢复") and resume_mode == "下一步":
                policy = self._normalize_defer_policy_label(token)
            elif token:
                target = token

        if len(parts) > 3 and parts[3]:
            policy = self._normalize_defer_policy_label(parts[3])

        if resume_mode == "指定步骤" and not target:
            target = self._get_default_defer_target(step_idx, step_names)
            if not target:
                resume_mode = "下一步"

        return {
            "seconds": max(0, seconds),
            "resume_mode": resume_mode,
            "target": target,
            "policy": policy,
            "step_names": step_names,
        }

    def _encode_defer_config(self, seconds, resume_mode, target, policy):
        seconds = max(0, int(seconds or 0))
        resume_mode = "指定步骤" if resume_mode == "指定步骤" else "下一步"
        target = str(target or "").strip() if resume_mode == "指定步骤" else ""
        policy = self._normalize_defer_policy_label(policy)
        return f"{seconds} | {resume_mode} | {target} | {policy}"

    def _update_defer_target_visibility(self, idx):
        w = self.action_table.cellWidget(idx, 3)
        if not w:
            return
        cb_resume = w.findChild(QComboBox, "defer_resume_mode")
        cb_target = w.findChild(QComboBox, "defer_target_step")
        if not cb_resume or not cb_target:
            return
        is_target_mode = cb_resume.currentText() == "指定步骤"
        cb_target.setVisible(is_target_mode)
        cb_target.setEnabled(is_target_mode and cb_target.count() > 0)

    def _update_defer_config(self, idx):
        w = self.action_table.cellWidget(idx, 3)
        acts = self.config['tasks'].get(self.current_task, [])
        if not w or idx >= len(acts):
            return
        sp_seconds = w.findChild(QSpinBox, "defer_seconds")
        cb_resume = w.findChild(QComboBox, "defer_resume_mode")
        cb_target = w.findChild(QComboBox, "defer_target_step")
        cb_policy = w.findChild(QComboBox, "defer_policy")
        if not sp_seconds or not cb_resume or not cb_policy:
            return
        self._update_defer_target_visibility(idx)
        target = cb_target.currentText().strip() if cb_target and cb_resume.currentText() == "指定步骤" else ""
        new_val = self._encode_defer_config(sp_seconds.value(), cb_resume.currentText(), target, cb_policy.currentText())
        acts[idx]['value'] = new_val
        save_config(self.config)

    def _build_defer_action_widget(self, idx, action):
        cfg = self._decode_defer_config(action.get('value', ''), idx)
        w = QWidget()
        l = QHBoxLayout(w)
        l.setContentsMargins(0, 0, 0, 0)
        l.setSpacing(4)

        sp_seconds = QSpinBox()
        sp_seconds.setObjectName("defer_seconds")
        sp_seconds.setRange(0, 99999)
        sp_seconds.setSuffix(" 秒")
        sp_seconds.setValue(cfg["seconds"])
        sp_seconds.setToolTip("挂起多久后允许恢复")

        cb_policy = QComboBox()
        cb_policy.setObjectName("defer_policy")
        cb_policy.addItems(["到时优先恢复", "整轮后再恢复"])
        cb_policy.setCurrentText(cfg["policy"])
        cb_policy.setToolTip("到时优先恢复：冷却到点后优先插回执行；整轮后再恢复：等当前这一轮普通组先跑完。")

        cb_resume = QComboBox()
        cb_resume.setObjectName("defer_resume_mode")
        cb_resume.addItems(["下一步", "指定步骤"])
        cb_resume.setCurrentText(cfg["resume_mode"])

        cb_target = QComboBox()
        cb_target.setObjectName("defer_target_step")
        cb_target.setMinimumWidth(120)
        if cfg["step_names"]:
            cb_target.addItems(cfg["step_names"])
            if cfg["target"] and cfg["target"] not in cfg["step_names"]:
                cb_target.addItem(cfg["target"])
            if cfg["target"]:
                cb_target.setCurrentText(cfg["target"])
        cb_target.setToolTip("指定恢复时从哪一步继续")

        for widget in (sp_seconds, cb_policy, cb_resume, cb_target):
            if hasattr(widget, "currentTextChanged"):
                widget.currentTextChanged.connect(lambda _=None, row=idx: self._update_defer_config(row))
            if hasattr(widget, "valueChanged"):
                widget.valueChanged.connect(lambda _=None, row=idx: self._update_defer_config(row))

        l.addWidget(sp_seconds, 0)
        l.addWidget(cb_policy, 0)
        l.addWidget(cb_resume, 0)
        l.addWidget(cb_target, 1)
        self.action_table.setCellWidget(idx, 3, w)
        self._update_defer_target_visibility(idx)
        return w

    def _clear_deferred_queue_panel(self):
        self._deferred_queue_data = []
        if hasattr(self, "_deferred_panel_timer"):
            self._deferred_panel_timer.stop()
        self._refresh_deferred_queue_panel()

    def _on_deferred_queue_update(self, items):
        self._deferred_queue_data = list(items or [])
        if hasattr(self, "_deferred_panel_timer"):
            if self._deferred_queue_data:
                self._deferred_panel_timer.start()
            else:
                self._deferred_panel_timer.stop()
        self._refresh_deferred_queue_panel()

    def _refresh_deferred_queue_panel(self):
        if not hasattr(self, "defer_queue_table"):
            return
        items = sorted(list(self._deferred_queue_data or []), key=lambda e: (e.get("due_at", 0), e.get("row_index", 0)))
        self.defer_queue_table.setRowCount(len(items))
        if not items:
            self.lbl_deferred_summary.setText("当前无挂起项")
            return

        ready_count = 0
        for row, item in enumerate(items):
            remaining = max(0.0, float(item.get("due_at", time.time())) - time.time())
            remaining_text = self._format_seconds(max(1, int(math.ceil(remaining)))) if remaining > 0 else "已到时"
            status_text = "冷却中" if remaining > 0 else "可恢复"
            if remaining <= 0:
                ready_count += 1

            values = [
                f"{item.get('loop_index', 0) + 1} / {item.get('row_index', 0) + 1}",
                item.get("origin_name", ""),
                item.get("resume_label", ""),
                item.get("policy_label", "整轮后再恢复"),
                remaining_text,
                status_text,
            ]
            for col, value in enumerate(values):
                cell = self.defer_queue_table.item(row, col)
                if not cell:
                    cell = QTableWidgetItem()
                    if col in (0, 4, 5):
                        cell.setTextAlignment(Qt.AlignCenter)
                    self.defer_queue_table.setItem(row, col, cell)
                cell.setText(str(value))
                cell.setToolTip(
                    f"第 {item.get('loop_index', 0) + 1} 轮 / 第 {item.get('row_index', 0) + 1} 组\n"
                    f"挂起步骤：{item.get('origin_name', '')}\n"
                    f"恢复步骤：{item.get('resume_label', '')}\n"
                    f"恢复策略：{item.get('policy_label', '整轮后再恢复')}"
                )
                if remaining <= 0:
                    cell.setBackground(QColor("#d1fae5"))
                    cell.setForeground(QColor("#065f46"))
                else:
                    cell.setBackground(QColor("#eff6ff"))
                    cell.setForeground(QColor("#1e3a8a"))

        self.lbl_deferred_summary.setText(f"当前挂起 {len(items)} 项，其中 {ready_count} 项已到时")

    def _update_url_config(self, idx, btn=None):
        w = self.action_table.cellWidget(idx, 3)
        if not w: return
        le = w.findChild(QLineEdit)
        if btn is None:
            btn = w.findChild(QPushButton)
        prof_id = btn.property("profile_id") if btn else ""
        prof_id = prof_id or ""
        url_text = le.text() if le else ""
        val = f"{url_text}|{prof_id}"
        self.config['tasks'][self.current_task][idx]['value'] = val
        save_config(self.config)
        self._log(f"💾 打开网址步骤已保存账户: [{get_profile_display_name(prof_id)}] ({prof_id})", "blue")

    def _update_drag_file_config(self, idx):
        w = self.action_table.cellWidget(idx, 3)
        if not w: return
        le = w.findChild(QLineEdit)
        chk = w.findChild(QCheckBox)
        acts = self.config['tasks'].get(self.current_task, [])
        if idx >= len(acts): return
        acts[idx]['value'] = le.text().strip() if le else ""
        acts[idx]['activate_target_before_drag'] = bool(chk.isChecked()) if chk else False
        save_config(self.config)

    def _update_if_config(self, idx):
        w = self.action_table.cellWidget(idx, 3)
        if not w: return
        le = w.findChild(QLineEdit)
        cbs = w.findChildren(QComboBox)
        if not le or len(cbs) < 2: return
        
        target = le.text().strip()
        ok_jump = cbs[0].currentText() if cbs[0].currentText() != "(顺序)" else ""
        fail_jump = cbs[1].currentText() if cbs[1].currentText() != "(顺序)" else ""
        
        val = f"{target} | {ok_jump} | {fail_jump}"
        self.config['tasks'][self.current_task][idx]['value'] = val
        save_config(self.config)
        # 同时同步刷新数据表格的表头（如果有 force_sync）
        self._refresh_data_table()

    def _pick_profile_for_action(self, idx, btn):
        # [增强] open_url 支持两种目标：
        # 1) 账号 profile（原逻辑）
        # 2) 已打开窗口（::hwnd=xxx），用于直接激活已打开的浏览器再打开网址
        menu = QMenu(self)
        act_acc = menu.addAction("选择账号…")
        act_win = menu.addAction("选择已打开窗口…")
        chosen = menu.exec_(btn.mapToGlobal(btn.rect().bottomLeft()))
        if chosen == act_acc:
            sel = ChromeProfileSelector(self); pid = sel.get_selection()
            if pid is not None:
                btn.setProperty("profile_id", pid); btn.setText(get_profile_display_name(pid))
                self._update_url_config(idx, btn)
        elif chosen == act_win:
            s = WindowSelector(self).get_selection()
            if s:
                btn.setProperty("profile_id", s); btn.setText(get_profile_display_name(s))
                self._update_url_config(idx, btn)

    def _show_cmd_presets(self, row, target_le):
        dlg = CommandPresetDialog(self)
        if dlg.exec_():
            cmd = dlg.get_command()
            if cmd:
                target_le.setText(cmd)
                self._update_config(row, 'value', cmd)

    def _show_clear_input_prefix_presets(self, row, target_le, save_cb=None):
        dlg = ClearInputPrefixPresetDialog(self)
        if dlg.exec_():
            prefix = dlg.get_prefix()
            if prefix is not None:
                target_le.setText(prefix)
                if callable(save_cb):
                    save_cb()
                else:
                    w = self.action_table.cellWidget(row, 3)
                    edits = w.findChildren(QLineEdit) if w else []
                    prefix_text = edits[0].text() if len(edits) > 0 else prefix
                    content_text = edits[1].text() if len(edits) > 1 else ""
                    self._update_config(row, 'value', f"{prefix_text}|{content_text}")

    def _open_prefix_library(self):
        dlg = ClearInputPrefixPresetDialog(self)
        dlg.exec_()

    def _update_config(self, idx, key, val):
        if not self.current_task: return
        if key == 'row_value_update':
            self.config['tasks'][self.current_task][idx]['value'] = val
            save_config(self.config)
            return
        acts = self.config['tasks'].get(self.current_task, [])
        if idx >= len(acts): return
        acts[idx][key] = val
        save_config(self.config)
        # 修复2：保存成功反馈 —— 当用户修改指令内容并离开输入框时，日志区闪烁确认
        if key == 'value':
            step_name = acts[idx].get('name', f'步骤{idx+1}')
            preview = str(val)[:40] + ('...' if len(str(val)) > 40 else '')
            self._log(f"💾 [已保存] 步骤「{step_name}」指令内容: {preview}", "blue")
        # 增强：如果是修改步骤名称，强制同步数据表头
        if key == 'name':
            self._refresh_data_table()
        
        # Only refresh UI when action TYPE changes (need to rebuild widgets for that row)
        # For value/delay changes the widget already reflects the new value — no refresh needed
        if key == 'action': 
            self._refresh_defer_target_options(persist_changes=True)
            self._refresh_actions()
            self._refresh_data_table()

    def _sync_data_headers(self):
        if not self.current_task: return
        reply = QMessageBox.question(
            self,
            '确认同步',
            "同步表头将根据当前流程更新批量数据列结构：\n"
            "1) 新增你流程里新加的步骤列（用该步骤默认值初始化）\n"
            "2) 移除已不存在的步骤列\n"
            "不会覆盖你已经填写过的内容。\n\n是否继续？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            # 同步前先把流程编辑区尚未失焦的最新输入写回 config（尤其是增强版前缀|内容）
            self._force_sync_action_widgets()
            self._save_data_table(flush=True)
            self._apply_task_data_schema(overwrite=False)
            self._refresh_data_table(force_sync=False)

    def _reset_data_to_presets(self):
        """将批量数据强制重置为流程编排的默认预设值（不只是同步表头）。"""
        if not self.current_task:
            return
        reply = QMessageBox.question(
            self,
            "确认重置",
            "将把当前任务的批量数据（可编辑步骤列）全部重置为流程编排中的默认值。\n"
            "你之前在批量数据里手动填的不同文本会被覆盖。\n\n是否继续？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return
        self._force_sync_action_widgets()
        self._save_data_table(flush=True)
        self._apply_task_data_schema(overwrite=True)
        self._refresh_data_table(force_sync=False)
        QMessageBox.information(self, "完成", "已重置为流程预设值。")

    def _apply_task_data_schema(self, overwrite=False):
        """同步批量数据的列 schema。
        - overwrite=False: 只补齐缺失列/清理无效列，不覆盖已有值
        - overwrite=True : 可编辑列全部写入流程默认值（强制重置）
        """
        if not self.current_task:
            return
        acts = self.config['tasks'].get(self.current_task, [])
        old_data = self.config['task_data'].get(self.current_task, [])

        def _is_keep_key(k: str) -> bool:
            # 允许保留的系统字段
            return k.startswith("_")

        new_data = []
        for row in old_data:
            row_dict = row if isinstance(row, dict) else {}
            out = {}
            # 保留系统字段（例如 _选中）
            for k, v in row_dict.items():
                if _is_keep_key(str(k)):
                    out[k] = v

            # 以“流程编排”为准构造允许字段集合
            for act in acts:
                name = act.get('name', 'Step')
                act_type = CMD_MAP.get(act.get('action'), "click")
                # 紧凑/坐标类列不应进入数据字典，否则可能触发“空值结束当前行”的逻辑
                if self._is_compact_data_action(act_type):
                    continue
                default_val = str(act.get('value', ''))

                if overwrite or name not in row_dict:
                    out[name] = default_val
                else:
                    out[name] = row_dict.get(name, default_val)

                # 延时：同步表头时保留已有；强制重置时清空（表示使用步骤默认延时）
                delay_key = f"{name}_延时"
                if overwrite:
                    if delay_key in row_dict:
                        out[delay_key] = ""
                else:
                    if delay_key in row_dict:
                        out[delay_key] = row_dict.get(delay_key, "")

                # 步骤跳过开关：同步表头保留；强制重置则不保留
                skip_key = f"{name}_跳过"
                if not overwrite and skip_key in row_dict:
                    out[skip_key] = row_dict.get(skip_key, False)

            new_data.append(out)

        self.config['task_data'][self.current_task] = new_data
        save_config(self.config)

    def _show_data_toolbar_menu(self):
        menu = QMenu(self)
        act_delay = menu.addAction("⏱️ 显示延时列")
        act_delay.setCheckable(True)
        act_delay.setChecked(self.btn_toggle_delay.isChecked())
        act_delay.triggered.connect(self.btn_toggle_delay.setChecked)
        act_delay.triggered.connect(self._toggle_delay_cols)

        act_noneditable = menu.addAction("👁️ 显示坐标步骤")
        act_noneditable.setCheckable(True)
        act_noneditable.setChecked(self.btn_toggle_noneditable.isChecked())
        act_noneditable.triggered.connect(self.btn_toggle_noneditable.setChecked)
        act_noneditable.triggered.connect(self._toggle_noneditable_cols)

        menu.addAction(f"📏 设置行高（当前 {self.data_row_height_spin.value()} px）", self._pick_data_row_height)
        menu.addSeparator()

        act_only_current = menu.addAction("仅执行当前任务")
        act_only_current.setCheckable(True)
        act_only_current.setChecked(self.chk_only_current.isChecked())
        act_only_current.triggered.connect(self.chk_only_current.setChecked)

        act_continue_on_fail = menu.addAction("❌ 失败也继续")
        act_continue_on_fail.setCheckable(True)
        act_continue_on_fail.setChecked(self.chk_continue_on_fail.isChecked())
        act_continue_on_fail.triggered.connect(self.chk_continue_on_fail.setChecked)

        act_auto_shutdown = menu.addAction("🏁 任务完关机")
        act_auto_shutdown.setCheckable(True)
        act_auto_shutdown.setChecked(self.chk_auto_shutdown.isChecked())
        act_auto_shutdown.triggered.connect(self.chk_auto_shutdown.setChecked)

        act_show_osd = menu.addAction("🖥️ 显示置顶进度条")
        act_show_osd.setCheckable(True)
        act_show_osd.setChecked(self.chk_show_osd.isChecked())
        act_show_osd.triggered.connect(self.chk_show_osd.setChecked)

        act_std_win = menu.addAction("📏 自动标准化窗口")
        act_std_win.setCheckable(True)
        act_std_win.setChecked(self.chk_std_win.isChecked())
        act_std_win.triggered.connect(self.chk_std_win.setChecked)

        act_multi_open = menu.addAction("👯 多账号并行模式")
        act_multi_open.setCheckable(True)
        act_multi_open.setChecked(self.chk_multi_open.isChecked())
        act_multi_open.triggered.connect(self.chk_multi_open.setChecked)

        menu.exec_(self.btn_more_data_ops.mapToGlobal(self.btn_more_data_ops.rect().bottomLeft()))

    def _pick_data_row_height(self):
        cur_val = self.data_row_height_spin.value() if hasattr(self, "data_row_height_spin") else 28
        value, ok = QInputDialog.getInt(self, "设置行高", "请输入批量数据统一行高：", cur_val, 24, 96, 1)
        if ok:
            self._set_data_row_height(value)

    def _toggle_delay_cols(self, checked):
        self.btn_toggle_delay.setText("⏲️ 隐藏延时列" if checked else "⏲️ 显示延时列")
        self._refresh_data_table()

    def _update_noneditable_toggle_text(self, checked):
        self.btn_toggle_noneditable.setText("👁️ 显示坐标步骤" if checked else "🙈 隐藏坐标步骤")
        self.btn_toggle_noneditable.setToolTip("切换后可显示或隐藏批量数据中不可直接编辑的步骤列，例如点击、移动、等待、截图等。")

    def _toggle_noneditable_cols(self, checked):
        self._update_noneditable_toggle_text(checked)
        if "layout" not in self.config:
            self.config["layout"] = {}
        self.config["layout"]["show_noneditable_steps"] = checked
        self._schedule_config_flush()
        self._refresh_data_table()

    def _set_data_row_height(self, value, save=True):
        row_height = min(max(24, int(value)), 96)
        self.config.setdefault("layout", {})
        self.config["layout"]["data_row_height"] = row_height

        if hasattr(self, "data_row_height_spin") and self.data_row_height_spin.value() != row_height:
            self.data_row_height_spin.blockSignals(True)
            self.data_row_height_spin.setValue(row_height)
            self.data_row_height_spin.blockSignals(False)

        if hasattr(self, "data_col_width_slider") and self.data_col_width_slider.value() != row_height:
            self.data_col_width_slider.blockSignals(True)
            self.data_col_width_slider.setValue(row_height)
            self.data_col_width_slider.blockSignals(False)
        if hasattr(self, "data_col_width_label"):
            self.data_col_width_label.setText(f"整体行高: {row_height} px")

        if hasattr(self, "data_table"):
            self._syncing_data_row_height = True
            try:
                self.data_table.verticalHeader().setDefaultSectionSize(row_height)
                for r in range(self.data_table.rowCount()):
                    self.data_table.setRowHeight(r, row_height)
            finally:
                self._syncing_data_row_height = False

        if save:
            self._schedule_config_flush()

    def _on_data_row_resized(self, logical_index, old_size, new_size):
        if getattr(self, "_syncing_data_row_height", False):
            return
        self._set_data_row_height(new_size, save=True)

    def _data_select_col(self):
        return 0

    def _data_run_col(self):
        return 1

    def _data_status_col(self):
        return 2

    def _data_first_value_col(self):
        return 3

    def _update_data_select_header(self):
        """同步第一列表头的全选状态提示。"""
        if not hasattr(self, "data_table") or self.data_table.columnCount() <= 0:
            return
        header_item = self.data_table.horizontalHeaderItem(self._data_select_col())
        if not header_item:
            return
        has_any = False
        all_checked = True
        for r in range(self.data_table.rowCount()):
            if not self._is_data_row_selectable(r):
                continue
            item = self.data_table.item(r, self._data_select_col())
            if not item:
                continue
            has_any = True
            if item.checkState() != Qt.Checked:
                all_checked = False
        header_item.setText("☑选择" if has_any and all_checked else "☐选择")

    def _on_data_header_clicked(self, logical_index):
        """点击“选择”表头时直接全选/取消全选。"""
        if logical_index != self._data_select_col():
            return
        has_any = False
        all_checked = True
        for r in range(self.data_table.rowCount()):
            if not self._is_data_row_selectable(r):
                continue
            item = self.data_table.item(r, self._data_select_col())
            if not item:
                continue
            has_any = True
            if item.checkState() != Qt.Checked:
                all_checked = False
        if not has_any:
            return
        self._set_all_row_check_state(not all_checked)

    def _set_all_row_check_state(self, checked):
        self.data_table.blockSignals(True)
        for r in range(self.data_table.rowCount()):
            item = self.data_table.item(r, self._data_select_col())
            if item:
                can_check = self._is_data_row_selectable(r)
                item.setCheckState(Qt.Checked if checked and can_check else Qt.Unchecked)
        self.data_table.blockSignals(False)
        self._update_data_select_header()
        self._save_data_table()

    def _invert_row_check_state(self):
        self.data_table.blockSignals(True)
        for r in range(self.data_table.rowCount()):
            item = self.data_table.item(r, self._data_select_col())
            if item:
                if not self._is_data_row_selectable(r):
                    item.setCheckState(Qt.Unchecked)
                else:
                    item.setCheckState(Qt.Unchecked if item.checkState() == Qt.Checked else Qt.Checked)
        self.data_table.blockSignals(False)
        self._update_data_select_header()
        self._save_data_table()

    def _handle_data_checkbox_shift_click(self, row, checked, anchor_row=None):
        """在“选择”列支持 Shift 连续勾选/取消勾选。"""
        if not hasattr(self, "data_table"):
            return False
        row_count = self.data_table.rowCount()
        if row < 0 or row >= row_count:
            return False
        if anchor_row is None or anchor_row < 0 or anchor_row >= row_count:
            anchor_row = row

        start, end = sorted((anchor_row, row))
        state = Qt.Checked if checked else Qt.Unchecked
        self.data_table.blockSignals(True)
        try:
            for r in range(start, end + 1):
                item = self.data_table.item(r, self._data_select_col())
                if item:
                    item.setCheckState(state if checked and self._is_data_row_selectable(r) else Qt.Unchecked)
        finally:
            self.data_table.blockSignals(False)
        self._update_data_select_header()
        self._save_data_table()
        return True

    def _toggle_select_all(self, checked):
        """兼容旧调用，内部仍转到独立的全选/取消全选逻辑。"""
        self._set_all_row_check_state(bool(checked))

    def _select_unsuccessful_rows(self):
        if not self.current_task:
            return
        statuses = self._row_statuses.get(self.current_task, {})
        if not statuses:
            QMessageBox.information(self, "提示", "当前还没有可用于筛选的执行结果。")
            return

        target_statuses = {ROW_STATUS_FAIL, ROW_STATUS_SKIP, ROW_STATUS_DEFER, ROW_STATUS_MANUAL}
        matched_rows = []
        self.data_table.blockSignals(True)
        try:
            for r in range(self.data_table.rowCount()):
                item = self.data_table.item(r, self._data_select_col())
                if not item:
                    continue
                should_check = (statuses.get(r) in target_statuses) and self._is_data_row_selectable(r)
                item.setCheckState(Qt.Checked if should_check else Qt.Unchecked)
                if should_check:
                    matched_rows.append(r)
        finally:
            self.data_table.blockSignals(False)

        self._update_data_select_header()
        self._save_data_table()
        if matched_rows:
            self._log(f"⚠️ 已选中 {len(matched_rows)} 行未成功数据", "orange")
        else:
            QMessageBox.information(self, "提示", "当前没有未成功的执行行。")

    def _get_checked_data_rows(self):
        """获取“批量数据”里勾选的行索引列表。"""
        if not hasattr(self, "data_table"):
            return []
        rows = []
        for r in range(self.data_table.rowCount()):
            item = self.data_table.item(r, self._data_select_col())
            if not item:
                continue
            if item.checkState() != Qt.Checked:
                continue
            if not self._is_data_row_selectable(r):
                continue
            rows.append(r)
        return rows

    def _build_subtask_entries_from_rows(self, rows):
        """把勾选的行转换为“子任务管理器”可识别的条目结构。"""
        out = []
        if not self.current_task:
            return out

        # 最近一次执行的行结果里，尝试提取窗口上下文（hwnd/title/class）
        latest_payload_by_row = {}
        try:
            for payload in reversed(getattr(self, "_last_run_row_results", []) or []):
                if not isinstance(payload, dict):
                    continue
                if payload.get("task_id") != self.current_task:
                    continue
                r_idx = int(payload.get("row_index", -1))
                if r_idx < 0:
                    continue
                if r_idx not in latest_payload_by_row:
                    latest_payload_by_row[r_idx] = payload
        except Exception:
            latest_payload_by_row = {}

        statuses = self._row_statuses.get(self.current_task, {}) or {}
        row_data_list = (self.config.get("task_data", {}) or {}).get(self.current_task, []) or []
        for r in rows:
            payload = latest_payload_by_row.get(int(r), {}) if isinstance(latest_payload_by_row, dict) else {}
            ctx = (payload.get("fail_ctx") or payload.get("row_ctx") or {}) if isinstance(payload, dict) else {}
            fg = (ctx.get("foreground") or {}) if isinstance(ctx, dict) else {}
            row_data = row_data_list[int(r)] if 0 <= int(r) < len(row_data_list) and isinstance(row_data_list[int(r)], dict) else {}

            hwnd = ctx.get("target_hwnd") or fg.get("hwnd") or row_data.get("_窗口hwnd") or None
            try:
                hwnd = int(hwnd) if hwnd else None
            except Exception:
                hwnd = None

            win_title = str(fg.get("title") or row_data.get("_窗口标题") or "")
            win_class = str(fg.get("class") or row_data.get("_窗口类名") or "")
            win_hint = str(row_data.get("_窗口提示") or "")
            is_browser = bool((win_class == "Chrome_WidgetWin_1") or row_data.get("_是否浏览器窗口", False))

            if sys.platform == "win32" and not hwnd:
                try:
                    target_hint = win_hint or win_title
                    if target_hint:
                        hwnd = find_browser_window_hwnd_by_hint(target_hint)
                except Exception:
                    hwnd = None
            if sys.platform == "win32" and hwnd:
                try:
                    hwnd = int(hwnd)
                    if not win_title:
                        win_title = str(get_window_text(hwnd) or "")
                    if not win_class:
                        win_class = str(get_window_class_name(hwnd) or "")
                    if not win_hint:
                        win_hint = str(build_window_display_text(win_title, hwnd) or win_title or "")
                    if win_class == "Chrome_WidgetWin_1":
                        is_browser = True
                except Exception:
                    pass

            entry = {
                "task_id": self.current_task,
                "task_display": self._get_task_display_text(self.current_task, with_folder=True) if hasattr(self, "_get_task_display_text") else str(self.current_task),
                "row_index": int(r),
                "status": str(statuses.get(int(r), "") or ""),
                "_checked": True,
                "step_index": int(ctx.get("step_index", 0) or 0) if isinstance(ctx, dict) else 0,
                "step_name": str(ctx.get("step_name", "") or "") if isinstance(ctx, dict) else "",
                "error": str(ctx.get("error", payload.get("last_error", "")) or "") if isinstance(payload, dict) else "",
                "hwnd": hwnd,
                "window_title": win_title,
                "window_class": win_class,
                "window_hint": win_hint,
                "is_browser": bool(is_browser),
            }
            out.append(entry)
        return out

    def _open_subtask_manager(self):
        """对勾选行打开“子任务管理器”。"""
        try:
            if not self.current_task:
                return
            rows = self._get_checked_data_rows()
            if not rows:
                QMessageBox.information(self, "提示", "请先在“批量数据”里勾选至少一行。")
                return
            entries = self._build_subtask_entries_from_rows(rows)
            dlg = FailureManagerDialog(self, entries)
            try:
                dlg.setWindowTitle("🧩 子任务管理器")
            except Exception:
                pass
            try:
                dlg.context_task_id = self.current_task
                if hasattr(dlg, "_load_parent_steps"):
                    dlg._load_parent_steps()
            except Exception:
                pass
            self._subtask_progress_dialog = dlg
            try:
                dlg.exec_()
            finally:
                if getattr(self, "_subtask_progress_dialog", None) is dlg:
                    self._subtask_progress_dialog = None
        except Exception as e:
            # 记录详细调用栈到日志区；弹窗只给出精简信息，避免噪音
            tb = traceback.format_exc()
            try:
                self._log(f"❌ 打开子任务管理器失败: {e}\n\n{tb}", "red")
            except Exception:
                pass
            tail = "\n".join((tb or "").strip().splitlines()[-12:]) if tb else ""
            QMessageBox.warning(
                self,
                "打开失败",
                f"{e}\n\n最近调用栈(末尾):\n{tail}"
            )

    def _sync_subtask_manager_runtime(self, entry=None, parent_step_idx=None, running=None):
        dlg = getattr(self, "_subtask_progress_dialog", None)
        if not dlg:
            return
        try:
            if running is not None and hasattr(dlg, "set_runtime_running"):
                dlg.set_runtime_running(bool(running))
            if hasattr(dlg, "set_runtime_progress"):
                dlg.set_runtime_progress(entry=entry, parent_step_idx=parent_step_idx)
        except RuntimeError:
            self._subtask_progress_dialog = None
        except Exception:
            pass

    def _sync_subtask_manager_entry_status(self, entry=None, status="", error=None):
        dlg = getattr(self, "_subtask_progress_dialog", None)
        if not dlg:
            return
        try:
            if hasattr(dlg, "update_entry_status"):
                dlg.update_entry_status(entry=entry, status=status, error=error)
        except RuntimeError:
            self._subtask_progress_dialog = None
        except Exception:
            pass

    def _update_task_row_status_cache(self, task_id, row_idx, status):
        if not task_id:
            return
        try:
            row_idx = int(row_idx)
        except Exception:
            return
        if row_idx < 0:
            return
        self._row_statuses.setdefault(task_id, {})[row_idx] = status
        if task_id == self.current_task:
            self._on_row_status(row_idx, status)

    def _persist_row_window_context(self, payload):
        """把每行最近一次捕获到的窗口上下文写回 task_data，重开软件后也能继续复用。"""
        try:
            if not isinstance(payload, dict):
                return
            task_id = str(payload.get("task_id") or self.current_task or "")
            row_idx = int(payload.get("row_index", -1))
            if not task_id or row_idx < 0:
                return
            data_rows = self.config.get("task_data", {}).get(task_id, [])
            if row_idx >= len(data_rows) or not isinstance(data_rows[row_idx], dict):
                return

            row_dict = data_rows[row_idx]
            ctx = payload.get("fail_ctx") or payload.get("row_ctx") or {}
            fg = ctx.get("foreground") or {}
            hwnd = ctx.get("target_hwnd") or fg.get("hwnd") or row_dict.get("_窗口hwnd") or None
            try:
                hwnd = int(hwnd) if hwnd else None
            except Exception:
                hwnd = None

            win_title = str(fg.get("title") or row_dict.get("_窗口标题") or "")
            win_class = str(fg.get("class") or row_dict.get("_窗口类名") or "")
            is_browser = bool((win_class == "Chrome_WidgetWin_1") or row_dict.get("_是否浏览器窗口", False))
            if sys.platform == "win32" and hwnd:
                try:
                    if not win_title:
                        win_title = str(get_window_text(hwnd) or "")
                    if not win_class:
                        win_class = str(get_window_class_name(hwnd) or "")
                    if win_class == "Chrome_WidgetWin_1":
                        is_browser = True
                except Exception:
                    pass

            try:
                win_hint = str(build_window_display_text(win_title, hwnd) or win_title or "")
            except Exception:
                win_hint = win_title

            row_dict["_窗口hwnd"] = int(hwnd) if hwnd else None
            row_dict["_窗口标题"] = win_title
            row_dict["_窗口类名"] = win_class
            row_dict["_窗口提示"] = win_hint
            row_dict["_是否浏览器窗口"] = bool(is_browser)
            self._schedule_config_flush()
        except Exception:
            pass

    def _on_subtask_row_status(self, status, entry=None):
        status_text = str(status or "")
        self._subtask_last_row_status = status_text
        target_entry = entry if isinstance(entry, dict) else getattr(self, "_current_subtask_entry", None)
        if isinstance(target_entry, dict):
            self._sync_subtask_manager_entry_status(entry=target_entry, status=status_text, error=None)
            self._update_task_row_status_cache(
                str(target_entry.get("task_id") or ""),
                target_entry.get("row_index", -1),
                status_text
            )

    def _on_subtask_row_result(self, payload, entry=None):
        self._subtask_last_row_result = payload if isinstance(payload, dict) else None
        if isinstance(payload, dict):
            self._persist_row_window_context(payload)
        if isinstance(payload, dict) and payload.get("status"):
            self._on_subtask_row_status(payload.get("status"), entry=entry)

    def _update_subtask_progress(self, percent):
        if getattr(self, "_ui_stop_reset_pending", False):
            return
        total_count = max(int(getattr(self, "_subtask_total_count", 0) or 0), 1)
        pending_count = len(getattr(self, "_subtask_queue", []) or [])
        completed_count = max(0, total_count - pending_count - 1)
        try:
            percent = int(percent)
        except Exception:
            percent = 0
        percent = max(0, min(100, percent))
        overall = int(round(((completed_count + percent / 100.0) / total_count) * 100))
        self.progress.setValue(max(0, min(100, overall)))
        self._update_osd_subtask(percent)

    def _on_subtask_engine_highlight(self, entry, parent_step_indices, auto_activate_inserted, row_idx, step_idx):
        parent_step_idx = None
        try:
            step_idx = int(step_idx)
        except Exception:
            step_idx = -1
        try:
            parent_step_indices = [int(i) for i in (parent_step_indices or [])]
        except Exception:
            parent_step_indices = []

        parent_start = 1 if auto_activate_inserted else 0
        parent_end = parent_start + len(parent_step_indices)
        if parent_start <= step_idx < parent_end:
            parent_step_idx = parent_step_indices[step_idx - parent_start]
        self._sync_subtask_manager_runtime(entry=entry, parent_step_idx=parent_step_idx, running=True)

    def _start_subtask_for_entries(self, entries, subtask_task_id, auto_activate_hwnd=True, parent_step_indices=None, continue_with_parent=False, context_task_id=""):
        """对条目列表顺序执行子任务（每个条目跑一次子任务流程）。"""
        try:
            if hasattr(self, '_engine') and self._engine.isRunning():
                QMessageBox.warning(self, "执行中", "请先等待当前执行结束或停止后再启动子任务。")
                return False
            if subtask_task_id and subtask_task_id not in self.config.get("tasks", {}):
                QMessageBox.warning(self, "提示", "子任务不存在。")
                return False
            if not subtask_task_id and not parent_step_indices:
                QMessageBox.warning(self, "提示", "没有可执行的前置步骤或子任务。")
                return False
            self._subtask_queue = list(entries or [])
            self._subtask_task_id = subtask_task_id
            self._subtask_auto_activate_hwnd = bool(auto_activate_hwnd)
            self._subtask_origin_task = self.current_task
            self._subtask_parent_step_indices = sorted({int(i) for i in (parent_step_indices or []) if int(i) >= 0})
            self._subtask_continue_with_parent = bool(continue_with_parent)
            self._subtask_context_task_id = str(context_task_id or self.current_task or "")
            self._subtask_total_count = len(self._subtask_queue)
            self._subtask_completed_count = 0
            self._subtask_stopped = False
            self._ui_stop_reset_pending = False
            self._subtask_last_row_status = ""
            self._subtask_last_row_result = None
            self._current_subtask_entry = None
            self.progress.setValue(0)
            subtask_label = self._get_task_display_text(subtask_task_id, with_folder=True) if subtask_task_id else "(无额外子任务)"
            self._log(f"🧩 准备执行子任务：共 {len(self._subtask_queue)} 项 | 子任务={subtask_label}", "purple")
            self._show_osd_for_subtask()
            self._sync_subtask_manager_runtime(entry=None, parent_step_idx=None, running=True)
            self._run_next_subtask_entry()
            return True
        except Exception as e:
            QMessageBox.warning(self, "子任务启动失败", str(e))
            self._sync_subtask_manager_runtime(entry=None, parent_step_idx=None, running=False)
            self.osd.hide()
            return False

    def _run_next_subtask_entry(self):
        """顺序执行子任务队列。"""
        try:
            if not getattr(self, "_subtask_queue", None):
                # 子任务完成：回到开始前所在主任务
                self._current_subtask_entry = None
                if not getattr(self, "_subtask_stopped", False):
                    self.progress.setValue(100)
                elif getattr(self, "_ui_stop_reset_pending", False):
                    self.progress.setValue(0)
                self._subtask_total_count = 0
                origin = getattr(self, "_subtask_origin_task", "")
                if origin:
                    self._activate_task_by_id(origin)
                self.btn_run.setEnabled(True); self.btn_run.setText("🚀 开始批量执行")
                self.btn_stop.setEnabled(False); self.btn_resume.setEnabled(False)
                self.btn_dry_run.setEnabled(True)
                self.btn_pause.setEnabled(False)
                if getattr(self, "_subtask_stopped", False):
                    self._log("🛑 子任务已停止，剩余队列已取消。", "orange")
                else:
                    self._log("✅ 子任务已全部执行完毕。", "green")
                self._sync_subtask_manager_runtime(entry=None, parent_step_idx=None, running=False)
                self.osd.hide()
                return

            entry = self._subtask_queue.pop(0)
            if not isinstance(entry, dict):
                QTimer.singleShot(0, self._run_next_subtask_entry)
                return
            self._current_subtask_entry = entry

            src_task_id = entry.get("task_id") or getattr(self, "_subtask_origin_task", "")
            row_idx = int(entry.get("row_index", 0) or 0)
            hwnd = entry.get("hwnd")
            try:
                hwnd = int(hwnd) if hwnd else None
            except Exception:
                hwnd = None
            if sys.platform == "win32" and not hwnd:
                try:
                    hint = str(entry.get("window_hint") or entry.get("window_title") or "").strip()
                    if hint:
                        hwnd = find_browser_window_hwnd_by_hint(hint)
                        if hwnd:
                            entry["hwnd"] = int(hwnd)
                except Exception:
                    hwnd = None

            row_data_list = (self.config.get("task_data", {}) or {}).get(src_task_id, []) or []
            row_data = {}
            if 0 <= row_idx < len(row_data_list):
                if isinstance(row_data_list[row_idx], dict):
                    row_data = row_data_list[row_idx]

            import copy
            src_task_actions = copy.deepcopy(self.config.get("tasks", {}).get(src_task_id, []) or [])
            subtask_actions = copy.deepcopy(self.config.get("tasks", {}).get(self._subtask_task_id, []) or []) if getattr(self, "_subtask_task_id", "") else []

            selected_parent_actions = []
            parent_step_indices = list(getattr(self, "_subtask_parent_step_indices", []) or [])
            if src_task_id == getattr(self, "_subtask_context_task_id", src_task_id):
                for idx in parent_step_indices:
                    if 0 <= idx < len(src_task_actions):
                        selected_parent_actions.append(copy.deepcopy(src_task_actions[idx]))

            actions = []
            has_parent_window_step = any(CMD_MAP.get(act.get("action"), "") in ["win_active", "open_url"] for act in selected_parent_actions)
            inserted_auto_activate = False
            # 如果本行已捕获到 hwnd，且前置步骤里没有窗口相关动作，则自动加一条激活窗口
            if getattr(self, "_subtask_auto_activate_hwnd", True) and hwnd and not has_parent_window_step:
                inserted_auto_activate = True
                actions.append({
                    "name": "[子任务] 激活窗口",
                    "action": "激活窗口",
                    "value": f"::hwnd={int(hwnd)}",
                    "x": 0, "y": 0,
                    "delay": 0,
                    "guard_enabled": False
                })

            actions.extend(selected_parent_actions)
            actions.extend(subtask_actions)
            if getattr(self, "_subtask_continue_with_parent", False):
                actions.extend(src_task_actions)

            self._subtask_last_row_status = ""
            self._subtask_last_row_result = None
            self._sync_subtask_manager_runtime(entry=entry, parent_step_idx=None, running=True)
            self._sync_subtask_manager_entry_status(entry=entry, status="", error=None)

            if not actions:
                self._log("⚠️ 当前条目没有可执行的步骤，已跳过。", "orange")
                self._sync_subtask_manager_entry_status(entry=entry, status=ROW_STATUS_SKIP, error="当前条目没有可执行的步骤")
                QTimer.singleShot(0, self._run_next_subtask_entry)
                return

            # UI 状态
            self.btn_run.setEnabled(False); self.btn_resume.setEnabled(False); self.btn_stop.setEnabled(True)
            self.btn_run.setText("🧩 子任务中...")
            self.btn_pause.setEnabled(True); self.btn_pause.setText("⏸️ 暂停"); self.btn_pause.setStyleSheet("background-color: #ff9800; color: white; font-weight: bold;")
            self._show_osd_for_subtask(entry=entry)

            label_task = str(entry.get("task_display", "") or src_task_id)
            label_row = row_idx + 1
            extra_desc = []
            if selected_parent_actions:
                extra_desc.append(f"前置{len(selected_parent_actions)}步")
            if subtask_actions:
                extra_desc.append(f"子任务{len(subtask_actions)}步")
            if getattr(self, "_subtask_continue_with_parent", False):
                extra_desc.append(f"继续父任务{len(src_task_actions)}步")
            extra_text = " | ".join(extra_desc) if extra_desc else "无额外说明"
            self._log(f"🧩 子任务执行：{label_task} | 第 {label_row} 行 | {extra_text}", "blue")

            self._engine = AutoEngine(
                actions, [row_data], 0, 0, 0, loops=1, start_l=0,
                retry_count=self.retry_spin.value(),
                on_error="fail_row",
                ignore_data=False,
                standardize_window=self.chk_std_win.isChecked()
            )
            subtask_name = self._get_task_display_text(self._subtask_task_id, with_folder=True) if getattr(self, "_subtask_task_id", "") else "前置步骤"
            self._engine._task_name = f"[子任务] {subtask_name}"
            self._engine._task_id = self._subtask_task_id or src_task_id
            self._engine.log_sig.connect(self._log)
            self._engine.prog_sig.connect(self._update_subtask_progress)
            self._engine.done_sig.connect(self._on_subtask_done)
            self._engine.row_status_sig.connect(lambda _row_idx, status, current_entry=entry: self._on_subtask_row_status(status, entry=current_entry))
            self._engine.row_result_sig.connect(lambda payload, current_entry=entry: self._on_subtask_row_result(payload, entry=current_entry))
            self._engine.highlight_sig.connect(
                lambda row_idx, step_idx, current_entry=entry, parent_rows=parent_step_indices, inserted_auto=inserted_auto_activate:
                    self._on_subtask_engine_highlight(current_entry, parent_rows, inserted_auto, row_idx, step_idx)
            )
            self._engine.detail_sig.connect(self.osd.update_detail)
            self._engine.start()
        except Exception as e:
            self._log(f"❌ 子任务异常: {e}", "red")
            self._sync_subtask_manager_entry_status(
                entry=getattr(self, "_current_subtask_entry", None),
                status=ROW_STATUS_FAIL,
                error=str(e)
            )
            QTimer.singleShot(0, self._run_next_subtask_entry)

    def _on_subtask_done(self):
        """单个子任务执行结束，继续下一个。"""
        try:
            final_status = getattr(getattr(self, '_engine', None), '_final_status', 'done')
            last_error = getattr(getattr(self, '_engine', None), '_last_error', '')
            current_entry = getattr(self, "_current_subtask_entry", None)
            if final_status == "stopped":
                # 用户停止：终止剩余队列
                self._subtask_stopped = True
                self._subtask_queue = []
                self._sync_subtask_manager_entry_status(entry=current_entry, status=ROW_STATUS_MANUAL, error="用户停止")
                if isinstance(current_entry, dict):
                    self._update_task_row_status_cache(
                        str(current_entry.get("task_id") or ""),
                        current_entry.get("row_index", -1),
                        ROW_STATUS_MANUAL
                    )
                QTimer.singleShot(50, self._run_next_subtask_entry)
                return
            if final_status == "failed":
                self._sync_subtask_manager_entry_status(entry=current_entry, status=ROW_STATUS_FAIL, error=last_error)
                if isinstance(current_entry, dict):
                    self._update_task_row_status_cache(
                        str(current_entry.get("task_id") or ""),
                        current_entry.get("row_index", -1),
                        ROW_STATUS_FAIL
                    )
                self._log(f"⚠️ 子任务执行失败: {last_error}", "orange")
            else:
                row_status = str(getattr(self, "_subtask_last_row_status", "") or "")
                if not row_status:
                    payload = getattr(self, "_subtask_last_row_result", None)
                    if isinstance(payload, dict):
                        row_status = str(payload.get("status", "") or "")
                if row_status not in {ROW_STATUS_OK, ROW_STATUS_FAIL, ROW_STATUS_SKIP, ROW_STATUS_DEFER, ROW_STATUS_MANUAL}:
                    row_status = ROW_STATUS_OK
                self._sync_subtask_manager_entry_status(entry=current_entry, status=row_status, error="")
                if isinstance(current_entry, dict):
                    self._update_task_row_status_cache(
                        str(current_entry.get("task_id") or ""),
                        current_entry.get("row_index", -1),
                        row_status
                    )
            total_count = max(int(getattr(self, "_subtask_total_count", 0) or 0), 1)
            self._subtask_completed_count = min(total_count, int(getattr(self, "_subtask_completed_count", 0) or 0) + 1)
            self.progress.setValue(int(round(self._subtask_completed_count * 100 / total_count)))
            self.btn_run.setText("🧩 子任务中...")
            QTimer.singleShot(120, self._run_next_subtask_entry)
        except Exception:
            QTimer.singleShot(120, self._run_next_subtask_entry)

    def _show_osd_for_subtask(self, entry=None):
        if not getattr(self, "chk_show_osd", None) or not self.chk_show_osd.isChecked():
            self.osd.hide()
            return

        current_entry = entry if isinstance(entry, dict) else getattr(self, "_current_subtask_entry", None)
        if isinstance(current_entry, dict):
            task_label = str(current_entry.get("task_display", "") or current_entry.get("task_id", "") or self.current_task)
            row_label = int(current_entry.get("row_index", 0) or 0) + 1
            self.osd.lbl_info.setText(f"🧩 准备执行子任务: {task_label} | 第 {row_label} 行")
        else:
            subtask_name = self._get_task_display_text(self._subtask_task_id, with_folder=True) if getattr(self, "_subtask_task_id", "") else "前置步骤"
            total_count = int(getattr(self, "_subtask_total_count", 0) or 0)
            self.osd.lbl_info.setText(f"🧩 准备启动子任务: {subtask_name} | 共 {total_count} 项")

        self.osd.bar.setValue(0)
        self.osd.lbl_pct.setText("0%")
        self.osd.lbl_detail.setText("准备就绪")
        self.osd.show()
        self.osd.raise_()
        if sys.platform == 'win32':
            self.osd._force_topmost(-1)

    def _update_osd_subtask(self, percent):
        if getattr(self, "_ui_stop_reset_pending", False):
            return
        if not self.chk_show_osd.isChecked():
            self.osd.hide()
            return

        if hasattr(self, '_engine') and self._engine.isRunning():
            e = self._engine
            cur_act = e.actions[e._cur_s] if e._cur_s < len(e.actions) else {}
            current_entry = getattr(self, "_current_subtask_entry", None) or {}
            total_groups = max(int(getattr(self, "_subtask_total_count", 0) or 0), 1)
            pending_groups = len(getattr(self, "_subtask_queue", []) or [])
            current_group = min(max(total_groups - pending_groups, 1), total_groups)
            task_label = str(current_entry.get("task_display", "") or current_entry.get("task_id", "") or self.current_task)
            row_label = int(current_entry.get("row_index", 0) or 0) + 1
            self.osd.update_progress(
                f"[子任务] {task_label} | 第 {row_label} 行",
                e._cur_l,
                current_group,
                total_groups,
                e._cur_s + 1,
                len(e.actions),
                cur_act.get('name', 'Step'),
                percent
            )

    def _schedule_config_flush(self, delay_ms=260):
        """合并短时间内的频繁保存请求，减少输入时卡顿。"""
        if hasattr(self, "_config_flush_timer"):
            self._config_flush_timer.start(max(0, int(delay_ms)))
        else:
            save_config(self.config)

    def _flush_config_now(self):
        if hasattr(self, "_config_flush_timer") and self._config_flush_timer.isActive():
            self._config_flush_timer.stop()
        save_config(self.config)

    def _is_compact_data_action(self, act_type):
        return act_type in ["click", "double_click", "right_click", "move", "hover_click", "scroll", "wait", "screenshot"]

    def _format_compact_data_value(self, action, row_dict=None, force_sync=False):
        """批量数据中的只读紧凑展示，既保留流程信息，又避免表格过宽。"""
        row_dict = row_dict or {}
        raw_action = action.get('action', '')
        act_type = CMD_MAP.get(raw_action, "click")

        if act_type in ["click", "double_click", "right_click", "move", "hover_click"]:
            return f"{raw_action} @{action.get('x', 0)},{action.get('y', 0)}"
        if act_type == "scroll":
            scroll_val = row_dict.get(action.get('name', 'Step'), action.get('value', ''))
            scroll_val = scroll_val if str(scroll_val).strip() else action.get('value', 0)
            return f"滚动 {scroll_val}"
        if act_type == "wait":
            delay_val = row_dict.get(f"{action.get('name', 'Step')}_延时", "")
            if str(delay_val).strip() == "":
                delay_val = action.get('delay', action.get('value', 1))
            return f"等待 {delay_val}秒"
        if act_type == "screenshot":
            save_to = row_dict.get(action.get('name', 'Step'), action.get('value', '')) if not force_sync else action.get('value', '')
            return f"截图 {os.path.basename(str(save_to))}" if str(save_to).strip() else "截图"
        return str(row_dict.get(action.get('name', 'Step'), action.get('value', '')))

    def _get_data_column_width(self, action=None, is_delay=False):
        if is_delay:
            return 62
        if not action:
            return 120

        act_type = CMD_MAP.get(action.get('action'), "click")
        if self._is_compact_data_action(act_type):
            return 116
        if act_type == "open_url":
            return 240
        if act_type in ["input", "clear_input", "clear_input_plus", "cmd", "if_image", "if_win", "defer"]:
            return 220
        if act_type in ["upload", "drag_file", "run_app"]:
            return 180
        if act_type in ["press", "hotkey", "win_active"]:
            return 160
        return 180

    def _apply_data_table_column_widths(self, acts, show_delay):
        """只恢复用户已调过的列宽；新列才使用默认宽度，不再自动适应内容。"""
        self._restore_column_widths()

    def _apply_data_table_column_visibility(self, acts, show_delay):
        """按开关隐藏不可编辑步骤，但不改变底层列顺序，避免数据错位。"""
        show_noneditable = self.btn_toggle_noneditable.isChecked() if hasattr(self, "btn_toggle_noneditable") else True
        col_idx = self._data_first_value_col()
        for action in acts:
            act_type = CMD_MAP.get(action.get('action'), "click")
            hide_this = self._is_compact_data_action(act_type) and not show_noneditable
            self.data_table.setColumnHidden(col_idx, hide_this)
            col_idx += 1
            if show_delay:
                self.data_table.setColumnHidden(col_idx, hide_this)
                col_idx += 1

    def _refresh_data_table(self, force_sync=False):
        """[深度优化] 采用分时渲染策略，先快速渲染文本框架，再通过定时器异步加载重量级控件。"""
        if not self.current_task: return
        
        # 取消之前的渲染任务，防止重叠
        if hasattr(self, '_render_timer') and self._render_timer.isActive():
            self._render_timer.stop()
            
        acts = self.config['tasks'].get(self.current_task, []); show_delay = self.btn_toggle_delay.isChecked()
        headers = ["选择", "执行", "状态"]
        for idx, a in enumerate(acts, 1):
            name = a.get('name', 'Step')
            headers.append(f"{idx}.{name}")
            if show_delay: headers.append(f"{idx}.{name}_延时")
        if len(headers) == self._data_first_value_col(): headers = ["选择", "执行", "状态", "(等待添加步骤)"]
        
        self.data_table.blockSignals(True)
        self.data_table.clearContents()
        old_data = self.config['task_data'].get(self.current_task, [])
        self.data_table.setColumnCount(len(headers))
        self.data_table.setHorizontalHeaderLabels(headers)
        default_delegate = QStyledItemDelegate(self.data_table)
        col_idx = self._data_first_value_col()
        for a in acts:
            act_type = CMD_MAP.get(a.get('action'), "click")
            if act_type in ["input", "clear_input"]:
                self.data_table.setItemDelegateForColumn(col_idx, MultiLineTextDelegate(self.data_table))
            else:
                self.data_table.setItemDelegateForColumn(col_idx, default_delegate)
            col_idx += 1
            if show_delay:
                self.data_table.setItemDelegateForColumn(col_idx, default_delegate)
                col_idx += 1
        rows_to_show = max(1, len(old_data))
        self.data_table.setRowCount(rows_to_show)
        
        # 第一阶段：极速渲染纯文本内容，让用户立即看到数据
        statuses = self._row_statuses.get(self.current_task, {})
        for r in range(rows_to_show):
            row_dict = old_data[r] if r < len(old_data) else {}
            s_item = QTableWidgetItem(statuses.get(r, ""))
            s_item.setFlags(s_item.flags() & ~Qt.ItemIsEditable); s_item.setTextAlignment(Qt.AlignCenter)
            self._apply_row_status_style(s_item, statuses.get(r, ""))
            self.data_table.setItem(r, self._data_status_col(), s_item)
            chk_item = QTableWidgetItem()
            chk_item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
            row_has_data = self._row_has_meaningful_data(row_dict, acts)
            is_checked = bool(row_dict.get("_选中", True)) if row_has_data else False
            chk_item.setCheckState(Qt.Checked if is_checked else Qt.Unchecked)
            self.data_table.setItem(r, self._data_select_col(), chk_item)

            self.data_table.removeCellWidget(r, self._data_run_col())
            # 用图标代替文字，避免按钮列较窄时文字不可见
            btn_run_row = QPushButton("")
            try:
                btn_run_row.setIcon(self.style().standardIcon(QStyle.SP_MediaPlay))
                btn_run_row.setIconSize(QSize(14, 14))
            except Exception:
                pass
            btn_run_row.setToolTip("单独执行这一行任务，不依赖“选择”勾选。")
            btn_run_row.setProperty("data_row", r)
            btn_run_row.setFixedHeight(max(24, self.config.get("layout", {}).get("data_row_height", 28) - 4))
            btn_run_row.setFixedWidth(30)
            btn_run_row.clicked.connect(lambda _checked=False, data_row=r: self._run_single_data_row(data_row))
            btn_run_row.setEnabled(r < len(old_data) and bool(acts))
            self.data_table.setCellWidget(r, self._data_run_col(), btn_run_row)

            col_idx = self._data_first_value_col()
            for a in acts:
                self.data_table.removeCellWidget(r, col_idx) # 清理旧控件
                name = a.get('name', 'Step'); raw_act_type = a.get('action'); act_type = CMD_MAP.get(raw_act_type, "click")
                if self._is_compact_data_action(act_type):
                    val = self._format_compact_data_value(a, row_dict, force_sync)
                else:
                    val = a.get('value', '') if force_sync else row_dict.get(name, a.get('value', ''))
                
                item = QTableWidgetItem(str(val))
                item.setToolTip(str(val))
                if self._is_compact_data_action(act_type):
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    item.setBackground(QColor("#f6f8fb"))
                    item.setForeground(QColor("#546e7a"))
                if row_dict.get(f"{name}_跳过", False):
                    item.setBackground(QColor("#eeeeee")); item.setForeground(QColor("#999999"))
                self.data_table.setItem(r, col_idx, item)
                col_idx += 1
                if show_delay:
                    d_item = QTableWidgetItem(str(row_dict.get(f"{name}_延时", "")))
                    d_item.setBackground(QColor("#fff3e0"))
                    self.data_table.setItem(r, col_idx, d_item)
                    col_idx += 1
            self.data_table.setRowHeight(r, self.config.get("layout", {}).get("data_row_height", 28))
                    
        self.data_table.blockSignals(False)
        self._update_data_select_header()
        
        # 第二阶段：启动异步渲染定时器，分批生成重量级控件
        self._render_row_idx = 0
        self._render_timer = QTimer(self)
        self._render_timer.timeout.connect(lambda: self._async_render_step(acts, old_data, show_delay, force_sync))
        self._render_timer.start(5) # 5ms 间隔，利用主线程空闲时间渲染

    def _async_render_step(self, acts, old_data, show_delay, force_sync=False):
        """异步渲染每一行的复杂控件。"""
        if self._render_row_idx >= self.data_table.rowCount():
            self._render_timer.stop()
            self._apply_data_table_column_widths(acts, show_delay)
            self._apply_data_table_column_visibility(acts, show_delay)
            self._set_data_row_height(self.config.get("layout", {}).get("data_row_height", 28), save=False)
            return
            
        r = self._render_row_idx
        row_dict = old_data[r] if r < len(old_data) else {}
        col_idx = self._data_first_value_col()
        
        for a in acts:
            raw_act_type = a.get('action'); act_type = CMD_MAP.get(raw_act_type, "click")
            name = a.get('name', 'Step'); val = row_dict.get(name, a.get('value', ''))
            
            if act_type == "win_active":
                _raw = str(val); _disp = _raw.split('::hwnd=')[0] if '::hwnd=' in _raw else _raw
                w = QWidget(); l = QHBoxLayout(w); l.setContentsMargins(0,0,0,0)
                le = QLineEdit(_disp); le.setStyleSheet("background-color: #e3f2fd; border: none;")
                le.editingFinished.connect(lambda r=r, c=col_idx: self._on_cell_widget_changed(r, c, self.sender().text()))
                btn = QPushButton("🔍"); btn.setFixedWidth(25); btn.setStyleSheet("border: none; background: #d1e9ff;")
                btn.clicked.connect(lambda chk, le=le, tr=r, tc=col_idx: [le.setText(s.split('::hwnd=')[0]), self._on_cell_widget_changed(tr, tc, s)] if (s := WindowSelector(self).get_selection()) else None)
                l.addWidget(le); l.addWidget(btn); self.data_table.setCellWidget(r, col_idx, w)
                col_idx += 1
                if show_delay:
                    col_idx += 1
            elif act_type == "open_url":
                w = QWidget(); l = QHBoxLayout(w); l.setContentsMargins(0,0,0,0)
                u_p = str(val).split('|')[0]; p_p = str(val).split('|')[1] if '|' in str(val) else ""
                le = QLineEdit(u_p); le.setStyleSheet("background-color: #e3f2fd; border: none;")
                le.editingFinished.connect(lambda r=r, c=col_idx: self._on_url_cell_changed(r, c))
                btn = QPushButton(get_profile_display_name(p_p)); btn.setProperty("profile_id", p_p)
                btn.setFixedWidth(92)
                btn.setStyleSheet("border: none; background: #d1e9ff; text-align: left; padding-left: 5px;")
                le.setToolTip(str(val))
                if isinstance(p_p, str) and "::hwnd=" in p_p:
                    btn.setToolTip(f"当前目标: 已打开窗口\n{p_p}\n\n点击可切换账号或选择已打开窗口")
                else:
                    btn.setToolTip(f"当前目标: {get_profile_display_name(p_p)}\n\n点击可切换账号或选择已打开窗口")
                btn.clicked.connect(lambda chk, r=r, c=col_idx, b=btn: self._pick_profile_for_data_cell(r, c, b))
                l.addWidget(le, 2); l.addWidget(btn, 1); self.data_table.setCellWidget(r, col_idx, w)
                col_idx += 1
                if show_delay:
                    col_idx += 1
            elif act_type == "clear_input_plus":
                w = QWidget(); l = QHBoxLayout(w); l.setContentsMargins(0,0,0,0); l.setSpacing(0)
                parts = str(val).split('|', 1); pre = parts[0] if len(parts)>0 else ""; con = parts[1] if len(parts)>1 else ""
                lbl = QLabel(f" {pre}"); lbl.setStyleSheet("background: #f5f5f5; color: #666; border-right: 1px solid #ccc; font-size: 11px;")
                le = MultiLineTextEdit(); le.setText(con); le.setStyleSheet("background: #fffde7; border: none;"); le.setFixedHeight(30)
                le.setToolTip(str(val))
                le.editingFinished.connect(lambda r=r, c=col_idx, p=pre, target=le: self._on_cell_widget_changed(r, c, f"{p}|{target.text()}"))
                if pre: l.addWidget(lbl, 1)
                l.addWidget(le, 3); self.data_table.setCellWidget(r, col_idx, w)
                self.data_table.setRowHeight(r, self.config.get("layout", {}).get("data_row_height", 28))
                col_idx += 1
                if show_delay:
                    col_idx += 1
            elif act_type in ["if_image", "if_win"]:
                # 使用文本展示，仅在需要时点击触发编辑器（或者保持文本以加速）
                col_idx += 1
                if show_delay:
                    col_idx += 1
            elif act_type in ["press", "hotkey"]:
                kr = KeyRecorder(); kr.setText(str(val))
                kr.setStyleSheet("background-color: #fffde7; border: none; color: #f57f17; font-weight: bold;")
                kr.setToolTip(str(val))
                kr.key_recorded.connect(lambda v, row=r, col=col_idx: self._on_cell_widget_changed(row, col, v))
                kr.setContextMenuPolicy(Qt.CustomContextMenu)
                kr.customContextMenuRequested.connect(lambda pos, w=kr, tr=r, tc=col_idx: self._forward_widget_context_menu(w, pos, tr, tc))
                self.data_table.setCellWidget(r, col_idx, kr)
                col_idx += 1
                if show_delay:
                    default_delay = int(a.get('delay', 1))
                    saved_val = row_dict.get(f"{name}_延时", "")
                    # Show saved value if set, otherwise show the step's default delay
                    display_val = str(saved_val) if str(saved_val).strip() != "" else str(default_delay)
                    d_item = QTableWidgetItem(display_val)
                    if str(saved_val).strip() != "" and str(saved_val).strip() != str(default_delay):
                        # Manually overridden — orange highlight
                        d_item.setBackground(QColor("#ffe0b2"))
                        d_item.setForeground(QColor("#e65100"))
                        d_item.setToolTip(f"⚠️ 已自定义: {saved_val}秒（步骤默认: {default_delay}秒）")
                    else:
                        # Showing default — subtle yellow, grey text
                        d_item.setBackground(QColor("#fff9c4"))
                        d_item.setForeground(QColor("#999999"))
                        d_item.setToolTip(f"当前使用步骤默认延时: {default_delay}秒，可填入数字覆盖")
                    self.data_table.setItem(r, col_idx, d_item); col_idx += 1
            else:
                col_idx += 1
                if show_delay:
                    col_idx += 1
        
        self.data_table.blockSignals(False)
        if force_sync: self._save_data_table(); self._log("✅ 已根据流程同步表头和默认数据", "green")
        
        self._render_row_idx += 1

    def _save_data_table(self, flush=False):
        if not self.current_task or self.data_table.signalsBlocked(): return

        rs = self.data_table.rowCount(); cs = self.data_table.columnCount()
        if rs == 0: return
        acts = self.config['tasks'].get(self.current_task, [])
        old_rows = self.config.get('task_data', {}).get(self.current_task, []) or []
        show_delay = self.btn_toggle_delay.isChecked()
        hs = [self.data_table.horizontalHeaderItem(i).text() for i in range(cs)]
        nd = []
        for r in range(rs):
            row_dict = {}
            old_row_dict = old_rows[r] if r < len(old_rows) and isinstance(old_rows[r], dict) else {}
            for k, v in old_row_dict.items():
                if isinstance(k, str) and k.startswith("_") and k != "_选中":
                    row_dict[k] = v
            col_idx = self._data_first_value_col()  # skip select/execute/status columns
            chk_item = self.data_table.item(r, self._data_select_col())
            row_dict["_选中"] = (chk_item.checkState() == Qt.Checked) if chk_item else True
            for a in acts:
                raw_act_type = a.get('action')
                
                act_type = CMD_MAP.get(raw_act_type, "click")
                is_coord_only = act_type in ["click", "double_click", "right_click", "move", "hover_click", "scroll"]
                widget = self.data_table.cellWidget(r, col_idx)
                w = self.data_table._resolve_widget(widget)
                if isinstance(w, QComboBox):
                    val = w.currentText()
                elif isinstance(w, KeyRecorder):
                    val = w.text()
                elif act_type == "open_url" and widget is not None:
                    # 核心修复：直接从 UI 控件抓取最新值，不依赖可能过时的 backing_item
                    le = _find_text_input(widget)
                    btn = widget.findChild(QPushButton)
                    url_text = le.text() if le else ""
                    prof_id = btn.property("profile_id") or "" if btn else ""
                    if str(url_text).strip() == "[SKIP_ROW]" or str(prof_id).strip() == "[SKIP_ROW]":
                        val = "[SKIP_ROW]"
                    else:
                        val = f"{url_text}|{prof_id}"
                    # 同时更新一下 backing_item 保持一致
                    bk = self.data_table.item(r, col_idx)
                    if bk: bk.setText(val)
                elif act_type == "clear_input_plus" and widget is not None:
                    # 核心修复：直接从 UI 控件抓取最新值
                    le = _find_text_input(widget)
                    lbl = widget.findChild(QLabel)
                    prefix = lbl.text().strip() if lbl else ""
                    content = le.text() if le else ""
                    if str(prefix).strip() == "[SKIP_ROW]" or str(content).strip() == "[SKIP_ROW]":
                        val = "[SKIP_ROW]"
                    else:
                        val = f"{prefix}|{content}"
                    bk = self.data_table.item(r, col_idx)
                    if bk: bk.setText(val)
                else:
                    item = self.data_table.item(r, col_idx)
                    val = item.text() if item else ""
                
                if not is_coord_only:
                    row_dict[a.get('name', f'步骤{col_idx-1}')] = val
                col_idx += 1
                if show_delay:
                    d_item = self.data_table.item(r, col_idx)
                    row_dict[f"{a.get('name', f'步骤{col_idx-1}')}_延时"] = d_item.text() if d_item else ""
                    col_idx += 1
            nd.append(row_dict)
        self.config['task_data'][self.current_task] = nd
        if flush:
            self._flush_config_now()
        else:
            self._schedule_config_flush()

    def _row_has_meaningful_data(self, row_dict, actions):
        if not isinstance(row_dict, dict):
            return False
        for action in actions:
            act_type = CMD_MAP.get(action.get('action'), "click")
            if act_type in ["click", "double_click", "right_click", "move", "hover_click", "scroll", "wait", "screenshot"]:
                continue
            step_name = action.get('name', '')
            raw_val = row_dict.get(step_name, "")
            s = str(raw_val or "").strip()
            default_raw = action.get('value', '')
            default_s = str(default_raw or "").strip()

            delay_key = f"{step_name}_延时"
            if str(row_dict.get(delay_key, "") or "").strip():
                return True

            skip_key = f"{step_name}_跳过"
            if bool(row_dict.get(skip_key, False)):
                return True

            if act_type == "clear_input_plus":
                if s == "[SKIP_ROW]":
                    return True
                parts = s.split("|", 1)
                prefix = parts[0].strip() if len(parts) > 0 else ""
                content = parts[1].strip() if len(parts) > 1 else ""
                def_parts = default_s.split("|", 1)
                def_prefix = def_parts[0].strip() if len(def_parts) > 0 else ""
                def_content = def_parts[1].strip() if len(def_parts) > 1 else ""
                if (prefix or content) and (prefix != def_prefix or content != def_content):
                    return True
            elif act_type == "open_url":
                if s == "[SKIP_ROW]":
                    return True
                parts = s.split("|", 1)
                url_part = parts[0].strip() if len(parts) > 0 else ""
                profile_part = parts[1].strip() if len(parts) > 1 else ""
                def_parts = default_s.split("|", 1)
                def_url_part = def_parts[0].strip() if len(def_parts) > 0 else ""
                def_profile_part = def_parts[1].strip() if len(def_parts) > 1 else ""
                if (url_part or profile_part) and (url_part != def_url_part or profile_part != def_profile_part):
                    return True
            else:
                if s and s != default_s:
                    return True
        return False

    def _is_data_row_selectable(self, row_index, actions=None, data_rows=None):
        if not self.current_task:
            return False
        if actions is None:
            actions = self.config.get('tasks', {}).get(self.current_task, [])
        if data_rows is None:
            data_rows = self.config.get('task_data', {}).get(self.current_task, [])
        if row_index < 0 or row_index >= len(data_rows):
            return False
        row_dict = data_rows[row_index] if isinstance(data_rows[row_index], dict) else {}
        return self._row_has_meaningful_data(row_dict, actions)

    def _auto_check_blank_rows(self, rows=None):
        if not self.current_task:
            return
        actions = self.config.get('tasks', {}).get(self.current_task, [])
        data_rows = self.config.get('task_data', {}).get(self.current_task, [])
        if not data_rows:
            return

        if rows is None:
            target_rows = range(len(data_rows))
        else:
            target_rows = []
            for row in rows:
                try:
                    row = int(row)
                except Exception:
                    continue
                if 0 <= row < len(data_rows):
                    target_rows.append(row)

        changed = False
        self.data_table.blockSignals(True)
        try:
            for row in target_rows:
                row_dict = data_rows[row] if row < len(data_rows) else {}
                should_check = self._row_has_meaningful_data(row_dict, actions)
                old_checked = bool(row_dict.get("_选中", False))
                row_dict["_选中"] = should_check
                if row < self.data_table.rowCount():
                    chk_item = self.data_table.item(row, self._data_select_col())
                    target_state = Qt.Checked if should_check else Qt.Unchecked
                    if chk_item and chk_item.checkState() != target_state:
                        chk_item.setCheckState(target_state)
                        changed = True
                    elif chk_item is None:
                        chk_item = QTableWidgetItem()
                        chk_item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
                        chk_item.setCheckState(target_state)
                        self.data_table.setItem(row, self._data_select_col(), chk_item)
                        changed = True
                if old_checked != should_check:
                    changed = True
        finally:
            self.data_table.blockSignals(False)

        if changed:
            self._update_data_select_header()
            self._save_data_table(flush=True)

    def _on_cell_widget_changed(self, row, col, val):
        item = self.data_table.item(row, col)
        if item:
            item.setText(val)
            self._save_data_table()

    def _on_url_cell_changed(self, row, col):
        w = self.data_table.cellWidget(row, col); le = w.findChild(QLineEdit); btn = w.findChild(QPushButton)
        val = f"{le.text()}|{btn.property('profile_id') or ''}"; self._on_cell_widget_changed(row, col, val)

    def _pick_profile_for_data_cell(self, row, col, btn):
        menu = QMenu(self)
        act_acc = menu.addAction("选择账号…")
        act_win = menu.addAction("选择已打开窗口…")
        chosen = menu.exec_(btn.mapToGlobal(btn.rect().bottomLeft()))
        if chosen == act_acc:
            sel = ChromeProfileSelector(self); pid = sel.get_selection()
            if pid is not None:
                btn.setProperty("profile_id", pid); btn.setText(get_profile_display_name(pid))
                self._on_url_cell_changed(row, col)
        elif chosen == act_win:
            s = WindowSelector(self).get_selection()
            if s:
                btn.setProperty("profile_id", s); btn.setText(get_profile_display_name(s))
                self._on_url_cell_changed(row, col)

    def _batch_assign_profiles(self):
        """批量处理中心：支持多步骤同时填充、分类过滤等。"""
        if not self.current_task: return
        acts = self.config['tasks'].get(self.current_task, [])
        show_delay = self.btn_toggle_delay.isChecked()
        
        # 1. 识别所有可批量编辑的列
        edit_cols = [] # [(col_idx, step_idx, step_name, act_type), ...]
        col_i = self._data_first_value_col()
        for s_idx, a in enumerate(acts):
            raw_act = a.get('action')
            act_type = CMD_MAP.get(raw_act, '')
            if act_type not in ["click", "double_click", "right_click", "move", "hover_click", "screenshot", "wait"]:
                edit_cols.append((col_i, s_idx, a.get('name', f'步骤{s_idx+1}'), act_type))
            col_i += 1
            if show_delay: col_i += 1
            
        if not edit_cols:
            QMessageBox.information(self, "提示", "当前任务没有可批量填充的步骤。")
            return

        # 2. 选择多个步骤
        dlg_select = QDialog(self)
        dlg_select.setWindowTitle("选择批量填充步骤")
        dlg_select.resize(400, 500)
        sel_ly = QVBoxLayout(dlg_select)
        sel_ly.addWidget(QLabel("请选择一个或多个要同时填充的步骤:"))
        step_list = QListWidget()
        step_list.setSelectionMode(QAbstractItemView.ExtendedSelection)
        for _, _, name, type_ in edit_cols:
            step_list.addItem(f"{name} ({type_})")
        step_list.selectAll()
        sel_ly.addWidget(step_list)
        
        btn_sel_ok = QPushButton("确定选择"); btn_sel_ok.setFixedHeight(40); btn_sel_ok.clicked.connect(dlg_select.accept)
        sel_ly.addWidget(btn_sel_ok)
        if dlg_select.exec_() != QDialog.Accepted: return
        
        selected_indices = [step_list.row(i) for i in step_list.selectedItems()]
        if not selected_indices: return
        
        # 目标列集合
        targets = [edit_cols[i] for i in selected_indices]
        
        selected_rows = sorted(set(i.row() for i in self.data_table.selectedItems()))
        if not selected_rows:
            selected_rows = list(range(self.data_table.rowCount()))
        if not selected_rows: return

        # 4. 弹出通用批量处理对话框
        dlg = QDialog(self)
        title_suffix = targets[0][2] if len(targets) == 1 else f"{len(targets)} 个步骤"
        dlg.setWindowTitle(f"🛠️ 批量填充中心 — {title_suffix}")
        dlg.setWindowFlags(Qt.Window | Qt.WindowContextHelpButtonHint | Qt.WindowMaximizeButtonHint | Qt.WindowMinimizeButtonHint | Qt.WindowCloseButtonHint)
        dlg.resize(1200, 850)
        main_ly = QVBoxLayout(dlg)
        
        main_splitter = QSplitter(Qt.Horizontal)
        # 允许拖拽调整左右区域宽度，并记忆用户的调整
        main_splitter.setHandleWidth(8)
        main_splitter.setChildrenCollapsible(False)
        main_splitter.setStretchFactor(0, 2)
        main_splitter.setStretchFactor(1, 1)
        
        # 左侧：待处理行与多步骤预览
        left_w = QWidget(); left_ly = QVBoxLayout(left_w); left_ly.setContentsMargins(0,0,0,0)
        left_header_v = QVBoxLayout()
        left_header_v.setContentsMargins(5, 5, 5, 5)
        left_header_v.setSpacing(6)

        left_header_row1 = QHBoxLayout()
        left_header_row1.setContentsMargins(0, 0, 0, 0)
        lbl_info = QLabel(f"待处理行（共 {len(selected_rows)} 行）: [💡 支持直接拖入外部文件或右侧工具项]")
        lbl_info.setStyleSheet("font-weight: bold; color: #555;")
        lbl_info.setWordWrap(False)
        lbl_info.setMinimumWidth(80)
        lbl_info.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        left_header_row1.addWidget(lbl_info)
        left_header_row1.addStretch()
        btn_batch_fill_menu = QPushButton("☰ 小菜单")
        btn_batch_fill_menu.setFixedHeight(28)
        btn_batch_fill_menu.setToolTip("打开批量填充中心的小菜单。")
        left_header_row1.addWidget(btn_batch_fill_menu)
        left_header_v.addLayout(left_header_row1)

        # 工具按钮行：放进横向滚动容器，避免按钮过多导致左侧最小宽度过大，从而“拖不动分隔条”
        tools_bar = QWidget()
        tools_h = QHBoxLayout(tools_bar)
        tools_h.setContentsMargins(0, 0, 0, 0)
        tools_h.setSpacing(6)
        
        # [核心修复] 重新设计加行按钮，确保在左侧面板顶部可见
        btn_add_row_dlg = QPushButton("➕ 临时增加行")
        btn_add_row_dlg.setToolTip("在任务末尾添加新数据行，方便在此界面继续填充。")
        btn_add_row_dlg.setMinimumWidth(96)
        btn_add_row_dlg.setFixedHeight(30)
        btn_add_row_dlg.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50; 
                color: white; 
                font-weight: bold; 
                border-radius: 4px;
                padding: 0 10px;
            }
            QPushButton:hover { background-color: #45a049; }
        """)

        def _capture_preview_state():
            try:
                table = preview_table
            except NameError:
                return {}
            state = {}
            for r in range(table.rowCount()):
                for c in range(table.columnCount()):
                    it = table.item(r, c)
                    if not it:
                        continue
                    key = (
                        it.data(Qt.UserRole),
                        it.data(Qt.UserRole + 1),
                        it.data(Qt.UserRole + 2)
                    )
                    state[key] = {
                        "text": it.text(),
                        "bg": it.background().color().name() if it.background().style() != Qt.NoBrush else None,
                        "role10": it.data(Qt.UserRole + 10)
                    }
            return state
        
        def _add_row_in_dlg():
            num, ok = QInputDialog.getInt(dlg, "添加数据行", "请输入要添加的行数:", 1, 1, 100)
            if not ok: return
            preview_state = _capture_preview_state()
            
            # [核心实装] 1. 在主配置中添加数据，并同步主表格 UI
            old_data = self.config['task_data'].get(self.current_task, [])
            start_idx = len(old_data)
            for _ in range(num):
                old_data.append({})
            self.config['task_data'][self.current_task] = old_data
            
            # 同步刷新主界面表格，确保“实装”
            self._refresh_data_table()
            self._save_data_table()
            
            # 2. 更新填充中心预览表
            for i in range(num):
                curr_real_row = start_idx + i
                # 将新行加入处理队列
                if curr_real_row not in selected_rows:
                    selected_rows.append(curr_real_row)
            _reload_preview_from_main(f"已实装新增 {num} 行", preserved_state=preview_state)
            self._auto_check_blank_rows(selected_rows)
            
        btn_add_row_dlg.clicked.connect(_add_row_in_dlg)
        tools_h.addWidget(btn_add_row_dlg)
        btn_paste_clipboard_dlg = QPushButton("📋 直贴表格")
        btn_paste_clipboard_dlg.setToolTip("把剪贴板里的表格内容直接贴到左侧当前选区；也支持在左侧预览表中按 Ctrl+V。")
        btn_paste_clipboard_dlg.setMinimumWidth(96)
        btn_paste_clipboard_dlg.setFixedHeight(30)
        btn_paste_clipboard_dlg.setStyleSheet("""
            QPushButton {
                background-color: #1e88e5;
                color: white;
                font-weight: bold;
                border-radius: 4px;
                padding: 0 10px;
            }
            QPushButton:hover { background-color: #1976d2; }
        """)
        tools_h.addWidget(btn_paste_clipboard_dlg)
        btn_sync_headers_dlg = QPushButton("🔄 同步表头")
        btn_sync_headers_dlg.setToolTip("在批量填充中心内直接同步批量数据列结构，并刷新左侧预览。")
        btn_sync_headers_dlg.setMinimumWidth(96)
        btn_sync_headers_dlg.setFixedHeight(30)
        tools_h.addWidget(btn_sync_headers_dlg)
        btn_reset_presets_dlg = QPushButton("🧹 重置预设")
        btn_reset_presets_dlg.setToolTip("在批量填充中心内直接按流程默认值重置批量数据，并刷新左侧预览。")
        btn_reset_presets_dlg.setMinimumWidth(96)
        btn_reset_presets_dlg.setFixedHeight(30)
        btn_reset_presets_dlg.setStyleSheet("background-color: #ffebee; border: 1px solid #ef9a9a;")
        tools_h.addWidget(btn_reset_presets_dlg)
        btn_del_row_dlg = QPushButton("🗑️ 删除选中行")
        btn_del_row_dlg.setToolTip("删除左侧预览表中选中的数据行，并同步到主表格。")
        btn_del_row_dlg.setMinimumWidth(96)
        btn_del_row_dlg.setFixedHeight(30)
        btn_del_row_dlg.setStyleSheet("background-color: #fff3e0; border: 1px solid #ffcc80;")
        tools_h.addWidget(btn_del_row_dlg)
        tools_h.addSpacing(8)
        tools_h.addWidget(QLabel("整体行高:"))
        preview_row_height_spin = QSpinBox()
        preview_row_height_spin.setRange(28, 160)
        preview_row_height_spin.setSuffix(" px")
        preview_row_height_spin.setToolTip("统一调整批量填充中心左侧所有行的高度。")
        tools_h.addWidget(preview_row_height_spin)
        tools_h.addStretch()

        tools_scroll = QScrollArea()
        tools_scroll.setFrameShape(QFrame.NoFrame)
        tools_scroll.setWidgetResizable(True)
        tools_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        tools_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        tools_scroll.setFixedHeight(38)
        tools_scroll.setMinimumWidth(0)
        tools_scroll.setWidget(tools_bar)
        left_header_v.addWidget(tools_scroll)

        _toolbar_auto_state = {"collapsed": None}

        def _apply_toolbar_collapsed(collapsed):
            tools_scroll.setVisible(not collapsed)
            btn_batch_fill_menu.setVisible(bool(collapsed))
            _toolbar_auto_state["collapsed"] = bool(collapsed)

        def _update_toolbar_auto_visibility(force=False):
            panel_width = left_w.width()
            collapsed = _toolbar_auto_state["collapsed"]
            if force or collapsed is None:
                should_collapse = panel_width < 760
            elif collapsed:
                should_collapse = panel_width < 860
            else:
                should_collapse = panel_width < 720
            if force or should_collapse != collapsed:
                _apply_toolbar_collapsed(should_collapse)

        _apply_toolbar_collapsed(False)

        left_ly.addLayout(left_header_v)
        
        # 定义支持拖拽的内部表格类
        class DragDropTable(QTableWidget):
            def __init__(self, r, c, parent_dlg):
                super().__init__(r, c, parent_dlg)
                self.setAcceptDrops(True)
            def dragEnterEvent(self, event):
                if event.mimeData().hasUrls() or event.mimeData().hasFormat("application/x-qabstractitemmodeldatalist"):
                    event.acceptProposedAction()
            def dragMoveEvent(self, event):
                event.acceptProposedAction()
            def dropEvent(self, event):
                pos = event.pos()
                row = self.rowAt(pos.y())
                col = self.columnAt(pos.x())
                if row == -1 or col == -1: return
                
                paths = []
                # 处理外部拖入的文件
                if event.mimeData().hasUrls():
                    for url in event.mimeData().urls():
                        local_path = url.toLocalFile()
                        if local_path: paths.append(os.path.normpath(local_path))
                # 处理内部拖拽 (例如从 file_tree 或 profile_list 拖过来)
                elif event.mimeData().hasFormat("application/x-qabstractitemmodeldatalist"):
                    # 如果是文件树拖拽
                    if file_tree.selectedItems():
                        for it in file_tree.selectedItems():
                            p = it.data(0, Qt.UserRole)
                            if p and not os.path.isdir(p): paths.append(p)
                    # 如果是账号列表拖拽
                    elif profile_list.selectedItems():
                        for it in profile_list.selectedItems():
                            pid = it.data(Qt.UserRole)
                            name = it.text().replace("🟢 ","").replace("⚪ ","")
                            # 账号填充需要特殊处理 val|pid
                            paths.append((name, f"|{pid}"))

                if paths:
                    # 从当前行开始向下填充
                    for i, p_info in enumerate(paths):
                        if row + i < self.rowCount():
                            item = self.item(row + i, col)
                            if item:
                                if isinstance(p_info, tuple): # 账号情况
                                    item.setText(p_info[0])
                                    item.setData(Qt.UserRole + 10, p_info[1])
                                else: # 文件路径情况
                                    item.setText(p_info)
                                item.setBackground(QColor("#fff9c4"))
                    event.acceptProposedAction()

        # --- 核心改进：预览表列拆分逻辑 ---
        preview_cols = [] # [(real_col_idx, step_idx, step_name, act_type, sub_type)] sub_type: "prefix" or "content" or "url" or "profile" or None
        for col_idx, s_idx, name, act_type in targets:
            if act_type == "clear_input_plus":
                preview_cols.append((col_idx, s_idx, name, act_type, "prefix"))
                preview_cols.append((col_idx, s_idx, name, act_type, "content"))
            elif act_type == "open_url":
                preview_cols.append((col_idx, s_idx, name, act_type, "url"))
                preview_cols.append((col_idx, s_idx, name, act_type, "profile"))
            else:
                preview_cols.append((col_idx, s_idx, name, act_type, None))

        # 核心数据预览表：显示选中的行和选中的步骤列
        preview_table = DragDropTable(len(selected_rows), len(preview_cols), dlg)
        
        headers = []
        for _, _, name, act_type, sub in preview_cols:
            if sub == "prefix": headers.append(f"{name}\n(前缀)")
            elif sub == "content": headers.append(f"{name}\n(内容)")
            elif sub == "url": headers.append(f"{name}\n(网址)")
            elif sub == "profile": headers.append(f"{name}\n(账号)")
            else: headers.append(f"{name}\n({act_type})")
        
        preview_table.setHorizontalHeaderLabels(headers)
        preview_table.setVerticalHeaderLabels([f"第{r+1}行" for r in selected_rows])
        preview_table.setSelectionBehavior(QAbstractItemView.SelectItems) # 恢复单元格选中模式
        preview_table.setSelectionMode(QAbstractItemView.ExtendedSelection) # 支持 Ctrl/Shift 多选单元格
        preview_table.setWordWrap(False) # 保持统一行高，不随内容自动撑高
        preview_table.verticalHeader().setMinimumSectionSize(28)
        preview_table.verticalHeader().setMaximumSectionSize(160)
        preview_table.verticalHeader().setSectionResizeMode(QHeaderView.Interactive)
        for pc_idx, (_, _, _, act_type, sub) in enumerate(preview_cols):
            if (act_type in ["input", "clear_input"] and sub is None) or (act_type == "clear_input_plus" and sub == "content"):
                preview_table.setItemDelegateForColumn(pc_idx, MultiLineTextDelegate(preview_table))

        base_dialog_title = f"🛠️ 批量填充中心 — {title_suffix}"
        _syncing_preview_row_height = [False]

        def _set_preview_row_height(value, save=True):
            row_height = min(max(28, int(value)), 160)
            self.config.setdefault("layout", {})
            self.config["layout"]["batch_fill_preview_row_height"] = row_height

            if preview_row_height_spin.value() != row_height:
                preview_row_height_spin.blockSignals(True)
                preview_row_height_spin.setValue(row_height)
                preview_row_height_spin.blockSignals(False)

            _syncing_preview_row_height[0] = True
            try:
                preview_table.verticalHeader().setDefaultSectionSize(row_height)
                for row in range(preview_table.rowCount()):
                    preview_table.setRowHeight(row, row_height)
            finally:
                _syncing_preview_row_height[0] = False

            if save:
                self._schedule_config_flush()

        def _on_preview_row_resized(_logical_index, _old_size, new_size):
            if _syncing_preview_row_height[0]:
                return
            _set_preview_row_height(new_size, save=True)

        preview_table.verticalHeader().sectionResized.connect(_on_preview_row_resized)
        preview_row_height_spin.valueChanged.connect(_set_preview_row_height)
        preview_row_height_spin.setValue(int(self.config.get("layout", {}).get(
            "batch_fill_preview_row_height",
            self.config.get("layout", {}).get("data_row_height", 56)
        )))

        def _update_batch_fill_caption(extra_text=""):
            if extra_text:
                lbl_info.setText(f"待处理行（共 {len(selected_rows)} 行）: [💡 {extra_text}]")
                dlg.setWindowTitle(f"{base_dialog_title} ({extra_text})")
            else:
                lbl_info.setText(f"待处理行（共 {len(selected_rows)} 行）: [💡 支持直接拖入外部文件或右侧工具项]")
                dlg.setWindowTitle(base_dialog_title)

        def _show_batch_fill_mini_menu():
            def _pick_preview_row_height():
                value, ok = QInputDialog.getInt(
                    dlg,
                    "设置行高",
                    "请输入预览区统一行高：",
                    preview_row_height_spin.value(),
                    28,
                    160,
                    1
                )
                if ok:
                    preview_row_height_spin.setValue(value)

            menu = QMenu(dlg)
            menu.addAction("➕ 临时增加行", _add_row_in_dlg)
            menu.addAction("📋 直贴表格", lambda: btn_paste_clipboard_dlg.click())
            menu.addAction("🔄 同步表头", lambda: btn_sync_headers_dlg.click())
            menu.addAction("🧹 重置预设", lambda: btn_reset_presets_dlg.click())
            menu.addAction("🗑️ 删除选中行", lambda: btn_del_row_dlg.click())
            menu.addSeparator()
            menu.addAction(f"📏 设置行高（当前 {preview_row_height_spin.value()} px）", _pick_preview_row_height)
            menu.exec_(btn_batch_fill_menu.mapToGlobal(btn_batch_fill_menu.rect().bottomLeft()))

        btn_batch_fill_menu.clicked.connect(_show_batch_fill_mini_menu)

        def _reload_preview_from_main(extra_text="", preserved_state=None):
            preview_table.blockSignals(True)
            preview_table.setRowCount(len(selected_rows))
            preview_table.setVerticalHeaderLabels([f"第{r+1}行" for r in selected_rows])
            for r_idx, real_row in enumerate(selected_rows):
                for pc_idx, (real_col, _step_idx, name, act_type, sub) in enumerate(preview_cols):
                    val = ""
                    item_backing = self.data_table.item(real_row, real_col)
                    if item_backing:
                        raw_val = item_backing.text()
                        if act_type == "clear_input_plus":
                            skip_token = "[SKIP_ROW]"
                            if raw_val.strip() == skip_token:
                                val = "" if sub == "prefix" else skip_token
                            else:
                                parts = raw_val.split('|', 1)
                                if sub == "prefix": val = parts[0] if len(parts) > 0 else ""
                                else: val = parts[1] if len(parts) > 1 else ""
                        elif act_type == "open_url":
                            skip_token = "[SKIP_ROW]"
                            if raw_val.strip() == skip_token:
                                val = skip_token if sub == "url" else ""
                            else:
                                parts = raw_val.split('|', 1)
                                if sub == "url":
                                    val = parts[0] if len(parts) > 0 else ""
                                else:
                                    p_id = parts[1] if len(parts) > 1 else ""
                                    val = get_profile_display_name(p_id) if p_id else ""
                        else:
                            val = raw_val
                    elif act_type in ["input", "clear_input", "upload", "run_app"]:
                        w = self.data_table.cellWidget(real_row, real_col)
                        if isinstance(w, QLineEdit): val = w.text()

                    it = QTableWidgetItem(val)
                    it.setData(Qt.UserRole, real_row)
                    it.setData(Qt.UserRole + 1, real_col)
                    it.setData(Qt.UserRole + 2, sub)
                    if act_type == "open_url" and sub == "profile":
                        p_id = ""
                        if item_backing:
                            raw_val = item_backing.text()
                            if raw_val.strip() != "[SKIP_ROW]":
                                parts = raw_val.split('|', 1)
                                p_id = parts[1] if len(parts) > 1 else ""
                        if p_id:
                            it.setData(Qt.UserRole + 10, f"|{p_id}")
                            it.setToolTip(p_id)
                        else:
                            it.setToolTip(val)
                    else:
                        it.setToolTip(val)
                    preview_table.setItem(r_idx, pc_idx, it)
            preview_table.blockSignals(False)
            if preserved_state:
                for r in range(preview_table.rowCount()):
                    for c in range(preview_table.columnCount()):
                        it = preview_table.item(r, c)
                        if not it:
                            continue
                        key = (
                            it.data(Qt.UserRole),
                            it.data(Qt.UserRole + 1),
                            it.data(Qt.UserRole + 2)
                        )
                        cell_state = preserved_state.get(key)
                        if not cell_state:
                            continue
                        it.setText(cell_state.get("text", ""))
                        it.setToolTip(cell_state.get("text", ""))
                        it.setData(Qt.UserRole + 10, cell_state.get("role10"))
                        bg = cell_state.get("bg")
                        if bg:
                            it.setBackground(QColor(bg))
                        else:
                            it.setBackground(QBrush(Qt.NoBrush))
            _set_preview_row_height(preview_row_height_spin.value(), save=False)
            _update_batch_fill_caption(extra_text)
        
        # [新增] 撤销功能支持 (Ctrl+Z)
        undo_stack = []
        def _save_undo_state():
            state = []
            for r in range(preview_table.rowCount()):
                row_data = []
                for c in range(preview_table.columnCount()):
                    it = preview_table.item(r, c)
                    if it:
                        row_data.append({
                            "text": it.text(),
                            "bg": it.background().color().name() if it.background().style() != Qt.NoBrush else None,
                            "role10": it.data(Qt.UserRole + 10)
                        })
                    else:
                        row_data.append(None)
                state.append(row_data)
            undo_stack.append(state)
            if len(undo_stack) > 50: undo_stack.pop(0) # 限制撤销步数

        def _perform_undo():
            if not undo_stack: return
            state = undo_stack.pop()
            preview_table.blockSignals(True)
            for r in range(len(state)):
                for c in range(len(state[r])):
                    data = state[r][c]
                    it = preview_table.item(r, c)
                    if it and data:
                        it.setText(data["text"])
                        if data["bg"]: it.setBackground(QColor(data["bg"]))
                        else: it.setBackground(QBrush(Qt.NoBrush))
                        if data["role10"]: it.setData(Qt.UserRole + 10, data["role10"])
            preview_table.blockSignals(False)

        def _set_preview_item_text(item_obj, new_val, bg_color="#fff9c4"):
            if not item_obj:
                return
            item_obj.setText(new_val)
            item_obj.setToolTip(str(new_val))
            item_obj.setData(Qt.UserRole + 10, None)
            item_obj.setBackground(QColor(bg_color))

        def _get_selected_preview_indexes():
            indexes = preview_table.selectedIndexes()
            if not indexes:
                curr = preview_table.currentIndex()
                if curr.isValid():
                    indexes = [curr]
            return indexes

        def _paste_matrix_to_preview(source_matrix=None):
            matrix = source_matrix or parse_spreadsheet_clipboard_text(pyperclip.paste())
            if not matrix:
                QMessageBox.warning(dlg, "提示", "剪贴板里没有可用的表格数据。")
                return False

            indexes = _get_selected_preview_indexes()
            if not indexes:
                QMessageBox.warning(dlg, "提示", "请先在左侧预览表中选择目标单元格或区域。")
                return False

            rows = sorted({idx.row() for idx in indexes})
            cols = sorted({idx.column() for idx in indexes})
            row_count = len(matrix)
            col_count = max((len(row) for row in matrix), default=0)
            if row_count == 0 or col_count == 0:
                QMessageBox.warning(dlg, "提示", "剪贴板表格为空。")
                return False

            _save_undo_state()
            is_full_rect = (
                bool(rows) and bool(cols) and
                len(indexes) == len(rows) * len(cols) and
                all(preview_table.item(r, c) for r in rows for c in cols)
            )

            if is_full_rect and (len(rows) > 1 or len(cols) > 1):
                for r_off, row_idx in enumerate(rows):
                    for c_off, col_idx in enumerate(cols):
                        it = preview_table.item(row_idx, col_idx)
                        new_val = matrix[r_off % row_count][c_off % col_count] if c_off % col_count < len(matrix[r_off % row_count]) else ""
                        _set_preview_item_text(it, new_val)
            else:
                anchor = sorted(indexes, key=lambda idx: (idx.row(), idx.column()))[0]
                start_row = anchor.row()
                start_col = anchor.column()
                for r_off, row in enumerate(matrix):
                    target_row = start_row + r_off
                    if target_row >= preview_table.rowCount():
                        break
                    for c_off, new_val in enumerate(row):
                        target_col = start_col + c_off
                        if target_col >= preview_table.columnCount():
                            break
                        it = preview_table.item(target_row, target_col)
                        _set_preview_item_text(it, new_val)

            preview_table.viewport().update()
            return True

        def _copy_matrix_from_preview():
            indexes = _get_selected_preview_indexes()
            if not indexes:
                return False

            rows = sorted({idx.row() for idx in indexes})
            cols = sorted({idx.column() for idx in indexes})
            if not rows or not cols:
                return False

            selected_keys = {(idx.row(), idx.column()) for idx in indexes}
            is_full_rect = len(selected_keys) == len(rows) * len(cols)

            lines = []
            if is_full_rect:
                for row in rows:
                    cells = []
                    for col in cols:
                        it = preview_table.item(row, col)
                        cells.append(it.text() if it else "")
                    lines.append("\t".join(cells))
            else:
                if len(cols) == 1:
                    for row in rows:
                        it = preview_table.item(row, cols[0])
                        lines.append(it.text() if it else "")
                elif len(rows) == 1:
                    cells = []
                    for col in cols:
                        it = preview_table.item(rows[0], col)
                        cells.append(it.text() if it else "")
                    lines.append("\t".join(cells))
                else:
                    min_row, max_row = rows[0], rows[-1]
                    min_col, max_col = cols[0], cols[-1]
                    for row in range(min_row, max_row + 1):
                        cells = []
                        for col in range(min_col, max_col + 1):
                            if (row, col) in selected_keys:
                                it = preview_table.item(row, col)
                                cells.append(it.text() if it else "")
                            else:
                                cells.append("")
                        lines.append("\t".join(cells))

            pyperclip.copy("\n".join(lines))
            return True

        # [新增] 监听键盘事件 (Delete / Ctrl+Z / Ctrl+C / Ctrl+V)
        def _on_key_press(event):
            if event.key() == Qt.Key_Delete:
                sel_items = preview_table.selectedItems()
                if not sel_items: return
                _save_undo_state() # 删除前保存状态
                for it in sel_items:
                    it.setText("")
                    it.setBackground(QColor(0, 0, 0, 0)) 
            elif event.key() == Qt.Key_C and (event.modifiers() & Qt.ControlModifier):
                _copy_matrix_from_preview()
            elif event.key() == Qt.Key_V and (event.modifiers() & Qt.ControlModifier):
                _paste_matrix_to_preview()
            elif event.key() == Qt.Key_Z and (event.modifiers() & Qt.ControlModifier):
                _perform_undo()
            else:
                QTableWidget.keyPressEvent(preview_table, event)
        preview_table.keyPressEvent = _on_key_press
        btn_paste_clipboard_dlg.clicked.connect(_paste_matrix_to_preview)

        def _sync_headers_in_dlg():
            self._sync_data_headers()
            self._refresh_data_table(force_sync=False)
            valid_rows = [r for r in selected_rows if 0 <= r < self.data_table.rowCount()]
            selected_rows[:] = valid_rows
            _reload_preview_from_main("已同步表头")

        def _reset_presets_in_dlg():
            self._reset_data_to_presets()
            self._refresh_data_table(force_sync=False)
            valid_rows = [r for r in selected_rows if 0 <= r < self.data_table.rowCount()]
            selected_rows[:] = valid_rows
            _reload_preview_from_main("已重置预设")

        def _delete_rows_in_dlg():
            row_indexes = sorted({idx.row() for idx in preview_table.selectedIndexes()})
            if not row_indexes:
                current_row = preview_table.currentRow()
                if current_row >= 0:
                    row_indexes = [current_row]
            if not row_indexes:
                QMessageBox.information(dlg, "提示", "请先在左侧预览表中选中要删除的数据行。")
                return
            real_rows = sorted({selected_rows[r] for r in row_indexes if 0 <= r < len(selected_rows)}, reverse=True)
            if not real_rows:
                return
            reply = QMessageBox.question(
                dlg,
                "确认删除",
                f"确定删除 {len(real_rows)} 行数据吗？\n删除后会同步更新主表格和当前预览。",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if reply != QMessageBox.Yes:
                return
            data_rows = self.config['task_data'].get(self.current_task, [])
            for real_row in real_rows:
                if 0 <= real_row < len(data_rows):
                    data_rows.pop(real_row)
            save_config(self.config)
            self._refresh_data_table(force_sync=False)
            deleted_set = set(real_rows)
            adjusted_rows = []
            for old_row in selected_rows:
                if old_row in deleted_set:
                    continue
                shift = sum(1 for r in real_rows if r < old_row)
                adjusted_rows.append(old_row - shift)
            selected_rows[:] = adjusted_rows
            undo_stack.clear()
            _reload_preview_from_main(f"已删除 {len(real_rows)} 行")

        btn_sync_headers_dlg.clicked.connect(_sync_headers_in_dlg)
        btn_reset_presets_dlg.clicked.connect(_reset_presets_in_dlg)
        btn_del_row_dlg.clicked.connect(_delete_rows_in_dlg)

        def _get_selected_preview_targets():
            """按当前选区形状返回有序目标格：
            单行选择 -> 按列从左到右
            单列选择 -> 按行从上到下
            矩形/多行多列 -> 宽优先按行展开；高优先按列展开
            """
            indexes = preview_table.selectedIndexes()
            if not indexes:
                curr = preview_table.currentIndex()
                if curr.isValid():
                    indexes = [curr]
            if not indexes:
                return []

            rows = sorted({idx.row() for idx in indexes})
            cols = sorted({idx.column() for idx in indexes})
            if len(rows) == 1:
                ordered = sorted(indexes, key=lambda idx: (idx.column(), idx.row()))
            elif len(cols) == 1:
                ordered = sorted(indexes, key=lambda idx: (idx.row(), idx.column()))
            else:
                row_span = rows[-1] - rows[0] + 1
                col_span = cols[-1] - cols[0] + 1
                if col_span >= row_span:
                    ordered = sorted(indexes, key=lambda idx: (idx.row(), idx.column()))
                else:
                    ordered = sorted(indexes, key=lambda idx: (idx.column(), idx.row()))

            targets = []
            seen = set()
            for idx in ordered:
                key = (idx.row(), idx.column())
                if key in seen:
                    continue
                seen.add(key)
                item = preview_table.item(idx.row(), idx.column())
                if item:
                    targets.append(item)
            return targets

        _reload_preview_from_main()
        
        left_ly.addWidget(preview_table)
        preview_selection_count = create_table_selection_label()
        left_ly.addWidget(preview_selection_count)
        bind_table_selection_label(preview_table, preview_selection_count, "左侧目标区：")
        main_splitter.addWidget(left_w)
        
        # 右侧：双并列工具箱模式
        right_w = QWidget(); right_ly = QVBoxLayout(right_w); right_ly.setContentsMargins(5,0,0,0); right_ly.setSpacing(10)
        
        # 优化后的顶部目标提示栏
        cur_target_label = QLabel("🎯 当前填充目标: [请点击左侧表格单元格]")
        cur_target_label.setStyleSheet("""
            QLabel { font-weight: 400; color: #222;
                color: #d32f2f; 
                font-weight: bold; 
                font-size: 14px; 
                padding: 10px; 
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #ffebee, stop:1 #fff); 
                border: 1px solid #ffcdd2;
                border-radius: 6px;
            }
        """)
        right_ly.addWidget(cur_target_label, 0) # 权重为 0，不自动拉伸

        # 前缀库面板：作为工具箱中的独立标签页使用
        prefix_group = QGroupBox("📚 前缀库面板")
        prefix_ly = QVBoxLayout(prefix_group)
        prefix_ly.setContentsMargins(8, 8, 8, 8)
        prefix_ly.setSpacing(6)

        prefix_cfg_path = os.path.join(BASE_DIR, "input_prefixes.json")
        prefix_presets = []

        def _load_prefix_presets():
            items = []
            if os.path.exists(prefix_cfg_path):
                try:
                    with open(prefix_cfg_path, "r", encoding="utf-8") as f:
                        raw = json.load(f)
                    if isinstance(raw, list):
                        for idx, p in enumerate(raw):
                            if not isinstance(p, dict):
                                continue
                            name = str(p.get("name", "")).strip() or f"前缀{idx+1}"
                            prefix = str(p.get("prefix", ""))
                            if name or prefix.strip():
                                items.append({"name": name, "prefix": prefix})
                except Exception as e:
                    log_internal_issue(f"批量填充中心加载前缀库失败: {prefix_cfg_path}", e)
            return items

        def _save_prefix_presets():
            try:
                with open(prefix_cfg_path, "w", encoding="utf-8") as f:
                    json.dump(prefix_presets, f, ensure_ascii=False, indent=4)
            except Exception as e:
                log_internal_issue(f"批量填充中心保存前缀库失败: {prefix_cfg_path}", e)
                QMessageBox.warning(dlg, "保存失败", f"前缀库保存失败：\n{e}")

        def _parse_prefix_import_text(raw_text):
            text = str(raw_text or "").lstrip("\ufeff").strip()
            if not text:
                return []

            imported = []

            # 1. 优先支持 JSON 数组：[{"name":"xxx","prefix":"yyy"}]
            try:
                obj = json.loads(text)
                if isinstance(obj, list):
                    for idx, p in enumerate(obj):
                        if not isinstance(p, dict):
                            continue
                        name = str(p.get("name", "")).strip() or f"导入{idx+1}"
                        prefix = str(p.get("prefix", ""))
                        imported.append({"name": name, "prefix": prefix})
                    if imported:
                        return imported
            except Exception:
                pass

            # 2. 逐行文本：支持 名称|前缀、名称<TAB>前缀、名称,前缀、仅前缀内容
            for line in text.splitlines():
                raw_line = line.rstrip("\r\n")
                s = raw_line.strip()
                if not s or s.startswith("#"):
                    continue

                compact = s.lower().replace(" ", "")
                if compact in {
                    "name|prefix", "名称|前缀", "名称|前缀内容",
                    "name,prefix", "名称,前缀", "名称,前缀内容",
                    "name\tprefix", "名称\t前缀", "名称\t前缀内容"
                }:
                    continue

                if "\t" in raw_line:
                    name, prefix = raw_line.split("\t", 1)
                elif "|" in raw_line:
                    name, prefix = raw_line.split("|", 1)
                elif "," in raw_line:
                    name, prefix = raw_line.split(",", 1)
                else:
                    name, prefix = "", raw_line

                name = str(name).strip() or f"导入{len(imported)+1}"
                prefix = str(prefix)
                if name or prefix.strip():
                    imported.append({"name": name, "prefix": prefix})

            return imported

        prefix_header = QHBoxLayout()
        prefix_header.addWidget(QLabel("🔍 搜索:"))
        prefix_search = QLineEdit()
        prefix_search.setPlaceholderText("输入名称或前缀内容...")
        prefix_header.addWidget(prefix_search)
        btn_prefix_refresh = QPushButton("🔄")
        btn_prefix_refresh.setToolTip("重新加载前缀库")
        prefix_header.addWidget(btn_prefix_refresh)
        prefix_ly.addLayout(prefix_header)

        prefix_table = QTableWidget()
        prefix_table.setColumnCount(2)
        prefix_table.setHorizontalHeaderLabels(["名称", "前缀内容"])
        prefix_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        prefix_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        prefix_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        prefix_table.setAlternatingRowColors(True)
        prefix_table.setWordWrap(True)
        prefix_table.verticalHeader().setDefaultSectionSize(34)
        prefix_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Interactive)
        prefix_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        prefix_table.setColumnWidth(0, 150)
        prefix_table.setItemDelegateForColumn(1, MultiLineTextDelegate(prefix_table))
        prefix_ly.addWidget(prefix_table, 1)
        prefix_selection_count = create_table_selection_label()
        prefix_ly.addWidget(prefix_selection_count)
        bind_table_selection_label(prefix_table, prefix_selection_count)
        prefix_table_updating = [False]
        prefix_tip = QLabel("提示：双击行或点“编辑”可弹出完整编辑窗口；支持“上移/下移”调整顺序。")
        prefix_tip.setStyleSheet("color: #666; padding: 2px 0 0 2px;")
        prefix_ly.addWidget(prefix_tip)

        def _refresh_prefix_table(selected_idx=None):
            prefix_table_updating[0] = True
            prefix_table.setRowCount(0)
            keyword = prefix_search.text().strip().lower()
            target_row = -1
            for idx, p in enumerate(prefix_presets):
                name = str(p.get("name", ""))
                prefix = str(p.get("prefix", ""))
                if keyword and keyword not in name.lower() and keyword not in prefix.lower():
                    continue
                row = prefix_table.rowCount()
                prefix_table.insertRow(row)
                it_name = QTableWidgetItem(name)
                it_name.setData(Qt.UserRole, idx)
                it_prefix = QTableWidgetItem(prefix)
                it_prefix.setData(Qt.UserRole, idx)
                prefix_table.setItem(row, 0, it_name)
                prefix_table.setItem(row, 1, it_prefix)
                if idx == selected_idx:
                    target_row = row
            prefix_table_updating[0] = False
            for row in range(prefix_table.rowCount()):
                prefix_table.resizeRowToContents(row)
                prefix_table.setRowHeight(row, min(max(prefix_table.rowHeight(row), 34), 120))
            if 0 <= target_row < prefix_table.rowCount():
                prefix_table.setCurrentCell(target_row, 0)
                prefix_table.selectRow(target_row)

        def _on_prefix_item_changed(item):
            if prefix_table_updating[0] or not item:
                return
            idx = item.data(Qt.UserRole)
            if idx is None:
                idx = _get_current_prefix_idx()
            try:
                idx = int(idx)
            except Exception:
                idx = -1
            if idx < 0 or idx >= len(prefix_presets):
                return

            row = item.row()
            name_item = prefix_table.item(row, 0)
            prefix_item = prefix_table.item(row, 1)
            name = name_item.text().strip() if name_item else ""
            prefix = prefix_item.text() if prefix_item else ""
            if not name:
                name = f"前缀{idx + 1}"
                prefix_table_updating[0] = True
                if name_item:
                    name_item.setText(name)
                else:
                    it_name = QTableWidgetItem(name)
                    it_name.setData(Qt.UserRole, idx)
                    prefix_table.setItem(row, 0, it_name)
                prefix_table_updating[0] = False
            prefix_presets[idx] = {"name": name, "prefix": prefix}
            _save_prefix_presets()
            prefix_table.resizeRowToContents(row)
            prefix_table.setRowHeight(row, min(max(prefix_table.rowHeight(row), 34), 120))

        def _get_current_prefix_idx():
            item = prefix_table.currentItem()
            if not item:
                return -1
            idx = item.data(Qt.UserRole)
            return int(idx) if idx is not None else -1

        def _select_prefix_idx(target_idx):
            if target_idx < 0:
                return
            for row in range(prefix_table.rowCount()):
                item = prefix_table.item(row, 0)
                idx = int(item.data(Qt.UserRole)) if item and item.data(Qt.UserRole) is not None else -1
                if idx == target_idx:
                    prefix_table.setCurrentCell(row, 0)
                    prefix_table.selectRow(row)
                    return

        def _import_prefix_presets():
            imp = QDialog(dlg)
            imp.setWindowTitle("📥 批量导入前缀库")
            imp.resize(760, 520)
            imp_ly = QVBoxLayout(imp)

            tip = QLabel(
                "导入前可直接在下表里改内容：第 1 列是名称，第 2 列是前缀内容。\n"
                "支持从文件载入或从剪贴板粘贴，载入后仍可继续手动编辑。"
            )
            tip.setStyleSheet("color: #666;")
            imp_ly.addWidget(tip)

            import_table = QTableWidget(0, 2)
            import_table.setHorizontalHeaderLabels(["名称", "前缀内容"])
            import_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Interactive)
            import_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
            import_table.setColumnWidth(0, 180)
            import_table.setSelectionBehavior(QAbstractItemView.SelectItems)
            import_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
            import_table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.SelectedClicked | QAbstractItemView.AnyKeyPressed)
            import_table.setAlternatingRowColors(True)
            import_table.setWordWrap(True)
            import_table.verticalHeader().setDefaultSectionSize(34)
            import_table.setItemDelegateForColumn(1, MultiLineTextDelegate(import_table))
            imp_ly.addWidget(import_table, 1)
            import_selection_count = create_table_selection_label()
            imp_ly.addWidget(import_selection_count)
            bind_table_selection_label(import_table, import_selection_count)

            def _resize_import_row(row):
                if row < 0:
                    return
                import_table.resizeRowToContents(row)
                import_table.setRowHeight(row, min(max(import_table.rowHeight(row), 34), 120))

            def _append_import_row(name="", prefix=""):
                row = import_table.rowCount()
                import_table.insertRow(row)
                import_table.setItem(row, 0, QTableWidgetItem(str(name)))
                import_table.setItem(row, 1, QTableWidgetItem(str(prefix)))
                return row

            def _set_import_rows(items):
                import_table.setRowCount(0)
                for p in items:
                    _append_import_row(p.get("name", ""), p.get("prefix", ""))
                if import_table.rowCount() == 0:
                    _append_import_row()
                for row in range(import_table.rowCount()):
                    _resize_import_row(row)

            def _collect_import_rows():
                rows = []
                for r in range(import_table.rowCount()):
                    name_item = import_table.item(r, 0)
                    prefix_item = import_table.item(r, 1)
                    name = name_item.text().strip() if name_item else ""
                    prefix = prefix_item.text() if prefix_item else ""
                    if not name and not prefix.strip():
                        continue
                    rows.append({"name": name or f"导入{len(rows) + 1}", "prefix": prefix})
                return rows

            _set_import_rows([])
            import_table.itemChanged.connect(lambda item: _resize_import_row(item.row()) if item else None)

            mode_ly = QHBoxLayout()
            mode_ly.addWidget(QLabel("导入模式:"))
            import_mode = QComboBox()
            import_mode.addItems(["同名覆盖", "仅追加", "先清空再导入"])
            import_mode.setCurrentText("同名覆盖")
            import_mode.setToolTip(
                "同名覆盖：名称相同则更新，不同则新增\n"
                "仅追加：全部追加到末尾，如重名会自动补序号\n"
                "先清空再导入：清空现有前缀库后再导入"
            )
            mode_ly.addWidget(import_mode)
            mode_ly.addStretch()
            imp_ly.addLayout(mode_ly)

            import_btns = QHBoxLayout()
            btn_add_import_row = QPushButton("➕ 添加行")
            btn_del_import_row = QPushButton("❌ 删除行")
            btn_load_file = QPushButton("📂 从文件载入")
            btn_load_clipboard = QPushButton("📋 从剪贴板粘贴")
            btn_merge_import = QPushButton("✅ 开始导入")
            btn_merge_import.setStyleSheet("background: #e8f5e9; font-weight: bold;")
            import_btns.addWidget(btn_add_import_row)
            import_btns.addWidget(btn_del_import_row)
            import_btns.addWidget(btn_load_file)
            import_btns.addWidget(btn_load_clipboard)
            import_btns.addStretch()
            import_btns.addWidget(btn_merge_import)
            imp_ly.addLayout(import_btns)

            def _load_import_file():
                path, _ = QFileDialog.getOpenFileName(
                    imp,
                    "选择前缀数据文件",
                    "",
                    "Data Files (*.json *.csv *.txt);;All Files (*.*)"
                )
                if not path:
                    return
                try:
                    items = []
                    lower_path = path.lower()
                    if lower_path.endswith(".csv"):
                        with open(path, "r", encoding="utf-8-sig", errors="ignore", newline="") as f:
                            reader = csv.reader(f)
                            for idx, row in enumerate(reader):
                                if not row:
                                    continue
                                header0 = str(row[0]).strip().lower() if len(row) > 0 else ""
                                header1 = str(row[1]).strip().lower() if len(row) > 1 else ""
                                if idx == 0 and header0 in {"name", "名称"} and header1 in {"prefix", "前缀", "前缀内容"}:
                                    continue
                                name = str(row[0]).strip() if len(row) > 0 else ""
                                prefix = str(row[1]) if len(row) > 1 else ""
                                if name or prefix.strip():
                                    items.append({"name": name or f"导入{len(items)+1}", "prefix": prefix})
                    else:
                        with open(path, "r", encoding="utf-8-sig", errors="ignore") as f:
                            items = _parse_prefix_import_text(f.read())
                    _set_import_rows(items)
                except Exception as e:
                    QMessageBox.warning(imp, "读取失败", f"读取文件失败：\n{e}")

            def _load_import_clipboard():
                text = QApplication.clipboard().text()
                items = _parse_prefix_import_text(text)
                if not items:
                    QMessageBox.warning(imp, "提示", "剪贴板里没有解析到可用的前缀数据。")
                    return
                _set_import_rows(items)

            def _do_import():
                nonlocal prefix_presets
                imported = _collect_import_rows()
                if not imported:
                    QMessageBox.warning(imp, "导入失败", "没有解析到可用的前缀数据。")
                    return

                mode = import_mode.currentText()
                if mode == "先清空再导入":
                    prefix_presets = []

                name_to_idx = {str(p.get("name", "")): i for i, p in enumerate(prefix_presets)}
                added = 0
                updated = 0
                for p in imported:
                    name = str(p.get("name", "")).strip() or f"导入{len(prefix_presets)+1}"
                    prefix = str(p.get("prefix", ""))

                    if mode == "仅追加":
                        base_name = name
                        suffix = 2
                        while name in name_to_idx:
                            name = f"{base_name}_{suffix}"
                            suffix += 1
                        name_to_idx[name] = len(prefix_presets)
                        prefix_presets.append({"name": name, "prefix": prefix})
                        added += 1
                    elif name in name_to_idx:
                        prefix_presets[name_to_idx[name]] = {"name": name, "prefix": prefix}
                        updated += 1
                    else:
                        name_to_idx[name] = len(prefix_presets)
                        prefix_presets.append({"name": name, "prefix": prefix})
                        added += 1

                _save_prefix_presets()
                _refresh_prefix_table()
                QMessageBox.information(imp, "导入完成", f"已导入完成：新增 {added} 条，更新 {updated} 条。")
                imp.accept()

            def _del_import_rows():
                rows = sorted({it.row() for it in import_table.selectedItems()}, reverse=True)
                if not rows and import_table.currentRow() >= 0:
                    rows = [import_table.currentRow()]
                for row in rows:
                    import_table.removeRow(row)
                if import_table.rowCount() == 0:
                    _append_import_row()

            btn_add_import_row.clicked.connect(lambda: _append_import_row())
            btn_del_import_row.clicked.connect(_del_import_rows)
            btn_load_file.clicked.connect(_load_import_file)
            btn_load_clipboard.clicked.connect(_load_import_clipboard)
            btn_merge_import.clicked.connect(_do_import)
            imp.exec_()

        def _export_prefix_presets():
            sel_model = prefix_table.selectionModel()
            sel_rows_model = sel_model.selectedRows() if sel_model else []
            export_items = []
            for m in sel_rows_model:
                row = m.row()
                it = prefix_table.item(row, 0)
                idx = int(it.data(Qt.UserRole)) if it and it.data(Qt.UserRole) is not None else -1
                if 0 <= idx < len(prefix_presets):
                    export_items.append({
                        "name": str(prefix_presets[idx].get("name", "")),
                        "prefix": str(prefix_presets[idx].get("prefix", ""))
                    })
            if not export_items:
                export_items = [
                    {"name": str(p.get("name", "")), "prefix": str(p.get("prefix", ""))}
                    for p in prefix_presets
                ]
            if not export_items:
                QMessageBox.warning(dlg, "提示", "前缀库为空，暂无可导出的内容。")
                return

            path, selected_filter = QFileDialog.getSaveFileName(
                dlg,
                "导出前缀库",
                "prefix_presets.json",
                "JSON Files (*.json);;CSV Files (*.csv);;Text Files (*.txt)"
            )
            if not path:
                return
            try:
                if path.lower().endswith(".csv") or "CSV" in selected_filter:
                    if not path.lower().endswith(".csv"):
                        path += ".csv"
                    with open(path, "w", encoding="utf-8-sig", newline="") as f:
                        writer = csv.writer(f)
                        writer.writerow(["name", "prefix"])
                        for p in export_items:
                            writer.writerow([p["name"], p["prefix"]])
                elif path.lower().endswith(".txt") or "Text" in selected_filter:
                    if not path.lower().endswith(".txt"):
                        path += ".txt"
                    with open(path, "w", encoding="utf-8") as f:
                        for p in export_items:
                            f.write(f"{p['name']}|{p['prefix']}\n")
                else:
                    if not path.lower().endswith(".json"):
                        path += ".json"
                    with open(path, "w", encoding="utf-8") as f:
                        json.dump(export_items, f, ensure_ascii=False, indent=4)
                QMessageBox.information(dlg, "导出完成", f"已导出 {len(export_items)} 条前缀。")
            except Exception as e:
                QMessageBox.warning(dlg, "导出失败", f"导出前缀库失败：\n{e}")

        def _add_prefix_preset():
            default_name = f"前缀{len(prefix_presets) + 1}"
            edit_dlg = PrefixPresetEditorDialog(dlg, name=default_name, prefix="", title="➕ 添加前缀")
            if edit_dlg.exec_() != QDialog.Accepted:
                return
            data = edit_dlg.get_data()
            prefix_presets.append({
                "name": data["name"] or default_name,
                "prefix": data["prefix"]
            })
            _save_prefix_presets()
            _refresh_prefix_table(selected_idx=len(prefix_presets) - 1)

        def _edit_prefix_preset():
            idx = _get_current_prefix_idx()
            if idx < 0 or idx >= len(prefix_presets):
                return
            current = prefix_presets[idx]
            edit_dlg = PrefixPresetEditorDialog(
                dlg,
                name=current.get("name", ""),
                prefix=current.get("prefix", ""),
                title=f"✏️ 编辑前缀 - {current.get('name', '') or f'前缀{idx + 1}'}"
            )
            if edit_dlg.exec_() != QDialog.Accepted:
                return
            data = edit_dlg.get_data()
            prefix_presets[idx] = {
                "name": data["name"] or f"前缀{idx + 1}",
                "prefix": data["prefix"]
            }
            _save_prefix_presets()
            _refresh_prefix_table(selected_idx=idx)

        def _move_prefix_preset(direction):
            idx = _get_current_prefix_idx()
            if idx < 0 or idx >= len(prefix_presets):
                return
            target_idx = idx + direction
            if target_idx < 0 or target_idx >= len(prefix_presets):
                return
            prefix_presets[idx], prefix_presets[target_idx] = prefix_presets[target_idx], prefix_presets[idx]
            _save_prefix_presets()
            _refresh_prefix_table(selected_idx=target_idx)
            _select_prefix_idx(target_idx)

        def _del_prefix_preset():
            idx = _get_current_prefix_idx()
            if idx < 0 or idx >= len(prefix_presets):
                return
            if QMessageBox.question(dlg, "确认删除", f"确定删除前缀「{prefix_presets[idx].get('name', '')}」吗？",
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No) != QMessageBox.Yes:
                return
            prefix_presets.pop(idx)
            _save_prefix_presets()
            _refresh_prefix_table(selected_idx=min(idx, len(prefix_presets) - 1))

        def _apply_prefix_presets_to_preview():
            sel_model = prefix_table.selectionModel()
            sel_rows_model = sel_model.selectedRows() if sel_model else []
            preset_values = []
            for m in sel_rows_model:
                row = m.row()
                it = prefix_table.item(row, 0)
                idx = int(it.data(Qt.UserRole)) if it and it.data(Qt.UserRole) is not None else -1
                if 0 <= idx < len(prefix_presets):
                    preset_values.append(str(prefix_presets[idx].get("prefix", "")))
            if not preset_values:
                idx = _get_current_prefix_idx()
                if 0 <= idx < len(prefix_presets):
                    preset_values = [str(prefix_presets[idx].get("prefix", ""))]
            if not preset_values:
                QMessageBox.warning(dlg, "提示", "请先在前缀库面板中选中至少一条前缀。")
                return

            target_items = _get_selected_preview_targets()
            if not target_items:
                QMessageBox.warning(dlg, "提示", "请先在左侧预览表中选择要填充的单元格。")
                return

            valid_targets = []
            for it in target_items:
                _, _, step_name, act_type, sub = preview_cols[it.column()]
                if act_type == "clear_input_plus" and sub == "prefix":
                    valid_targets.append(it)
            if not valid_targets:
                QMessageBox.warning(dlg, "提示", "当前选区里没有“前缀”单元格，请先选择某个“(前缀)”格子。")
                return

            _save_undo_state()
            for i, it in enumerate(valid_targets):
                it.setText(preset_values[i % len(preset_values)])
                it.setBackground(QColor("#fff9c4"))

        prefix_btns = QHBoxLayout()
        btn_prefix_add = QPushButton("➕ 添加")
        btn_prefix_edit = QPushButton("✏️ 编辑")
        btn_prefix_up = QPushButton("⬆️ 上移")
        btn_prefix_down = QPushButton("⬇️ 下移")
        btn_prefix_del = QPushButton("❌ 删除")
        btn_prefix_import = QPushButton("📥 导入")
        btn_prefix_export = QPushButton("📤 导出")
        btn_prefix_apply = QPushButton("⬅️ 填入当前前缀列")
        btn_prefix_apply.setStyleSheet("background: #e8f5e9; font-weight: bold;")
        prefix_btns.addWidget(btn_prefix_add)
        prefix_btns.addWidget(btn_prefix_edit)
        prefix_btns.addWidget(btn_prefix_up)
        prefix_btns.addWidget(btn_prefix_down)
        prefix_btns.addWidget(btn_prefix_del)
        prefix_btns.addWidget(btn_prefix_import)
        prefix_btns.addWidget(btn_prefix_export)
        prefix_btns.addStretch()
        prefix_btns.addWidget(btn_prefix_apply)
        prefix_ly.addLayout(prefix_btns)

        prefix_presets = _load_prefix_presets()
        _refresh_prefix_table()
        prefix_search.textChanged.connect(_refresh_prefix_table)
        btn_prefix_refresh.clicked.connect(lambda: [prefix_presets.clear(), prefix_presets.extend(_load_prefix_presets()), _refresh_prefix_table()])
        prefix_table.itemChanged.connect(_on_prefix_item_changed)
        prefix_table.itemDoubleClicked.connect(lambda *_: _edit_prefix_preset())
        btn_prefix_add.clicked.connect(_add_prefix_preset)
        btn_prefix_edit.clicked.connect(_edit_prefix_preset)
        btn_prefix_up.clicked.connect(lambda: _move_prefix_preset(-1))
        btn_prefix_down.clicked.connect(lambda: _move_prefix_preset(1))
        btn_prefix_del.clicked.connect(_del_prefix_preset)
        btn_prefix_import.clicked.connect(_import_prefix_presets)
        btn_prefix_export.clicked.connect(_export_prefix_presets)
        btn_prefix_apply.clicked.connect(_apply_prefix_presets_to_preview)

        tool_mode_row = QHBoxLayout()
        tool_mode_row.addStretch()
        btn_toggle_tool_mode = QPushButton("⇆ 切到单窗口")
        btn_toggle_tool_mode.setCheckable(True)
        btn_toggle_tool_mode.setFixedHeight(30)
        btn_toggle_tool_mode.setStyleSheet("""
            QPushButton {
                background: #eef5ff;
                border: 1px solid #90caf9;
                border-radius: 6px;
                padding: 0 12px;
                font-weight: bold;
                color: #1565c0;
            }
            QPushButton:hover { background: #e3f2fd; }
            QPushButton:checked {
                background: #f5f5f5;
                border: 1px solid #bdbdbd;
                color: #424242;
            }
        """)
        btn_toggle_tool_mode.setToolTip("点击即可在单窗口和双窗口之间切换。")
        tool_mode_row.addWidget(btn_toggle_tool_mode)
        right_ly.addLayout(tool_mode_row)

        # 使用 Splitter 实现可切换的单/双工具箱
        tool_splitter = QSplitter(Qt.Horizontal)
        tool_splitter.setStyleSheet("""
            QSplitter::handle { background: #d6d6d6; width: 8px; }
            QSplitter::handle:hover { background: #90caf9; }
        """)
        tool_splitter.setChildrenCollapsible(True)
        tool_splitter.setOpaqueResize(True)
        right_ly.addWidget(tool_splitter, 1) # 权重为 1，占据所有剩余高度

        def create_tool_box(box_id):
            # 创建一个完整的工具箱副本
            tabs = QTabWidget()
            
            # --- 工具1: 窗口选择 ---
            win_tool = QWidget(); win_ly = QVBoxLayout(win_tool)
            win_list = QListWidget(); win_list.setSelectionMode(QAbstractItemView.ExtendedSelection)
            win_ly.addWidget(QLabel("选择当前打开的窗口/标签页:"))
            search_win = QLineEdit(); search_win.setPlaceholderText("搜索窗口标题...")
            win_ly.addWidget(search_win)
            lbl_win_selection_count = create_table_selection_label()
            
            def _get_wins():
                import pygetwindow as pgw
                wins = []
                profile_meta = self.config.get("profile_meta", {}) or {}
                try:
                    clear_window_profile_caches()
                    for w in pgw.getAllWindows():
                        if w.title and w.visible and "批量填充中心" not in w.title:
                            hwnd = getattr(w, '_hWnd', None)
                            wins.append((build_window_display_text(w.title, hwnd, "[软件] ", profile_meta), w.title, hwnd))
                except Exception as e:
                    log_internal_issue("批量填充中心扫描窗口列表失败", e)
                wins.sort(key=lambda x: x[0])
                return wins

            def _refresh_win():
                all_w = _get_wins()
                win_list.clear()
                txt = search_win.text().lower()
                for d, r, hwnd in all_w:
                    if not txt or txt in d.lower():
                        it = QListWidgetItem(d)
                        it.setData(Qt.UserRole, r)          # 纯标题
                        it.setData(Qt.UserRole + 1, hwnd)   # hwnd 唯一标识
                        win_list.addItem(it)
            
            search_win.textChanged.connect(_refresh_win)
            btn_ref_win = QPushButton("🔄 刷新窗口"); btn_ref_win.clicked.connect(_refresh_win)
            win_ly.addWidget(win_list)
            win_ly.addWidget(lbl_win_selection_count)
            bind_item_view_selection_label(win_list, lbl_win_selection_count, kind_text="个窗口")
            win_ly.addWidget(btn_ref_win)
            _refresh_win()
            # 双击窗口列表自动填充
            win_list.itemDoubleClicked.connect(lambda: _do_fill())
            tabs.addTab(win_tool, "🪟 窗口")

            # --- 工具2: 账号选择 ---
            prof_tool = QWidget(); prof_ly = QVBoxLayout(prof_tool)
            prof_list = QListWidget(); prof_list.setSelectionMode(QAbstractItemView.ExtendedSelection)
            prof_list.setStyleSheet("""
                QListWidget::item { border-bottom: 1px solid #f0f0f0; padding: 4px; }
                QListWidget::item:selected { background: #e3f2fd; color: #0d47a1; }
            """)
            
            prof_ly.addWidget(QLabel("选择 Chrome 账户 (支持备注搜索/排序/标签):"))
            
            # 顶部操作栏：搜索 + 排序
            prof_header = QHBoxLayout()
            search_prof = QLineEdit(); search_prof.setPlaceholderText("🔍 搜索名称/备注/邮箱...")
            prof_header.addWidget(search_prof, 3)
            
            sort_prof = QComboBox()
            sort_prof.addItems(["默认排序", "按数字排序", "已激活优先", "按名称 A-Z", "按备注排序", "已标记优先"])
            prof_header.addWidget(sort_prof, 1)
            prof_ly.addLayout(prof_header)

            # 标签过滤栏
            filter_lay = QHBoxLayout()
            chk_hide_bad = QCheckBox("隐藏已失效账号")
            chk_hide_bad.setChecked(True)
            filter_lay.addWidget(chk_hide_bad)
            prof_ly.addLayout(filter_lay)
            
            # [新增] 账号数据缓存，避免搜索时重复深度扫描磁盘
            prof_data_cache = []

            def _refresh_prof(force_rescan=False):
                nonlocal prof_data_cache
                prof_list.clear()
                txt = search_prof.text().lower()
                hide_bad = chk_hide_bad.isChecked()
                sort_mode = sort_prof.currentText()
                
                # 只有点击“刷新”按钮或首次加载时，才执行耗时的磁盘扫描
                if force_rescan or not prof_data_cache:
                    if force_rescan:
                        clear_chrome_profile_cache()
                    prof_data_cache = get_chrome_profiles(force_refresh=force_rescan)
                
                # 获取账号状态元数据
                prof_meta = self.config.get("profile_meta", {})
                # 获取当前激活的账号路径
                active_paths = get_active_chrome_profiles()
                
                filtered_profs = []
                for pid, pname, pemail, premark, prawid in prof_data_cache:
                    meta = prof_meta.get(pid, {})
                    status = meta.get("status", "normal")
                    tag = meta.get("tag", "")
                    
                    # 搜索过滤逻辑 (基于缓存)
                    if txt:
                        match = (txt in pname.lower() or 
                                 txt in pid.lower() or 
                                 (pemail and txt in pemail.lower()) or 
                                 (premark and txt in premark.lower()) or 
                                 (tag and txt in tag.lower()))
                        if not match: continue
                    
                    # 状态过滤
                    if hide_bad and status == "bad": continue
                    
                    # [修复] 判断是否已激活：使用统一的 is_same_path 逻辑
                    is_active = any(is_same_path(pid, ap) for ap in active_paths)
                    
                    filtered_profs.append({
                        "pid": pid, "name": pname, "email": pemail, "remark": premark, 
                        "status": status, "tag": tag, "prawid": prawid, "is_active": is_active
                    })
                
                # 排序逻辑 (基于过滤后的结果)
                def _natural_sort_key(s):
                    return [int(text) if text.isdigit() else text.lower() for text in re.split('([0-9]+)', s)]

                if sort_mode == "按名称 A-Z":
                    filtered_profs.sort(key=lambda x: x["name"].lower())
                elif sort_mode == "按数字排序":
                    filtered_profs.sort(key=lambda x: natural_sort_key(x["name"]))
                elif sort_mode == "按备注排序":
                    filtered_profs.sort(key=lambda x: (x["remark"] or "").lower())
                elif sort_mode == "已标记优先":
                    filtered_profs.sort(key=lambda x: (0 if x["tag"] else 1, x["name"].lower()))
                elif sort_mode == "已激活优先":
                    filtered_profs.sort(key=lambda x: (0 if x["is_active"] else 1, natural_sort_key(x["name"])))
                else: # 默认排序也建议使用自然排序，防止 1, 11, 2 出现
                    filtered_profs.sort(key=lambda x: natural_sort_key(x["name"]))
                
                # [优化] 渲染列表，智能去重备注显示
                for i, p in enumerate(filtered_profs):
                    disp = p["name"]
                    # 如果备注名和显示名完全一样，就不再额外显示备注
                    if p["remark"] and p["remark"].lower() not in disp.lower():
                        disp += f" [{p['remark']}]"
                    
                    if p["tag"]: disp = f"🏷️{p['tag']} | {disp}"
                    
                    # 激活状态标识
                    if p["is_active"]: disp = f"🟢 {disp}"
                    
                    # 添加序号显示
                    disp = f"{i+1}. {disp}"
                    
                    it = QListWidgetItem(disp)
                    it.setData(Qt.UserRole, p["pid"])
                    
                    if p["status"] == "bad":
                        it.setText(f"❌ {disp} (已失效)")
                        it.setForeground(QColor("#9e9e9e"))
                    elif p["tag"]:
                        it.setForeground(QColor("#00695c"))
                    
                    prof_list.addItem(it)

            def _show_prof_menu(pos):
                item = prof_list.itemAt(pos)
                if not item: return
                pid = item.data(Qt.UserRole)
                menu = QMenu()
                
                # 标签管理
                tag_act = menu.addAction("🏷️ 设置/修改标签")
                status_act = menu.addAction("❌ 标记为已失效" if "已失效" not in item.text() else "✅ 恢复为正常")
                menu.addSeparator()
                copy_path = menu.addAction("📋 复制路径")
                
                act = menu.exec_(prof_list.mapToGlobal(pos))
                if not act: return
                
                if "profile_meta" not in self.config: self.config["profile_meta"] = {}
                if pid not in self.config["profile_meta"]: self.config["profile_meta"][pid] = {}
                
                if act == tag_act:
                    old_tag = self.config["profile_meta"][pid].get("tag", "")
                    new_tag, ok = QInputDialog.getText(dlg, "账号标签", "请输入标签内容 (如: 备用, 常用, 异常):", text=old_tag)
                    if ok:
                        self.config["profile_meta"][pid]["tag"] = new_tag.strip()
                        save_config(self.config); _refresh_prof()
                elif act == status_act:
                    current = self.config["profile_meta"][pid].get("status", "normal")
                    self.config["profile_meta"][pid]["status"] = "bad" if current == "normal" else "normal"
                    save_config(self.config); _refresh_prof()
                elif act == copy_path:
                    pyperclip.copy(pid)

            prof_list.setContextMenuPolicy(Qt.CustomContextMenu)
            prof_list.customContextMenuRequested.connect(_show_prof_menu)
            # 搜索、排序和过滤：仅操作缓存数据，不扫描磁盘，极速响应
            search_prof.textChanged.connect(lambda: _refresh_prof(force_rescan=False))
            sort_prof.currentIndexChanged.connect(lambda: _refresh_prof(force_rescan=False))
            chk_hide_bad.stateChanged.connect(lambda: _refresh_prof(force_rescan=False))
            
            # 仅在点击此按钮时执行深度磁盘扫描
            btn_ref_prof = QPushButton("🔄 深度扫描并刷新账号"); btn_ref_prof.clicked.connect(lambda: _refresh_prof(force_rescan=True))
            prof_ly.addWidget(prof_list)
            lbl_prof_selection_count = create_table_selection_label()
            prof_ly.addWidget(lbl_prof_selection_count)
            bind_item_view_selection_label(prof_list, lbl_prof_selection_count, kind_text="个账户")
            prof_ly.addWidget(btn_ref_prof)
            _refresh_prof(force_rescan=True) # 首次打开时扫描一次
            # 双击账号列表自动填充
            prof_list.itemDoubleClicked.connect(lambda: _do_fill())
            tabs.addTab(prof_tool, "👤 账号")

            # --- 工具3: 文件/图片 ---
            file_tool = QWidget(); file_ly = QVBoxLayout(file_tool)
            file_tree = QTreeWidget(); file_tree.setHeaderLabel("文件夹与文件"); file_tree.setSelectionMode(QAbstractItemView.ExtendedSelection)
            file_ly.addWidget(QLabel("扫描文件夹获取文件路径:"))
            
            search_file = QLineEdit(); search_file.setPlaceholderText("🔍 搜索文件名...")
            file_ly.addWidget(search_file)
            
            filter_bar = QHBoxLayout()
            filter_bar.addWidget(QLabel("过滤:"))
            EXT_MAP = {
                "img": {".png", ".jpg", ".jpeg", ".bmp", ".gif", ".webp", ".ico", ".tiff"},
                "vid": {".mp4", ".mov", ".avi", ".mkv", ".flv", ".wmv"},
                "aud": {".mp3", ".wav", ".flac", ".aac", ".ogg", ".m4a", ".wma"},
                "doc": {".txt", ".md", ".pdf", ".doc", ".docx", ".xls", ".xlsx", ".csv"},
            }
            filter_btns = {}
            for k, icon, tip in [("img", "🖼️ 图片", "图片"), ("vid", "🎬 视频", "视频"), ("aud", "🎵 音频", "音频"), ("doc", "📄 文档", "文档"), ("oth", "⚙️ 其他", "其他")]:
                btn = QPushButton(icon); btn.setCheckable(True); btn.setMinimumWidth(60); btn.setFixedHeight(30)
                btn.setStyleSheet("QPushButton:checked { background: #bbdefb; border: 1px solid #2196f3; border-radius: 4px; }")
                filter_btns[k] = btn; filter_bar.addWidget(btn)
            filter_bar.addStretch(); file_ly.addLayout(filter_bar)
            
            btn_scan = QPushButton("📂 扫描文件夹"); btn_scan.setStyleSheet("background:#e3f2fd; padding:5px;"); file_ly.addWidget(btn_scan)
            smart_fill_bar = QHBoxLayout()
            btn_smart_fill_append = QPushButton("➕ 填充到空行")
            btn_smart_fill_append.setStyleSheet("background:#ede7f6; padding:6px; font-weight:bold;")
            btn_smart_fill_append.setToolTip("自动寻找后续空白行进行填充，不覆盖已有内容。")
            smart_fill_bar.addWidget(btn_smart_fill_append)
            btn_smart_fill_selected = QPushButton("🎯 覆盖选区")
            btn_smart_fill_selected.setStyleSheet("background:#fce4ec; padding:6px; font-weight:bold;")
            btn_smart_fill_selected.setToolTip("仅覆盖你当前选中的目标单元格/行。")
            smart_fill_bar.addWidget(btn_smart_fill_selected)
            btn_smart_fill_rule = QPushButton("⚙️ 填充规则")
            btn_smart_fill_rule.setToolTip("打开通用智能填充规则设计窗口，自定义素材分组、文件名匹配、扩展名和缺料行为。")
            smart_fill_bar.addWidget(btn_smart_fill_rule)
            file_ly.addLayout(smart_fill_bar)
            file_ly.addWidget(file_tree)
            lbl_file_selection_count = create_table_selection_label()
            file_ly.addWidget(lbl_file_selection_count)
            bind_item_view_selection_label(file_tree, lbl_file_selection_count, kind_text="个文件/项")

            # --- 智能预览区域 ---
            preview_container = QGroupBox("🔍 智能预览")
            preview_container.setMinimumHeight(250)
            p_ly = QVBoxLayout(preview_container)
            
            # 使用堆栈布局切换不同预览模式
            preview_stack = QStackedWidget()
            
            # 模式1: 图片
            lbl_preview_img = QLabel("选中文件即可预览")
            lbl_preview_img.setAlignment(Qt.AlignCenter)
            lbl_preview_img.setStyleSheet("background: #f5f5f5; color: #999;")
            preview_stack.addWidget(lbl_preview_img)
            
            # 模式2: 文本
            text_preview = QTextEdit()
            text_preview.setReadOnly(True)
            text_preview.setFont(QFont("Consolas", 9) if sys.platform == 'win32' else QFont("Monospace", 9))
            text_preview.setStyleSheet("background: #fafafa; border: none;")
            preview_stack.addWidget(text_preview)
            
            p_ly.addWidget(preview_stack)
            
            # 底部信息行
            lbl_file_info = QLabel("")
            lbl_file_info.setStyleSheet("font-size: 10px; color: #666;")
            p_ly.addWidget(lbl_file_info)
            
            file_ly.addWidget(preview_container)
            
            _last_path = [None]

            def _scan(re_path=None):
                p = re_path or QFileDialog.getExistingDirectory(dlg, "选择文件夹")
                if not p: return
                _last_path[0] = p
                file_tree.clear()
                root = QTreeWidgetItem(file_tree, [os.path.basename(p)]); root.setData(0, Qt.UserRole, p)
                root.setIcon(0, dlg.style().standardIcon(QStyle.SP_DirIcon))
                active_types = [k for k, b in filter_btns.items() if b.isChecked()]
                
                def _add(parent, path):
                    has_visible_child = False
                    try:
                        for e in sorted_scandir_entries(path):
                            if e.is_dir():
                                c = QTreeWidgetItem(parent, [e.name]); c.setData(0, Qt.UserRole, e.path)
                                c.setIcon(0, dlg.style().standardIcon(QStyle.SP_DirIcon))
                                if _add(c, e.path): has_visible_child = True
                                else: # 如果子目录里没东西且开启了过滤，就隐藏这个子目录
                                    if active_types: c.setHidden(True)
                            else:
                                matched = True
                                if active_types:
                                    _, ext = os.path.splitext(e.name.lower())
                                    matched = False
                                    for t in active_types:
                                        if t == "oth":
                                            if not any(ext in exts for exts in EXT_MAP.values()): matched = True; break
                                        elif ext in EXT_MAP.get(t, set()): matched = True; break
                                if matched:
                                    c = QTreeWidgetItem(parent, [e.name]); c.setData(0, Qt.UserRole, e.path)
                                    c.setIcon(0, dlg.style().standardIcon(QStyle.SP_FileIcon))
                                    has_visible_child = True
                    except Exception as e:
                        log_internal_issue(f"扫描文件树失败: {path}", e)
                    # 只要有匹配的文件，就自动展开这个文件夹
                    if has_visible_child: parent.setExpanded(True)
                    return has_visible_child
                _add(root, p); root.setExpanded(True)
            
            btn_scan.clicked.connect(lambda: _scan())
            for b in filter_btns.values(): b.clicked.connect(lambda: _scan(_last_path[0]) if _last_path[0] else None)

            def _filter_files(txt):
                txt = txt.lower()
                def _toggle(item):
                    match = txt in item.text(0).lower()
                    any_child_match = False
                    for i in range(item.childCount()):
                        if _toggle(item.child(i)): any_child_match = True
                    show = match or any_child_match
                    item.setHidden(not show)
                    if txt and show: item.setExpanded(True)
                    return show
                for i in range(file_tree.topLevelItemCount()): _toggle(file_tree.topLevelItem(i))
            search_file.textChanged.connect(_filter_files)

            def _make_smart_fill_step_key(step_idx, act_type, sub=None):
                sub_key = sub if sub else "main"
                return f"{int(step_idx)}|{act_type}|{sub_key}"

            def _make_smart_fill_legacy_step_key(real_col, act_type, sub=None):
                sub_key = sub if sub else "main"
                return f"{int(real_col)}|{act_type}|{sub_key}"

            def _get_smart_fill_target_columns():
                target_columns = []
                seen = set()
                for pc_idx, (real_col, step_idx, step_name, act_type, sub) in enumerate(preview_cols):
                    is_target = (
                        (act_type in ["upload", "drag_file", "run_app"] and sub is None) or
                        (act_type in ["input", "clear_input"] and sub is None) or
                        (act_type == "clear_input_plus" and sub == "content")
                    )
                    if not is_target:
                        continue
                    step_key = _make_smart_fill_step_key(step_idx, act_type, sub)
                    if step_key in seen:
                        continue
                    seen.add(step_key)
                    target_columns.append((pc_idx, real_col, step_idx, step_name, act_type, sub, step_key))
                return target_columns

            def _build_default_step_rule(step_name, act_type, sub, smart_rules):
                if act_type in ["upload", "drag_file", "run_app"]:
                    if act_type in ["upload", "drag_file"]:
                        default_exts = smart_rules.get("image_exts_text", SMART_FILL_IMAGE_EXTS_TEXT)
                    else:
                        default_exts = smart_rules.get("file_exts_text", SMART_FILL_FILE_EXTS_TEXT)
                    return {
                        "step_name": step_name,
                        "enabled": True,
                        "target_kind": "file",
                        "exts_text": str(default_exts),
                        "consume_mode": "sequential",
                        "text_fill_mode": "content",
                        "shortage_action": "blank",
                    }
                text_default_exts = (
                    SMART_FILL_TXT_ONLY_EXTS_TEXT
                    if act_type == "clear_input_plus" and sub == "content"
                    else smart_rules.get("text_exts_text", SMART_FILL_TEXT_EXTS_TEXT)
                )
                return {
                    "step_name": step_name,
                    "enabled": True,
                    "target_kind": "text",
                    "exts_text": str(text_default_exts),
                    "consume_mode": "sequential",
                    "text_fill_mode": str(smart_rules.get("text_fill_mode", "content")),
                    "shortage_action": str(smart_rules.get("text_shortage_action", "skip_row")),
                }

            def _get_effective_step_rule(step_key, step_name, act_type, sub, smart_rules, legacy_step_key=None):
                base = _build_default_step_rule(step_name, act_type, sub, smart_rules)
                raw_map = smart_rules.get("step_rules", {})
                if isinstance(raw_map, dict):
                    raw = raw_map.get(step_key)
                    if not isinstance(raw, dict) and legacy_step_key:
                        raw = raw_map.get(legacy_step_key)
                    if isinstance(raw, dict):
                        for k in ["enabled", "exts_text", "consume_mode", "text_fill_mode", "shortage_action"]:
                            if k in raw:
                                base[k] = raw[k]
                base["parsed_exts"] = _parse_exts_text(base.get("exts_text"), base.get("exts_text"))
                return base

            def _open_step_rule_picker(target_mode, folder_path, smart_rules):
                target_columns = _get_smart_fill_target_columns()
                if not target_columns:
                    return None

                dlg_step = QDialog(dlg)
                dlg_step.setWindowTitle("🧩 本次智能填充步骤规则")
                dlg_step.resize(980, 560)
                step_ly = QVBoxLayout(dlg_step)

                intro = QLabel(
                    f"当前目录：{folder_path}\n"
                    f"当前模式：{'填充到空行' if target_mode == 'append_empty' else '覆盖选区'}\n"
                    "这里可以给每个步骤单独设置本次使用的素材规则。点击“开始填充”后，这些规则会保存为该步骤下次的默认值。"
                )
                intro.setWordWrap(True)
                intro.setStyleSheet("color:#555;")
                step_ly.addWidget(intro)

                rule_table = QTableWidget(len(target_columns), 6)
                rule_table.setHorizontalHeaderLabels(["步骤", "目标类型", "扩展名过滤", "消耗方式", "文本填充值", "缺料时"])
                rule_table.verticalHeader().setVisible(False)
                rule_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
                rule_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
                rule_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
                rule_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
                for col in [3, 4, 5]:
                    rule_table.horizontalHeader().setSectionResizeMode(col, QHeaderView.ResizeToContents)
                step_ly.addWidget(rule_table, 1)

                row_widgets = []
                for row_idx, (_, real_col, step_idx, step_name, act_type, sub, step_key) in enumerate(target_columns):
                    legacy_step_key = _make_smart_fill_legacy_step_key(real_col, act_type, sub)
                    rule = _get_effective_step_rule(step_key, step_name, act_type, sub, smart_rules, legacy_step_key=legacy_step_key)
                    type_text = "文件上传/拖拽" if rule["target_kind"] == "file" else "文本填充"
                    if act_type == "clear_input_plus":
                        type_text = "增强文本内容"

                    chk_enabled = QCheckBox(step_name)
                    chk_enabled.setChecked(bool(rule.get("enabled", True)))
                    chk_enabled.setToolTip(f"步骤序号: {step_idx + 1} | 列索引: {real_col} | 动作类型: {act_type}")
                    step_widget = QWidget()
                    step_widget_ly = QHBoxLayout(step_widget)
                    step_widget_ly.setContentsMargins(4, 0, 4, 0)
                    step_widget_ly.setSpacing(0)
                    step_widget_ly.addWidget(chk_enabled)
                    step_widget_ly.addStretch()
                    it_type = QTableWidgetItem(type_text)
                    rule_table.setCellWidget(row_idx, 0, step_widget)
                    rule_table.setItem(row_idx, 1, it_type)

                    cb_exts = _create_exts_combo(
                        current_text=str(rule.get("exts_text", "")),
                        target_kind=rule.get("target_kind", "file"),
                        act_type=act_type,
                        sub=sub,
                        placeholder=".png 或 .txt,.md"
                    )
                    rule_table.setCellWidget(row_idx, 2, cb_exts)

                    cb_consume = QComboBox()
                    cb_consume.addItem("顺序消耗", "sequential")
                    cb_consume.addItem("首个素材重复使用", "repeat_first")
                    idx = cb_consume.findData(rule.get("consume_mode", "sequential"))
                    if idx >= 0:
                        cb_consume.setCurrentIndex(idx)
                    rule_table.setCellWidget(row_idx, 3, cb_consume)

                    cb_text_mode = QComboBox()
                    cb_text_mode.addItem("填入内容", "content")
                    cb_text_mode.addItem("填入路径", "path")
                    idx = cb_text_mode.findData(rule.get("text_fill_mode", "content"))
                    if idx >= 0:
                        cb_text_mode.setCurrentIndex(idx)
                    cb_text_mode.setEnabled(rule["target_kind"] == "text")
                    rule_table.setCellWidget(row_idx, 4, cb_text_mode)

                    cb_shortage = QComboBox()
                    cb_shortage.addItem("留空", "blank")
                    cb_shortage.addItem("写入 [SKIP_ROW]", "skip_row")
                    idx = cb_shortage.findData(rule.get("shortage_action", "blank"))
                    if idx >= 0:
                        cb_shortage.setCurrentIndex(idx)
                    rule_table.setCellWidget(row_idx, 5, cb_shortage)

                    row_data = {
                        "step_key": step_key,
                        "step_name": step_name,
                        "act_type": act_type,
                        "sub": sub,
                        "target_kind": rule["target_kind"],
                        "chk_enabled": chk_enabled,
                        "cb_exts": cb_exts,
                        "cb_consume": cb_consume,
                        "cb_text_mode": cb_text_mode,
                        "cb_shortage": cb_shortage,
                    }

                    def _sync_step_row_enabled(rd):
                        enabled = rd["chk_enabled"].isChecked()
                        rd["cb_exts"].setEnabled(enabled)
                        rd["cb_consume"].setEnabled(enabled)
                        rd["cb_shortage"].setEnabled(enabled)
                        rd["cb_text_mode"].setEnabled(enabled and rd["target_kind"] == "text")

                    _sync_step_row_enabled(row_data)
                    chk_enabled.toggled.connect(lambda _checked, rd=row_data: _sync_step_row_enabled(rd))
                    row_widgets.append(row_data)

                tip = QLabel(
                    "规则说明：\n"
                    "1. 文件上传/拖拽步骤主要看“扩展名过滤”，例如步骤1写 `.png`，步骤2写 `.txt`。\n"
                    "2. “首个素材重复使用”适合 1 张图配多条文案；图会重复用，文案仍可按顺序一条一条消耗。\n"
                    "3. 文本步骤可以选择填入文件内容，或只填入文件路径。\n"
                    "4. “填充到空行”时，会先跳过已经有同名视频成品的文案；例如同目录下有 `1.txt` 和 `1.mp4`，则 `1.txt` 不再参与空行填充。\n"
                    "5. 只有还没有同名视频的文案，才会和该素材组里的图片一起继续填充到空行。"
                )
                tip.setWordWrap(True)
                tip.setStyleSheet("color:#666;")
                step_ly.addWidget(tip)

                btn_row = QHBoxLayout()
                btn_reset = QPushButton("恢复当前步骤默认")
                btn_ok = QPushButton("开始填充")
                btn_cancel = QPushButton("取消")
                btn_ok.setStyleSheet("background:#2196f3; color:white; font-weight:bold;")
                btn_row.addWidget(btn_reset)
                btn_row.addStretch()
                btn_row.addWidget(btn_ok)
                btn_row.addWidget(btn_cancel)
                step_ly.addLayout(btn_row)

                def _reset_step_rule_table():
                    for row_data in row_widgets:
                        default_rule = _build_default_step_rule(
                            row_data["step_name"],
                            row_data["act_type"],
                            row_data["sub"],
                            get_default_smart_fill_rules()
                        )
                        row_data["chk_enabled"].setChecked(bool(default_rule.get("enabled", True)))
                        row_data["cb_exts"].setEditText(str(default_rule.get("exts_text", "")))
                        idx = row_data["cb_consume"].findData(default_rule.get("consume_mode", "sequential"))
                        if idx >= 0:
                            row_data["cb_consume"].setCurrentIndex(idx)
                        idx = row_data["cb_text_mode"].findData(default_rule.get("text_fill_mode", "content"))
                        if idx >= 0:
                            row_data["cb_text_mode"].setCurrentIndex(idx)
                        idx = row_data["cb_shortage"].findData(default_rule.get("shortage_action", "blank"))
                        if idx >= 0:
                            row_data["cb_shortage"].setCurrentIndex(idx)

                def _confirm_step_rules():
                    new_step_rules = {}
                    enabled_count = 0
                    for row_data in row_widgets:
                        enabled = row_data["chk_enabled"].isChecked()
                        if enabled:
                            enabled_count += 1
                        exts_text = _get_combo_text(row_data["cb_exts"])
                        if enabled and not exts_text:
                            QMessageBox.warning(dlg_step, "提示", f"步骤「{row_data['step_name']}」请至少填写一个扩展名。")
                            return
                        new_step_rules[row_data["step_key"]] = {
                            "enabled": enabled,
                            "exts_text": exts_text,
                            "consume_mode": row_data["cb_consume"].currentData(),
                            "text_fill_mode": row_data["cb_text_mode"].currentData(),
                            "shortage_action": row_data["cb_shortage"].currentData(),
                        }
                    if enabled_count <= 0:
                        QMessageBox.warning(dlg_step, "提示", "请至少勾选一个需要参与智能填充的步骤。")
                        return
                    dlg_step._step_rules_result = new_step_rules
                    dlg_step.accept()

                btn_reset.clicked.connect(_reset_step_rule_table)
                btn_ok.clicked.connect(_confirm_step_rules)
                btn_cancel.clicked.connect(dlg_step.reject)

                if dlg_step.exec_() != QDialog.Accepted:
                    return None
                return getattr(dlg_step, "_step_rules_result", None)

            def _open_smart_fill_rules_dialog():
                rules = get_smart_fill_rules(self.config)
                dlg_rule = QDialog(dlg)
                dlg_rule.setWindowTitle("⚙️ 智能填充规则设计")
                dlg_rule.resize(760, 620)
                rule_ly = QVBoxLayout(dlg_rule)

                intro = QLabel(
                    "这里可以自定义目录智能填充规则。保存后，后续“填充到空行”和“覆盖选区”都会按这套规则执行。"
                )
                intro.setWordWrap(True)
                intro.setStyleSheet("color:#555;")
                rule_ly.addWidget(intro)

                form = QFormLayout()

                cb_bundle_mode = QComboBox()
                cb_bundle_mode.addItem("自动判断：根目录素材 + 一级子目录素材", "auto")
                cb_bundle_mode.addItem("整个当前目录算 1 组", "root_only")
                cb_bundle_mode.addItem("每个一级子目录算 1 组", "subdirs_only")
                idx = cb_bundle_mode.findData(rules.get("bundle_mode", "auto"))
                if idx >= 0:
                    cb_bundle_mode.setCurrentIndex(idx)
                form.addRow("素材分组方式:", cb_bundle_mode)

                chk_recursive = QCheckBox("子目录内递归扫描")
                chk_recursive.setChecked(bool(rules.get("scan_subdirs_recursive", True)))
                form.addRow("目录扫描:", chk_recursive)

                cb_name_mode = QComboBox()
                cb_name_mode.addItem("只认纯序号文件名", "pure_number")
                cb_name_mode.addItem("只要包含数字就算", "contains_number")
                cb_name_mode.addItem("非中文且含字母", "non_chinese")
                cb_name_mode.addItem("非中文且无编号", "non_chinese_no_number")
                cb_name_mode.addItem("所有文本文件都参与", "all")
                cb_name_mode.addItem("使用正则表达式", "regex")
                idx = cb_name_mode.findData(rules.get("text_filename_mode", "pure_number"))
                if idx >= 0:
                    cb_name_mode.setCurrentIndex(idx)
                form.addRow("文本文件名规则:", cb_name_mode)

                le_name_regex = QLineEdit(str(rules.get("custom_text_regex", "")))
                le_name_regex.setPlaceholderText(r"例如: ^\d+$ 或 ^[A-Za-z_-]+$")
                form.addRow("文件名正则:", le_name_regex)

                cb_upload_source = QComboBox()
                cb_upload_source.addItem("上传类步骤只用图片", "images_only")
                cb_upload_source.addItem("上传类先用图片，没图再退回文件", "images_then_files")
                idx = cb_upload_source.findData(rules.get("upload_source", "images_only"))
                if idx >= 0:
                    cb_upload_source.setCurrentIndex(idx)
                form.addRow("上传类来源:", cb_upload_source)

                cb_file_source = QComboBox()
                cb_file_source.addItem("文件类步骤优先文件，没文件再退回图片", "files_then_images")
                cb_file_source.addItem("文件类步骤只用文件", "files_only")
                cb_file_source.addItem("文件类步骤只用图片", "images_only")
                idx = cb_file_source.findData(rules.get("file_source", "files_then_images"))
                if idx >= 0:
                    cb_file_source.setCurrentIndex(idx)
                form.addRow("文件类来源:", cb_file_source)

                cb_text_fill_mode = QComboBox()
                cb_text_fill_mode.addItem("文本步骤填入文件内容", "content")
                cb_text_fill_mode.addItem("文本步骤填入文件路径", "path")
                idx = cb_text_fill_mode.findData(rules.get("text_fill_mode", "content"))
                if idx >= 0:
                    cb_text_fill_mode.setCurrentIndex(idx)
                form.addRow("文本填充值:", cb_text_fill_mode)

                cb_text_shortage = QComboBox()
                cb_text_shortage.addItem("文本不够时写入 [SKIP_ROW]", "skip_row")
                cb_text_shortage.addItem("文本不够时留空", "blank")
                idx = cb_text_shortage.findData(rules.get("text_shortage_action", "skip_row"))
                if idx >= 0:
                    cb_text_shortage.setCurrentIndex(idx)
                form.addRow("文本不足时:", cb_text_shortage)

                cb_image_exts = _create_exts_combo(
                    current_text=str(rules.get("image_exts_text", SMART_FILL_IMAGE_EXTS_TEXT)),
                    target_kind="file",
                    act_type="upload",
                    placeholder=".png,.jpg,.jpeg"
                )
                form.addRow("图片扩展名:", cb_image_exts)

                cb_text_exts = _create_exts_combo(
                    current_text=str(rules.get("text_exts_text", SMART_FILL_TEXT_EXTS_TEXT)),
                    target_kind="text",
                    act_type="input",
                    placeholder=".txt,.md,.csv"
                )
                form.addRow("文本扩展名:", cb_text_exts)

                cb_file_exts = _create_exts_combo(
                    current_text=str(rules.get("file_exts_text", SMART_FILL_FILE_EXTS_TEXT)),
                    target_kind="file",
                    act_type="run_app",
                    placeholder=".png,.jpg,.txt,.zip,.mp4"
                )
                form.addRow("通用文件扩展名:", cb_file_exts)

                rule_ly.addLayout(form)

                hint = QLabel(
                    "提示：\n"
                    "1. “填充到空行”永远不会覆盖已有值。\n"
                    "2. “覆盖选区”只会覆盖你当前选中的目标格。\n"
                    "3. “非中文且含字母”不会再匹配 1、2、3 这类纯数字文件名。\n"
                    "4. “非中文且无编号”更严格，适合只抓纯英文文件名，不要数字。\n"
                    "5. “填充到空行”会自动跳过已经有同名视频的文案，只保留待生成视频的文案继续填充。\n"
                    "6. 如果你想做更细命名规则，继续用正则表达式。"
                )
                hint.setStyleSheet("color:#666;")
                hint.setWordWrap(True)
                rule_ly.addWidget(hint)

                btn_row = QHBoxLayout()
                btn_reset_rule = QPushButton("恢复默认规则")
                btn_save_rule = QPushButton("保存规则")
                btn_close_rule = QPushButton("关闭")
                btn_save_rule.setStyleSheet("background:#2196f3; color:white; font-weight:bold;")
                btn_row.addWidget(btn_reset_rule)
                btn_row.addStretch()
                btn_row.addWidget(btn_save_rule)
                btn_row.addWidget(btn_close_rule)
                rule_ly.addLayout(btn_row)

                def _apply_rule_form(rule_dict):
                    idx = cb_bundle_mode.findData(rule_dict.get("bundle_mode", "auto"))
                    if idx >= 0: cb_bundle_mode.setCurrentIndex(idx)
                    chk_recursive.setChecked(bool(rule_dict.get("scan_subdirs_recursive", True)))
                    idx = cb_name_mode.findData(rule_dict.get("text_filename_mode", "pure_number"))
                    if idx >= 0: cb_name_mode.setCurrentIndex(idx)
                    le_name_regex.setText(str(rule_dict.get("custom_text_regex", "")))
                    idx = cb_upload_source.findData(rule_dict.get("upload_source", "images_only"))
                    if idx >= 0: cb_upload_source.setCurrentIndex(idx)
                    idx = cb_file_source.findData(rule_dict.get("file_source", "files_then_images"))
                    if idx >= 0: cb_file_source.setCurrentIndex(idx)
                    idx = cb_text_fill_mode.findData(rule_dict.get("text_fill_mode", "content"))
                    if idx >= 0: cb_text_fill_mode.setCurrentIndex(idx)
                    idx = cb_text_shortage.findData(rule_dict.get("text_shortage_action", "skip_row"))
                    if idx >= 0: cb_text_shortage.setCurrentIndex(idx)
                    cb_image_exts.setEditText(str(rule_dict.get("image_exts_text", SMART_FILL_IMAGE_EXTS_TEXT)))
                    cb_text_exts.setEditText(str(rule_dict.get("text_exts_text", SMART_FILL_TEXT_EXTS_TEXT)))
                    cb_file_exts.setEditText(str(rule_dict.get("file_exts_text", SMART_FILL_FILE_EXTS_TEXT)))

                def _collect_rule_form():
                    return {
                        "bundle_mode": cb_bundle_mode.currentData(),
                        "scan_subdirs_recursive": chk_recursive.isChecked(),
                        "text_filename_mode": cb_name_mode.currentData(),
                        "custom_text_regex": le_name_regex.text().strip(),
                        "upload_source": cb_upload_source.currentData(),
                        "file_source": cb_file_source.currentData(),
                        "text_fill_mode": cb_text_fill_mode.currentData(),
                        "text_shortage_action": cb_text_shortage.currentData(),
                        "image_exts_text": _get_combo_text(cb_image_exts),
                        "text_exts_text": _get_combo_text(cb_text_exts),
                        "file_exts_text": _get_combo_text(cb_file_exts),
                        "step_rules": dict(rules.get("step_rules", {})) if isinstance(rules.get("step_rules", {}), dict) else {},
                    }

                def _reset_rule_form():
                    _apply_rule_form(get_default_smart_fill_rules())

                def _save_rule_form():
                    new_rules = _collect_rule_form()
                    if new_rules["text_filename_mode"] == "regex" and not new_rules["custom_text_regex"]:
                        QMessageBox.warning(dlg_rule, "提示", "当你选择“使用正则表达式”时，请填写一个文件名正则。")
                        return
                    if new_rules["text_filename_mode"] == "regex":
                        try:
                            re.compile(new_rules["custom_text_regex"])
                        except Exception as e:
                            QMessageBox.warning(dlg_rule, "正则无效", f"文件名正则表达式无效：\n{e}")
                            return
                    self.config["smart_fill_rules"] = new_rules
                    save_config(self.config)
                    QMessageBox.information(dlg_rule, "成功", "智能填充规则已保存。")
                    dlg_rule.accept()

                btn_reset_rule.clicked.connect(_reset_rule_form)
                btn_save_rule.clicked.connect(_save_rule_form)
                btn_close_rule.clicked.connect(dlg_rule.reject)
                dlg_rule.exec_()

            def _is_smart_fill_target_cell(item):
                if not item:
                    return False
                _, _, _, cell_act_type, cell_sub = preview_cols[item.column()]
                return (
                    (cell_act_type in ["upload", "drag_file", "run_app"] and cell_sub is None) or
                    (cell_act_type in ["input", "clear_input"] and cell_sub is None) or
                    (cell_act_type == "clear_input_plus" and cell_sub == "content")
                )

            def _is_enabled_smart_fill_cell(item, smart_rules=None):
                if not _is_smart_fill_target_cell(item):
                    return False
                if not isinstance(smart_rules, dict):
                    return True
                real_col, step_idx, step_name, cell_act_type, cell_sub = preview_cols[item.column()]
                step_key = _make_smart_fill_step_key(step_idx, cell_act_type, cell_sub)
                legacy_step_key = _make_smart_fill_legacy_step_key(real_col, cell_act_type, cell_sub)
                step_rule = _get_effective_step_rule(
                    step_key,
                    step_name,
                    cell_act_type,
                    cell_sub,
                    smart_rules,
                    legacy_step_key=legacy_step_key
                )
                return bool(step_rule.get("enabled", True))

            def _iter_preview_targets_for_smart_fill(target_mode="append_empty", smart_rules=None):
                selected_indexes = preview_table.selectedIndexes()
                targets = []
                if target_mode == "selected_only":
                    for idx in selected_indexes:
                        it = preview_table.item(idx.row(), idx.column())
                        if it and _is_enabled_smart_fill_cell(it, smart_rules=smart_rules):
                            targets.append(it)
                else:
                    candidate_rows = []
                    for rr in range(preview_table.rowCount()):
                        row_items = []
                        row_has_value = False
                        for cc in range(preview_table.columnCount()):
                            it = preview_table.item(rr, cc)
                            if not it or not _is_enabled_smart_fill_cell(it, smart_rules=smart_rules):
                                continue
                            row_items.append(it)
                            if str(it.text()).strip():
                                row_has_value = True
                        if row_items and not row_has_value:
                            candidate_rows.extend(row_items)
                    targets = candidate_rows

                row_map = {}
                for it in targets:
                    row_map.setdefault(it.row(), []).append(it)
                for row_items in row_map.values():
                    row_items.sort(key=lambda x: x.column())
                row_targets = [(row_idx, row_map[row_idx]) for row_idx in sorted(row_map.keys())]
                return row_targets

            def _take_smart_value(items, cursor_holder, repeat_single=False):
                if not items:
                    return None
                if repeat_single and len(items) == 1:
                    return items[0]
                idx = cursor_holder[0]
                if idx >= len(items):
                    return None
                cursor_holder[0] += 1
                return items[idx]

            def _peek_smart_value(items, cursor_holder, repeat_single=False):
                if not items:
                    return None
                if repeat_single and len(items) == 1:
                    return items[0]
                idx = cursor_holder[0]
                if idx >= len(items):
                    return None
                return items[idx]

            def _build_bundle_step_items(bundle, rule, pending_texts=None):
                parsed_exts = set(rule.get("parsed_exts") or [])
                target_kind = str(rule.get("target_kind", "text"))
                result = []
                seen = set()

                def _append(path, payload):
                    if not path:
                        return
                    norm_key = os.path.normcase(os.path.normpath(str(path)))
                    if norm_key in seen:
                        return
                    ext = os.path.splitext(str(path).lower())[1]
                    if parsed_exts and ext not in parsed_exts:
                        return
                    seen.add(norm_key)
                    result.append(payload)

                if target_kind == "text":
                    text_source_list = pending_texts if pending_texts is not None else bundle.get("texts", [])
                    for text_data in text_source_list:
                        _append(text_data.get("path", ""), text_data)
                    return result

                for path in bundle.get("images", []):
                    _append(path, path)
                for path in bundle.get("files", []):
                    _append(path, path)
                for text_data in bundle.get("texts", []):
                    _append(text_data.get("path", ""), text_data.get("path", ""))
                return result

            def _prepare_bundle_runtime_state(bundle, target_items, smart_rules, pending_text_only=False):
                pending_texts = _get_pending_texts_from_bundle(bundle) if pending_text_only else None
                runtime = {
                    "bundle": bundle,
                    "step_items": {},
                    "step_cursors": {},
                    "step_rules": {},
                    "pending_text_only": pending_text_only,
                    "pending_texts": pending_texts if pending_texts is not None else [],
                }
                for it in target_items:
                    real_col, step_idx, step_name, cell_act_type, cell_sub = preview_cols[it.column()]
                    step_key = _make_smart_fill_step_key(step_idx, cell_act_type, cell_sub)
                    if step_key in runtime["step_rules"]:
                        continue
                    legacy_step_key = _make_smart_fill_legacy_step_key(real_col, cell_act_type, cell_sub)
                    rule = _get_effective_step_rule(
                        step_key, step_name, cell_act_type, cell_sub, smart_rules, legacy_step_key=legacy_step_key
                    )
                    runtime["step_rules"][step_key] = rule
                    if not bool(rule.get("enabled", True)):
                        runtime["step_items"][step_key] = []
                        runtime["step_cursors"][step_key] = [0]
                        continue
                    runtime["step_items"][step_key] = _build_bundle_step_items(bundle, rule, pending_texts=pending_texts)
                    runtime["step_cursors"][step_key] = [0]
                return runtime

            def _describe_bundle_exhaust_reason(bundle, bundle_runtime, used_text_files=None):
                if not bundle_runtime:
                    return "未命中当前步骤规则"

                used_name_set = {
                    str(name) for name in (used_text_files or []) if str(name).strip()
                }

                bundle_name = str(bundle.get("name", "未命名") or "未命名")
                all_texts = list(bundle.get("texts", []) or [])
                pending_texts = bundle_runtime.get("pending_texts") if bundle_runtime.get("pending_text_only") else all_texts
                pending_name_set = {
                    os.path.basename(str(text_data.get("path", "")))
                    for text_data in pending_texts
                    if isinstance(text_data, dict) and text_data.get("path")
                }
                pending_unused_names = [name for name in pending_name_set if name and name not in used_name_set]

                if bundle_runtime.get("pending_text_only") and all_texts and not pending_texts:
                    return "文案都有同名视频"

                text_step_keys = []
                file_step_keys = []
                for step_key, rule in (bundle_runtime.get("step_rules") or {}).items():
                    if not bool(rule.get("enabled", True)):
                        continue
                    if str(rule.get("target_kind", "text")) == "file":
                        file_step_keys.append(step_key)
                    else:
                        text_step_keys.append(step_key)

                if not text_step_keys and not file_step_keys:
                    return "未启用任何步骤"

                for step_key in text_step_keys:
                    if not bundle_runtime.get("step_items", {}).get(step_key):
                        return "文案规则未命中"
                for step_key in file_step_keys:
                    if not bundle_runtime.get("step_items", {}).get(step_key):
                        return "图片规则未命中"

                for step_key in text_step_keys:
                    step_rule = (bundle_runtime.get("step_rules") or {}).get(step_key, {})
                    if str(step_rule.get("consume_mode", "sequential")) != "sequential":
                        continue
                    step_items = bundle_runtime.get("step_items", {}).get(step_key, []) or []
                    step_cursor = (bundle_runtime.get("step_cursors", {}).get(step_key) or [0])[0]
                    if step_cursor >= len(step_items) and pending_unused_names:
                        return "文案数量不足"

                for step_key in file_step_keys:
                    step_rule = (bundle_runtime.get("step_rules") or {}).get(step_key, {})
                    if str(step_rule.get("consume_mode", "sequential")) != "sequential":
                        continue
                    step_items = bundle_runtime.get("step_items", {}).get(step_key, []) or []
                    step_cursor = (bundle_runtime.get("step_cursors", {}).get(step_key) or [0])[0]
                    if step_cursor >= len(step_items):
                        if pending_unused_names and used_text_files:
                            return f"图片已被 {used_text_files[-1]} 占用"
                        return "图片数量不足"

                if pending_unused_names:
                    return "未通过当前步骤规则"
                if all_texts:
                    return f"素材组「{bundle_name}」已用完"
                return "没有可用文案"

            def _collect_bundle_skip_log_lines(bundle, bundle_runtime, used_text_files, remaining_row_count):
                if not bundle_runtime:
                    return []

                bundle_name = str(bundle.get("name", "未命名") or "未命名")
                all_texts = list(bundle.get("texts", []) or [])
                if not all_texts:
                    return []

                pending_texts = bundle_runtime.get("pending_texts") if bundle_runtime.get("pending_text_only") else all_texts
                pending_name_set = {
                    os.path.basename(str(text_data.get("path", "")))
                    for text_data in pending_texts
                    if isinstance(text_data, dict) and text_data.get("path")
                }
                used_name_set = {str(name) for name in used_text_files if str(name).strip()}
                skipped_lines = []

                reason_for_pending = "没有空行" if remaining_row_count <= 0 else _describe_bundle_exhaust_reason(
                    bundle, bundle_runtime, used_text_files
                )

                for text_data in all_texts:
                    if not isinstance(text_data, dict):
                        continue
                    text_path = str(text_data.get("path", "") or "")
                    if not text_path:
                        continue
                    text_name = os.path.basename(text_path)
                    if not text_name or text_name in used_name_set:
                        continue
                    if text_name not in pending_name_set:
                        skipped_lines.append(f"{bundle_name}: {text_name} 未填充，原因：已有同名视频")
                    else:
                        skipped_lines.append(f"{bundle_name}: {text_name} 未填充，原因：{reason_for_pending}")
                return skipped_lines

            def _bundle_has_more_rows(bundle_runtime, target_items, pending_text_only=False):
                text_step_keys = []
                for it in target_items:
                    real_col, step_idx, _, cell_act_type, cell_sub = preview_cols[it.column()]
                    if (
                        (cell_act_type in ["input", "clear_input"] and cell_sub is None) or
                        (cell_act_type == "clear_input_plus" and cell_sub == "content")
                    ):
                        step_key = _make_smart_fill_step_key(step_idx, cell_act_type, cell_sub)
                        if step_key not in text_step_keys:
                            text_step_keys.append(step_key)
                if pending_text_only and text_step_keys:
                    for step_key in text_step_keys:
                        step_rule = bundle_runtime["step_rules"].get(step_key, {})
                        if not bool(step_rule.get("enabled", True)):
                            continue
                        items = bundle_runtime["step_items"].get(step_key, [])
                        cursor_holder = bundle_runtime["step_cursors"].setdefault(step_key, [0])
                        if cursor_holder[0] < len(items):
                            break
                    else:
                        return False
                has_enabled_step = False
                has_sequential = False
                for it in target_items:
                    real_col, step_idx, step_name, cell_act_type, cell_sub = preview_cols[it.column()]
                    step_key = _make_smart_fill_step_key(step_idx, cell_act_type, cell_sub)
                    rule = bundle_runtime["step_rules"].get(step_key) or _get_effective_step_rule(
                        step_key,
                        step_name,
                        cell_act_type,
                        cell_sub,
                        smart_rules,
                        legacy_step_key=_make_smart_fill_legacy_step_key(real_col, cell_act_type, cell_sub)
                    )
                    if not bool(rule.get("enabled", True)):
                        continue
                    has_enabled_step = True
                    items = bundle_runtime["step_items"].get(step_key, [])
                    cursor_holder = bundle_runtime["step_cursors"].setdefault(step_key, [0])
                    if str(rule.get("consume_mode", "sequential")) == "sequential":
                        has_sequential = True
                        if cursor_holder[0] < len(items):
                            return True
                if not has_enabled_step:
                    return False
                if has_sequential:
                    return False
                for it in target_items:
                    real_col, step_idx, step_name, cell_act_type, cell_sub = preview_cols[it.column()]
                    step_key = _make_smart_fill_step_key(step_idx, cell_act_type, cell_sub)
                    rule = bundle_runtime["step_rules"].get(step_key) or _get_effective_step_rule(
                        step_key,
                        step_name,
                        cell_act_type,
                        cell_sub,
                        smart_rules,
                        legacy_step_key=_make_smart_fill_legacy_step_key(real_col, cell_act_type, cell_sub)
                    )
                    if not bool(rule.get("enabled", True)):
                        continue
                    if bundle_runtime["step_items"].get(step_key):
                        return True
                return False

            def _fill_row_with_bundle(target_items, bundle_runtime, smart_rules, pending_text_only=False):
                changed = 0
                skip_count = 0
                skip_token = "[SKIP_ROW]"
                used_text_files = []
                used_file_sources = []
                used_real_source = False
                reserved_text_sources = {}
                reserved_file_sources = {}
                row_anchor_text_data = None

                if pending_text_only:
                    text_targets = []
                    file_targets = []
                    for it in target_items:
                        real_col, step_idx, step_name, cell_act_type, cell_sub = preview_cols[it.column()]
                        step_key = _make_smart_fill_step_key(step_idx, cell_act_type, cell_sub)
                        step_rule = bundle_runtime["step_rules"].get(step_key) or _get_effective_step_rule(
                            step_key,
                            step_name,
                            cell_act_type,
                            cell_sub,
                            smart_rules,
                            legacy_step_key=_make_smart_fill_legacy_step_key(real_col, cell_act_type, cell_sub)
                        )
                        if not bool(step_rule.get("enabled", True)):
                            continue
                        if (
                            (cell_act_type in ["input", "clear_input"] and cell_sub is None) or
                            (cell_act_type == "clear_input_plus" and cell_sub == "content")
                        ):
                            text_targets.append((it, step_key, step_name, cell_act_type, cell_sub))
                        elif cell_act_type in ["upload", "drag_file", "run_app"] and cell_sub is None:
                            file_targets.append((it, step_key, step_name, cell_act_type, cell_sub))

                    if text_targets:
                        first_it, step_key, step_name, cell_act_type, cell_sub = text_targets[0]
                        step_rule = bundle_runtime["step_rules"].get(step_key) or _get_effective_step_rule(
                            step_key,
                            step_name,
                            cell_act_type,
                            cell_sub,
                            smart_rules,
                            legacy_step_key=_make_smart_fill_legacy_step_key(real_col, cell_act_type, cell_sub)
                        )
                        consume_mode = str(step_rule.get("consume_mode", "sequential"))
                        repeat_single = (consume_mode == "repeat_first")
                        step_items = bundle_runtime["step_items"].get(step_key, [])
                        step_cursor = bundle_runtime["step_cursors"].setdefault(step_key, [0])
                        text_preview = _peek_smart_value(step_items, step_cursor, repeat_single=repeat_single)
                        if text_preview is None:
                            return changed, skip_count, used_text_files, used_file_sources, used_real_source
                    else:
                        text_preview = None

                    for _it, step_key, step_name, cell_act_type, cell_sub in file_targets:
                        step_rule = bundle_runtime["step_rules"].get(step_key) or _get_effective_step_rule(
                            step_key,
                            step_name,
                            cell_act_type,
                            cell_sub,
                            smart_rules,
                            legacy_step_key=_make_smart_fill_legacy_step_key(real_col, cell_act_type, cell_sub)
                        )
                        consume_mode = str(step_rule.get("consume_mode", "sequential"))
                        repeat_single = (consume_mode == "repeat_first")
                        step_items = bundle_runtime["step_items"].get(step_key, [])
                        step_cursor = bundle_runtime["step_cursors"].setdefault(step_key, [0])
                        source_preview = _peek_smart_value(step_items, step_cursor, repeat_single=repeat_single)
                        if source_preview is None:
                            return changed, skip_count, used_text_files, used_file_sources, used_real_source

                    if text_targets:
                        _it, step_key, step_name, cell_act_type, cell_sub = text_targets[0]
                        step_rule = bundle_runtime["step_rules"].get(step_key) or _get_effective_step_rule(
                            step_key,
                            step_name,
                            cell_act_type,
                            cell_sub,
                            smart_rules,
                            legacy_step_key=_make_smart_fill_legacy_step_key(real_col, cell_act_type, cell_sub)
                        )
                        consume_mode = str(step_rule.get("consume_mode", "sequential"))
                        repeat_single = (consume_mode == "repeat_first")
                        step_items = bundle_runtime["step_items"].get(step_key, [])
                        step_cursor = bundle_runtime["step_cursors"].setdefault(step_key, [0])
                        text_data = _take_smart_value(step_items, step_cursor, repeat_single=repeat_single)
                        if text_data is None:
                            return changed, skip_count, used_text_files, used_file_sources, used_real_source
                        reserved_text_sources[step_key] = text_data
                        row_anchor_text_data = text_data

                    for _it, step_key, step_name, cell_act_type, cell_sub in file_targets:
                        step_rule = bundle_runtime["step_rules"].get(step_key) or _get_effective_step_rule(
                            step_key,
                            step_name,
                            cell_act_type,
                            cell_sub,
                            smart_rules,
                            legacy_step_key=_make_smart_fill_legacy_step_key(real_col, cell_act_type, cell_sub)
                        )
                        consume_mode = str(step_rule.get("consume_mode", "sequential"))
                        repeat_single = (consume_mode == "repeat_first")
                        step_items = bundle_runtime["step_items"].get(step_key, [])
                        step_cursor = bundle_runtime["step_cursors"].setdefault(step_key, [0])
                        source_path = _take_smart_value(step_items, step_cursor, repeat_single=repeat_single)
                        if source_path is None:
                            return 0, 0, [], [], False
                        reserved_file_sources[step_key] = source_path

                    if text_targets and row_anchor_text_data is None:
                        return changed, skip_count, used_text_files, used_file_sources, used_real_source

                for it in target_items:
                    real_col, step_idx, step_name, cell_act_type, cell_sub = preview_cols[it.column()]
                    step_key = _make_smart_fill_step_key(step_idx, cell_act_type, cell_sub)
                    step_rule = bundle_runtime["step_rules"].get(step_key) or _get_effective_step_rule(
                        step_key,
                        step_name,
                        cell_act_type,
                        cell_sub,
                        smart_rules,
                        legacy_step_key=_make_smart_fill_legacy_step_key(real_col, cell_act_type, cell_sub)
                    )
                    if not bool(step_rule.get("enabled", True)):
                        continue
                    consume_mode = str(step_rule.get("consume_mode", "sequential"))
                    repeat_single = (consume_mode == "repeat_first")
                    step_items = bundle_runtime["step_items"].get(step_key, [])
                    step_cursor = bundle_runtime["step_cursors"].setdefault(step_key, [0])
                    fill_val = None

                    if cell_act_type in ["upload", "drag_file", "run_app"] and cell_sub is None:
                        source_path = reserved_file_sources.pop(step_key, None)
                        if source_path is None:
                            preferred_item = None
                            source_path = _take_smart_value_with_preferred(
                                step_items, step_cursor, preferred_item=preferred_item, repeat_single=repeat_single
                            )
                        if source_path:
                            fill_val = source_path
                            used_file_sources.append(os.path.basename(str(source_path)))
                            used_real_source = True
                        else:
                            fill_val = skip_token if step_rule.get("shortage_action", "blank") == "skip_row" else ""
                    elif cell_act_type in ["input", "clear_input"] and cell_sub is None:
                        text_data = reserved_text_sources.pop(step_key, None)
                        if text_data is None:
                            text_data = _take_smart_value(step_items, step_cursor, repeat_single=repeat_single)
                        if text_data:
                            fill_val = text_data["path"] if step_rule.get("text_fill_mode") == "path" else text_data["content"]
                            if text_data.get("path"):
                                used_text_files.append(os.path.basename(text_data["path"]))
                            used_real_source = True
                        else:
                            fill_val = skip_token if step_rule.get("shortage_action", "skip_row") == "skip_row" else ""
                    elif cell_act_type == "clear_input_plus" and cell_sub == "content":
                        text_data = reserved_text_sources.pop(step_key, None)
                        if text_data is None:
                            text_data = _take_smart_value(step_items, step_cursor, repeat_single=repeat_single)
                        if text_data:
                            fill_val = text_data["path"] if step_rule.get("text_fill_mode") == "path" else text_data["content"]
                            if text_data.get("path"):
                                used_text_files.append(os.path.basename(text_data["path"]))
                            used_real_source = True
                        else:
                            fill_val = skip_token if step_rule.get("shortage_action", "skip_row") == "skip_row" else ""

                    if fill_val is None:
                        continue

                    it.setText(fill_val)
                    it.setToolTip(str(fill_val))
                    it.setData(Qt.UserRole + 10, None)
                    it.setBackground(QColor("#fff9c4") if fill_val != skip_token else QColor("#fff3e0"))
                    changed += 1
                    if fill_val == skip_token:
                        skip_count += 1

                return changed, skip_count, used_text_files, used_file_sources, used_real_source

            def _smart_fill_from_directory(target_mode="append_empty"):
                folder_path = None
                current_item = file_tree.currentItem()
                if current_item:
                    current_path = current_item.data(0, Qt.UserRole)
                    if current_path and os.path.isdir(current_path):
                        folder_path = current_path

                if not folder_path and _last_path[0] and os.path.isdir(_last_path[0]):
                    folder_path = _last_path[0]

                if not folder_path:
                    folder_path = QFileDialog.getExistingDirectory(dlg, "选择智能填充目录")
                if not folder_path:
                    return

                smart_rules = get_smart_fill_rules(self.config)
                current_step_rules = _open_step_rule_picker(target_mode, folder_path, smart_rules)
                if current_step_rules is None:
                    return
                if current_step_rules:
                    merged_rules = get_smart_fill_rules(self.config)
                    merged_step_rules = dict(merged_rules.get("step_rules", {})) if isinstance(merged_rules.get("step_rules", {}), dict) else {}
                    merged_step_rules.update(current_step_rules)
                    merged_rules["step_rules"] = merged_step_rules
                    self.config["smart_fill_rules"] = merged_rules
                    save_config(self.config)
                    smart_rules = merged_rules

                enabled_step_count = 0
                for _pc_idx, real_col, step_idx, step_name, act_type, sub, step_key in _get_smart_fill_target_columns():
                    rule = _get_effective_step_rule(
                        step_key,
                        step_name,
                        act_type,
                        sub,
                        smart_rules,
                        legacy_step_key=_make_smart_fill_legacy_step_key(real_col, act_type, sub)
                    )
                    if bool(rule.get("enabled", True)):
                        enabled_step_count += 1
                if enabled_step_count <= 0:
                    QMessageBox.warning(dlg, "提示", "请至少勾选一个需要参与智能填充的步骤。")
                    return

                bundles = collect_smart_fill_bundles(folder_path, smart_rules)
                if not bundles:
                    QMessageBox.warning(dlg, "提示", "所选目录中没有找到可用于智能填充的图片或文本文件。")
                    return

                row_targets = _iter_preview_targets_for_smart_fill(target_mode=target_mode, smart_rules=smart_rules)
                if not row_targets:
                    if target_mode == "selected_only":
                        QMessageBox.warning(dlg, "提示", "请先选中要覆盖的目标单元格或整行，再使用“覆盖选区”。")
                    else:
                        QMessageBox.warning(dlg, "提示", "没有找到可填充的空行。若要覆盖指定位置，请先手动选中那些单元格/行再执行智能填充。")
                    return

                _save_undo_state()
                row_text_hit_lines = []
                row_ptr = 0
                last_bundle_name = None
                last_bundle_end_reason = None
                for bundle in bundles:
                    if row_ptr >= len(row_targets):
                        break
                    bundle_name = bundle.get("name", "未命名")
                    bundle_runtime = None
                    bundle_used_text_files = []
                    while row_ptr < len(row_targets):
                        target_row_idx, target_items = row_targets[row_ptr]
                        if bundle_runtime is None:
                            bundle_runtime = _prepare_bundle_runtime_state(
                                bundle,
                                target_items,
                                smart_rules,
                                pending_text_only=(target_mode == "append_empty")
                            )
                        elif not _bundle_has_more_rows(
                            bundle_runtime,
                            target_items,
                            pending_text_only=(target_mode == "append_empty")
                        ):
                            break

                        changed, skipped, hit_text_files, hit_file_sources, used_real_source = _fill_row_with_bundle(
                            target_items,
                            bundle_runtime,
                            smart_rules,
                            pending_text_only=(target_mode == "append_empty")
                        )
                        if not changed and not used_real_source:
                            break

                        header_item = preview_table.verticalHeaderItem(target_row_idx)
                        row_label = header_item.text() if header_item and header_item.text() else f"第{target_row_idx + 1}行"
                        row_dedup_text_files = []
                        row_seen_text_files = set()
                        for name in hit_text_files:
                            if not name or name in row_seen_text_files:
                                continue
                            row_seen_text_files.add(name)
                            row_dedup_text_files.append(name)
                        row_dedup_file_sources = []
                        row_seen_file_sources = set()
                        for name in hit_file_sources:
                            if not name or name in row_seen_file_sources:
                                continue
                            row_seen_file_sources.add(name)
                            row_dedup_file_sources.append(name)
                        if row_dedup_text_files and row_dedup_file_sources:
                            row_text_hit_lines.append(
                                f"{row_label}: 使用 {'、'.join(row_dedup_text_files)} + {'、'.join(row_dedup_file_sources)}"
                            )
                        elif row_dedup_text_files:
                            row_text_hit_lines.append(f"{row_label}: 使用 {'、'.join(row_dedup_text_files)}")
                        elif row_dedup_file_sources:
                            row_text_hit_lines.append(f"{row_label}: 使用 {'、'.join(row_dedup_file_sources)}")
                        else:
                            row_text_hit_lines.append(f"{row_label}: 使用素材组「{bundle_name}」")
                        bundle_used_text_files.extend(row_dedup_text_files)
                        row_ptr += 1

                    row_text_hit_lines.extend(
                        _collect_bundle_skip_log_lines(
                            bundle,
                            bundle_runtime,
                            bundle_used_text_files,
                            len(row_targets) - row_ptr
                        )
                    )
                    last_bundle_name = str(bundle_name or "未命名")
                    last_bundle_end_reason = _describe_bundle_exhaust_reason(
                        bundle,
                        bundle_runtime,
                        bundle_used_text_files
                    )

                if row_ptr < len(row_targets):
                    first_row_idx, _target_items = row_targets[row_ptr]
                    last_row_idx, _target_items = row_targets[-1]
                    first_header_item = preview_table.verticalHeaderItem(first_row_idx)
                    last_header_item = preview_table.verticalHeaderItem(last_row_idx)
                    first_label = first_header_item.text() if first_header_item and first_header_item.text() else f"第{first_row_idx + 1}行"
                    last_label = last_header_item.text() if last_header_item and last_header_item.text() else f"第{last_row_idx + 1}行"
                    remaining_count = len(row_targets) - row_ptr
                    reason_text = "所有素材组都已用完"
                    if last_bundle_name and last_bundle_end_reason:
                        reason_text += f"（最后停在「{last_bundle_name}」：{last_bundle_end_reason}）"
                    if remaining_count == 1:
                        row_text_hit_lines.append(f"{first_label}: 未填充，原因：{reason_text}")
                    else:
                        row_text_hit_lines.append(
                            f"{first_label} - {last_label}: 共 {remaining_count} 行未填充，原因：{reason_text}"
                        )

                preview_table.viewport().update()
                self._log("🧩 智能填充逐行结果：", "blue")
                if row_text_hit_lines:
                    for line in row_text_hit_lines:
                        self._log(f"   {line}", "gray")
                else:
                    self._log("   未产生逐行文本文件对应关系", "gray")

            btn_smart_fill_append.clicked.connect(lambda: _smart_fill_from_directory("append_empty"))
            btn_smart_fill_selected.clicked.connect(lambda: _smart_fill_from_directory("selected_only"))
            btn_smart_fill_rule.clicked.connect(_open_smart_fill_rules_dialog)

            # 点击预览功能
            def _on_file_clicked(item):
                path = item.data(0, Qt.UserRole)
                if not path or not os.path.isfile(path):
                    lbl_file_info.setText("")
                    preview_stack.setCurrentIndex(0)
                    lbl_preview_img.setPixmap(QPixmap())
                    lbl_preview_img.setText("选中文件即可预览")
                    return
                
                # 获取基础信息
                try:
                    size_kb = os.path.getsize(path) / 1024
                    mtime = datetime.fromtimestamp(os.path.getmtime(path)).strftime('%Y-%m-%d %H:%M')
                    lbl_file_info.setText(f"大小: {size_kb:.1f} KB | 修改: {mtime}")
                except: lbl_file_info.setText("")

                _, ext = os.path.splitext(path.lower())
                
                # 1. 图片预览
                if ext in {'.png', '.jpg', '.jpeg', '.bmp', '.gif', '.webp', '.ico'}:
                    preview_stack.setCurrentIndex(0)
                    try:
                        pix = QPixmap(path)
                        if not pix.isNull():
                            scaled = pix.scaled(lbl_preview_img.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation)
                            lbl_preview_img.setPixmap(scaled)
                            lbl_preview_img.setText("")
                            lbl_preview_img.setStyleSheet("background: #000; border: 1px solid #444;")
                        else:
                            lbl_preview_img.setPixmap(QPixmap())
                            lbl_preview_img.setText("⚠️ 格式不支持")
                    except: lbl_preview_img.setText("⚠️ 加载失败")
                
                # 2. 文本/文档预览
                elif ext in {'.txt', '.log', '.md', '.csv', '.ini', '.py', '.bat', '.json', '.xml'}:
                    preview_stack.setCurrentIndex(1)
                    try:
                        content = _read_text_for_smart_fill(path)
                        if content is None:
                            raise RuntimeError("文本解码失败")
                        content = content[:2000]
                        if len(content) >= 2000:
                            content += "\n\n...(内容较长，仅显示预览)..."
                        text_preview.setPlainText(content)
                    except Exception as e:
                        text_preview.setPlainText(f"读取失败: {e}")
                
                # 3. 其他文件
                else:
                    preview_stack.setCurrentIndex(0)
                    lbl_preview_img.setPixmap(QPixmap())
                    lbl_preview_img.setText(f"该类型 ({ext}) 不支持预览\n仅显示文件信息")
                    lbl_preview_img.setStyleSheet("background: #f5f5f5; color: #666;")
            
            file_tree.itemClicked.connect(_on_file_clicked)

            def _on_file_double_clicked(item):
                path = item.data(0, Qt.UserRole)
                if os.path.isfile(path): _do_fill()
            file_tree.itemDoubleClicked.connect(_on_file_double_clicked)
            tabs.addTab(file_tool, "📂 图片/文件")
            # --- 工具4: 文字 ---
            text_tool = QWidget(); text_ly = QVBoxLayout(text_tool)
            text_tip = QLabel(
                "支持直接从 Google 表格 / Excel / 飞书表格复制后粘贴到下方表格；"
                "单元格里的多行文案会保留，不再按换行强行拆成多条。"
            )
            text_tip.setWordWrap(True)
            text_tip.setStyleSheet("color:#555; background:#f5f5f5; border:1px solid #e0e0e0; border-radius:6px; padding:8px;")
            text_ly.addWidget(text_tip)

            text_bar = QHBoxLayout()
            text_bar.addWidget(QLabel("<b>文字表格:</b>"))
            text_bar.addStretch()
            btn_prefix_lib = QPushButton("📚 前缀库")
            btn_prefix_lib.setToolTip("将所选前缀填入当前表格单元格。")
            btn_prefix_lib.setFixedHeight(28)
            btn_text_paste = QPushButton("📋 粘贴表格")
            btn_text_add_row = QPushButton("➕ 行")
            btn_text_add_col = QPushButton("➕ 列")
            btn_text_clear = QPushButton("🧹 清空")
            for _btn in [btn_text_paste, btn_text_add_row, btn_text_add_col, btn_text_clear]:
                _btn.setFixedHeight(28)
            text_bar.addWidget(btn_prefix_lib)
            text_bar.addWidget(btn_text_paste)
            text_bar.addWidget(btn_text_add_row)
            text_bar.addWidget(btn_text_add_col)
            text_bar.addWidget(btn_text_clear)
            text_ly.addLayout(text_bar)

            default_text_cols = max(4, min(8, len(preview_cols) if preview_cols else 4))
            text_grid = SpreadsheetPasteTable(12, default_text_cols, dlg)
            for col in range(text_grid.columnCount()):
                text_grid.setColumnWidth(col, 180)
            text_ly.addWidget(text_grid)
            text_grid_selection_count = create_table_selection_label()
            text_ly.addWidget(text_grid_selection_count)
            bind_table_selection_label(text_grid, text_grid_selection_count, "右侧源表：")

            def _pick_prefix_from_library():
                dlg_prefix = ClearInputPrefixPresetDialog(dlg)
                if dlg_prefix.exec_():
                    prefix = dlg_prefix.get_prefix()
                    if prefix is not None:
                        cur_row = max(0, text_grid.currentRow())
                        cur_col = max(0, text_grid.currentColumn())
                        text_grid.ensure_size(cur_row + 1, cur_col + 1)
                        target_item = text_grid.item(cur_row, cur_col)
                        if not target_item:
                            target_item = QTableWidgetItem("")
                            text_grid.setItem(cur_row, cur_col, target_item)
                        target_item.setText(prefix)
                        text_grid.resizeRowToContents(cur_row)
                        text_grid.setRowHeight(cur_row, min(max(text_grid.rowHeight(cur_row), 42), 140))
                        _update_text_preview()

            btn_prefix_lib.clicked.connect(_pick_prefix_from_library)

            # --- 文字实时预览同步 ---
            def _update_text_preview():
                if tabs.currentWidget() != text_tool: return
                matrix = text_grid.get_active_matrix()

                if not matrix:
                    preview_stack.setCurrentIndex(0)
                    lbl_preview_img.setText("请先在右侧表格中输入或粘贴文字")
                    lbl_file_info.setText("")
                    return

                preview_stack.setCurrentIndex(1)
                preview_lines = []
                row_count = len(matrix)
                col_count = max((len(row) for row in matrix), default=0)

                for r_idx, row in enumerate(matrix[:6]):
                    cells = []
                    for c_idx, cell in enumerate(row[:4]):
                        display = str(cell).replace("\n", " ↵ ")
                        if len(display) > 36:
                            display = display[:33] + "..."
                        cells.append(f"{make_excel_column_name(c_idx)}{r_idx+1}: {display}")
                    preview_lines.append(" | ".join(cells))

                if row_count > 6 or col_count > 4:
                    preview_lines.append("...")
                text_preview.setPlainText("\n".join(preview_lines))
                source_desc = "当前选区" if text_grid.get_selected_matrix() else "整张表"
                lbl_file_info.setText(f"表格预览: 使用{source_desc}，共 {row_count} 行 × {col_count} 列")

            text_grid.itemChanged.connect(lambda *_: _update_text_preview())
            text_grid.itemSelectionChanged.connect(_update_text_preview)
            btn_text_paste.clicked.connect(lambda: text_grid.smart_paste() and _update_text_preview())
            btn_text_add_row.clicked.connect(lambda: (text_grid.ensure_size(text_grid.rowCount() + 1, text_grid.columnCount()), _update_text_preview()))
            btn_text_add_col.clicked.connect(lambda: (text_grid.ensure_size(text_grid.rowCount(), text_grid.columnCount() + 1), text_grid.setColumnWidth(text_grid.columnCount() - 1, 180), _update_text_preview()))
            btn_text_clear.clicked.connect(lambda: (text_grid.clear_all_contents(), _update_text_preview()))
            tabs.addTab(text_tool, "✍️ 文字")

            # --- 工具5: 特殊指令 ---
            spec_tool = QWidget(); spec_ly = QVBoxLayout(spec_tool)
            spec_ly.addWidget(QLabel("<b>特殊执行指令:</b>"))
            btn_skip_row = QPushButton("⏭️ 结束当前行 (SKIP_ROW)")
            btn_skip_row.setToolTip("执行到此步骤时，立即停止当前行的后续步骤，直接跳转到下一行数据。\n在批量填充中心留空也具有同样效果。")
            btn_skip_row.setFixedHeight(45)
            btn_skip_row.setStyleSheet("background: #fff3e0; border: 1px solid #ff9800; font-weight: bold; color: #e65100;")
            
            def _insert_skip():
                _save_undo_state() # 修改前保存状态
                c_idx = preview_table.currentColumn()
                if c_idx < 0: 
                    QMessageBox.warning(dlg, "提示", "请先在左侧预览表中点击要填充的单元格。")
                    return
                sel_items = preview_table.selectedItems()
                if not sel_items:
                    curr = preview_table.currentItem()
                    if curr: sel_items = [curr]
                if not sel_items: return
                for it in sel_items:
                    it.setText("[SKIP_ROW]")
                    it.setBackground(QColor("#fff3e0"))
            
            btn_skip_row.clicked.connect(_insert_skip)
            spec_ly.addWidget(btn_skip_row)
            spec_ly.addStretch()
            tabs.addTab(spec_tool, "✨ 特殊")

            # 前缀库放到工具箱标签页里，紧跟“特殊”后面，避免挤占右侧主布局
            if box_id == 1:
                tabs.addTab(prefix_group, "📚 前缀库")

            # 标签页切换时重置预览
            def _on_tab_changed(idx):
                cur = tabs.widget(idx)
                if cur == text_tool:
                    _update_text_preview()
                elif cur == file_tool:
                    preview_stack.setCurrentIndex(0)
                    lbl_preview_img.setText("在上方选择文件以预览")
                    lbl_file_info.setText("")
                else:
                    preview_stack.setCurrentIndex(0)
                    lbl_preview_img.setText("当前分类不支持预览")
                    lbl_file_info.setText("")
            tabs.currentChanged.connect(_on_tab_changed)

            # --- 统一填充按钮 ---
            btn_fill = QPushButton(f"⚡ 批量填充 (从工具箱 {box_id})")
            btn_fill.setFixedHeight(40)
            btn_fill.setStyleSheet("background: #e8f5e9; font-weight: bold; border: 2px solid #4caf50;")
            
            def _do_fill():
                _save_undo_state() # 修改前保存状态
                cur_tab = tabs.currentWidget()

                target_items = _get_selected_preview_targets()
                if not target_items:
                    QMessageBox.warning(dlg, "提示", "请先在左侧预览表中选择要填充的单元格。")
                    return

                # 获取源数据（支持多选）
                source_vals = []
                if cur_tab == win_tool:
                    # [修复] 将 hwnd 一并带入，构成“标题::hwnd=12345”格式
                    for it in win_list.selectedItems():
                        _title = it.data(Qt.UserRole) or it.text()
                        _hwnd  = it.data(Qt.UserRole + 1)
                        if _hwnd:
                            source_vals.append(f"{_title}::hwnd={_hwnd}")
                        else:
                            source_vals.append(_title)
                elif cur_tab == prof_tool:
                    # [修复] 填充账号时，同时记录显示名称和内部 profile_id，以元组形式传递
                    source_vals = [(it.text(), it.data(Qt.UserRole)) for it in prof_list.selectedItems()]
                elif cur_tab == file_tool:
                    for it in file_tree.selectedItems():
                        path = it.data(0, Qt.UserRole)
                        if path and os.path.isfile(path):
                            _, ext = os.path.splitext(path.lower())
                            # 1. 图片类：必须填充完整路径
                            if ext in {'.png', '.jpg', '.jpeg', '.bmp', '.gif', '.webp', '.ico'}:
                                source_vals.append(path)
                            # 2. 文档类：读取文本内容填充
                            elif ext in {'.txt', '.log', '.md', '.csv', '.ini', '.json', '.xml', '.py', '.bat'}:
                                try:
                                    # 增加一个大小检查，防止双击超大文件导致卡死（限制 1MB）
                                    if os.path.getsize(path) > 1024 * 1024:
                                        source_vals.append(path) # 太大了就填路径
                                    else:
                                        content = _read_text_for_smart_fill(path, max_bytes=1024 * 1024)
                                        source_vals.append(content if content is not None else path)
                                except: source_vals.append(path) # 失败则退回路径
                            # 3. 其他（视频、压缩包等）：填充路径
                            else:
                                source_vals.append(path)
                elif cur_tab == text_tool:
                    source_matrix = text_grid.get_active_matrix()
                    if not source_matrix:
                        QMessageBox.warning(dlg, "提示", "请先在文字表格中输入或粘贴内容。")
                        return

                    selected_indexes = preview_table.selectedIndexes()
                    if not selected_indexes:
                        cur = preview_table.currentIndex()
                        if cur.isValid():
                            selected_indexes = [cur]

                    rows = sorted({idx.row() for idx in selected_indexes})
                    cols = sorted({idx.column() for idx in selected_indexes})
                    is_full_rect = (
                        bool(rows) and bool(cols) and
                        len(selected_indexes) == len(rows) * len(cols) and
                        all(preview_table.item(r, c) for r in rows for c in cols)
                    )

                    source_row_count = len(source_matrix)
                    source_col_count = max((len(row) for row in source_matrix), default=0)

                    if is_full_rect and (source_row_count > 1 or source_col_count > 1):
                        for r_off, row_idx in enumerate(rows):
                            for c_off, col_idx in enumerate(cols):
                                it = preview_table.item(row_idx, col_idx)
                                new_val = source_matrix[r_off % source_row_count][c_off % source_col_count]
                                _set_preview_item_text(it, new_val)
                    else:
                        flat_vals = [cell for row in source_matrix for cell in row]
                        if not flat_vals:
                            QMessageBox.warning(dlg, "提示", "文字表格中没有可用内容。")
                            return
                        for i, it in enumerate(target_items):
                            _set_preview_item_text(it, flat_vals[i % len(flat_vals)])
                    return

                if not source_vals: return
                
                # 核心批量填充逻辑：将选中的源数据按当前选区顺序填入目标格
                for i, it in enumerate(target_items):
                    val = source_vals[i % len(source_vals)]
                    _, _, _, cell_act_type, cell_sub = preview_cols[it.column()]
                    # [修复] 账号类型：val 是 (显示名, profile_id) 元组，需拆分处理
                    if isinstance(val, tuple):
                        disp_name, pid = val
                        it.setText(disp_name)
                        it.setData(Qt.UserRole + 10, f"|{pid}")
                    elif "::hwnd=" in str(val) and cell_act_type in ["win_active", "open_url"]:
                        # [修复] 窗口类目标显示纯标题，完整标识存入额外字段
                        _disp = val.split('::hwnd=')[0] or "已打开窗口"
                        it.setText(_disp)
                        if cell_act_type == "open_url" and cell_sub == "profile":
                            it.setData(Qt.UserRole + 10, f"|{val}")
                        else:
                            it.setData(Qt.UserRole + 10, val)
                    else:
                        it.setText(val)
                        it.setData(Qt.UserRole + 10, None)
                    it.setToolTip(str(val))
                    it.setBackground(QColor("#fff9c4"))

            btn_fill.clicked.connect(_do_fill)
            
            container = QWidget(); c_ly = QVBoxLayout(container); c_ly.setContentsMargins(0,0,0,0)
            container.setMinimumWidth(180)
            container.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            c_ly.addWidget(tabs); c_ly.addWidget(btn_fill)
            return container

        # 创建两个并列的工具箱
        tool_box_1 = create_tool_box(1)
        tool_box_2 = create_tool_box(2)
        tool_splitter.addWidget(tool_box_1)
        tool_splitter.addWidget(tool_box_2)
        tool_splitter.setStretchFactor(0, 1)
        tool_splitter.setStretchFactor(1, 1)

        def _normalize_tool_splitter_sizes(sizes):
            if not sizes or len(sizes) < 2:
                return [1, 1]
            left = max(int(sizes[0]), 180)
            right = max(int(sizes[1]), 180)
            return [left, right]

        def _save_tool_splitter_sizes():
            if tool_box_2.isVisible():
                self.config.setdefault("layout", {})["batch_fill_toolbox_sizes"] = tool_splitter.sizes()

        tool_splitter.splitterMoved.connect(lambda *_: _save_tool_splitter_sizes())

        def _apply_toolbox_mode(is_dual, save_mode=False):
            if is_dual:
                saved_sizes = self.config.get("layout", {}).get("batch_fill_toolbox_sizes", [420, 420])
                tool_splitter.setHandleWidth(8)
                tool_box_2.show()
                tool_splitter.setSizes(_normalize_tool_splitter_sizes(saved_sizes))
                btn_toggle_tool_mode.setText("⇆ 切到单窗口")
                btn_toggle_tool_mode.setChecked(False)
            else:
                _save_tool_splitter_sizes()
                tool_splitter.setHandleWidth(0)
                tool_box_2.hide()
                tool_splitter.setSizes([1, 0])
                btn_toggle_tool_mode.setText("⇆ 切到双窗口")
                btn_toggle_tool_mode.setChecked(True)
            if save_mode:
                self.config.setdefault("layout", {})["batch_fill_toolbox_mode"] = "dual" if is_dual else "single"
                save_config(self.config)

        saved_mode = self.config.get("layout", {}).get("batch_fill_toolbox_mode", "dual")
        btn_toggle_tool_mode.clicked.connect(lambda checked: _apply_toolbox_mode(not checked, save_mode=True))
        _apply_toolbox_mode(saved_mode == "dual", save_mode=False)

        def _on_cell_clicked(row, col):
            real_col, _step_idx, name, act_type, sub = preview_cols[col]
            display_type = f"{act_type}-{sub}" if sub else act_type
            cur_target_label.setText(f"🎯 当前填充目标: {name} ({display_type})")
            cur_target_label.setStyleSheet("""
                QLabel { font-weight: 400; color: #222;
                    color: #2e7d32; 
                    font-weight: bold; 
                    font-size: 14px; 
                    padding: 10px; 
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #e8f5e9, stop:1 #fff); 
                    border: 1px solid #c8e6c9;
                    border-radius: 6px;
                }
            """)
            
        preview_table.cellClicked.connect(_on_cell_clicked)
        if len(targets) > 0: QTimer.singleShot(100, lambda: _on_cell_clicked(0, 0))

        main_splitter.addWidget(right_w)

        def _normalize_batch_fill_splitter_sizes(sizes):
            if not sizes or len(sizes) < 2:
                return [520, 680]
            left = max(int(sizes[0]), 360)
            right = max(int(sizes[1]), 420)
            return [left, right]

        def _save_batch_fill_splitter_sizes():
            self.config.setdefault("layout", {})["batch_fill_main_splitter_sizes"] = main_splitter.sizes()
            self._schedule_config_flush()

        try:
            saved_sizes = self.config.get("layout", {}).get("batch_fill_main_splitter_sizes")
            if saved_sizes:
                main_splitter.setSizes(_normalize_batch_fill_splitter_sizes(saved_sizes))
            else:
                main_splitter.setSizes(_normalize_batch_fill_splitter_sizes([520, 680]))
        except Exception:
            pass
        main_splitter.splitterMoved.connect(lambda *_: (_save_batch_fill_splitter_sizes(), _update_toolbar_auto_visibility()))
        main_ly.addWidget(main_splitter)
        _orig_dlg_resize_event = dlg.resizeEvent
        def _batch_fill_dlg_resize_event(event):
            if callable(_orig_dlg_resize_event):
                _orig_dlg_resize_event(event)
            QTimer.singleShot(0, lambda: _update_toolbar_auto_visibility())
        dlg.resizeEvent = _batch_fill_dlg_resize_event
        QTimer.singleShot(0, lambda: _update_toolbar_auto_visibility(force=True))
        
        # 底部应用按钮
        btn_apply = QPushButton("✅ 应用所有更改到主表格"); btn_apply.setFixedHeight(50); btn_apply.setStyleSheet("background: #2196f3; color: white; font-size: 16px; font-weight: bold;")
        main_ly.addWidget(btn_apply)
        
        def _apply_to_main():
            """核心逻辑：将预览表拆分的数据重新合并并填回主表格。"""
            # row_col_data 结构：(行, 列) -> { 子类型: 值 }
            row_col_data = {} 
            
            for r in range(preview_table.rowCount()):
                for c in range(preview_table.columnCount()):
                    it = preview_table.item(r, c)
                    if not it: continue
                    real_row = it.data(Qt.UserRole)
                    real_col = it.data(Qt.UserRole + 1)
                    sub = it.data(Qt.UserRole + 2) # "prefix", "content", "url", "profile", None
                    val = it.text()
                    extra = it.data(Qt.UserRole + 10) # 针对账号 PID 等特殊数据
                    
                    key = (real_row, real_col)
                    if key not in row_col_data: 
                        row_col_data[key] = {"prefix":"", "content":"", "url":"", "profile":"", "val":"", "p_id":""}
                    
                    if sub == "prefix": row_col_data[key]["prefix"] = val
                    elif sub == "content": row_col_data[key]["content"] = val
                    elif sub == "url": row_col_data[key]["url"] = val
                    elif sub == "profile": 
                        row_col_data[key]["profile"] = val
                        if extra and extra.startswith("|"): row_col_data[key]["p_id"] = extra[1:]
                    else: 
                        row_col_data[key]["val"] = val
                    
                    # 如果是普通填充带了 extra 数据，也要记录
                    if extra and not row_col_data[key]["p_id"]:
                        row_col_data[key]["extra_raw"] = extra

            # [关键修复] 应用更改前确保主表格信号不被阻塞，否则界面不会刷新
            self.data_table.blockSignals(False)
            
            for (real_row, real_col), data in row_col_data.items():
                # 寻找该列对应的动作类型
                act_type = ""
                for rc, _, _, at, _ in preview_cols:
                    if rc == real_col:
                        act_type = at
                        break
                
                # 1. 组装最终存入主表的字符串值
                skip_token = "[SKIP_ROW]"
                has_skip_token = any(str(data.get(k, "")).strip() == skip_token for k in ["prefix", "content", "url", "profile", "val"])
                if has_skip_token:
                    final_val = skip_token
                elif act_type == "clear_input_plus":
                    final_val = f"{data['prefix']}|{data['content']}"
                elif act_type == "open_url":
                    p_id = data["p_id"]
                    if not p_id:
                        typed_profile = str(data.get("profile", "") or "").strip()
                        if typed_profile and typed_profile != "[SKIP_ROW]":
                            p_id = typed_profile
                    if not p_id:
                        orig_it = self.data_table.item(real_row, real_col)
                        if orig_it and "|" in orig_it.text():
                            p_id = orig_it.text().split("|", 1)[1]
                    final_val = f"{data['url']}|{p_id}"
                elif act_type == "win_active":
                    # [修复] 如果 extra_raw 包含完整 hwnd 标识，直接使用；否则用 val
                    _extra = data.get("extra_raw", "")
                    if _extra and "::hwnd=" in _extra:
                        final_val = _extra  # extra_raw 存的就是完整标识
                    else:
                        final_val = data['val'] or _extra
                else:
                    final_val = data['val']
                    if "extra_raw" in data: final_val += data["extra_raw"]
                
                # 2. 更新主表格界面
                # 无论是否有 backing item，都先更新它，再同步刷新可见 cellWidget
                it_main = self.data_table.item(real_row, real_col)
                if it_main:
                    it_main.setText(final_val)
                else:
                    new_it = QTableWidgetItem(final_val)
                    self.data_table.setItem(real_row, real_col, new_it)
                    it_main = new_it

                # 同步刷新 cellWidget（open_url / clear_input_plus / win_active 等均有可见控件）
                w = self.data_table.cellWidget(real_row, real_col)
                if w:
                    if act_type == "open_url":
                        le = _find_text_input(w)
                        btn = w.findChild(QPushButton)
                        if final_val == skip_token:
                            url_part = skip_token
                            p_id_part = ""
                        else:
                            url_part = data["url"]
                            p_id_part = data["p_id"]
                            if not p_id_part and "|" in final_val:
                                p_id_part = final_val.split("|", 1)[1]
                        if le:
                            le.blockSignals(True)
                            le.setText(url_part)
                            le.blockSignals(False)
                        if btn:
                            btn.setProperty("profile_id", p_id_part)
                            btn.setText(get_profile_display_name(p_id_part))
                    elif act_type == "clear_input_plus":
                        le = _find_text_input(w)
                        lbl = w.findChild(QLabel)
                        if final_val == skip_token:
                            if le:
                                le.blockSignals(True)
                                le.setText(skip_token)
                                le.blockSignals(False)
                            if lbl:
                                lbl.setText(" ")
                        else:
                            if le:
                                le.blockSignals(True)
                                le.setText(data["content"])
                                le.blockSignals(False)
                            if lbl:
                                lbl.setText(f" {data['prefix']}")
                    elif act_type == "win_active":
                        # [修复] win_active 的 cellWidget 显示纯标题，不显示 hwnd
                        le = _find_text_input(w)
                        if le:
                            _disp = final_val.split('::hwnd=')[0] if '::hwnd=' in final_val else final_val
                            le.blockSignals(True)
                            le.setText(_disp)
                            le.setToolTip(final_val)
                            le.blockSignals(False)
                    else:
                        le = _find_text_input(w)
                        if le:
                            le.blockSignals(True)
                            le.setText(final_val)
                            le.blockSignals(False)
                        elif isinstance(w, (QLineEdit, MultiLineTextEdit)):
                            w.blockSignals(True)
                            w.setText(final_val)
                            w.blockSignals(False)
            dlg.accept()
            # [关键修复] 强制触发一次界面重绘
            self.data_table.viewport().update()
            self._save_data_table()
            self._auto_check_blank_rows(selected_rows)
            QMessageBox.information(self, "成功", "所有更改已应用。")
            
        btn_apply.clicked.connect(_apply_to_main)
        dlg.exec_()
        self._auto_check_blank_rows(selected_rows)

    def _add_data_row(self):
        """添加数据行，支持批量添加。"""
        if not self.current_task: return
        num, ok = QInputDialog.getInt(self, "添加数据行", "请输入要添加的行数:", 1, 1, 1000)
        if not ok: return
        
        old_data = self.config['task_data'].get(self.current_task, [])
        for _ in range(num):
            old_data.append({})
        self.config['task_data'][self.current_task] = old_data
        
        # 性能优化：如果添加的行数较多，或者总行数较多，提示用户正在处理
        if num > 50 or len(old_data) > 200:
            self._log(f"⏳ 正在批量创建 {num} 行数据并渲染界面，请稍候...", "blue")
            QApplication.setOverrideCursor(Qt.WaitCursor)
            try:
                self._refresh_data_table()
            finally:
                QApplication.restoreOverrideCursor()
        else:
            self._refresh_data_table()
            
        self._save_data_table()
        self._log(f"✅ 已成功添加 {num} 行数据", "green")
    def _forward_widget_context_menu(self, widget, local_pos, table_row, table_col):
        """把嵌入单元格的 widget（QComboBox / KeyRecorder）的右键事件转发给数据表右键菜单。"""
        # 把 widget 本地坐标 → 全局坐标 → data_table viewport 本地坐标
        global_pos = widget.mapToGlobal(local_pos)
        viewport_pos = self.data_table.viewport().mapFromGlobal(global_pos)
        self._show_data_menu(viewport_pos)

    def _on_data_item_changed(self, item):
        if self.data_table.signalsBlocked(): return
        if item:
            try:
                row_height = int(self.config.get("layout", {}).get("data_row_height", 28))
                self.data_table.setRowHeight(item.row(), row_height)
            except Exception:
                pass
        self._update_data_select_header()
        self._save_data_table()

    # [新增] 图片路径自动预览
    _IMG_EXTS = {".png", ".jpg", ".jpeg", ".bmp", ".gif", ".webp", ".tiff", ".ico"}

    def _on_data_cell_clicked_preview(self, row, col):
        """点击数据表单元格：如果内容是图片路径，自动弹出悬浮预览窗口。"""
        item = self.data_table.item(row, col)
        if not item:
            # 尝试从 cellWidget 中获取文本
            w = self.data_table.cellWidget(row, col)
            if w:
                le = w.findChild(QLineEdit)
                path = le.text().strip() if le else ""
            else:
                return
        else:
            path = item.text().strip()

        # 如果是 open_url 格式（url|profile_id），只取 url 部分
        if "|" in path:
            path = path.split("|")[0].strip()

        # 判断是否为图片文件
        _, ext = os.path.splitext(path.lower())
        if ext not in self._IMG_EXTS:
            return
        if not os.path.isfile(path):
            return

        self._show_image_preview(path, row, col)

    def _show_image_preview(self, img_path, row=None, col=None):
        """弹出图片预览悬浮窗口（始终置顶，可拖动，自动适应大小）。"""
        # 如果已有预览窗口则关闭
        if hasattr(self, '_img_preview_win') and self._img_preview_win and self._img_preview_win.isVisible():
            self._img_preview_win.close()

        win = QDialog(self)
        win.setWindowTitle(f"🖼️ 图片预览")
        win.setWindowFlags(Qt.WindowStaysOnTopHint | Qt.Tool | Qt.WindowCloseButtonHint)
        win.resize(420, 380)
        win.setStyleSheet("QDialog { background: #1e1e1e; } QLabel { font-weight: 400; color: #222; color: #eee; }")
        self._img_preview_win = win

        ly = QVBoxLayout(win)
        ly.setContentsMargins(6, 6, 6, 6)
        ly.setSpacing(4)

        # 文件名和路径信息
        fname = os.path.basename(img_path)
        lbl_name = QLabel(f"<b>{fname}</b>")
        lbl_name.setStyleSheet("font-size: 12px; color: #90caf9; padding: 2px;")
        lbl_name.setWordWrap(True)
        ly.addWidget(lbl_name)

        lbl_path = QLabel(img_path)
        lbl_path.setStyleSheet("font-size: 10px; color: #78909c; padding: 2px;")
        lbl_path.setWordWrap(True)
        lbl_path.setTextInteractionFlags(Qt.TextSelectableByMouse)
        ly.addWidget(lbl_path)

        # 图片显示区域
        lbl_img = QLabel()
        lbl_img.setAlignment(Qt.AlignCenter)
        lbl_img.setStyleSheet("background: #111; border-radius: 4px; padding: 4px;")
        lbl_img.setMinimumSize(200, 200)
        ly.addWidget(lbl_img, 1)

        # 加载图片并缩放适应
        def _load_img():
            try:
                pixmap = QPixmap(img_path)
                if pixmap.isNull():
                    lbl_img.setText("⚠️ 图片无法加载")
                    return
                # 保持比例缩放到最大 600x500
                scaled = pixmap.scaled(600, 500, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                lbl_img.setPixmap(scaled)
                lbl_img.resize(scaled.width(), scaled.height())
                # 显示尺寸信息
                if row is not None and col is not None:
                    lbl_size.setText(f"尺寸: {pixmap.width()} × {pixmap.height()} px  |  第 {row+1} 行第 {col+1} 列")
                else:
                    lbl_size.setText(f"尺寸: {pixmap.width()} × {pixmap.height()} px")
            except Exception as e:
                lbl_img.setText(f"⚠️ 加载失败: {e}")

        # 尺寸信息
        lbl_size = QLabel("加载中...")
        lbl_size.setStyleSheet("font-size: 10px; color: #aaa; padding: 2px;")
        lbl_size.setAlignment(Qt.AlignCenter)
        ly.addWidget(lbl_size)

        # 底部按钮行
        btn_row = QHBoxLayout()
        btn_open_folder = QPushButton("📂 打开所在文件夹")
        btn_open_folder.setStyleSheet("background:#37474f;color:white;padding:5px 12px;")
        btn_open_folder.clicked.connect(lambda: (
            os.startfile(os.path.dirname(img_path)) if sys.platform == 'win32'
            else os.system(f'xdg-open "{os.path.dirname(img_path)}"')
        ))
        btn_close = QPushButton("❌ 关闭")
        btn_close.setStyleSheet("background:#c62828;color:white;padding:5px 12px;")
        btn_close.clicked.connect(win.close)
        btn_row.addWidget(btn_open_folder)
        btn_row.addStretch()
        btn_row.addWidget(btn_close)
        ly.addLayout(btn_row)

        # 窗口定位：默认在屏幕右下角，但如果是手动触发，确保在视野内
        screen_geo = QApplication.primaryScreen().availableGeometry()
        target_x = screen_geo.right() - 450
        target_y = screen_geo.bottom() - 420
        
        # 如果是文件树点击（row/col为None），可以考虑稍微偏向中心一点
        if row is None:
            target_x = (screen_geo.width() - win.width()) // 2
            target_y = (screen_geo.height() - win.height()) // 2
            
        win.move(target_x, target_y)
        win.show()
        win.raise_()
        win.activateWindow()
        QTimer.singleShot(50, _load_img)  # 延迟加载确保窗口已显示

    def _show_data_menu(self, pos):
        row = self.data_table.rowAt(pos.y())
        col = self.data_table.columnAt(pos.x())
        if row < 0: return
        menu = QMenu(self)
        menu.addAction(f"▶️ 单独执行此行：第{row+1}行").triggered.connect(lambda: self._run_single_data_row(row))
        menu.addSeparator()
        step_idx = self._col_to_step_idx(col)
        if step_idx >= 0:
            acts = self.config['tasks'].get(self.current_task, [])
            name = acts[step_idx].get('name', f'步骤{step_idx+1}')
            ds = self.config['task_data'].get(self.current_task, [])
            row_data = ds[row] if row < len(ds) else {}
            is_skipped = row_data.get(f"{name}_跳过", False)
            
            toggle_act = menu.addAction("✅ 启用此步骤" if is_skipped else "🚫 跳过此步骤")
            toggle_act.triggered.connect(lambda: self._toggle_step_skip(row, name, not is_skipped))
            menu.addSeparator()
            menu.addAction(f"🎯 仅执行此格：第{row+1}行 × [{name}]").triggered.connect(
                lambda _c, si=step_idx, dr=row: self._run_single_step_with_row(si, dr))
        
        menu.addAction("📋 复制此行").triggered.connect(lambda: self._copy_data_row(row))
        menu.addAction("👯 批量克隆此行").triggered.connect(lambda: self._batch_copy_data_row(row))
        menu.addSeparator()
        menu.addAction("➕ 批量添加空行").triggered.connect(self._add_data_row)
        menu.addAction("❌ 删除选中行").triggered.connect(self._del_data_row)
        menu.addAction("🧹 删除已完成行").triggered.connect(self._del_completed_data_rows)
        menu.exec_(self.data_table.viewport().mapToGlobal(pos))

    def _toggle_step_skip(self, row, step_name, should_skip):
        ds = self.config['task_data'].get(self.current_task, [])
        if row < len(ds):
            ds[row][f"{step_name}_跳过"] = should_skip
            self._save_data_table()
            self._refresh_data_table()

    def _col_to_step_idx(self, col):
        if col < self._data_first_value_col(): return -1
        acts = self.config['tasks'].get(self.current_task, [])
        show_delay = self.btn_toggle_delay.isChecked()
        curr_col = self._data_first_value_col()
        for i in range(len(acts)):
            if curr_col == col: return i
            curr_col += 1
            if show_delay:
                if curr_col == col: return -1
                curr_col += 1
        return -1

    def _copy_data_row(self, row):
        if not self.current_task: return
        self._save_data_table(); import copy; new_row = copy.deepcopy(self.config['task_data'][self.current_task][row])
        self.config['task_data'][self.current_task].insert(row + 1, new_row); save_config(self.config); self._refresh_data_table()
    def _batch_copy_data_row(self, row):
        if not self.current_task: return
        self._save_data_table(); num, ok = QInputDialog.getInt(self, "批量复制数据行", "请输入要复制的份数:", 5, 1, 1000)
        if ok:
            import copy; base_row = self.config['task_data'][self.current_task][row]
            for i in range(num): self.config['task_data'][self.current_task].insert(row + i + 1, copy.deepcopy(base_row))
            save_config(self.config); self._refresh_data_table()
    def _del_data_row(self):
        if not self.current_task: return
        rows = sorted(list(set([i.row() for i in self.data_table.selectedItems()])), reverse=True)
        if not rows: return
        for r in rows: self.config['task_data'][self.current_task].pop(r)
        if hasattr(self, '_row_statuses') and self.current_task in self._row_statuses:
            old_statuses = self._row_statuses.get(self.current_task, {})
            removed = set(rows)
            new_statuses = {}
            shift = 0
            for idx in range(len(self.config['task_data'][self.current_task]) + len(rows)):
                if idx in removed:
                    shift += 1
                    continue
                if idx in old_statuses:
                    new_statuses[idx - shift] = old_statuses[idx]
            self._row_statuses[self.current_task] = new_statuses
        save_config(self.config); self._refresh_data_table()

    def _del_completed_data_rows(self):
        if not self.current_task:
            return
        self._save_data_table(flush=True)
        data_rows = self.config.get('task_data', {}).get(self.current_task, [])
        statuses = self._row_statuses.get(self.current_task, {})
        rows = []
        for idx in range(len(data_rows)):
            if statuses.get(idx) != ROW_STATUS_OK:
                continue
            row_dict = data_rows[idx] if isinstance(data_rows[idx], dict) else {}
            if not bool(row_dict.get("_选中", False)):
                continue
            rows.append(idx)
        if not rows:
            QMessageBox.information(self, "提示", "当前没有“已勾选”且“已完成”的行可删除。")
            return
        reply = QMessageBox.question(
            self,
            "删除已完成行",
            f"确定删除当前任务里 {len(rows)} 行“已勾选”且“已完成”的数据吗？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return
        removed = set(rows)
        for r in sorted(rows, reverse=True):
            self.config['task_data'][self.current_task].pop(r)
        new_statuses = {}
        shift = 0
        for idx in range(len(data_rows)):
            if idx in removed:
                shift += 1
                continue
            if idx in statuses:
                new_statuses[idx - shift] = statuses[idx]
        self._row_statuses[self.current_task] = new_statuses
        save_config(self.config)
        self._refresh_data_table()
        self._log(f"🧹 已删除 {len(rows)} 行已完成数据", "green")
    def _refresh_win_combos(self):
        """刷新数据表中所有「激活窗口」列的窗口下拉列表。"""
        titles = sorted(list(set([t for t in gw.getAllTitles() if t.strip()])))
        for r in range(self.data_table.rowCount()):
            for c in range(self.data_table.columnCount()):
                widget = self.data_table.cellWidget(r, c)
                if isinstance(widget, QComboBox):
                    cur = widget.currentText()
                    widget.blockSignals(True)
                    widget.clear()
                    widget.addItems(titles)
                    widget.setCurrentText(cur)
                    widget.blockSignals(False)
        self._log("🪟 窗口列表已刷新", "blue")

    def _select_window(self, idx):
        """[修复] 保存带 hwnd 的唯一标识，按鈕显示纯标题。"""
        s = WindowSelector(self); t = s.get_selection()
        if t:
            self.config['tasks'][self.current_task][idx]['value'] = t
            save_config(self.config)
            self._refresh_actions()
    def _pick_default_file(self, idx, line_edit, save_cb=None):
        f, _ = QFileDialog.getOpenFileName(self, "选择默认文件", "", "All Files (*.*)")
        if f:
            line_edit.setText(os.path.normpath(f))
            if callable(save_cb):
                save_cb(idx)
    def _add_action(self):
        if not self.current_task: return
        idx = self.action_table.currentRow()
        new_name = self._make_unique_action_name(f"步骤{len(self.config['tasks'][self.current_task])+1}")
        new_act = {"name": new_name, "action": "左键点击", "x": 0, "y": 0, "value": "", "delay": 1}
        if idx >= 0: self.config['tasks'][self.current_task].insert(idx + 1, new_act)
        else: self.config['tasks'][self.current_task].append(new_act)
        save_config(self.config); self._refresh_actions(); self._refresh_defer_target_options(persist_changes=True)
        new_idx = idx + 1 if idx >= 0 else self.action_table.rowCount() - 1; self.action_table.setCurrentCell(new_idx, 0)
    def _del_action(self):
        if not self.current_task: return
        rows = sorted(list(set([i.row() for i in self.action_table.selectedItems()])), reverse=True)
        if not rows: return
        removed_names = []
        for r in rows:
            acts = self.config['tasks'][self.current_task]
            if 0 <= r < len(acts):
                removed_names.append(acts[r].get('name', f'步骤{r+1}'))
                acts.pop(r)
        self._refresh_defer_target_options(removed_names=removed_names, persist_changes=False)
        save_config(self.config); self._refresh_actions(); self._refresh_defer_target_options(persist_changes=False)
    def _start_record(self, idx):
        """[独立吸管版] 坐标拾取逻辑：
        1. 弹出『独立顶级』吸管工具窗，不受主窗口最小化影响。
        2. 自动最小化主窗口，帮用户腾出屏幕空间。
        3. 用户拖拽吸管，松开即拾取，拾取后主窗口自动恢复。"""
        self.recording_idx = idx
        acts = self.config['tasks'].get(self.current_task, [])
        step_name = acts[idx].get('name', f'步骤{idx+1}') if idx < len(acts) else f'步骤{idx+1}'

        # 清理旧资源
        def _cleanup():
            if hasattr(self, '_coord_timer') and self._coord_timer:
                self._coord_timer.stop()
            if hasattr(self, '_picker_hint') and self._picker_hint:
                self._picker_hint.close()
            # 自动恢复主窗口
            self.showNormal()
            self.raise_()
            self.activateWindow()

        # [新增] 自动最小化主窗口，方便用户看到底层软件
        self.showMinimized()

        # 1. 创建吸管工具窗 (解除 self 绑定，设为独立顶级窗口)
        self._picker_hint = QDialog(None) 
        self._picker_hint.setWindowFlags(Qt.WindowStaysOnTopHint | Qt.FramelessWindowHint | Qt.Tool)
        self._picker_hint.setAttribute(Qt.WA_DeleteOnClose) 
        self._picker_hint.setStyleSheet("""
            QDialog { background: #1a1a2e; border: 3px solid #00e5ff; border-radius: 15px; }
            QLabel { font-weight: 400; color: #222; color: white; font-family: 'Microsoft YaHei'; }
            QPushButton#close_btn { 
                background: #ff5252; color: white; border: none; border-radius: 10px; 
                font-weight: bold; font-size: 12px;
            }
            QPushButton#close_btn:hover { background: #ff1744; }
        """)
        self._picker_hint.resize(320, 200)
        
        main_ly = QVBoxLayout(self._picker_hint); main_ly.setContentsMargins(15, 10, 15, 15)
        
        # 标题栏：包含标题和关闭按钮
        title_h = QHBoxLayout()
        lbl_title = QLabel(f"🎨 吸管拾取器：{step_name}")
        lbl_title.setStyleSheet("font-size: 12px; font-weight: bold; color: #00e5ff;")
        title_h.addWidget(lbl_title)
        
        btn_close = QPushButton("×")
        btn_close.setObjectName("close_btn")
        btn_close.setFixedSize(24, 24)
        btn_close.clicked.connect(_cleanup)
        title_h.addWidget(btn_close)
        main_ly.addLayout(title_h)
        
        # 核心：吸管图标（可交互区域）
        self._dropper_btn = QLabel("🧪") # 使用实验瓶图标作为吸管
        self._dropper_btn.setStyleSheet("""
            QLabel { font-weight: 400; color: #222; 
                font-size: 50px; background: #252545; border-radius: 40px; 
                padding: 10px; border: 2px dashed #00e5ff;
            }
            QLabel:hover { background: #303060; border: 2px solid #00e5ff; }
        """)
        self._dropper_btn.setFixedSize(90, 90)
        self._dropper_btn.setAlignment(Qt.AlignCenter)
        self._dropper_btn.setCursor(Qt.PointingHandCursor)
        main_ly.addWidget(self._dropper_btn, 0, Qt.AlignCenter)
        
        self._lbl_live_coord = QLabel("📍 准备拾取...")
        self._lbl_live_coord.setStyleSheet("font-size: 16px; color: #a5d6a7; font-weight: bold; margin-top: 5px;")
        self._lbl_live_coord.setAlignment(Qt.AlignCenter)
        main_ly.addWidget(self._lbl_live_coord)
        
        lbl_info = QLabel("<b>按住</b>图标拖到目标点，<b>松开</b>即录制")
        lbl_info.setStyleSheet("color: #b0bec5; font-size: 11px; margin-top: 5px;"); lbl_info.setAlignment(Qt.AlignCenter)
        main_ly.addWidget(lbl_info)

        # 2. 吸管核心逻辑：按住拖拽拾取
        self._is_picking = False
        
        def _dropper_mousePress(event):
            if event.button() == Qt.LeftButton:
                self._is_picking = True
                self._dropper_btn.setText("🎯")
                self._dropper_btn.setStyleSheet(self._dropper_btn.styleSheet().replace("#252545", "#004d40"))
                # 开启全局追踪
                self._picker_hint.grabMouse() 
                self._log("🧪 吸管已激活，请拖动到目标位置...", "blue")
                event.accept()

        def _dropper_mouseMove(event):
            if self._is_picking:
                # 获取全局坐标
                pos = QCursor.pos()
                self._lbl_live_coord.setText(f"📍 当前: ({pos.x()}, {pos.y()})")
                event.accept()

        def _dropper_mouseRelease(event):
            if self._is_picking:
                self._is_picking = False
                self._picker_hint.releaseMouse()
                # 记录最终坐标
                final_pos = QCursor.pos()
                self._log(f"✅ 吸管拾取成功: ({final_pos.x()}, {final_pos.y()})", "green")
                
                # 延迟关闭，让用户看清坐标
                QTimer.singleShot(100, lambda: (
                    _cleanup(),
                    self._do_record_at(final_pos.x(), final_pos.y())
                ))
                event.accept()

        # 绑定事件到吸管图标（和整个窗口以防滑出）
        self._picker_hint.mousePressEvent = _dropper_mousePress
        self._picker_hint.mouseMoveEvent = _dropper_mouseMove
        self._picker_hint.mouseReleaseEvent = _dropper_mouseRelease
        
        # 3. 实时坐标刷新（未按住时也显示，方便定位）
        self._coord_timer = QTimer()
        def _refresh():
            if not self._is_picking:
                pos = QCursor.pos()
                self._lbl_live_coord.setText(f"📍 预览: ({pos.x()}, {pos.y()})")
        self._coord_timer.timeout.connect(_refresh)
        self._coord_timer.start(50)

        # 4. 全窗口拖拽逻辑（排除吸管图标）
        self._win_drag_pos = None
        
        # 重写窗口事件以支持拖拽和 Esc
        def _picker_press_event(event):
            if event.button() == Qt.LeftButton:
                # 如果点击的是吸管图标，触发吸管逻辑
                child = self._picker_hint.childAt(event.pos())
                if child == self._dropper_btn:
                    _dropper_mousePress(event)
                else:
                    # 否则触发窗口拖拽
                    self._win_drag_pos = event.globalPos() - self._picker_hint.pos()
                    event.accept()

        def _picker_move_event(event):
            if self._is_picking:
                _dropper_mouseMove(event)
            elif self._win_drag_pos:
                self._picker_hint.move(event.globalPos() - self._win_drag_pos)
                event.accept()

        def _picker_release_event(event):
            if self._is_picking:
                _dropper_mouseRelease(event)
            else:
                self._win_drag_pos = None
                event.accept()

        def _picker_key_event(event):
            if event.key() == Qt.Key_Escape:
                _cleanup()
                self._log("🚫 坐标拾取已取消", "orange")

        self._picker_hint.mousePressEvent = _picker_press_event
        self._picker_hint.mouseMoveEvent = _picker_move_event
        self._picker_hint.mouseReleaseEvent = _picker_release_event
        self._picker_hint.keyPressEvent = _picker_key_event

        # 5. 显示并激活
        self._picker_hint.show()
        self._picker_hint.raise_()
        self._picker_hint.activateWindow()

        # Esc 退出
        def _check_esc(event):
            if event.key() == Qt.Key_Escape:
                _cleanup()
                self._log("🚫 坐标拾取已取消", "orange")
        self._picker_hint.keyPressEvent = _check_esc

    def _do_record(self):
        """[兼容旧逻辑] 不再使用，保留以防其他地方引用。"""
        x, y = pyautogui.position()
        self._do_record_at(x, y)

    def _do_record_at(self, x, y):
        """[单屏优化] 实际写入坐标的核心逻辑。"""
        self._log(f"🔍 准备写入坐标: ({x}, {y}) 到步骤索引 {self.recording_idx}", "blue")
        acts = self.config['tasks'].get(self.current_task, [])
        if self.recording_idx < len(acts):
            try:
                # 1. 更新内存数据
                acts[self.recording_idx].update({'x': x, 'y': y})
                step_name = acts[self.recording_idx].get('name', f'步骤{self.recording_idx+1}')
                try:
                    guard_info = create_action_guard_snapshot(self.current_task, step_name, x, y)
                    acts[self.recording_idx].update(guard_info)
                    if guard_info.get("guard_image"):
                        self._log(f"🧭 已同步保存界面守卫快照: {os.path.basename(guard_info['guard_image'])}", "gray")
                except Exception as guard_err:
                    log_internal_issue(f"录制界面守卫失败: {self.current_task} / {self.recording_idx}", guard_err)
                    self._log(f"⚠️ 坐标已保存，但界面守卫快照保存失败: {guard_err}", "orange")
                self._log(f"💾 内存数据已更新: {acts[self.recording_idx]}", "gray")
                
                # 2. 物理保存到文件
                save_config(self.config)
                self._log(f"💾 配置文件已物理保存到硬盘", "gray")
                
                self._log(f"✅ 坐标已录制: [{step_name}] -> ({x}, {y})", "green")
                
                # 3. 同步更新 UI：流程编排表
                self._refresh_actions()
                self._log(f"🔄 流程编排表 UI 已刷新", "gray")
                
                # 4. 同步更新数据表格中的显示
                show_delay = self.btn_toggle_delay.isChecked()
                target_col = 1
                for i in range(self.recording_idx):
                    target_col += 1
                    if show_delay: target_col += 1
                
                self.data_table.blockSignals(True)
                updated_count = 0
                for r in range(self.data_table.rowCount()):
                    item = self.data_table.item(r, target_col)
                    if item:
                        item.setText(f"{x}, {y}")
                        updated_count += 1
                self.data_table.blockSignals(False)
                self._log(f"🔄 数据表 {updated_count} 行坐标显示已更新", "gray")
            except Exception as e:
                self._log(f"❌ 写入坐标过程出错: {str(e)}", "red")
                import traceback
                print(traceback.format_exc())
        else:
            self._log(f"⚠️ 录制失败：步骤索引 {self.recording_idx} 越界 (总数 {len(acts)})", "orange")
    def _import_csv(self):
        p, _ = QFileDialog.getOpenFileName(self, "导入 CSV", "", "CSV Files (*.csv)")
        if p:
            with open(p, 'r', encoding='utf-8-sig') as f: self.config['task_data'][self.current_task] = list(csv.DictReader(f)); save_config(self.config); self._sync_data_headers()

    def _import_xlsx(self):
        if not HAS_OPENPYXL:
            QMessageBox.warning(self, "缺少依赖", "请先安装 openpyxl：\npip install openpyxl"); return
        p, _ = QFileDialog.getOpenFileName(self, "导入 Excel", "", "Excel Files (*.xlsx *.xls)")
        if not p: return
        try:
            wb = openpyxl.load_workbook(p, data_only=True)
            # Let user pick sheet if multiple
            sheet_name = wb.sheetnames[0]
            if len(wb.sheetnames) > 1:
                sheet_name, ok = QInputDialog.getItem(self, "选择工作表", "Sheet:", wb.sheetnames, 0, False)
                if not ok: return
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows: QMessageBox.warning(self, "空表", "所选工作表为空"); return
            headers = [str(c) if c is not None else "" for c in rows[0]]
            data = []
            for row in rows[1:]:
                row_dict = {headers[i]: (str(v) if v is not None else "") for i, v in enumerate(row) if i < len(headers)}
                data.append(row_dict)
            self.config['task_data'][self.current_task] = data
            save_config(self.config); self._sync_data_headers()
            self._log(f"✅ 已从 {os.path.basename(p)} 导入 {len(data)} 行数据", "green")
        except Exception as e:
            QMessageBox.critical(self, "导入失败", str(e))

    def _apply_row_status_style(self, item, status):
        if not item:
            return
        item.setText(status)
        item.setTextAlignment(Qt.AlignCenter)
        if status == ROW_STATUS_OK:
            item.setBackground(QColor("#c8e6c9"))
        elif status == ROW_STATUS_FAIL:
            item.setBackground(QColor("#ffcdd2"))
        elif status == ROW_STATUS_SKIP:
            item.setBackground(QColor("#fff9c4"))
        elif status == ROW_STATUS_DEFER:
            item.setBackground(QColor("#dbeafe"))
        elif status == ROW_STATUS_MANUAL:
            item.setBackground(QColor("#ffe0b2"))
        else:
            item.setBackground(QColor("#ffffff"))

    def _on_row_status(self, row_idx, status):
        """Called from engine thread via signal — update status cell and cache."""
        if self.current_task not in self._row_statuses:
            self._row_statuses[self.current_task] = {}
        self._row_statuses[self.current_task][row_idx] = status
        if row_idx < self.data_table.rowCount():
            item = self.data_table.item(row_idx, self._data_status_col())
            if not item:
                item = QTableWidgetItem(); item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.data_table.setItem(row_idx, self._data_status_col(), item)
            self._apply_row_status_style(item, status)

    def _on_row_result(self, payload):
        """汇总每一行执行结果，便于执行结束后快速定位失败项/失败窗口。"""
        try:
            if not isinstance(payload, dict):
                return
            self._last_run_row_results.append(payload)
            self._persist_row_window_context(payload)
            if payload.get("status") != ROW_STATUS_FAIL:
                return

            ctx = payload.get("fail_ctx") or payload.get("row_ctx") or {}
            fg = ctx.get("foreground") or {}
            task_id = payload.get("task_id") or self.current_task
            hwnd = ctx.get("target_hwnd") or fg.get("hwnd") or None

            win_title = fg.get("title") or ""
            win_class = fg.get("class") or ""
            if sys.platform == "win32" and hwnd and not win_class:
                try:
                    win_class = get_window_class_name(int(hwnd)) or ""
                except Exception:
                    win_class = ""
            is_browser = (str(win_class) == "Chrome_WidgetWin_1")

            self._last_run_failures.append({
                "task_id": task_id,
                "task_display": self._get_task_display_text(task_id, with_folder=True) if hasattr(self, "_get_task_display_text") else str(task_id),
                "row_index": int(payload.get("row_index", 0)),
                "step_index": int(ctx.get("step_index", payload.get("step_index", 0) or 0)),
                "step_name": str(ctx.get("step_name", "") or ""),
                "error": str(ctx.get("error", payload.get("last_error", "")) or ""),
                "hwnd": int(hwnd) if hwnd else None,
                "window_title": str(win_title or ""),
                "window_class": str(win_class or ""),
                "is_browser": bool(is_browser),
            })
        except Exception:
            pass

    def _on_highlight(self, row_idx, step_idx):
        """Highlight the currently-executing data row; clear when row_idx==-1."""
        for r in range(self.data_table.rowCount()):
            for c in range(self._data_first_value_col(), self.data_table.columnCount()):
                item = self.data_table.item(r, c)
                if item:
                    acts = self.config['tasks'].get(self.current_task, [])
                    show_delay = self.btn_toggle_delay.isChecked()
                    # Recompute base color
                    col_data_idx = c - self._data_first_value_col()  # 0-based, skipping固定控制列
                    a_idx = col_data_idx // (2 if show_delay else 1)
                    if 0 <= a_idx < len(acts):
                        act_type = CMD_MAP.get(acts[a_idx].get('action'), "click")
                        is_coord = act_type in ["click", "double_click", "right_click", "move", "hover_click", "scroll"]
                        base_color = "#eeeeee" if is_coord else "#e3f2fd"
                    else:
                        base_color = "#e3f2fd"
                    if r == row_idx:
                        item.setBackground(QColor("#fff176"))   # bright yellow highlight
                    else:
                        # Restore cached status color or base
                        statuses = self._row_statuses.get(self.current_task, {})
                        st = statuses.get(r, "")
                        if st == ROW_STATUS_OK:   item.setBackground(QColor("#c8e6c9"))
                        elif st == ROW_STATUS_FAIL: item.setBackground(QColor("#ffcdd2"))
                        elif st == ROW_STATUS_SKIP: item.setBackground(QColor("#fff9c4"))
                        elif st == ROW_STATUS_DEFER: item.setBackground(QColor("#dbeafe"))
                        elif st == ROW_STATUS_MANUAL: item.setBackground(QColor("#ffe0b2"))
                        else: item.setBackground(QColor(base_color))
        if row_idx >= 0:
            self.data_table.scrollToItem(self.data_table.item(row_idx, self._data_first_value_col()) or self.data_table.item(row_idx, self._data_status_col()))

    def _move_row_action(self, delta):
        rows = self._get_selected_action_rows(fallback_row=self.action_table.currentRow())
        self._move_selected_actions_by_delta(delta, rows=rows)

    def _move_row_action_by_idx(self, row, delta):
        rows = self._get_selected_action_rows(fallback_row=row)
        if row in rows and len(rows) > 1:
            self._move_selected_actions_by_delta(delta, rows=rows)
        else:
            self._move_selected_actions_by_delta(delta, rows=[row])

    def _get_selected_action_rows(self, fallback_row=None):
        if not hasattr(self, "action_table"):
            return []
        selection_model = self.action_table.selectionModel()
        rows = sorted({idx.row() for idx in selection_model.selectedRows()}) if selection_model else []
        if not rows and fallback_row is not None and fallback_row >= 0:
            rows = [fallback_row]
        return rows

    def _select_action_rows(self, rows):
        if not hasattr(self, "action_table"):
            return
        rows = sorted({int(r) for r in rows if int(r) >= 0})
        self.action_table.clearSelection()
        selection_model = self.action_table.selectionModel()
        if not rows or not selection_model:
            return
        flags = QItemSelectionModel.Select | QItemSelectionModel.Rows
        for row in rows:
            index = self.action_table.model().index(row, 0)
            if index.isValid():
                selection_model.select(index, flags)
        self.action_table.setCurrentCell(rows[-1], 0, QItemSelectionModel.NoUpdate)

    def _ensure_action_row_selected(self, row):
        rows = self._get_selected_action_rows()
        if row < 0:
            return rows
        if row not in rows:
            self._select_action_rows([row])
            return [row]
        return rows or [row]

    def _reorder_action_rows(self, rows, insert_at, log_text=None):
        if not self.current_task:
            return []
        acts = self.config['tasks'].get(self.current_task, [])
        normalized_rows = sorted({int(r) for r in rows if 0 <= int(r) < len(acts)})
        if not normalized_rows:
            return []

        row_set = set(normalized_rows)
        moved_items = [acts[r] for r in normalized_rows]
        remaining = [act for idx, act in enumerate(acts) if idx not in row_set]

        insert_at = max(0, min(int(insert_at), len(acts)))
        adjusted_insert = insert_at - sum(1 for r in normalized_rows if r < insert_at)
        adjusted_insert = max(0, min(adjusted_insert, len(remaining)))

        new_acts = remaining[:adjusted_insert] + moved_items + remaining[adjusted_insert:]
        if len(new_acts) == len(acts) and all(old is new for old, new in zip(acts, new_acts)):
            return []

        self.config['tasks'][self.current_task] = new_acts
        save_config(self.config)
        self._refresh_actions()
        self._refresh_defer_target_options(persist_changes=True)
        self._refresh_data_table()

        new_rows = list(range(adjusted_insert, adjusted_insert + len(moved_items)))
        self._select_action_rows(new_rows)
        if log_text:
            self._log(log_text, "blue")
        return new_rows

    def _move_selected_actions_by_delta(self, delta, rows=None):
        if not self.current_task or delta not in (-1, 1):
            return
        acts = self.config['tasks'].get(self.current_task, [])
        if not acts:
            return

        if rows is None:
            rows = self._get_selected_action_rows(fallback_row=self.action_table.currentRow())
        normalized_rows = sorted({int(r) for r in rows if 0 <= int(r) < len(acts)})
        if not normalized_rows:
            return

        row_set = set(normalized_rows)
        changed = False
        if delta < 0:
            for row in normalized_rows:
                if row > 0 and (row - 1) not in row_set:
                    acts[row - 1], acts[row] = acts[row], acts[row - 1]
                    row_set.remove(row)
                    row_set.add(row - 1)
                    changed = True
        else:
            for row in reversed(normalized_rows):
                if row < len(acts) - 1 and (row + 1) not in row_set:
                    acts[row + 1], acts[row] = acts[row], acts[row + 1]
                    row_set.remove(row)
                    row_set.add(row + 1)
                    changed = True

        if not changed:
            return

        self.config['tasks'][self.current_task] = acts
        save_config(self.config)
        self._refresh_actions()
        self._refresh_defer_target_options(persist_changes=True)
        self._refresh_data_table()

        new_rows = sorted(row_set)
        self._select_action_rows(new_rows)
        if len(normalized_rows) > 1:
            direction_text = "上移" if delta < 0 else "下移"
            self._log(f"↕️ 已批量{direction_text} {len(normalized_rows)} 个步骤", "blue")
        else:
            old_row = normalized_rows[0]
            new_row = new_rows[0]
            self._log(f"↕️ 步骤已移动: 第{old_row+1}行 → 第{new_row+1}行", "blue")

    def _copy_selected_actions(self, rows=None, copies_each=1):
        if not self.current_task:
            return
        acts = self.config['tasks'].get(self.current_task, [])
        if not acts:
            return

        if isinstance(rows, bool):
            rows = None
        if rows is None:
            rows = self._get_selected_action_rows(fallback_row=self.action_table.currentRow())
        normalized_rows = sorted({int(r) for r in rows if 0 <= int(r) < len(acts)})
        if not normalized_rows:
            return

        copies_each = max(1, int(copies_each))
        import copy
        insert_at = normalized_rows[-1] + 1
        created_items = []
        for copy_idx in range(copies_each):
            for row in normalized_rows:
                base_act = acts[row]
                new_act = copy.deepcopy(base_act)
                base_name = base_act.get('name', '步骤')
                suffix = "_副本" if copies_each == 1 else f"_复{copy_idx + 1}"
                new_act['name'] = self._make_unique_action_name(f"{base_name}{suffix}")
                created_items.append(new_act)

        for offset, new_act in enumerate(created_items):
            acts.insert(insert_at + offset, new_act)

        save_config(self.config)
        self._refresh_actions()
        self._refresh_defer_target_options(persist_changes=True)
        self._refresh_data_table()

        new_rows = list(range(insert_at, insert_at + len(created_items)))
        self._select_action_rows(new_rows)
        if len(normalized_rows) > 1:
            self._log(f"📋 已复制 {len(normalized_rows)} 个选中步骤，共新增 {len(created_items)} 个步骤", "blue")
        else:
            self._log(f"📋 已复制步骤，共新增 {len(created_items)} 个步骤", "blue")

    def _move_selected_actions_to(self, rows=None):
        if not self.current_task:
            return
        acts = self.config['tasks'].get(self.current_task, [])
        if not acts:
            return

        if isinstance(rows, bool):
            rows = None
        if rows is None:
            rows = self._get_selected_action_rows(fallback_row=self.action_table.currentRow())
        normalized_rows = sorted({int(r) for r in rows if 0 <= int(r) < len(acts)})
        if not normalized_rows:
            return

        default_pos = min(len(acts) + 1, normalized_rows[-1] + 2)
        target_pos, ok = QInputDialog.getInt(
            self,
            "批量移动步骤",
            f"已选中 {len(normalized_rows)} 个步骤。\n请输入要移动到第几步前面（输入 {len(acts)+1} 表示挪到末尾）:",
            default_pos,
            1,
            len(acts) + 1,
        )
        if not ok:
            return

        target_text = "末尾" if target_pos == len(acts) + 1 else f"第 {target_pos} 步前"
        self._reorder_action_rows(normalized_rows, target_pos - 1, log_text=f"↕️ 已将 {len(normalized_rows)} 个步骤挪到{target_text}")

    def _move_row_data(self, delta):
        row = self.data_table.currentRow()
        data = self.config['task_data'].get(self.current_task, [])
        if row < 0 or row >= len(data): return
        new_row = row + delta
        if new_row < 0 or new_row >= len(data): return
        
        item = data.pop(row)
        data.insert(new_row, item)
        self.config['task_data'][self.current_task] = data
        save_config(self.config)
        self._refresh_data_table()
        self.data_table.selectRow(new_row)
        self._log(f"↕️ 数据行已移动: 第{row+1}行 → 第{new_row+1}行", "blue")
    def _export_csv(self):
        p, _ = QFileDialog.getSaveFileName(self, "导出 CSV", "", "CSV Files (*.csv)")
        if p:
            d = self.config['task_data'].get(self.current_task, [])
            if d:
                with open(p, 'w', encoding='utf-8-sig', newline='') as f: w = csv.DictWriter(f, fieldnames=d[0].keys()); w.writeheader(); w.writerows(d)
    def _show_action_header_menu(self, pos):
        row = self.action_table.verticalHeader().logicalIndexAt(pos.y())
        if row < 0: return
        selected_rows = self._ensure_action_row_selected(row)
        multi_selected = len(selected_rows) > 1
        menu = QMenu(self)
        menu.addAction("⬆️ 上移选中步骤" if multi_selected else "⬆️ 向上移动步骤").triggered.connect(lambda: self._move_row_action_by_idx(row, -1))
        menu.addAction("⬇️ 下移选中步骤" if multi_selected else "⬇️ 向下移动步骤").triggered.connect(lambda: self._move_row_action_by_idx(row, 1))
        menu.addAction("↕️ 批量挪到指定位置..." if multi_selected else "↕️ 挪到指定位置...").triggered.connect(lambda: self._move_selected_actions_to(selected_rows))
        menu.addSeparator()
        menu.addAction("📋 复制选中步骤" if multi_selected else "📋 复制此步骤").triggered.connect(lambda: self._copy_action(row))
        menu.addAction("👯 批量克隆选中步骤" if multi_selected else "👯 批量克隆此步骤").triggered.connect(lambda: self._batch_copy_action(row))
        menu.addAction("❌ 删除选中步骤" if multi_selected else "❌ 删除此步骤").triggered.connect(self._del_action)
        menu.exec_(self.action_table.verticalHeader().mapToGlobal(pos))

    def _show_data_header_menu(self, pos):
        row = self.data_table.verticalHeader().logicalIndexAt(pos.y())
        if row < 0: return
        menu = QMenu(self)
        menu.addAction("⬆️ 向上移动数据行").triggered.connect(lambda: self._move_row_data_by_idx(row, -1))
        menu.addAction("⬇️ 向下移动数据行").triggered.connect(lambda: self._move_row_data_by_idx(row, 1))
        menu.addSeparator()
        menu.addAction("🗑️ 删除此行").triggered.connect(self._del_data_row)
        menu.exec_(self.data_table.verticalHeader().mapToGlobal(pos))

    def _show_action_menu(self, pos):
        row = self.action_table.rowAt(pos.y())
        if row < 0: return
        selected_rows = self._ensure_action_row_selected(row)
        multi_selected = len(selected_rows) > 1
        acts = self.config['tasks'].get(self.current_task, [])
        step_name = acts[row].get('name', f'步骤{row+1}') if row < len(acts) else f'步骤{row+1}'
        menu = QMenu(self)
        run_step_act = menu.addAction(f"🎯 仅执行此步骤：[{step_name}]（用第1行数据）"); run_step_act.triggered.connect(lambda _checked, r=row: self._run_single_step(r))
        menu.addSeparator()
        run_from_act = menu.addAction(f"▶️ 从此步骤开始执行（仅当前任务）"); run_from_act.triggered.connect(lambda _checked, r=row: self._run_from_step(r, only_current=True))
        run_from_all_act = menu.addAction(f"▶️▶️ 从此步骤开始执行（含后续所有任务）"); run_from_all_act.triggered.connect(lambda _checked, r=row: self._run_from_step(r, only_current=False))
        menu.addSeparator()
        menu.addAction("⬆️ 上移选中步骤" if multi_selected else "⬆️ 向上移动步骤").triggered.connect(lambda _checked, r=row: self._move_row_action_by_idx(r, -1))
        menu.addAction("⬇️ 下移选中步骤" if multi_selected else "⬇️ 向下移动步骤").triggered.connect(lambda _checked, r=row: self._move_row_action_by_idx(r, 1))
        menu.addAction("↕️ 批量挪到指定位置..." if multi_selected else "↕️ 挪到指定位置...").triggered.connect(lambda _checked: self._move_selected_actions_to(selected_rows))
        menu.addSeparator()
        copy_act = menu.addAction("📋 复制选中步骤" if multi_selected else "📋 复制此步骤"); copy_act.triggered.connect(lambda _checked, r=row: self._copy_action(r))
        clone_act = menu.addAction("👯 批量克隆选中步骤" if multi_selected else "👯 批量克隆此步骤"); clone_act.triggered.connect(lambda _checked, r=row: self._batch_copy_action(r))
        insert_act = menu.addAction("➕ 在下方插入新步骤"); insert_act.triggered.connect(self._add_action)
        del_act = menu.addAction("❌ 删除选中步骤" if multi_selected else "❌ 删除此步骤"); del_act.triggered.connect(self._del_action)
        menu.exec_(self.action_table.viewport().mapToGlobal(pos))
    def _copy_action(self, row):
        rows = self._get_selected_action_rows(fallback_row=row)
        if row not in rows:
            rows = [row]
        self._copy_selected_actions(rows=rows, copies_each=1)
    def _batch_copy_action(self, row):
        if not self.current_task: return
        rows = self._get_selected_action_rows(fallback_row=row)
        if row not in rows:
            rows = [row]
        count_text = f"已选中 {len(rows)} 个步骤。\n" if len(rows) > 1 else ""
        num, ok = QInputDialog.getInt(self, "批量克隆步骤", f"{count_text}请输入要克隆的份数:", 3, 1, 100)
        if ok:
            self._copy_selected_actions(rows=rows, copies_each=num)
    def _run_from_step(self, row, only_current=False, data_row=0):
        """Start execution from the given step and data row; optionally chain subsequent tasks."""
        self._force_sync_action_widgets() # 运行前强制同步 UI
        if not self.current_task: return
        if hasattr(self, '_engine') and self._engine.isRunning():
            QMessageBox.warning(self, "执行中", "请先停止当前执行再使用此功能。")
            return
        acts = self.config['tasks'].get(self.current_task, [])
        if not acts or row >= len(acts): return
        step_name = acts[row].get('name', f'第{row+1}步')

        # 核心逻辑：如果在“流程编排”标签页触发，强制进入“测试模式”，只跑单组数据
        is_test_run = (self.tabs.currentIndex() == 0)

        if only_current:
            # 核心修复：确保不触发任务链队列
            self._task_queue = [] 
            if is_test_run:
                self._log(f"🧪 [流程测试模式] 从步骤 [{step_name}] 开始执行（仅跑当前选中的单组数据）", "blue")
            else:
                self._log(f"▶️ 从第 {data_row+1} 组数据 / 步骤 [{step_name}] 开始执行（仅当前任务）", "blue")
        else:
            all_tasks = self._get_task_names()
            cur_idx = all_tasks.index(self.current_task) if self.current_task in all_tasks else -1
            self._task_queue = all_tasks[cur_idx + 1:] if cur_idx >= 0 else []
            if self._task_queue:
                queue_text = " → ".join(self._get_task_display_text(task_id, with_folder=True) for task_id in self._task_queue)
                self._log(f"▶️ 从第 {data_row+1} 组数据 / 步骤 [{step_name}] 开始执行，完成后将依次执行: {queue_text}", "blue")
            else:
                self._log(f"▶️ 从第 {data_row+1} 组数据 / 步骤 [{step_name}] 开始执行...", "blue")

        self._row_statuses[self.current_task] = {}
        self._refresh_data_table()
        self.resume_point = (0, data_row, row)
        self._execute(0, data_row, row, is_test=is_test_run)

    def _run_single_step(self, row):
        """单步测试：只执行步骤表里第row步，使用第1行数据。"""
        self._force_sync_action_widgets() # 运行前强制同步 UI
        acts = self.config['tasks'].get(self.current_task, [])
        if not acts or row >= len(acts): return
        
        # 核心修复：确保单步测试不触发任务链队列
        self._task_queue = [] 
        # [修复] 避免沿用上一次批量执行的失败列表，导致执行完自动弹窗（被误认为“自动打开子任务”）
        self._last_run_row_results = []
        self._last_run_failures = []
        self._suppress_failure_dialog_once = True
        
        single_act = [acts[row]]
        # 修复：如果是从流程编排页（Index 0）触发，使用步骤默认参数
        # 如果是从批量数据页（Index 1）触发，使用选中的行数据
        is_test_run = (self.tabs.currentIndex() == 0)
        if is_test_run:
            dummy_data = [{}] # 此时 ignore_data=True 会让引擎取 acts[row]['value']
        else:
            # 批量数据页右键，取当前选中的行数据
            cur_data_row = self.data_table.currentRow()
            real_data = self.config['task_data'].get(self.current_task, [])
            if cur_data_row >= 0 and cur_data_row < len(real_data):
                dummy_data = [real_data[cur_data_row]]
            elif real_data:
                dummy_data = [real_data[0]]
            else:
                dummy_data = [{}]
        step_name = acts[row].get('name', f'第{row+1}步')
        
        # 流程编排页的单步测试也应该忽略数据表
        if is_test_run:
            self._log(f"🎯 单步测试: 执行步骤 [{step_name}]（优先使用步骤默认值）", "blue")
        else:
            self._log(f"🎯 单步测试: 执行步骤 [{step_name}]（使用第1行数据）", "blue")
        
        self.btn_run.setEnabled(False); self.btn_stop.setEnabled(True)
        # 确保 loops=1, 且不触发后续任务
        self._engine = AutoEngine(single_act, dummy_data, 0, 0, 0, loops=1, ignore_data=is_test_run)
        self._engine.log_sig.connect(self._log); self._engine.done_sig.connect(self._on_done); self._engine.start()

    def _run_single_step_with_row(self, step_idx, data_row):
        """单格测试：只执行数据表第data_row行 × 步骤step_idx，不继续执行后续步骤。"""
        self._force_sync_action_widgets() # 运行前强制同步 UI
        self._save_data_table()
        acts = self.config['tasks'].get(self.current_task, [])
        if not acts or step_idx >= len(acts): return
        
        # 核心修复：确保单格测试不触发任务链队列
        self._task_queue = [] 
        self._last_run_row_results = []
        self._last_run_failures = []
        self._suppress_failure_dialog_once = True
        
        all_data = self.config['task_data'].get(self.current_task, [{}])
        row_data = all_data[data_row] if data_row < len(all_data) else {}
        single_act = [acts[step_idx]]
        single_data = [row_data]
        step_name = acts[step_idx].get('name', f'第{step_idx+1}步')
        self._log(f"🎯 单格测试: 第{data_row+1}行 × 步骤 [{step_name}]", "blue")
        self.btn_run.setEnabled(False); self.btn_stop.setEnabled(True)
        # 确保 loops=1
        self._engine = AutoEngine(single_act, single_data, 0, 0, 0, loops=1)
        self._engine.log_sig.connect(self._log)
        self._engine.done_sig.connect(self._on_done)
        self._engine.start()

    def _run_single_data_row(self, data_row):
        """单独执行批量数据中的某一行，忽略该行是否被勾选。"""
        if not self.current_task:
            return
        if hasattr(self, '_engine') and self._engine.isRunning():
            QMessageBox.warning(self, "执行中", "请先停止当前执行再单独运行这一行。")
            return

        self._force_sync_action_widgets()
        self._refresh_defer_target_options(persist_changes=True)
        self._save_data_table()

        acts = self.config['tasks'].get(self.current_task, [])
        all_data = self.config.get('task_data', {}).get(self.current_task, [])
        if not acts:
            QMessageBox.information(self, "提示", "当前任务还没有步骤可执行。")
            return
        if data_row < 0 or data_row >= len(all_data):
            QMessageBox.information(self, "提示", "这一行还没有可执行的数据。")
            return

        import copy
        row_data = copy.deepcopy(all_data[data_row] if isinstance(all_data[data_row], dict) else {})
        row_data["_选中"] = True

        self._task_queue = []
        self._current_on_finished = None
        self._last_run_row_results = []
        self._last_run_failures = []
        self._suppress_failure_dialog_once = True
        self._ui_stop_reset_pending = False
        self._row_statuses.setdefault(self.current_task, {})[data_row] = ""
        self._refresh_data_table()
        self.resume_point = (0, data_row, 0)

        if self.chk_show_osd.isChecked():
            self.osd.lbl_info.setText(f"🚀 准备单独执行第 {data_row+1} 行...")
            self.osd.bar.setValue(0)
            self.osd.show()
            self.osd.raise_()

        self._log(f"▶️ 单独执行第 {data_row+1} 行任务（共 {len(acts)} 步）", "blue")
        self.btn_run.setEnabled(False); self.btn_resume.setEnabled(False); self.btn_stop.setEnabled(True); self.btn_run.setText("正在执行...")
        self.btn_pause.setEnabled(True); self.btn_pause.setText("⏸️ 暂停"); self.btn_pause.setStyleSheet("background-color: #ff9800; color: white; font-weight: bold;")

        on_error_map = {0: "stop", 1: "skip", 2: "fail_row"}
        on_error = on_error_map.get(self.error_combo.currentIndex(), "stop")
        self._engine = AutoEngine(
            acts, [row_data], self.delay_spin.value(), 0, 0, 1, 0,
            retry_count=self.retry_spin.value(), on_error=on_error,
            ignore_data=False, standardize_window=self.chk_std_win.isChecked()
        )
        self._engine._task_name = self.current_task
        self._engine._task_id = self.current_task
        self._engine.log_sig.connect(self._log)
        self._engine.prog_sig.connect(self.progress.setValue)
        self._engine.prog_sig.connect(lambda p: self._update_osd(p))
        self._engine.pause_sig.connect(self._on_pause)
        self._engine.done_sig.connect(self._on_done)
        self._engine.row_status_sig.connect(lambda _row_idx, status, target_row=data_row: self._on_row_status(target_row, status))
        self._engine.row_result_sig.connect(self._on_row_result)
        self._engine.highlight_sig.connect(lambda row_idx, step_idx, target_row=data_row: self._on_highlight(target_row if row_idx >= 0 else -1, step_idx))
        self._engine.hotkey_paused_sig.connect(self._on_hotkey_paused)
        self._engine.detail_sig.connect(self.osd.update_detail)
        self._engine.deferred_queue_sig.connect(self._on_deferred_queue_update)
        self._engine.start()
    # [v3] OSD 新增按钮处理方法 ─────────────────────────────────────────────
    def _osd_skip_step(self):
        """跳过当前步骤。"""
        if not hasattr(self, '_engine') or not self._engine.isRunning(): return
        self._engine.skip_step()
        self._log("⏭ [手动] 请求跳过当前步骤", "orange")

    def _osd_next_row(self):
        """放弃当前行剩余步骤，跳到下一行。"""
        if not hasattr(self, '_engine') or not self._engine.isRunning(): return
        self._engine.next_row()
        self._log("⏩ [手动] 请求跳到下一行", "orange")

    def _osd_retry_step(self):
        """重试当前步骤。"""
        if not hasattr(self, '_engine') or not self._engine.isRunning(): return
        self._engine.retry_step()
        self._log("🔁 [手动] 请求重试当前步骤", "blue")

    def _toggle_pause(self):
        if not hasattr(self, '_engine') or not self._engine.isRunning(): return
        if self._engine._paused:
            self._engine.resume()
            self.btn_pause.setText("⏸️ 暂停")
            self.btn_pause.setStyleSheet("background-color: #ff9800; color: white; font-weight: bold;")
            self.osd.btn_pause.setText("⏸ 暂停")
            self.osd.btn_pause.setStyleSheet("QPushButton { background-color: #ff9800; color: white; border-radius: 3px; font-size: 11px; font-weight: bold; padding: 1px 4px; }")
            self._log("▶️ 已恢复执行", "green")
        else:
            self._engine.pause()
            self.btn_pause.setText("▶️ 继续")
            self.btn_pause.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold;")
            self.osd.btn_pause.setText("▶ 继续")
            self.osd.btn_pause.setStyleSheet("QPushButton { background-color: #4CAF50; color: white; border-radius: 3px; font-size: 11px; font-weight: bold; padding: 1px 4px; }")
            self._log("⏸️ 已暂停，点击「继续」恢复执行", "orange")

    def _stop_execution(self):
        if hasattr(self, '_engine') and self._engine.isRunning():
            self._ui_stop_reset_pending = True
            self._engine.stop()
            for sig in (getattr(self._engine, "prog_sig", None), getattr(self._engine, "detail_sig", None)):
                if sig is None:
                    continue
                try:
                    sig.disconnect()
                except Exception:
                    pass
            self._log("🛑 正在停止执行...", "red")
        if hasattr(self, "_subtask_queue"):
            self._subtask_stopped = True
            self._subtask_queue = []
            self._current_subtask_entry = None
            self._sync_subtask_manager_runtime(entry=None, parent_step_idx=None, running=False)
        if hasattr(self, "_repair_queue"):
            self._repair_stopped = True
            self._repair_queue = []
        self.osd.hide() # 停止执行，隐藏悬浮窗
        self.osd.lbl_detail.setText("")
        self.osd.bar.setValue(0)
        self.osd.lbl_pct.setText("0%")
        self._clear_deferred_queue_panel()
        self._task_queue = []
        self._current_on_finished = None
        self.progress.setValue(0)
        self.btn_run.setEnabled(True); self.btn_run.setText("🚀 开始批量执行")
        self.btn_pause.setEnabled(False); self.btn_pause.setText("⏸️ 暂停"); self.btn_pause.setStyleSheet("background-color: #ff9800; color: white; font-weight: bold;")
        self.btn_stop.setEnabled(False); self.btn_resume.setEnabled(False)
        self.btn_dry_run.setEnabled(True)
    def _run_all(self, on_finished=None):
        # 保存回调
        self._current_on_finished = on_finished
        # 清理上一次“批量执行”的失败汇总（跨任务）
        self._last_run_row_results = []
        self._last_run_failures = []
        # 在日志文件写入任务开始分隔线
        _start_banner = (
            f"\n{'='*60}\n"
            f"►►► 任务开始: {self._get_task_display_text(self.current_task, with_folder=True)}  "
            f"时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            f"{'='*60}"
        )
        self._log(_start_banner, "purple")
        
        # Build task queue unless "only current task" is checked
        if on_finished:
            self._task_queue = [] # 排程模式下不使用默认的任务队列逻辑
        elif self.chk_only_current.isChecked():
            self._task_queue = []
            self._log(f"🚀 开始执行当前任务: [{self._get_task_display_text(self.current_task, with_folder=True)}]", "blue")
        else:
            all_tasks = self._get_task_names()
            cur_idx = all_tasks.index(self.current_task) if self.current_task in all_tasks else -1
            self._task_queue = all_tasks[cur_idx + 1:] if cur_idx >= 0 else []
            if self._task_queue:
                queue_text = " → ".join(self._get_task_display_text(task_id, with_folder=True) for task_id in self._task_queue)
                self._log(f"🚀 开始执行，将依次运行: [{self._get_task_display_text(self.current_task, with_folder=True)}] → {queue_text}", "purple")
            else:
                self._log(f"🚀 开始执行当前任务: [{self._get_task_display_text(self.current_task, with_folder=True)}]", "blue")
        # Clear previous run statuses
        self._row_statuses[self.current_task] = {}
        self._refresh_data_table()
        is_test_run = (self.tabs.currentIndex() == 0)
        self.resume_point = (0, 0, 0); self._execute(0, 0, 0, is_test=is_test_run)
    def _resume_execution(self):
        is_test_run = (self.tabs.currentIndex() == 0)
        self._execute(*self.resume_point, is_test=is_test_run)
    def _force_sync_action_widgets(self):
        """修复3：运行前强制从 UI 表格抓取最新内容写入 config，防止幽灵输入框问题。"""
        if not self.current_task: return
        acts = self.config['tasks'].get(self.current_task, [])
        for i in range(self.action_table.rowCount()):
            if i >= len(acts): break
            w = self.action_table.cellWidget(i, 3)
            if w is None: continue
            # 直接是 QLineEdit（非容器）
            if isinstance(w, QLineEdit) and w.isEnabled():
                new_val = w.text()
                if acts[i].get('value', '') != new_val:
                    acts[i]['value'] = new_val
                    self._log(f"🔒 [运行前同步] 步骤「{acts[i].get('name', f'步骤{i+1}')}」已从界面抓取最新值", "blue")
            # 容器 QWidget（如 CMD/运行程序/上传文件/打开网址/条件判断）
            elif isinstance(w, QWidget):
                act_type = acts[i].get('action', '')
                if act_type == "🌐 打开网址":
                    le = w.findChild(QLineEdit)
                    btn = w.findChild(QPushButton)
                    if le and btn:
                        url_text = le.text()
                        prof_id = btn.property("profile_id") or ""
                        new_val = f"{url_text}|{prof_id}"
                        if acts[i].get('value', '') != new_val:
                            acts[i]['value'] = new_val
                            self._log(f"🔒 [运行前同步] 步骤「{acts[i].get('name', f'步骤{i+1}')}」网址/账户已同步", "blue")
                elif act_type == "✨ 清空并输入(增强版)":
                    # 这里的内容区可能是 MultiLineTextEdit（而不是 QLineEdit）
                    edits = w.findChildren(QLineEdit)
                    prefix_text = edits[0].text() if len(edits) > 0 else ""
                    content_text = edits[1].text() if len(edits) > 1 else ""
                    if content_text == "":
                        try:
                            mles = w.findChildren(MultiLineTextEdit)
                            if mles and hasattr(mles[0], "text"):
                                content_text = mles[0].text()
                        except Exception:
                            pass
                    if content_text == "":
                        try:
                            tes = w.findChildren(QTextEdit)
                            if tes and hasattr(tes[0], "toPlainText"):
                                content_text = tes[0].toPlainText()
                        except Exception:
                            pass
                    new_val = f"{prefix_text}|{content_text}"
                    if acts[i].get('value', '') != new_val:
                        acts[i]['value'] = new_val
                        self._log(f"🔒 [运行前同步] 步骤「{acts[i].get('name', f'步骤{i+1}')}」前缀/内容已同步", "blue")
                elif act_type == "⏸️ 延后执行":
                    sp_seconds = w.findChild(QSpinBox, "defer_seconds")
                    cb_resume = w.findChild(QComboBox, "defer_resume_mode")
                    cb_target = w.findChild(QComboBox, "defer_target_step")
                    cb_policy = w.findChild(QComboBox, "defer_policy")
                    if sp_seconds and cb_resume and cb_policy:
                        target_text = cb_target.currentText().strip() if cb_target and cb_resume.currentText() == "指定步骤" else ""
                        new_val = self._encode_defer_config(sp_seconds.value(), cb_resume.currentText(), target_text, cb_policy.currentText())
                        if acts[i].get('value', '') != new_val:
                            acts[i]['value'] = new_val
                            self._log(f"🔒 [运行前同步] 步骤「{acts[i].get('name', f'步骤{i+1}')}」延后配置已同步", "blue")
                elif any(x in act_type for x in ["如果找图成功", "如果窗口存在"]):
                    le = w.findChild(QLineEdit)
                    cbs = w.findChildren(QComboBox)
                    if le and len(cbs) >= 2:
                        target = le.text().strip()
                        ok_jump = cbs[0].currentText() if cbs[0].currentText() != "(顺序)" else ""
                        fail_jump = cbs[1].currentText() if cbs[1].currentText() != "(顺序)" else ""
                        new_val = f"{target} | {ok_jump} | {fail_jump}"
                        if acts[i].get('value', '') != new_val:
                            acts[i]['value'] = new_val
                            self._log(f"🔒 [运行前同步] 步骤「{acts[i].get('name', f'步骤{i+1}')}」判断逻辑已同步", "blue")
                else:
                    # 通用容器（CMD/运行程序/上传文件等）
                    le = w.findChild(QLineEdit)
                    if le and le.isEnabled():
                        new_val = le.text()
                        if acts[i].get('value', '') != new_val:
                            acts[i]['value'] = new_val
                            self._log(f"🔒 [运行前同步] 步骤「{acts[i].get('name', f'步骤{i+1}')}」内容已同步", "blue")
        save_config(self.config)

    def _execute(self, l, t, s, is_test=False):
        # 核心修复：强制结束表格当前的编辑状态，确保正在输入的单元格内容被提交
        if self.data_table.is_editing():
            self.data_table.clearSelection()
            self.data_table.setCurrentCell(-1, -1)
            
        self._force_sync_action_widgets()  # 修复3：运行前强制同步 UI → config
        self._refresh_defer_target_options(persist_changes=True)  # 运行前修复旧配置里的挂起目标
        self._save_data_table() # 内部会从 UI 控件重新抓取最新值并保存到 config
        
        # 重新从配置加载数据，确保传递给引擎的是最新快照
        as_ = self.config['tasks'].get(self.current_task, [])
        ds = self.config['task_data'].get(self.current_task, [{}])
        
        # 立即显示 OSD 悬浮窗（如果勾选了）
        if self.chk_show_osd.isChecked():
            self.osd.lbl_info.setText(f"🚀 准备启动: {self._get_task_display_text(self.current_task, with_folder=True)}...")
            self.osd.bar.setValue(0)
            self.osd.show()
            # 强制提升到最顶层
            self.osd.raise_()
        
        as_ = self.config['tasks'].get(self.current_task, [])
        ds = self.config['task_data'].get(self.current_task, [{}])
        
        # 核心修复：如果是测试模式（从流程编排触发）
        loops_to_run = self.loop_spin.value()
        if is_test:
            # 测试模式下，组索引强制归零，循环次数强制为1
            t = 0 
            loops_to_run = 1
            self._log(f"💡 测试模式：单次运行当前流程（忽略批量数据）", "gray")

        if as_:
            self._clear_deferred_queue_panel()
            self._ui_stop_reset_pending = False
            self.btn_run.setEnabled(False); self.btn_resume.setEnabled(False); self.btn_stop.setEnabled(True); self.btn_run.setText("正在执行...")
            self.btn_pause.setEnabled(True); self.btn_pause.setText("⏸️ 暂停"); self.btn_pause.setStyleSheet("background-color: #ff9800; color: white; font-weight: bold;")
            on_error_map = {0: "stop", 1: "skip", 2: "fail_row"}
            on_error = on_error_map.get(self.error_combo.currentIndex(), "stop")
            self._engine = AutoEngine(as_, ds, self.delay_spin.value(), t, s, loops_to_run, l,
                                          retry_count=self.retry_spin.value(), on_error=on_error, ignore_data=is_test,
                                          standardize_window=self.chk_std_win.isChecked())
            self._engine._task_name = self.current_task  # 注入任务名称供进度文件使用
            self._engine._task_id = self.current_task
            self._engine.log_sig.connect(self._log)
            self._engine.prog_sig.connect(self.progress.setValue)
            self._engine.prog_sig.connect(lambda p: self._update_osd(p)) # 同步更新 OSD
            self._engine.pause_sig.connect(self._on_pause)
            self._engine.done_sig.connect(self._on_done)
            self._engine.row_status_sig.connect(self._on_row_status)
            self._engine.row_result_sig.connect(self._on_row_result)
            self._engine.highlight_sig.connect(self._on_highlight)
            self._engine.hotkey_paused_sig.connect(self._on_hotkey_paused)
            self._engine.detail_sig.connect(self.osd.update_detail)  # [v3] 修复: 延时倒计时同步到 OSD
            self._engine.deferred_queue_sig.connect(self._on_deferred_queue_update)
            self._engine.start()
    def _on_pause(self, l, t, s): 
        self.resume_point = (l, t, s); self.btn_run.setEnabled(True); self.btn_resume.setEnabled(True); self.btn_stop.setEnabled(False); self.btn_run.setText("🚀 重新开始"); self._log(f"🛑 暂停: 轮{l+1}, 组{t+1}, 步{s+1}", "orange")
        self.osd.lbl_info.setText(f"⏸ <b>已暂停</b> | 组 {t+1} | 步 {s+1}")

    def _update_osd(self, percent):
        if getattr(self, "_ui_stop_reset_pending", False):
            return
        if not self.chk_show_osd.isChecked():
            self.osd.hide()
            return
            
        if hasattr(self, '_engine') and self._engine.isRunning():
            e = self._engine
            cur_act = e.actions[e._cur_s] if e._cur_s < len(e.actions) else {}
            self.osd.update_progress(
                self.current_task, e._cur_l, e._cur_t + 1, len(e.data_list),
                e._cur_s + 1, len(e.actions), cur_act.get('name', 'Step'), percent
            )

    def _log(self, m, c="white"):
        time_str = datetime.now().strftime('%H:%M:%S')
        log_entry = f"[{time_str}] {m}"
        # 界面日志区：根据颜色应用 HTML 样式
        color_map = {
            "red": "#ff5252", "green": "#69f0ae", "blue": "#82b1ff",
            "orange": "#ffab40", "purple": "#ea80fc", "gray": "#aaaaaa",
            "black": "#d4d4d4", "white": "#d4d4d4"
        }
        html_color = color_map.get(c, "#d4d4d4")
        self.log_area.append(f'<span style="color:{html_color}">{log_entry}</span>')
        
        # 同步日志到 OSD 微型控制台
        if hasattr(self, 'osd') and self.osd.isVisible():
            self.osd.add_log(m, c)
            
        # 日志文件：写入纯文本
        log_dir = os.path.join(BASE_DIR, "logs")
        if not os.path.exists(log_dir): os.makedirs(log_dir)
        log_file = os.path.join(log_dir, f"log_{datetime.now().strftime('%Y%m%d')}.txt")
        try:
            with open(log_file, "a", encoding="utf-8") as f: f.write(log_entry + "\n")
        except Exception as e:
            log_internal_issue(f"写入运行日志失败: {log_file}", e)
    def _show_config_load_warning(self):
        msg = f"检测到配置文件读取失败，程序已回退为空配置以避免直接崩溃。\n\n错误信息：{self._config_load_error}"
        if self._config_load_backup_path:
            msg += f"\n\n已自动备份损坏配置到：\n{self._config_load_backup_path}"
        QMessageBox.warning(self, "配置读取失败", msg)

    def _start_schedule_timer_runtime(self):
        if not self._timer_config:
            return
        if not hasattr(self, 'schedule_timer'):
            self.schedule_timer = QTimer(self)
            self.schedule_timer.timeout.connect(self._check_schedule)
        if not self.schedule_timer.isActive():
            self.schedule_timer.start(1000)

    def _apply_loaded_schedule_timer(self):
        if hasattr(self, 'schedule_timer') and self.schedule_timer.isActive():
            self.schedule_timer.stop()
        self.btn_timer.blockSignals(True)
        self.btn_timer.setChecked(bool(self._timer_enabled and self._timer_config))
        self.btn_timer.blockSignals(False)
        if self._timer_enabled and self._timer_config:
            self._start_schedule_timer_runtime()
            self._check_schedule()
        else:
            self.btn_timer.setText("⏰ 计划时间")

    def _toggle_timer(self, checked):
        if checked:
            dialog = ScheduleDialog(self, self._timer_config)
            if dialog.exec_() == QDialog.Accepted:
                self._timer_config = dialog.get_config()
                self._timer_enabled = True
                self._last_schedule_trigger_key = None
                self._start_schedule_timer_runtime()
                self._save_current_schedule_bundle()
                self._log(f"⏰ 调度已开启: {self._timer_config['value']}", "purple")
            else:
                self.btn_timer.blockSignals(True)
                self.btn_timer.setChecked(bool(self._timer_enabled and self._timer_config))
                self.btn_timer.blockSignals(False)
        else:
            if hasattr(self, 'schedule_timer'):
                self.schedule_timer.stop()
            self._timer_enabled = False
            self.btn_timer.setText("⏰ 计划时间")
            self._save_current_schedule_bundle()
            self._log("⏰ 调度已关闭", "gray")

    def _check_schedule(self):
        if not self._timer_config:
            return
        if hasattr(self, '_engine') and self._engine.isRunning():
            return
        now_dt = datetime.now()
        mode = self._timer_config['mode']
        val = self._timer_config['value']
        
        if mode == 0 or mode == 2: # Once or Specific Date
            target_dt = datetime.strptime(val, "%Y-%m-%d %H:%M")
            if now_dt >= target_dt:
                if getattr(self, '_last_schedule_trigger_key', None) == val:
                    return
                self._last_schedule_trigger_key = val
                if hasattr(self, 'schedule_timer'):
                    self.schedule_timer.stop()
                self._timer_enabled = False
                self.btn_timer.blockSignals(True); self.btn_timer.setChecked(False); self.btn_timer.blockSignals(False); self.btn_timer.setText("⏰ 计划时间")
                self._save_current_schedule_bundle()
                self._log("⏰ 调度时间到，开始执行...", "purple"); self._run_all()
                return
            diff = target_dt - now_dt
        else: # Daily
            target_time = datetime.strptime(val, "%H:%M").time()
            target_dt = datetime.combine(now_dt.date(), target_time)
            if now_dt.time() >= target_time:
                trigger_key = f"{now_dt.date().isoformat()} {val}"
                if getattr(self, '_last_schedule_trigger_key', None) != trigger_key:
                    self._last_schedule_trigger_key = trigger_key
                    self._log("⏰ 每日调度时间到，开始执行...", "purple")
                    self._run_all()
                # Push target to tomorrow so countdown shows time until next trigger
                target_dt += timedelta(days=1)
            diff = target_dt - now_dt

        # Smart Countdown
        seconds = int(diff.total_seconds())
        days = seconds // 86400
        hours = (seconds % 86400) // 3600
        minutes = (seconds % 3600) // 60
        secs = seconds % 60
        
        if days > 0: countdown = f"{days}天 {hours}时{minutes}分 后"
        else: countdown = f"{hours:02d}:{minutes:02d}:{secs:02d} 后"
        self.btn_timer.setText(f"⏰ {countdown}")

    def _on_done(self):
        self.btn_run.setEnabled(True); self.btn_resume.setEnabled(False)
        self.btn_stop.setEnabled(False); self.btn_dry_run.setEnabled(True)
        self.btn_run.setText("🚀 开始批量执行")
        self.btn_pause.setEnabled(False); self.btn_pause.setText("⏸️ 暂停"); self.btn_pause.setStyleSheet("background-color: #ff9800; color: white; font-weight: bold;")
        self.osd.hide() # 任务完成，隐藏悬浮窗
        final_status = getattr(getattr(self, '_engine', None), '_final_status', 'done')
        last_error = getattr(getattr(self, '_engine', None), '_last_error', '')
        if final_status == "done":
            self.progress.setValue(100)
        elif final_status in ("failed", "stopped"):
            if final_status == "stopped" and getattr(self, "_ui_stop_reset_pending", False):
                self.progress.setValue(0)
            else:
                self.progress.setValue(max(0, int(getattr(getattr(self, '_engine', None), '_last_percent', self.progress.value() or 0))))
        else:
            self.progress.setValue(0)

        if final_status == "stopped":
            self._ui_stop_reset_pending = False
            self._task_queue = []
            self._current_on_finished = None
            self._log("🛑 当前任务已停止，未继续执行后续任务。", "orange")
            return

        if final_status == "failed":
            fail_msg = "❌ 当前任务执行失败，已停止后续任务。"
            if last_error:
                fail_msg += f"\n错误详情：{last_error}"
            self._log(fail_msg, "red")
            # 若开启“失败也继续”，则不弹窗打断流程，继续跑后续任务；失败项会在最后统一汇总
            if getattr(self, '_task_queue', []) and hasattr(self, "chk_continue_on_fail") and self.chk_continue_on_fail.isChecked():
                next_task = self._task_queue.pop(0)
                if next_task in self.config.get('tasks', {}):
                    self._log(
                        f"⚠️ 任务 [{self._get_task_display_text(self.current_task, with_folder=True)}] 失败，但已开启“失败也继续”，将切到下一个任务: [{self._get_task_display_text(next_task, with_folder=True)}]",
                        "orange"
                    )
                    self._activate_task_by_id(next_task)
                    self._row_statuses[next_task] = {}
                    self._refresh_data_table()
                    self.resume_point = (0, 0, 0)
                    self._execute(0, 0, 0)
                    return
            # 默认行为：停止后续任务
            self._task_queue = []
            self._current_on_finished = None
            QMessageBox.warning(self, "执行失败", fail_msg)
            return

        # 在日志文件写入任务完成分隔线
        _done_banner = (
            f"{'='*60}\n"
            f"◄◄◄ 任务完成: {self._get_task_display_text(self.current_task, with_folder=True)}  "
            f"时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            f"{'='*60}\n"
        )
        self._log(_done_banner, "green")
        
        # 优先处理排程回调
        if hasattr(self, '_current_on_finished') and self._current_on_finished:
            callback = self._current_on_finished
            self._current_on_finished = None
            callback()
            return

        # If there are more tasks in the sequential queue, run the next one
        if getattr(self, '_task_queue', []):
            next_task = self._task_queue.pop(0)
            if next_task in self.config['tasks']:
                self._log(
                    f"✅ 任务 [{self._get_task_display_text(self.current_task, with_folder=True)}] 完成，自动切换到下一个任务: [{self._get_task_display_text(next_task, with_folder=True)}]",
                    "purple"
                )
                self._activate_task_by_id(next_task)
                self._row_statuses[next_task] = {}
                self._refresh_data_table()
                self.resume_point = (0, 0, 0)
                self._execute(0, 0, 0)
                return
        self._task_queue = []
        
        suppress_dialog = bool(getattr(self, "_suppress_failure_dialog_once", False))
        self._suppress_failure_dialog_once = False

        # 全部任务结束后：失败项仅记录日志，不再自动弹出失败管理器，避免打断批量执行收尾
        try:
            if (not suppress_dialog) and getattr(self, "_last_run_failures", []):
                self._log(f"⚠️ 本次执行有 {len(self._last_run_failures)} 个失败项；如需处理，可手动打开失败管理器。", "orange")
        except Exception:
            pass

        # 自动关机逻辑
        if self.chk_auto_shutdown.isChecked():
            self._log("🏁 全部任务执行完毕，即将按设置进入自动关机倒计时", "green")
            ShutdownDialog(self).exec_()
        else:
            self._log("🏁 全部任务执行完毕", "green")

    def _jump_to_failure(self, failure_item):
        """从失败管理器跳回对应任务/行，方便人工处理或重跑。"""
        try:
            task_id = failure_item.get("task_id", "") if isinstance(failure_item, dict) else ""
            row_idx = int(failure_item.get("row_index", 0)) if isinstance(failure_item, dict) else 0
            if task_id:
                self._activate_task_by_id(task_id)
            # 切到“批量数据”页（默认 index=1）
            try:
                self.tabs.setCurrentIndex(1)
            except Exception:
                pass
            if hasattr(self, "data_table") and 0 <= row_idx < self.data_table.rowCount():
                self.data_table.selectRow(row_idx)
                self.data_table.setCurrentCell(row_idx, self._data_status_col())
                it = self.data_table.item(row_idx, self._data_first_value_col()) or self.data_table.item(row_idx, self._data_status_col())
                if it:
                    self.data_table.scrollToItem(it)
        except Exception:
            pass

    def _start_repair_for_failures(self, failures, repair_task_id):
        """批量对失败项执行“修复子任务”，修复步骤与原流程独立。"""
        try:
            if hasattr(self, '_engine') and self._engine.isRunning():
                QMessageBox.warning(self, "执行中", "请先等待当前执行结束或停止后再启动修复子任务。")
                return
            if not repair_task_id or repair_task_id not in self.config.get("tasks", {}):
                QMessageBox.warning(self, "提示", "修复子任务不存在或未选择。")
                return
            self._repair_queue = list(failures or [])
            self._repair_stopped = False
            self._repair_task_id = repair_task_id
            self._repair_origin_task = self.current_task
            self._log(f"🛠️ 准备批量修复：共 {len(self._repair_queue)} 个失败项 | 修复任务={self._get_task_display_text(repair_task_id, with_folder=True)}", "purple")
            self._run_next_repair_item()
        except Exception as e:
            QMessageBox.warning(self, "修复启动失败", str(e))

    def _create_repair_task_from_failure(self, failure_item):
        """基于某个失败项快速创建一个新的“修复任务”，供你现场录制并保存。"""
        try:
            if not isinstance(failure_item, dict):
                return

            # 1) 尽量把失败窗口先激活，便于你马上录制修复动作
            hwnd = failure_item.get("hwnd")
            if hwnd:
                try:
                    force_activate_window(int(hwnd))
                except Exception:
                    pass

            # 2) 生成任务名称（尽量可读且不冲突）
            base_name = "修复"
            step_name = str(failure_item.get("step_name", "") or "").strip()
            if step_name:
                step_name = step_name.replace("/", "_").replace("\\", "_")
                base_name = f"修复_{step_name}"
            ts = datetime.now().strftime("%m%d_%H%M%S")
            name = f"{base_name}_{ts}"

            # 3) 创建任务（建议放到“修复”文件夹）
            self._sync_tree_structure_to_config()
            task_id = _new_task_id(set(self.config.get("tasks", {}).keys()))
            self.config.setdefault('tasks', {})
            self.config.setdefault('task_data', {})
            self.config['tasks'][task_id] = []
            self.config['task_data'][task_id] = []

            repair_folder = "修复"
            self.config.setdefault('folders', [])
            if repair_folder not in self.config['folders']:
                self.config['folders'].append(repair_folder)
            self._set_task_location(task_id, folder=repair_folder, name=name)

            # 4) 第 1 步：激活失败窗口（如果捕获到 hwnd）
            if hwnd:
                self.config['tasks'][task_id].append({
                    "name": "[修复] 激活失败窗口",
                    "action": "激活窗口",
                    "value": f"::hwnd={int(hwnd)}",
                    "x": 0, "y": 0,
                    "delay": 0,
                    "guard_enabled": False
                })

            save_config(self.config)
            self._reload_task_combo_after_config_change(task_id)
            self._refresh_schedule_task_options()

            # 5) 自动切换到新任务，并切到“流程编排”页方便录制
            try:
                self._activate_task_by_id(task_id)
            except Exception:
                pass
            try:
                self.tabs.setCurrentIndex(0)
            except Exception:
                pass

            self._log(f"✅ 已创建修复任务: [{self._get_task_display_text(task_id, with_folder=True)}]，现在你可以继续录制修复步骤并保存。", "green")
            QMessageBox.information(self, "已创建修复任务",
                "已为你创建一个新的修复任务，并自动写入「激活失败窗口(hwnd)」作为第1步。\n\n"
                "接下来你只需要像平时一样新增步骤、录制坐标/输入，即可把修复动作存起来。")
        except Exception as e:
            QMessageBox.warning(self, "创建修复任务失败", str(e))

    def _run_next_repair_item(self):
        """顺序执行修复队列，每个失败项跑一次修复流程。"""
        try:
            if not getattr(self, "_repair_queue", None):
                # 修复完成：回到修复前所在主任务
                origin = getattr(self, "_repair_origin_task", "")
                if origin:
                    self._activate_task_by_id(origin)
                # 恢复主按钮状态
                self.btn_run.setEnabled(True); self.btn_run.setText("🚀 开始批量执行")
                self.btn_stop.setEnabled(False); self.btn_resume.setEnabled(False)
                self.btn_dry_run.setEnabled(True)
                self.btn_pause.setEnabled(False); self.btn_pause.setText("⏸️ 暂停"); self.btn_pause.setStyleSheet("background-color: #ff9800; color: white; font-weight: bold;")
                if getattr(self, "_repair_stopped", False):
                    self._log("🛑 修复子任务已停止，剩余队列已取消。", "orange")
                else:
                    self._log("✅ 修复子任务已全部执行完毕。你可以在失败管理器里选择失败项进行重跑。", "green")
                return

            entry = self._repair_queue.pop(0)
            hwnd = entry.get("hwnd") if isinstance(entry, dict) else None
            if not hwnd:
                self._log("⚠️ 跳过一个失败项：未捕获到窗口句柄(hwnd)。", "orange")
                QTimer.singleShot(0, self._run_next_repair_item)
                return

            import copy
            base_actions = copy.deepcopy(self.config.get("tasks", {}).get(self._repair_task_id, []) or [])
            activate_action = {
                "name": "[修复] 激活失败窗口",
                "action": "激活窗口",
                "value": f"::hwnd={int(hwnd)}",
                "x": 0, "y": 0,
                "delay": 0,
                "guard_enabled": False
            }
            actions = [activate_action] + base_actions

            # UI 状态
            self.btn_run.setEnabled(False); self.btn_resume.setEnabled(False); self.btn_stop.setEnabled(True)
            self.btn_run.setText("🛠️ 修复中...")
            self.btn_pause.setEnabled(True); self.btn_pause.setText("⏸️ 暂停"); self.btn_pause.setStyleSheet("background-color: #ff9800; color: white; font-weight: bold;")

            self._engine = AutoEngine(
                actions, [{}], 0, 0, 0, loops=1, start_l=0,
                retry_count=self.retry_spin.value(),
                on_error="fail_row",
                ignore_data=True,
                standardize_window=self.chk_std_win.isChecked()
            )
            self._engine._task_name = f"[修复] {self._get_task_display_text(self._repair_task_id, with_folder=True)}"
            self._engine._task_id = self._repair_task_id
            self._engine.log_sig.connect(self._log)
            self._engine.prog_sig.connect(self.progress.setValue)
            self._engine.done_sig.connect(self._on_repair_done)
            self._engine.start()
        except Exception as e:
            self._log(f"❌ 修复子任务异常: {e}", "red")
            QTimer.singleShot(0, self._run_next_repair_item)

    def _on_repair_done(self):
        """单个失败项修复结束，继续下一个。"""
        try:
            final_status = getattr(getattr(self, '_engine', None), '_final_status', 'done')
            if final_status == "stopped":
                self._repair_stopped = True
                self._repair_queue = []
                QTimer.singleShot(100, self._run_next_repair_item)
                return
            if final_status == "failed":
                err = getattr(getattr(self, '_engine', None), '_last_error', '')
                self._log(f"⚠️ 修复子任务执行失败: {err}", "orange")
            # 恢复按钮文案（仍在修复队列中就继续）
            self.btn_run.setText("🛠️ 修复中...")
            QTimer.singleShot(100, self._run_next_repair_item)
        except Exception:
            QTimer.singleShot(100, self._run_next_repair_item)

    # ── Feature 8: Global Hotkeys ─────────────────────────────────────────────
    def _default_hotkeys(self):
        return {
            "pause": "F9",
            "stop": "F10",
            "capture": "F8",
            "skip_step": "F6",
            "next_row": "F7",
            "retry_step": "F5"
        }

    def _register_hotkeys(self):
        """[增强] 注册全局控制热键：F9启动/暂停，F10停止。"""
        if not HAS_KEYBOARD:
            self.lbl_hotkey.setText("⚠️ 未安装 keyboard 库，执行控制全局热键不可用")
            return
        
        try:
            # 清理旧热键
            keyboard.unhook_all()
            self._hotkey_hooks.clear()
            
            hk = self.config.get("hotkeys", self._default_hotkeys())
            pause_key = hk.get("pause", "F9")
            stop_key  = hk.get("stop",  "F10")
            cap_key   = hk.get("capture", "F8")
            skip_key  = hk.get("skip_step", "F6")
            next_key  = hk.get("next_row", "F7")
            retry_key = hk.get("retry_step", "F5")

            # 注册新热键
            h1 = keyboard.add_hotkey(pause_key, self._hotkey_pause_toggle)
            h2 = keyboard.add_hotkey(stop_key,  self._hotkey_stop)
            h3 = keyboard.add_hotkey(cap_key,   self._hotkey_capture)
            h4 = keyboard.add_hotkey(skip_key, self._hotkey_skip_step)
            h5 = keyboard.add_hotkey(next_key, self._hotkey_next_row)
            h6 = keyboard.add_hotkey(retry_key, self._hotkey_retry_step)
            self._hotkey_hooks = [h1, h2, h3, h4, h5, h6]
            
            self.lbl_hotkey.setText(
                f"⌨️ 全局热键: [{pause_key}] 暂停/继续  [{skip_key}] 跳步  "
                f"[{next_key}] 下一行  [{retry_key}] 重试  [{stop_key}] 停止"
            )
            if hasattr(self, 'osd'):
                self.osd.btn_skip_step.setToolTip(f"跳过当前步骤，直接执行下一步\n快捷键: {skip_key}")
                self.osd.btn_next_row.setToolTip(f"放弃当前行剩余步骤，跳到下一行开始执行\n快捷键: {next_key}")
                self.osd.btn_retry.setToolTip(f"重新执行当前步骤（不跳过，重来一次）\n快捷键: {retry_key}")
                self.osd.btn_stop.setToolTip(f"终止整个任务\n快捷键: {stop_key}")
        except Exception as e:
            self.lbl_hotkey.setText(f"⚠️ 热键注册失败: {e}")

    def _hotkey_pause_toggle(self):
        """Called from keyboard listener thread — use QTimer to touch UI safely."""
        QTimer.singleShot(0, self._do_hotkey_pause_toggle)

    def _do_hotkey_pause_toggle(self):
        if not hasattr(self, '_engine') or not self._engine.isRunning(): return
        if self._engine._paused:
            self._engine.resume()
            self._log("▶️ [热键] 已恢复执行", "green")
            self.btn_stop.setEnabled(True)
            self.btn_pause.setText("⏸️ 暂停"); self.btn_pause.setStyleSheet("background-color: #ff9800; color: white; font-weight: bold;")
            self.osd.btn_pause.setText("⏸ 暂停")
            self.osd.btn_pause.setStyleSheet("QPushButton { background-color: #ff9800; color: white; border-radius: 3px; font-size: 11px; font-weight: bold; padding: 1px 4px; }")
        else:
            self._engine.pause()
            self._log("⏸️ [热键] 已暂停，再按一次继续", "orange")
            self.btn_resume.setEnabled(True)
            self.btn_pause.setText("▶️ 继续"); self.btn_pause.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold;")
            self.osd.btn_pause.setText("▶ 继续")
            self.osd.btn_pause.setStyleSheet("QPushButton { background-color: #4CAF50; color: white; border-radius: 3px; font-size: 11px; font-weight: bold; padding: 1px 4px; }")

    def _hotkey_stop(self):
        QTimer.singleShot(0, self._stop_execution)

    def _hotkey_capture(self):
        QTimer.singleShot(0, self._start_region_capture)

    def _hotkey_skip_step(self):
        QTimer.singleShot(0, self._osd_skip_step)

    def _hotkey_next_row(self):
        QTimer.singleShot(0, self._osd_next_row)

    def _hotkey_retry_step(self):
        QTimer.singleShot(0, self._osd_retry_step)

    def _on_hotkey_paused(self, l, t, s):
        self.resume_point = (l, t, s)
        self.btn_resume.setEnabled(True)
        self.btn_run.setText("⏸️ 已暂停")

    def _show_hotkey_settings(self):
        hk = self.config.get("hotkeys", self._default_hotkeys())
        dlg = QDialog(self); dlg.setWindowTitle("⌨️ 全局热键设置"); dlg.resize(420, 380)
        ly = QVBoxLayout(dlg)
        ly.addWidget(QLabel("修改后点击保存，重启程序或手动触发注册生效："))
        form_rows = [
            ("暂停/继续 热键:", "pause"),
            ("停止 热键:", "stop"),
            ("框选截图 热键:", "capture"),
            ("跳过当前步骤 热键:", "skip_step"),
            ("切到下一行 热键:", "next_row"),
            ("重试当前步骤 热键:", "retry_step")
        ]
        edits = {}
        for label, key in form_rows:
            row = QHBoxLayout(); row.addWidget(QLabel(label))
            ed = QLineEdit(hk.get(key, "")); row.addWidget(ed); edits[key] = ed; ly.addLayout(row)
        ly.addWidget(QLabel("示例: F5  F6  F7  F8  F9  F10  ctrl+shift+p  alt+s"))
        btns = QHBoxLayout()
        btn_ok = QPushButton("💾 保存并应用"); btn_cancel = QPushButton("取消")
        btn_ok.clicked.connect(dlg.accept); btn_cancel.clicked.connect(dlg.reject)
        btns.addWidget(btn_ok); btns.addWidget(btn_cancel); ly.addLayout(btns)
        if dlg.exec_() == QDialog.Accepted:
            new_hk = {k: edits[k].text().strip() for k in edits}
            self.config["hotkeys"] = new_hk; save_config(self.config)
            self._register_hotkeys()
            self._log(f"⌨️ 热键已更新: {new_hk}", "blue")

    # ── Feature 9: Region Capture ─────────────────────────────────────────────
    def _start_region_capture(self, target_step_idx=None):
        cap_dir = os.path.join(BASE_DIR, "captures")
        self._capture_target_idx = target_step_idx  # None = standalone, int = fill step value
        self._overlay = ScreenshotOverlay(cap_dir, callback=None)
        self._overlay.captured.connect(self._on_region_captured)
        self._overlay.show()

    def _on_region_captured(self, fpath):
        self._log(f"📸 框选截图已保存: {fpath}", "green")
        # If triggered from a step's capture button, fill the value
        idx = getattr(self, '_capture_target_idx', None)
        if idx is not None and self.current_task:
            acts = self.config['tasks'].get(self.current_task, [])
            if 0 <= idx < len(acts):
                acts[idx]['value'] = fpath
                save_config(self.config)
                self._refresh_actions()
                self._log(f"✅ 已将截图路径填入步骤 [{acts[idx].get('name','')}]", "green")
        self._capture_target_idx = None
        # Ask if user wants to open containing folder
        reply = QMessageBox.question(self, "截图完成",
            f"截图已保存到:\n{fpath}\n\n是否打开所在文件夹?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            folder = os.path.dirname(fpath)
            if sys.platform == 'win32': os.startfile(folder)
            else: os.system(f'xdg-open "{folder}"')

    # ── Feature 10: Dry Run ───────────────────────────────────────────────────
    def _run_dry(self):
        self._row_statuses[self.current_task] = {}
        self._refresh_data_table()
        self._log("🧪 开始试运行（只打印步骤，不操作鼠标/键盘）...", "purple")
        self._save_data_table()
        as_ = self.config['tasks'].get(self.current_task, [])
        ds  = self.config['task_data'].get(self.current_task, [{}])
        if not as_: self._log("⚠️ 没有步骤可执行", "orange"); return
        self.btn_run.setEnabled(False); self.btn_dry_run.setEnabled(False)
        self.btn_stop.setEnabled(True)
        on_error_map = {0: "stop", 1: "skip", 2: "fail_row"}
        on_error = on_error_map.get(self.error_combo.currentIndex(), "stop")
        self._engine = AutoEngine(as_, ds, 0, 0, 0, self.loop_spin.value(), 0,
                                      retry_count=0, on_error=on_error, dry_run=True)
        self._engine._task_name = self.current_task  # 注入任务名称供进度文件使用
        self._engine._task_id = self.current_task
        self._engine.log_sig.connect(self._log)
        self._engine.prog_sig.connect(self.progress.setValue)
        self._engine.done_sig.connect(self._on_done)
        self._engine.row_status_sig.connect(self._on_row_status)
        self._engine.row_result_sig.connect(self._on_row_result)
        self._engine.highlight_sig.connect(self._on_highlight)
        self._engine.start()

    def _manual_save_data(self):
        """强制同步并保存当前表格中的所有手动改动。"""
        if not self.current_task: return
        
        # 1. 强制结束表格编辑状态，确保正在输入的文字被提交
        if self.data_table.is_editing():
            self.data_table.setCurrentCell(self.data_table.currentRow(), self.data_table.currentColumn())
            self.data_table.clearFocus()
            # 再次确认编辑状态已结束
            self.data_table.setFocus()
            self.data_table.clearFocus()

        # 2. 调用现有的保存逻辑
        # 注意：_save_data_table 内部已经实现了从 UI 控件抓取最新值的逻辑
        self._save_data_table(flush=True)
        
        # 3. 界面反馈
        self._log("✅ 数据已强制同步并保存到本地配置", "#4CAF50")
        QMessageBox.information(self, "保存成功", "当前批量数据已成功强制同步并保存。")

    def _on_skip_checkbox_changed(self, row, key, state):
        # 状态改变时立即保存，防止丢失
        self._save_data_table()

    def closeEvent(self, event):
        if HAS_KEYBOARD:
            for h in self._hotkey_hooks:
                try:
                    keyboard.remove_hotkey(h)
                except Exception as e:
                    log_internal_issue(f"移除热键失败: {h}", e)
        hdr = self.action_table.horizontalHeader()
        if "layout" not in self.config: self.config["layout"] = {}
        self.config["layout"].update({
            "size": [self.size().width(), self.size().height()],
            "pos": [self.pos().x(), self.pos().y()],
            "splitter_sizes": self.main_splitter.sizes(),
            "v_splitter_sizes": self.right_splitter.sizes(),
            "action_col_widths": [hdr.sectionSize(i) for i in range(hdr.count())],
            "last_task": self.current_task # [新增] 退出时确保记录当前任务
        })
        self._flush_config_now()
        super().closeEvent(event)

if __name__ == "__main__":
    # [修复] 统一 DPI 策略：必须在 QApplication 实例化之前完成所有 DPI 相关设置
    # 步骤1：通知 Windows 该进程已自行处理 DPI（Per-Monitor V2，Win10 1703+）
    if sys.platform == 'win32':
        try:
            import ctypes
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("SimuOps.AutoManager")
            ctypes.windll.user32.SetProcessDpiAwarenessContext(-4)
        except Exception as e:
            log_internal_issue("设置 Per-Monitor V2 DPI 失败，回退到旧接口", e)
            try:
                ctypes.windll.shcore.SetProcessDpiAwareness(2)  # PROCESS_PER_MONITOR_DPI_AWARE
            except Exception as e2:
                log_internal_issue("设置旧版 DPI 感知接口失败", e2)

    # 步骤2：启用 Qt 高 DPI 支持（必须在 QApplication 实例化前调用）
    if hasattr(Qt, 'AA_EnableHighDpiScaling'):
        QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    if hasattr(Qt, 'AA_UseHighDpiPixmaps'):
        QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)

    # 步骤3：[关键修复] 使用 RoundPreferFloor 策略
    # Round 在 125%/150% 等非整数缩放下会产生半像素偏移，导致文字重影/模糊
    # RoundPreferFloor 向下取整，确保像素级对齐，消除重影
    if hasattr(Qt, 'HighDpiScaleFactorRoundingPolicy'):
        QApplication.setHighDpiScaleFactorRoundingPolicy(
            Qt.HighDpiScaleFactorRoundingPolicy.RoundPreferFloor
        )

    app = QApplication(sys.argv)
    icon_candidates = [
        os.path.join(BASE_DIR, "app_icon.ico"),
        os.path.join(BASE_DIR, "app.ico"),
        os.path.join(BASE_DIR, "app_icon.png"),
    ]
    app_icon = None
    for icon_path in icon_candidates:
        if os.path.exists(icon_path):
            app_icon = QIcon(icon_path)
            app.setWindowIcon(app_icon)
            break

    # 步骤4：设置全局字体，使用清晰的渲染策略
    # PreferFullHinting：让字体渲染器使用完整 hinting，像素级对齐，边缘锐利
    # 去掉 PreferQuality（它会强制抗锯齿，与 FullHinting 冲突，叠加后反而晕染）
    from PyQt5.QtGui import QFont
    main_font = QFont("Microsoft YaHei", 10)
    main_font.setWeight(QFont.Normal)
    main_font.setStyleStrategy(QFont.PreferAntialias | QFont.PreferFullHinting)
    QApplication.setFont(main_font)
    window = AutoManager()
    if app_icon is not None:
        window.setWindowIcon(app_icon)

    def save_layout_on_quit():
        hdr = window.action_table.horizontalHeader()
        if "layout" not in window.config: window.config["layout"] = {}
        window.config["layout"].update({
            "size": [window.size().width(), window.size().height()],
            "pos": [window.pos().x(), window.pos().y()],
            "action_col_widths": [hdr.sectionSize(i) for i in range(hdr.count())],
            "last_task": window.current_task # [新增] 退出时确保记录当前任务
        })
        save_config(window.config)

    app.aboutToQuit.connect(save_layout_on_quit)
    window.show()
    # Restore layout after show() so Qt has finished laying out the window
    layout_cfg = window.config.get("layout", {})
    # if layout_cfg.get("splitter_sizes"):
    #     window.main_splitter.setSizes(layout_cfg["splitter_sizes"])
    if layout_cfg.get("action_col_widths"):
        for i, w in enumerate(layout_cfg["action_col_widths"]):
            window.action_table.setColumnWidth(i, w)

    sys.exit(app.exec())
